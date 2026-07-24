import os
import re
import tempfile
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

from flask import (
    Blueprint, render_template, request, jsonify
)
from flask_login import login_required

from app import db
from app.models.boletim_financeiro import BoletimFinanceiro
from app.utils.audit import registrar_log
from app.models.boletim_financeiro import BoletimFinanceiro
from app.models.estrutura_boletim import EstruturaBoletim
import unicodedata

boletim_financeiro_bp = Blueprint(
    'boletim_financeiro', __name__, url_prefix='/boletim-financeiro'
)

# =========================================================================
# CONSTANTES DE PARSE
# =========================================================================
# Tons de verde usados como TOTAL/subtotal no Boletim (não capturar).
_VERDES_CONHECIDOS = {'FF339966', 'FFCCFFCC'}

# Ordem fixa dos meses -> nº do mês (JANEIRO = 1 ... DEZEMBRO = 12).
_MESES_ORDEM = [
    'JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO',
    'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO',
]

# Rótulos de total conhecidos (dupla checagem: além da cor, pelo nome).
_TOTAIS_NOME = {
    'SALDO FINAL DE CAIXA',
    'DISPONIBILIDADE TOTAL',
    'DISPONÍVEL INICIAL',
    'BANCOS - CONTAS CORRENTES',
    'APLICAÇÕES FINANCEIRAS',
    'BLOQUEIOS JUDICIAIS - FUNDO DE INVESTIMENTOS',
}

_MARCADOR_INICIO = 'SALDO FINAL DE CAIXA'
_MARCADOR_FIM = 'NATUREZAS MOVIMENTADAS'

def _chave_comparacao(nome):
    """
    Chave 'à prova de formatação': remove acentos (imune a NFC/NFD), pontos de
    recuo, espaços e pontuação, e passa para UPPER.
    Ex.: '....APLICAÇÕES REALIZADAS' -> 'APLICACOESREALIZADAS'
    """
    s = unicodedata.normalize('NFKD', str(nome or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'[^A-Za-z0-9]', '', s).upper()


# Trechos que, se presentes no nome, fazem a linha ser DESCARTADA
# (movimentação interna dos fundos, não é natureza do Boletim).
_EXCLUIR_CHAVES = (
    'APLICACOESREALIZADAS',
    'RESGATESEFETUADOS',
)


def _deve_excluir(nome):
    """True se a linha é de movimentação interna dos fundos (não capturar)."""
    chave = _chave_comparacao(nome)
    return any(k in chave for k in _EXCLUIR_CHAVES)

# =========================================================================
# HELPERS DE PARSE
# =========================================================================
def _texto_limpo(valor):
    """Normaliza espaços e faz trim. Retorna '' se None."""
    if valor is None:
        return ''
    return re.sub(r'\s+', ' ', str(valor)).strip()


def _eh_celula_verde(cell):
    """
    True se a célula tem preenchimento sólido em tom de verde.
    Checa os códigos conhecidos e, como reforço, uma heurística de
    'canal verde dominante' para pegar qualquer outro tom de verde.
    """
    fill = cell.fill
    if fill is None or fill.patternType != 'solid':
        return False

    rgb = getattr(fill.fgColor, 'rgb', None)
    if not isinstance(rgb, str):
        return False

    if rgb.upper() in _VERDES_CONHECIDOS:
        return True

    hexc = rgb[-6:]
    try:
        r = int(hexc[0:2], 16)
        g = int(hexc[2:4], 16)
        b = int(hexc[4:6], 16)
    except ValueError:
        return False
    return g > r and g > b and g >= 128


def _to_decimal_2(valor):
    """Converte para Decimal(18,2). Trata None/NaN/vazio como 0.00."""
    if valor is None or valor == '':
        return Decimal('0.00')
    try:
        import math
        if isinstance(valor, float) and math.isnan(valor):
            return Decimal('0.00')
    except Exception:
        pass
    try:
        return Decimal(str(valor)).quantize(
            Decimal('0.01'), rounding=ROUND_HALF_UP
        )
    except (InvalidOperation, ValueError):
        return Decimal('0.00')

def _chave_natureza(nome):
    """
    Normaliza o nome da natureza para comparação: remove os pontos/espaços de
    recuo do início (ex.: '....RESGATES EFETUADOS'), colapsa espaços e passa
    para UPPER. Mesma normalização usada na FIN_TB019.
    """
    base = re.sub(r'^[.\s]+', '', (nome or ''))
    return re.sub(r'\s+', ' ', base).strip().upper()

# =========================================================================
# PROCESSAMENTO DO EXCEL DO BOLETIM (ETAPA 1)
# =========================================================================
def _processar_excel_boletim(caminho_arquivo):
    """
    Lê o Excel do Boletim e grava em BDG.FIN_TB020_BOLETIM_FINANCEIRO.

    NU_LINHA: vem da FIN_TB019_ESTRUTURA_BOLETIM comparando a NATUREZA. Quando
    a mesma natureza aparece mais de uma vez no Boletim (fundos que constam em
    APLICAÇÕES FINANCEIRAS e novamente em BLOQUEIOS JUDICIAIS), as ocorrências
    consomem os NU_LINHA em ordem — assim cada bloco recebe o seu.

    Não são capturadas: linhas verdes (totais), rótulos de total conhecidos e
    as linhas de movimentação interna dos fundos ('....APLICAÇÕES REALIZADAS'
    e '....RESGATES EFETUADOS'), descartadas por _deve_excluir() — comparação
    imune a pontos de recuo, espaços, acentos e forma Unicode.

    Idempotência: apaga apenas as competências (MES_EXECUCAO) presentes no
    arquivo e reinsere.
    """
    from openpyxl import load_workbook

    wb = load_workbook(caminho_arquivo, data_only=True)
    ws = wb.active

    # 1. ANO ('ANO: 2026' -> '2026')
    ano = None
    for r in range(1, 9):
        for c in range(1, ws.max_column + 1):
            m = re.search(r'ANO[:\s]*?(\d{4})',
                          _texto_limpo(ws.cell(r, c).value), re.IGNORECASE)
            if m:
                ano = m.group(1)
                break
        if ano:
            break
    if not ano:
        raise Exception("Não foi possível localizar o ANO (ex.: 'ANO: 2026') na planilha.")

    # 2. Cabeçalho -> coluna da NATUREZA e mapa mês -> coluna
    col_natureza = None
    header_row = None
    mapa_mes = {}
    for r in range(1, 15):
        rotulos = {
            _texto_limpo(ws.cell(r, c).value).upper(): c
            for c in range(1, ws.max_column + 1)
        }
        if 'NATUREZA' in rotulos and 'JANEIRO' in rotulos:
            header_row = r
            col_natureza = rotulos['NATUREZA']
            for i, mes in enumerate(_MESES_ORDEM, start=1):
                if mes in rotulos:
                    mapa_mes[i] = rotulos[mes]
            break
    if header_row is None or not mapa_mes:
        raise Exception("Cabeçalho (NATUREZA / meses) não encontrado na planilha.")

    # 3. Marcador de início (SALDO FINAL DE CAIXA)
    marcador = None
    for r in range(header_row + 1, ws.max_row + 1):
        if _texto_limpo(ws.cell(r, col_natureza).value).upper() == _MARCADOR_INICIO:
            marcador = r
            break
    if marcador is None:
        raise Exception("Marcador ' SALDO FINAL DE CAIXA' não encontrado na planilha.")

    # 4. Captura das naturezas (guardando o GRUPO = último cabeçalho verde acima)
    capturadas = []   # (grupo, nome, linha_excel)
    excluidas = []    # nomes descartados por _deve_excluir()
    grupo_atual = _texto_limpo(ws.cell(marcador, col_natureza).value)
    for r in range(marcador + 1, ws.max_row + 1):
        celula = ws.cell(r, col_natureza)
        nome = _texto_limpo(celula.value)
        if not nome:
            continue
        if nome.upper().startswith(_MARCADOR_FIM):
            break
        if _eh_celula_verde(celula) or nome.upper() in _TOTAIS_NOME:
            grupo_atual = nome  # cabeçalho verde vira o grupo das linhas seguintes
            continue
        if _deve_excluir(nome):
            excluidas.append(nome)
            continue
        capturadas.append((grupo_atual, nome, r))
    if not capturadas:
        raise Exception("Nenhuma natureza capturada abaixo de 'SALDO FINAL DE CAIXA'.")

    # 5. Último mês com dados (não grava meses futuros zerados)
    #    >>> Para gravar SEMPRE os 12 meses: ultimo_mes = max(mapa_mes.keys())
    ultimo_mes = 0
    for mes, col in mapa_mes.items():
        for (_g, _n, r) in capturadas:
            if (ws.cell(r, col).value or 0) not in (0, None):
                ultimo_mes = max(ultimo_mes, mes)
                break
    if ultimo_mes == 0:
        ultimo_mes = max(mapa_mes.keys())
    meses_do_arquivo = [f"{ano}{m:02d}" for m in range(1, ultimo_mes + 1)]

    # 6. Mapa NATUREZA -> [NU_LINHA, ...] (FIN_TB019, em ordem de NU_LINHA)
    mapa_nu_linha = EstruturaBoletim.carregar_mapa_natureza_nu_linha()
    if not mapa_nu_linha:
        raise Exception('FIN_TB019_ESTRUTURA_BOLETIM está vazia — não há NU_LINHA para vincular.')

    # 6b. Naturezas repetidas no Excel (informativo)
    contagem = {}
    for (_g, nome, _r) in capturadas:
        k = _chave_natureza(nome)
        contagem[k] = contagem.get(k, 0) + 1
    ambiguas = sorted({nome for (_g, nome, _r) in capturadas
                       if contagem[_chave_natureza(nome)] > 1})

    # 7. Resolve NU_LINHA consumindo as ocorrências em ordem
    #    (1ª ocorrência -> 1º NU_LINHA; 2ª ocorrência -> 2º NU_LINHA; ...)
    ocorrencia = {}
    nao_encontradas = []
    resolvidas = []  # (nu_linha, nome, linha_excel)
    for (_g, nome, r) in capturadas:
        chave = _chave_natureza(nome)
        lista = mapa_nu_linha.get(chave, [])
        idx = ocorrencia.get(chave, 0)
        if idx >= len(lista):
            nao_encontradas.append(nome)
            continue
        ocorrencia[chave] = idx + 1
        resolvidas.append((lista[idx], nome, r))

    # 8. Idempotência: apaga as competências do arquivo antes de reinserir
    apagados = BoletimFinanceiro.query.filter(
        BoletimFinanceiro.MES_EXECUCAO.in_(meses_do_arquivo)
    ).delete(synchronize_session=False)
    db.session.flush()

    # 9. Monta os registros, protegendo a PK (NU_LINHA, MES_EXECUCAO)
    registros = []
    vistos = set()
    conflitos = 0
    for (nu, nome, r) in resolvidas:
        for mes in range(1, ultimo_mes + 1):
            mes_exec = f"{ano}{mes:02d}"
            if (nu, mes_exec) in vistos:
                conflitos += 1
                continue
            vistos.add((nu, mes_exec))
            registros.append(BoletimFinanceiro(
                NU_LINHA=nu,
                NATUREZA=nome,
                MES_EXECUCAO=mes_exec,
                VR_EXECUTADO=_to_decimal_2(ws.cell(r, mapa_mes[mes]).value),
            ))

    db.session.bulk_save_objects(registros)
    db.session.commit()

    return {
        'ano': ano,
        'capturadas': len(capturadas),
        'resolvidas': len(resolvidas),
        'excluidas': len(excluidas),
        'meses': ultimo_mes,
        'apagados': int(apagados or 0),
        'inseridos': len(registros),
        'nao_encontradas': nao_encontradas,
        'ambiguas': ambiguas,
        'conflitos': conflitos,
    }


# =========================================================================
# PÁGINA PRINCIPAL
# =========================================================================
@boletim_financeiro_bp.route('/')
@login_required
def index():
    """Página principal do Boletim Financeiro (upload da ETAPA 1)."""
    total_registros = BoletimFinanceiro.query.count()
    return render_template(
        'boletim_financeiro/index.html',
        total_registros=total_registros,
    )


# =========================================================================
# UPLOAD DO EXCEL DO BOLETIM (ETAPA 1)
# =========================================================================
@boletim_financeiro_bp.route('/upload', methods=['POST'])
@login_required
def upload():
    """Recebe o Excel do Boletim e grava em FIN_TB020_BOLETIM_FINANCEIRO."""
    if 'arquivo' not in request.files:
        return jsonify({'success': False, 'message': 'Nenhum arquivo enviado.'}), 400

    arquivo = request.files['arquivo']
    nome_arquivo = (arquivo.filename or '').strip()

    if not nome_arquivo:
        return jsonify({'success': False, 'message': 'Nome do arquivo está vazio.'}), 400

    if not nome_arquivo.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'success': False, 'message': 'Arquivo deve ser .xlsx ou .xlsm.'}), 400

    caminho_tmp = os.path.join(
        tempfile.gettempdir(),
        f'boletim_upload_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{nome_arquivo}'
    )
    try:
        arquivo.save(caminho_tmp)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Falha ao salvar arquivo temporário: {e}'}), 500

    try:
        resumo = _processar_excel_boletim(caminho_tmp)

        partes = [
            f'Boletim {resumo["ano"]}: {resumo["inseridos"]} registro(s) inserido(s) '
            f'({resumo["resolvidas"]} natureza(s) × {resumo["meses"]} mês(es); '
            f'{resumo["apagados"]} substituído(s)).'
        ]
        if resumo['ambiguas']:
            partes.append(
                f'⚠ {len(resumo["ambiguas"])} natureza(s) com nome repetido no Excel — '
                f'o NU_LINHA por nome não as distingue: '
                + '; '.join(resumo['ambiguas']) + '.'
            )
        if resumo['nao_encontradas']:
            partes.append(
                f'⚠ {len(resumo["nao_encontradas"])} natureza(s) sem correspondência na '
                f'FIN_TB019 (ignoradas): ' + '; '.join(resumo['nao_encontradas']) + '.'
            )
        if resumo['conflitos']:
            partes.append(
                f'⚠ {resumo["conflitos"]} registro(s) ignorado(s) por colisão de '
                f'(NU_LINHA, MES_EXECUCAO).'
            )
        mensagem = ' '.join(partes)

        registrar_log(
            acao='carga',
            entidade='boletim_financeiro',
            entidade_id=None,
            descricao=f'Upload do Boletim Financeiro ({nome_arquivo})',
            dados_novos={'arquivo': nome_arquivo, **resumo},
        )

        return jsonify({'success': True, 'message': mensagem, **resumo})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro ao processar planilha: {str(e)}'}), 500

    finally:
        try:
            if os.path.exists(caminho_tmp):
                os.remove(caminho_tmp)
        except Exception:
            pass