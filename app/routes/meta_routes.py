# app/routes/meta_routes.py
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, session, send_file
from app.models.metas_redistribuicao import MetasPercentuaisDistribuicao, Metas, MetasPeriodoAvaliativo
from app.models.meta_avaliacao import MetaAvaliacao, MetaSemestral
from app.models.edital import Edital
from app.models.periodo import PeriodoAvaliacao
from app.models.empresa_participante import EmpresaParticipante
from app.utils.meta_calculator import MetaCalculator
from datetime import datetime
from flask_login import login_required, current_user
from app import db
from app.utils.audit import registrar_log
from sqlalchemy import text
from decimal import Decimal

meta_bp = Blueprint('meta', __name__, url_prefix='/credenciamento')


@meta_bp.context_processor
def inject_current_year():
    return {'current_year': datetime.utcnow().year}


@meta_bp.route('/metas')
@login_required
def lista_metas():
    """Lista de metas usando a tabela DCA_TB019_DISTRIBUICAO_MENSAL"""
    # Parâmetros para filtro
    edital_id = request.args.get('edital_id', type=int)
    periodo_id = request.args.get('periodo_id', type=int)
    empresa_id = request.args.get('empresa_id', type=int)
    competencia = request.args.get('competencia', type=str)

    # Query usando DCA_TB019_DISTRIBUICAO_MENSAL - CORRIGIDA OS JOINS
    query = """
        SELECT 
            dm.MES_COMPETENCIA,
            dm.VR_META_MES,
            dm.ID_EMPRESA,
            dm.ID_EDITAL,
            dm.ID_PERIODO,
            ds.NO_EMPRESA_ABREVIADA,
            ed.DESCRICAO as DESCRICAO_EDITAL,
            ed.NU_EDITAL,
            ed.ANO as ANO_EDITAL,
            p.ID_PERIODO as NUM_PERIODO,
            p.DT_INICIO,
            p.DT_FIM,
            p.ID as PERIODO_PK
        FROM BDG.DCA_TB019_DISTRIBUICAO_MENSAL dm
        INNER JOIN BDG.DCA_TB017_DISTRIBUICAO_SUMARIO ds
            ON dm.ID_DISTRIBUICAO_SUMARIO = ds.ID
            AND ds.DELETED_AT IS NULL
        LEFT JOIN BDG.DCA_TB008_EDITAIS ed 
            ON ed.ID = ds.ID_EDITAL
        LEFT JOIN BDG.DCA_TB001_PERIODO_AVALIACAO p 
            ON p.ID_PERIODO = ds.ID_PERIODO
            AND p.ID_EDITAL = ds.ID_EDITAL
        WHERE dm.DELETED_AT IS NULL
    """

    params = {}

    # Aplicar filtros
    if edital_id:
        query += " AND ds.ID_EDITAL = :edital_id"
        params['edital_id'] = edital_id
    if periodo_id:
        query += " AND p.ID = :periodo_id"
        params['periodo_id'] = periodo_id
    if empresa_id:
        query += " AND ds.ID_EMPRESA = :empresa_id"
        params['empresa_id'] = empresa_id
    if competencia:
        query += " AND dm.MES_COMPETENCIA = :competencia"
        params['competencia'] = competencia

    # Ordenar por maior edital e período primeiro
    query += " ORDER BY ds.ID_EDITAL DESC, p.ID_PERIODO DESC, dm.MES_COMPETENCIA, ds.NO_EMPRESA_ABREVIADA"

    result = db.session.execute(text(query), params)

    metas = []
    for row in result:
        meta = {
            'MES_COMPETENCIA': row[0],
            'VR_META_MES': float(row[1]) if row[1] else 0,
            'ID_EMPRESA': row[2],
            'ID_EDITAL': row[3],
            'ID_PERIODO': row[4],
            'NO_EMPRESA_ABREVIADA': row[5],
            'DESCRICAO_EDITAL': row[6],
            'NU_EDITAL': row[7],
            'ANO_EDITAL': row[8],
            'NUM_PERIODO': row[9],
            'DT_INICIO': row[10],
            'DT_FIM': row[11],
            'PERIODO_PK': row[12]
        }
        metas.append(meta)

    # Buscar dados para os filtros - ordenar por maior ID
    editais = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).all()
    periodos = PeriodoAvaliacao.query.filter(PeriodoAvaliacao.DELETED_AT == None).order_by(
        PeriodoAvaliacao.ID_EDITAL.desc(),
        PeriodoAvaliacao.ID_PERIODO.desc()
    ).all()

    # Buscar empresas distintas da tabela DCA_TB017_DISTRIBUICAO_SUMARIO
    sql_empresas = text("""
        SELECT DISTINCT ds.ID_EMPRESA, ds.NO_EMPRESA_ABREVIADA
        FROM BDG.DCA_TB017_DISTRIBUICAO_SUMARIO ds
        WHERE ds.DELETED_AT IS NULL
        ORDER BY ds.NO_EMPRESA_ABREVIADA
    """)
    result_empresas = db.session.execute(sql_empresas)
    empresas = []
    for row in result_empresas:
        empresas.append({
            'ID': row[0],
            'NO_EMPRESA_ABREVIADA': row[1]
        })

    # Buscar competências distintas da tabela DCA_TB019_DISTRIBUICAO_MENSAL
    sql_comp = text("""
        SELECT DISTINCT MES_COMPETENCIA 
        FROM BDG.DCA_TB019_DISTRIBUICAO_MENSAL 
        WHERE DELETED_AT IS NULL 
        ORDER BY MES_COMPETENCIA DESC
    """)
    competencias = [row[0] for row in db.session.execute(sql_comp)]

    return render_template('credenciamento/lista_metas.html',
                           metas=metas,
                           editais=editais,
                           periodos=periodos,
                           empresas=empresas,
                           competencias=competencias,
                           filtro_edital_id=edital_id,
                           filtro_periodo_id=periodo_id,
                           filtro_empresa_id=empresa_id,
                           filtro_competencia=competencia)

@meta_bp.route('/metas/visualizar-redistribuicao')
@login_required
def visualizar_redistribuicao():
    """Página para visualizar cálculo de redistribuição de metas"""
    editais = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).all()
    periodos = PeriodoAvaliacao.query.filter(PeriodoAvaliacao.DELETED_AT == None).order_by(
        PeriodoAvaliacao.ID_EDITAL.desc(),
        PeriodoAvaliacao.ID_PERIODO.desc()
    ).all()

    return render_template('credenciamento/visualizar_redistribuicao.html',
                           editais=editais,
                           periodos=periodos)


@meta_bp.route('/metas/visualizar-calculo')
@login_required
def visualizar_calculo():
    """API para retornar dados do cálculo de redistribuição"""
    try:
        edital_id = request.args.get('edital_id', type=int)
        periodo_id = request.args.get('periodo_id', type=int)

        if not edital_id or not periodo_id:
            return jsonify({'sucesso': False, 'erro': 'Parâmetros inválidos'})

        from app.utils.visualizador_redistribuicao import VisualizadorRedistribuicao

        visualizador = VisualizadorRedistribuicao(edital_id, periodo_id)
        dados = visualizador.calcular_redistribuicao_completa()

        return jsonify({
            'sucesso': True,
            'dados': dados
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'erro': str(e)})


@meta_bp.route('/metas/redistribuicao')
@login_required
def redistribuicao_metas():
    """Página para redistribuição de metas"""
    editais = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).all()
    periodos = PeriodoAvaliacao.query.filter(PeriodoAvaliacao.DELETED_AT == None).order_by(
        PeriodoAvaliacao.ID_EDITAL.desc(),
        PeriodoAvaliacao.ID_PERIODO.desc()
    ).all()

    return render_template('credenciamento/form_redistribuicao.html',
                           editais=editais,
                           periodos=periodos)


@meta_bp.route('/metas/buscar-empresas-ativas')
@login_required
def buscar_empresas_ativas():
    """
    Busca empresas que podem ser descredenciadas.
    A fonte de dados agora é a TB018, para garantir consistência.
    """
    try:
        edital_id = request.args.get('edital_id', type=int)
        periodo_pk_id = request.args.get('periodo_id', type=int)

        if not edital_id or not periodo_pk_id:
            return jsonify({'sucesso': False, 'erro': 'Edital e Período são obrigatórios.'})

        # Buscar a chave de negócio do período
        periodo = PeriodoAvaliacao.query.get(periodo_pk_id)
        if not periodo:
            return jsonify({'sucesso': False, 'erro': f'Período com ID {periodo_pk_id} não encontrado.'})
        periodo_business_id = periodo.ID_PERIODO

        # CORREÇÃO: A query agora busca da TB018, a mesma fonte do cálculo principal.
        sql = text("""
            SELECT DISTINCT
                m.ID_EMPRESA,
                m.NO_EMPRESA_ABREVIADA,
                m.PERCENTUAL_SALDO_DEVEDOR
            FROM BDG.DCA_TB018_METAS_REDISTRIBUIDAS_MENSAL m
            WHERE m.ID_EDITAL = :edital_id
            AND m.ID_PERIODO = :periodo_id
            AND m.DELETED_AT IS NULL
            AND m.PERCENTUAL_SALDO_DEVEDOR > 0  -- Apenas empresas com percentual a ser redistribuído
            AND m.DT_REFERENCIA = (
                SELECT MAX(DT_REFERENCIA)
                FROM BDG.DCA_TB018_METAS_REDISTRIBUIDAS_MENSAL
                WHERE ID_EDITAL = :edital_id
                AND ID_PERIODO = :periodo_id
                AND DELETED_AT IS NULL
            )
            ORDER BY m.NO_EMPRESA_ABREVIADA
        """)

        result = db.session.execute(sql, {
            'edital_id': edital_id,
            'periodo_id': periodo_business_id
        }).fetchall()

        if not result:
            return jsonify({'sucesso': False, 'erro': 'Nenhuma empresa ativa encontrada na última distribuição salva.'})

        empresas = []
        for row in result:
            empresas.append({
                'id_empresa': row.ID_EMPRESA,
                'nome_empresa': row.NO_EMPRESA_ABREVIADA,
                'percentual': float(row.PERCENTUAL_SALDO_DEVEDOR)
            })

        return jsonify({'sucesso': True, 'empresas': empresas})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'erro': str(e)})


@meta_bp.route('/metas/calcular-nova-redistribuicao', methods=['POST'])
@login_required
def calcular_nova_redistribuicao():
    """Calcula uma nova redistribuição de metas"""
    try:
        data = request.json
        from app.utils.redistribuicao_dinamica import RedistribuicaoDinamica

        redistribuidor = RedistribuicaoDinamica(
            edital_id=data['edital_id'],
            periodo_id=data['periodo_id'],
            empresa_saindo_id=data['empresa_id'],
            data_descredenciamento=data['data_descredenciamento'],
            incremento_meta=data.get('incremento_meta', 1.0)
        )
        resultado = redistribuidor.calcular_redistribuicao()
        return jsonify(resultado)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)})


@meta_bp.route('/metas/salvar-redistribuicao', methods=['POST'])
@login_required
def salvar_redistribuicao():
    """Salva a redistribuição calculada"""
    try:
        data = request.json
        from app.utils.redistribuicao_dinamica import RedistribuicaoDinamica

        redistribuidor = RedistribuicaoDinamica(
            edital_id=data['edital_id'],
            periodo_id=data['periodo_id'],
            empresa_saindo_id=data['empresa_id'],
            data_descredenciamento=data['data_descredenciamento'],
            incremento_meta=data.get('incremento_meta', 1.0)
        )
        # O cálculo é feito novamente dentro do método salvar para garantir consistência
        sucesso = redistribuidor.salvar_redistribuicao()

        if sucesso:
            registrar_log(
                acao='redistribuicao_metas', entidade='meta',
                entidade_id=f"{data['edital_id']}-{data['periodo_id']}",
                descricao=f"Redistribuição de metas - Empresa {data['empresa_id']} descredenciada em {data['data_descredenciamento']}",
                dados_novos=data
            )
            return jsonify({'sucesso': True, 'mensagem': 'Redistribuição salva com sucesso!'})
        else:
            return jsonify({'sucesso': False, 'erro': 'Erro desconhecido ao salvar redistribuição'})

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'erro': str(e)})


@meta_bp.route('/metas/nova', methods=['GET', 'POST'])
@login_required
def nova_meta():
    """Página para cálculo automático de metas"""
    # Ordenar por maior edital e período
    editais = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).all()
    periodos = PeriodoAvaliacao.query.filter(PeriodoAvaliacao.DELETED_AT == None).order_by(
        PeriodoAvaliacao.ID_EDITAL.desc(),
        PeriodoAvaliacao.ID_PERIODO.desc()
    ).all()

    return render_template('credenciamento/form_meta.html',
                           editais=editais,
                           periodos=periodos)


# Localizar a função calcular_metas e garantir que ela pegue o fator_incremento do formulário
@meta_bp.route('/metas/calcular', methods=['POST'])
@login_required
def calcular_metas():
    """Calcula metas usando a classe MetaCalculator"""
    try:
        edital_id = int(request.form['edital_id'])
        periodo_id = int(request.form['periodo_id'])
        fator_incremento = float(request.form.get('fator_incremento', 1.00))

        # Usar MetaCalculator com fator_incremento
        calculator = MetaCalculator(edital_id, periodo_id, fator_incremento)
        metas_calculadas = calculator.calcular_metas_completas()

        # Obter período para retornar as datas
        periodo = PeriodoAvaliacao.query.get(periodo_id)

        return jsonify({
            'sucesso': True,
            'metas': metas_calculadas,
            'periodo': {
                'inicio': periodo.DT_INICIO.strftime('%d/%m/%Y'),
                'fim': periodo.DT_FIM.strftime('%d/%m/%Y')
            }
        })

    except ValueError as e:
        return jsonify({'erro': str(e)}), 404
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@meta_bp.route('/metas/salvar-calculadas', methods=['POST'])
@login_required
def salvar_metas_calculadas():
    """Salva as metas calculadas após confirmação do usuário"""
    try:
        data = request.json
        metas_data = data['metas']
        edital_id = int(data['edital_id'])
        periodo_id = int(data['periodo_id'])

        # Criar calculador e salvar
        calculator = MetaCalculator(edital_id, periodo_id)
        calculator.salvar_metas(metas_data)

        # Registrar log
        registrar_log(
            acao='calcular_metas',
            entidade='meta',
            entidade_id=f"{edital_id}-{periodo_id}",
            descricao=f'Cálculo e salvamento de metas para edital {edital_id} período {periodo_id}',
            dados_novos={'edital_id': edital_id, 'periodo_id': periodo_id}
        )

        return jsonify({
            'sucesso': True,
            'mensagem': 'Metas salvas com sucesso!'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 500


@meta_bp.route('/metas/excluir/<int:id>')
@login_required
def excluir_meta(id):
    try:
        meta = MetaAvaliacao.query.get_or_404(id)
        meta.DELETED_AT = datetime.utcnow()
        db.session.commit()

        registrar_log(
            acao='excluir',
            entidade='meta',
            entidade_id=meta.ID,
            descricao=f'Exclusão de meta de avaliação para {meta.COMPETENCIA}',
            dados_antigos={'deleted_at': None},
            dados_novos={'deleted_at': meta.DELETED_AT.strftime('%Y-%m-%d %H:%M:%S')}
        )

        flash('Meta de avaliação removida com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir meta: {str(e)}', 'danger')

    return redirect(url_for('meta.lista_metas'))


@meta_bp.route('/metas/buscar-empresas-redistribuicao')
@login_required
def buscar_empresas_redistribuicao():
    """Busca empresas disponíveis para redistribuição agrupadas por data"""
    try:
        edital_id = request.args.get('edital_id', type=int)
        periodo_id = request.args.get('periodo_id', type=int)

        # Buscar todas as distribuições do período
        sql = text("""
            SELECT DISTINCT
                mpd.DT_REFERENCIA,
                mpd.ID_EMPRESA,
                mpd.NO_EMPRESA_ABREVIADA,
                mpd.VR_SALDO_DEVEDOR_DISTRIBUIDO,
                mpd.PERCENTUAL_SALDO_DEVEDOR
            FROM BDG.DCA_TB018_METAS_REDISTRIBUIDAS_MENSAL mpd
            WHERE mpd.ID_EDITAL = :edital_id
            AND mpd.ID_PERIODO = :periodo_id
            AND mpd.DELETED_AT IS NULL
            ORDER BY mpd.DT_REFERENCIA, mpd.NO_EMPRESA_ABREVIADA
        """)

        periodo = PeriodoAvaliacao.query.get(periodo_id)

        result = db.session.execute(sql, {
            'edital_id': edital_id,
            'periodo_id': periodo.ID_PERIODO
        })

        # Agrupar por data de referência
        empresas_por_data = {}

        for row in result:
            dt_ref = row.DT_REFERENCIA.strftime('%Y-%m-%d')

            if dt_ref not in empresas_por_data:
                empresas_por_data[dt_ref] = []

            empresas_por_data[dt_ref].append({
                'id_empresa': row.ID_EMPRESA,
                'nome_empresa': row.NO_EMPRESA_ABREVIADA,
                'saldo_devedor': float(row.VR_SALDO_DEVEDOR_DISTRIBUIDO) if row.VR_SALDO_DEVEDOR_DISTRIBUIDO else 0,
                'percentual': float(row.PERCENTUAL_SALDO_DEVEDOR) if row.PERCENTUAL_SALDO_DEVEDOR else 0
            })

        return jsonify({
            'sucesso': True,
            'empresas_por_data': empresas_por_data
        })

    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)})


@meta_bp.route('/metas/calcular-redistribuicao', methods=['POST'])
@login_required
def calcular_redistribuicao():
    """Calcula preview da redistribuição de metas"""
    try:
        data = request.json

        from app.utils.redistribuicao_dinamica import RedistribuicaoDinamica

        redistribuidor = RedistribuicaoDinamica(
            edital_id=int(data['edital_id']),
            periodo_id=int(data['periodo_id']),
            empresa_saindo_id=int(data['empresa_id']),
            data_descredenciamento=data['data_descredenciamento'],
            incremento_meta=float(data['incremento_meta'])
        )

        resultado = redistribuidor.calcular_redistribuicao()

        return jsonify(resultado)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@meta_bp.route('/metas/nova-distribuicao')
@login_required
def nova_meta_distribuicao():
    """Página para criar nova distribuição inicial de metas"""
    editais = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).all()
    # Não vamos mais passar os períodos aqui, eles serão carregados dinamicamente
    return render_template('credenciamento/form_nova_meta.html',
                           editais=editais)


@meta_bp.route('/metas/calcular-distribuicao-inicial', methods=['POST'])
@login_required
def calcular_distribuicao_inicial():
    """Calcula a distribuição inicial de metas"""
    try:
        data = request.json
        edital_id = int(data['edital_id'])
        periodo_id = int(data['periodo_id'])
        incremento_meta = float(data.get('incremento_meta', 1.0))

        from app.utils.distribuicao_inicial import DistribuicaoInicial

        distribuidor = DistribuicaoInicial(edital_id, periodo_id, incremento_meta)
        resultado = distribuidor.calcular_distribuicao()

        # Guardar na sessão para posterior salvamento
        session['distribuicao_calculada'] = resultado
        session['distribuicao_edital'] = edital_id
        session['distribuicao_periodo'] = periodo_id
        session['distribuicao_incremento'] = incremento_meta # Salva o incremento na sessão

        return jsonify({
            'sucesso': True,
            'resultado': resultado
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'erro': str(e)})


@meta_bp.route('/metas/salvar-distribuicao-inicial', methods=['POST'])
@login_required
def salvar_distribuicao_inicial():
    """Salva a distribuição inicial calculada"""
    try:
        data = request.json
        edital_id = int(data['edital_id'])
        periodo_id = int(data['periodo_id'])

        # Verificar se os dados de sessão existem e correspondem à requisição
        if (session.get('distribuicao_edital') != edital_id or
                session.get('distribuicao_periodo') != periodo_id or
                'distribuicao_calculada' not in session):
            return jsonify({
                'sucesso': False,
                'erro': 'Dados de cálculo não encontrados na sessão. Por favor, calcule a distribuição novamente.'
            })

        # ====================================================================
        # REMOVIDA A VALIDAÇÃO QUE IMPEDIA SALVAR SE JÁ EXISTISSE
        # Agora o soft delete é feito automaticamente na função salvar_distribuicao()
        # ====================================================================

        # Buscar dados calculados da sessão
        dados_calculados = session.get('distribuicao_calculada')
        incremento_salvar = session.get('distribuicao_incremento', 1.0)

        periodo = PeriodoAvaliacao.query.get(periodo_id)

        from app.utils.distribuicao_inicial import DistribuicaoInicial
        distribuidor = DistribuicaoInicial(edital_id, periodo_id, incremento_salvar)

        # Salvar distribuição (com soft delete automático dentro)
        if distribuidor.salvar_distribuicao(dados_calculados):
            # Limpar dados da sessão após o salvamento
            session.pop('distribuicao_calculada', None)
            session.pop('distribuicao_edital', None)
            session.pop('distribuicao_periodo', None)
            session.pop('distribuicao_incremento', None)

            # Registrar log
            registrar_log(
                acao='criar_distribuicao_inicial',
                entidade='meta',
                entidade_id=f"{edital_id}-{periodo.ID_PERIODO}",
                descricao=f'Criação de distribuição inicial de metas para edital {edital_id} período {periodo.ID_PERIODO} com incremento de {incremento_salvar}',
                dados_novos={'edital_id': edital_id, 'periodo_id': periodo_id, 'incremento_meta': incremento_salvar}
            )

            return jsonify({
                'sucesso': True,
                'mensagem': 'Distribuição salva com sucesso!'
            })
        else:
            return jsonify({
                'sucesso': False,
                'erro': 'Erro ao salvar distribuição'
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'erro': str(e)})


# NOVA ROTA PARA BUSCAR PERÍODOS FILTRADOS
@meta_bp.route('/metas/buscar-periodos-por-edital')
@login_required
def buscar_periodos_por_edital():
    """Busca períodos de avaliação filtrados por edital."""
    try:
        edital_id = request.args.get('edital_id', type=int)
        if not edital_id:
            return jsonify([])

        # Query para buscar apenas os períodos do edital selecionado
        periodos = PeriodoAvaliacao.query.filter(
            PeriodoAvaliacao.ID_EDITAL == edital_id,
            PeriodoAvaliacao.DELETED_AT == None
        ).order_by(PeriodoAvaliacao.ID_PERIODO.desc()).all()

        # Formata os dados para o frontend
        periodos_json = [
            {
                'id': p.ID,
                'texto': f"Período {p.ID_PERIODO} - {p.DT_INICIO.strftime('%d/%m/%Y')} a {p.DT_FIM.strftime('%d/%m/%Y')}"
            } for p in periodos
        ]
        return jsonify(periodos_json)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@meta_bp.route('/metas/exportar-distribuicao-excel', methods=['POST'])
@login_required
def exportar_distribuicao_excel():
    """Exporta a distribuição calculada para Excel"""
    try:
        import io
        import pandas as pd
        from flask import send_file

        data = request.json
        resultado = data['resultado']

        # Preparar dados para o DataFrame
        dados_excel = []

        # Adicionar dados das empresas
        for empresa in resultado['empresas']:
            linha = {
                'Empresa': empresa['nome'],
                'Saldo Devedor': empresa['saldo_devedor'],
                'Percentual': empresa['percentual']
            }

            # Adicionar colunas dos meses
            for mes in resultado['meses']:
                linha[f"{mes['nome']}/{mes['competencia'].split('-')[0]}"] = empresa['metas'][mes['competencia']]

            linha['Total Meta'] = empresa['total']
            dados_excel.append(linha)

        # Adicionar linha de total SISCOR
        linha_siscor = {
            'Empresa': 'TOTAL SISCOR (100%)',
            'Saldo Devedor': resultado['total_saldo_devedor'],
            'Percentual': 100.0
        }
        for mes in resultado['meses']:
            linha_siscor[f"{mes['nome']}/{mes['competencia'].split('-')[0]}"] = resultado['totais_siscor'][
                mes['competencia']]
        linha_siscor['Total Meta'] = resultado['total_geral_siscor']
        dados_excel.append(linha_siscor)

        # Adicionar linha de total com incremento
        incremento_perc = resultado['incremento_meta'] * 100
        linha_incremento = {
            'Empresa': f'TOTAL META ({incremento_perc:.0f}%)',
            'Saldo Devedor': resultado['total_saldo_devedor'],
            'Percentual': 100.0
        }
        for mes in resultado['meses']:
            linha_incremento[f"{mes['nome']}/{mes['competencia'].split('-')[0]}"] = \
            resultado['totais_meta_incrementada'][mes['competencia']]
        linha_incremento['Total Meta'] = resultado['total_geral_incrementado']
        dados_excel.append(linha_incremento)

        # Criar DataFrame
        df = pd.DataFrame(dados_excel)

        # Criar arquivo Excel em memória
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Distribuição de Metas', index=False)

            # Obter workbook e worksheet
            workbook = writer.book
            worksheet = writer.sheets['Distribuição de Metas']

            # Formatar cabeçalho
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'center',
                'fg_color': '#6c63ff',
                'font_color': 'white',
                'border': 1
            })

            # Formatar valores monetários
            money_format = workbook.add_format({
                'num_format': 'R$ #,##0.00',
                'border': 1
            })

            # Formatar percentual
            percent_format = workbook.add_format({
                'num_format': '0.00000000%',
                'border': 1
            })

            # Aplicar formatação
            for idx, col in enumerate(df.columns):
                worksheet.write(0, idx, col, header_format)

                # Ajustar largura das colunas
                if col == 'Empresa':
                    worksheet.set_column(idx, idx, 40)
                elif col in ['Saldo Devedor', 'Total Meta'] or '/' in col:
                    worksheet.set_column(idx, idx, 15)
                    # Aplicar formato monetário
                    for row in range(1, len(df) + 1):
                        if col != 'Percentual':
                            worksheet.write(row, idx, df.iloc[row - 1][col], money_format)
                elif col == 'Percentual':
                    worksheet.set_column(idx, idx, 12)
                    # Aplicar formato percentual
                    for row in range(1, len(df) - 2 + 1):  # Não aplicar nas linhas de total
                        worksheet.write(row, idx, df.iloc[row - 1][col] / 100, percent_format)

        output.seek(0)

        # Gerar nome do arquivo
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        periodo_info = resultado.get('periodo_info', 'periodo')
        filename = f'distribuicao_metas_{periodo_info.replace(" ", "_")}_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


# =====================================================
# NOVA FUNCIONALIDADE: META TERMO
# =====================================================

@meta_bp.route('/metas/meta-termo')
@login_required
def meta_termo():
    """Página para visualizar cálculo de Meta Termo"""
    # Buscar editais e períodos para os filtros
    editais = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).all()
    periodos = PeriodoAvaliacao.query.filter(PeriodoAvaliacao.DELETED_AT == None).order_by(
        PeriodoAvaliacao.ID_EDITAL.desc(),
        PeriodoAvaliacao.ID_PERIODO.desc()
    ).all()

    return render_template('credenciamento/meta_termo.html',
                           editais=editais,
                           periodos=periodos)


@meta_bp.route('/metas/buscar-meta-termo')
@login_required
def buscar_meta_termo():
    """Busca e calcula os dados de Meta Termo"""
    try:
        # Buscar o edital e período mais atual da tabela DCA_TB021_PERCENTUAL_TERMO
        sql_query = text("""
            SELECT TOP 1
                EDITAL,
                PERIODO
            FROM BDDASHBOARDBI.BDG.DCA_TB021_PERCENTUAL_TERMO
            ORDER BY EDITAL DESC, PERIODO DESC
        """)

        resultado_filtro = db.session.execute(sql_query).fetchone()

        if not resultado_filtro:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum dado encontrado na tabela DCA_TB021_PERCENTUAL_TERMO'
            })

        edital_atual = resultado_filtro[0]
        periodo_atual = resultado_filtro[1]

        # Buscar os dados filtrados pelo edital e período mais atual
        # JOIN com a tabela DCA_TB017_DISTRIBUICAO_SUMARIO para pegar o nome da empresa
        sql_dados = text("""
            SELECT 
                PT.EDITAL,
                PT.PERIODO,
                PT.COD_EMPRESA_COBRANCA,
                COALESCE(DS.NO_EMPRESA_ABREVIADA, 'Empresa ' + CAST(PT.COD_EMPRESA_COBRANCA AS VARCHAR)) AS NO_EMPRESA,
                PT.VR_SD_DEVEDOR,
                PT.PERC,
                PT.META_PERC
            FROM BDDASHBOARDBI.BDG.DCA_TB021_PERCENTUAL_TERMO PT
            LEFT JOIN BDG.DCA_TB017_DISTRIBUICAO_SUMARIO DS
                ON PT.COD_EMPRESA_COBRANCA = DS.ID_EMPRESA
                AND PT.EDITAL = DS.ID_EDITAL
                AND PT.PERIODO = DS.ID_PERIODO
                AND DS.DELETED_AT IS NULL
            WHERE PT.EDITAL = :edital
                AND PT.PERIODO = :periodo
            ORDER BY PT.COD_EMPRESA_COBRANCA
        """)

        resultado_dados = db.session.execute(
            sql_dados,
            {'edital': edital_atual, 'periodo': periodo_atual}
        ).fetchall()

        # Processar os dados e calcular META_TERMO
        dados_processados = []
        for row in resultado_dados:
            edital = row[0]
            periodo = row[1]
            cod_empresa = row[2]
            nome_empresa = row[3]
            vr_sd_devedor = float(row[4]) if row[4] else 0
            perc = float(row[5]) if row[5] else 0
            meta_perc = float(row[6]) if row[6] else 0

            # Calcular META_TERMO = (META_PERC / VR_SD_DEVEDOR) * 100
            # Tratamento para divisão por zero
            if vr_sd_devedor != 0:
                meta_termo = (meta_perc / vr_sd_devedor) * 100
            else:
                meta_termo = 0

            dados_processados.append({
                'edital': edital,
                'periodo': periodo,
                'cod_empresa_cobranca': cod_empresa,
                'nome_empresa': nome_empresa,
                'vr_sd_devedor': vr_sd_devedor,
                'perc': perc,
                'meta_perc': meta_perc,
                'meta_termo': round(meta_termo, 10)  # 10 casas decimais
            })

        return jsonify({
            'sucesso': True,
            'dados': dados_processados,
            'edital': edital_atual,
            'periodo': periodo_atual
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        })


@meta_bp.route('/metas/exportar-meta-termo')
@login_required
def exportar_meta_termo():
    """Exporta os dados de Meta Termo para Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        # Buscar o edital e período mais atual
        sql_query = text("""
            SELECT TOP 1
                EDITAL,
                PERIODO
            FROM BDDASHBOARDBI.BDG.DCA_TB021_PERCENTUAL_TERMO
            ORDER BY EDITAL DESC, PERIODO DESC
        """)

        resultado_filtro = db.session.execute(sql_query).fetchone()

        if not resultado_filtro:
            flash('Nenhum dado encontrado para exportação.', 'warning')
            return redirect(url_for('meta.meta_termo'))

        edital_atual = resultado_filtro[0]
        periodo_atual = resultado_filtro[1]

        # Buscar os dados com JOIN para pegar o nome da empresa
        sql_dados = text("""
            SELECT 
                PT.EDITAL,
                PT.PERIODO,
                PT.COD_EMPRESA_COBRANCA,
                COALESCE(DS.NO_EMPRESA_ABREVIADA, 'Empresa ' + CAST(PT.COD_EMPRESA_COBRANCA AS VARCHAR)) AS NO_EMPRESA,
                PT.VR_SD_DEVEDOR,
                PT.PERC,
                PT.META_PERC
            FROM BDDASHBOARDBI.BDG.DCA_TB021_PERCENTUAL_TERMO PT
            LEFT JOIN BDG.DCA_TB017_DISTRIBUICAO_SUMARIO DS
                ON PT.COD_EMPRESA_COBRANCA = DS.ID_EMPRESA
                AND PT.EDITAL = DS.ID_EDITAL
                AND PT.PERIODO = DS.ID_PERIODO
                AND DS.DELETED_AT IS NULL
            WHERE PT.EDITAL = :edital
                AND PT.PERIODO = :periodo
            ORDER BY PT.COD_EMPRESA_COBRANCA
        """)

        resultado_dados = db.session.execute(
            sql_dados,
            {'edital': edital_atual, 'periodo': periodo_atual}
        ).fetchall()

        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Meta Termo"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="5b52e5", end_color="5b52e5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws.merge_cells('A1:G1')
        titulo_cell = ws['A1']
        titulo_cell.value = f'Meta Termo - Edital {edital_atual} - Período {periodo_atual}'
        titulo_cell.font = Font(bold=True, size=14)
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
        titulo_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Data de geração
        ws.merge_cells('A2:G2')
        data_cell = ws['A2']
        data_cell.value = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        data_cell.alignment = Alignment(horizontal="center")

        # Cabeçalhos
        headers = ['Edital', 'Período', 'Empresa', 'Saldo Devedor', 'Perc', 'Meta Perc', 'Meta Termo (%)']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Dados
        for row_num, row_data in enumerate(resultado_dados, 5):
            edital = row_data[0]
            periodo = row_data[1]
            cod_empresa = row_data[2]
            nome_empresa = row_data[3]
            vr_sd_devedor = float(row_data[4]) if row_data[4] else 0
            perc = float(row_data[5]) if row_data[5] else 0
            meta_perc = float(row_data[6]) if row_data[6] else 0

            # Calcular META_TERMO
            if vr_sd_devedor != 0:
                meta_termo = (meta_perc / vr_sd_devedor) * 100
            else:
                meta_termo = 0

            # Escrever dados
            ws.cell(row=row_num, column=1).value = edital
            ws.cell(row=row_num, column=2).value = periodo
            ws.cell(row=row_num, column=3).value = nome_empresa
            ws.cell(row=row_num, column=4).value = vr_sd_devedor
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=5).value = perc
            ws.cell(row=row_num, column=5).number_format = '#,##0.0000000000'
            ws.cell(row=row_num, column=6).value = meta_perc
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            ws.cell(row=row_num, column=7).value = meta_termo
            ws.cell(row=row_num, column=7).number_format = '#,##0.0000000000'

            # Aplicar bordas
            for col_num in range(1, 8):
                ws.cell(row=row_num, column=col_num).border = border

        # Ajustar largura das colunas
        column_widths = [12, 12, 30, 18, 15, 18, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Registrar no log
        registrar_log(
            acao='exportar_meta_termo',
            entidade='meta',
            entidade_id=f"{edital_atual}-{periodo_atual}",
            descricao=f'Exportação de Meta Termo para Edital {edital_atual} Período {periodo_atual}',
            dados_novos={'edital': edital_atual, 'periodo': periodo_atual, 'total_registros': len(resultado_dados)}
        )

        # Nome do arquivo
        nome_arquivo = f'meta_termo_{edital_atual}_{periodo_atual}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except ImportError:
        flash('É necessário instalar a biblioteca openpyxl para exportar dados: pip install openpyxl', 'warning')
        return redirect(url_for('meta.meta_termo'))
    except Exception as e:
        flash(f'Erro ao exportar dados: {str(e)}', 'danger')
        import traceback
        traceback.print_exc()
        return redirect(url_for('meta.meta_termo'))


@meta_bp.route('/metas/exportar-relatorio-excel', methods=['POST'])
@login_required
def exportar_relatorio_metas_excel():
    """Exporta relatório consolidado de metas para Excel"""
    try:
        import io
        import pandas as pd

        data = request.json
        edital_id = data.get('edital_id')
        periodo_id = data.get('periodo_id')

        if not edital_id or not periodo_id:
            return jsonify({'erro': 'Edital e Período são obrigatórios'}), 400

        # Buscar dados do edital e período
        edital = Edital.query.get(edital_id)
        periodo = PeriodoAvaliacao.query.get(periodo_id)

        if not edital or not periodo:
            return jsonify({'erro': 'Edital ou Período não encontrado'}), 404

        # Query para buscar todas as metas do edital/período
        query = text("""
            SELECT 
                ds.NO_EMPRESA_ABREVIADA,
                dm.MES_COMPETENCIA,
                dm.VR_META_MES,
                ds.VR_SALDO_DEVEDOR_DISTRIBUIDO,
                ds.PERCENTUAL_SALDO_DEVEDOR
            FROM BDG.DCA_TB019_DISTRIBUICAO_MENSAL dm
            INNER JOIN BDG.DCA_TB017_DISTRIBUICAO_SUMARIO ds
                ON dm.ID_DISTRIBUICAO_SUMARIO = ds.ID
                AND ds.DELETED_AT IS NULL
            WHERE ds.ID_EDITAL = :edital_id
            AND ds.ID_PERIODO = :periodo_id
            AND dm.DELETED_AT IS NULL
            ORDER BY ds.NO_EMPRESA_ABREVIADA, dm.MES_COMPETENCIA
        """)

        result = db.session.execute(query, {
            'edital_id': edital_id,
            'periodo_id': periodo.ID_PERIODO
        })

        # Estruturar dados
        dados_empresas = {}
        meses_unicos = set()

        for row in result:
            empresa = row[0]
            mes_comp = row[1]
            valor_meta = float(row[2]) if row[2] else 0
            saldo_devedor = float(row[3]) if row[3] else 0
            percentual = float(row[4]) if row[4] else 0

            meses_unicos.add(mes_comp)

            if empresa not in dados_empresas:
                dados_empresas[empresa] = {
                    'saldo_devedor': saldo_devedor,
                    'percentual': percentual,
                    'metas': {},
                    'total': 0
                }

            dados_empresas[empresa]['metas'][mes_comp] = valor_meta
            dados_empresas[empresa]['total'] += valor_meta

        # Ordenar meses
        meses_ordenados = sorted(list(meses_unicos))

        # Preparar dados para DataFrame
        dados_excel = []

        # Adicionar dados das empresas
        for empresa, info in sorted(dados_empresas.items()):
            linha = {
                'Empresa': empresa,
                'Saldo Devedor': info['saldo_devedor'],
                'Percentual (%)': info['percentual']
            }

            # Adicionar colunas dos meses
            for mes in meses_ordenados:
                # Formatar mês como JAN/2025
                ano = mes.split('-')[0]
                mes_num = int(mes.split('-')[1])
                meses_nomes = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN',
                               'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
                nome_mes = f"{meses_nomes[mes_num - 1]}/{ano}"

                linha[nome_mes] = info['metas'].get(mes, 0)

            linha['Total Meta'] = info['total']
            dados_excel.append(linha)

        # Adicionar linha de TOTAL
        linha_total = {
            'Empresa': 'TOTAL GERAL',
            'Saldo Devedor': sum(emp['saldo_devedor'] for emp in dados_empresas.values()),
            'Percentual (%)': sum(emp['percentual'] for emp in dados_empresas.values())
        }

        for mes in meses_ordenados:
            ano = mes.split('-')[0]
            mes_num = int(mes.split('-')[1])
            meses_nomes = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN',
                           'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
            nome_mes = f"{meses_nomes[mes_num - 1]}/{ano}"

            linha_total[nome_mes] = sum(
                emp['metas'].get(mes, 0) for emp in dados_empresas.values()
            )

        linha_total['Total Meta'] = sum(emp['total'] for emp in dados_empresas.values())
        dados_excel.append(linha_total)

        # Criar DataFrame
        df = pd.DataFrame(dados_excel)

        # Criar arquivo Excel em memória
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Metas Mensais', index=False)

            # Obter workbook e worksheet
            workbook = writer.book
            worksheet = writer.sheets['Metas Mensais']

            # Formatar cabeçalho
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'center',
                'align': 'center',
                'fg_color': '#6c63ff',
                'font_color': 'white',
                'border': 1
            })

            # Formatar valores monetários
            money_format = workbook.add_format({
                'num_format': 'R$ #,##0.00',
                'border': 1
            })

            # Formatar percentuais
            percent_format = workbook.add_format({
                'num_format': '0.00000000"%"',
                'border': 1
            })

            # Formatar linha de total
            total_format = workbook.add_format({
                'bold': True,
                'bg_color': '#f0f0f0',
                'num_format': 'R$ #,##0.00',
                'border': 1
            })

            # Aplicar formatação ao cabeçalho
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)

            # Aplicar formatação às colunas
            for row_num in range(1, len(df) + 1):
                # Empresa (texto)
                worksheet.write(row_num, 0, df.iloc[row_num - 1]['Empresa'])

                # Saldo Devedor (moeda)
                worksheet.write(row_num, 1, df.iloc[row_num - 1]['Saldo Devedor'], money_format)

                # Percentual
                worksheet.write(row_num, 2, df.iloc[row_num - 1]['Percentual (%)'] / 100, percent_format)

                # Metas mensais (moeda)
                for col_num in range(3, len(df.columns)):
                    formato = total_format if df.iloc[row_num - 1]['Empresa'] == 'TOTAL GERAL' else money_format
                    worksheet.write(row_num, col_num, df.iloc[row_num - 1, col_num], formato)

            # Ajustar largura das colunas
            worksheet.set_column('A:A', 20)  # Empresa
            worksheet.set_column('B:B', 18)  # Saldo Devedor
            worksheet.set_column('C:C', 18)  # Percentual
            worksheet.set_column('D:Z', 15)  # Meses

        output.seek(0)

        # Nome do arquivo
        nome_arquivo = f"metas_{edital.NU_EDITAL}_{edital.ANO}_periodo_{periodo.ID_PERIODO}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Registrar log - CORREÇÃO AQUI
        registrar_log(
            acao='exportar',
            entidade='distribuicao_mensal',
            entidade_id=None,
            descricao=f'Exportação de metas - Edital {edital.NU_EDITAL}/{edital.ANO} - Período {periodo.ID_PERIODO}'
        )

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500