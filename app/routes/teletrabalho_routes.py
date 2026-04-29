# app/routes/teletrabalho_routes.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from app import db
from app.models.teletrabalho import Teletrabalho, Feriado, BloqueioDia
from app.models.usuario import Usuario, Empregado
from app.utils.audit import registrar_log
from datetime import datetime, timedelta, date
from sqlalchemy import and_, or_, func, text, extract
import calendar
import math
from app.models.teletrabalho import ConfigAreaTeletrabalho
from flask import send_file
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from app.models.teletrabalho import Teletrabalho, ConfigAreaTeletrabalho, Feriado, BloqueioDia, Calendario



teletrabalho_bp = Blueprint('teletrabalho', __name__, url_prefix='/teletrabalho')



def verificar_nivel_acesso_teletrabalho(usuario):
    """
    Determina o nível de acesso do usuário no sistema de teletrabalho

    PRIORIDADE:
    1. GERENTE (cargo) → retorna 'gerente' (vê apenas sua área)
    2. ADMIN (perfil) → retorna 'admin' (vê tudo)
    3. MODERADOR (perfil, não gerente) → retorna 'moderador' (vê tudo)
    4. USUÁRIO → retorna 'usuario' (vê apenas próprios dados)

    IMPORTANTE: Gerente prevalece sobre moderador!
    Se alguém é gerente E moderador, será tratado como gerente.
    """
    # 1. PRIORIDADE: Verificar se é GERENTE (pelo cargo)
    if usuario.empregado and usuario.empregado.dsCargo:
        if 'GERENTE' in usuario.empregado.dsCargo.upper():
            return 'gerente'

    # 2. Se não é gerente, verificar perfil ADMIN
    if usuario.PERFIL == 'admin':
        return 'admin'

    # 3. Se não é gerente nem admin, verificar MODERADOR
    if usuario.PERFIL == 'moderador':
        return 'moderador'

    # 4. Usuário comum
    return 'usuario'


def eh_gestor_ou_admin(usuario):
    """Verifica se o usuário é gestor, admin ou moderador"""
    nivel = verificar_nivel_acesso_teletrabalho(usuario)
    return nivel in ['gerente', 'admin', 'moderador']


def pode_ver_todas_areas(usuario):
    """Verifica se o usuário pode ver todas as áreas (admin ou moderador não-gerente)"""
    nivel = verificar_nivel_acesso_teletrabalho(usuario)
    return nivel in ['admin', 'moderador']

def calcular_limite_area(area, tipo_area, percentual=30.0):
    """
    Calcula limite baseado em ConfigAreaTeletrabalho (se existir)
    ou conta pessoas reais na área
    """
    from app.models.teletrabalho import ConfigAreaTeletrabalho

    # Buscar configuração salva
    config = ConfigAreaTeletrabalho.query.filter_by(
        AREA=area,
        TIPO_AREA=tipo_area,
        DELETED_AT=None
    ).first()

    if config:
        # Usar configuração personalizada
        qtd = config.QTD_TOTAL_PESSOAS
        percentual = float(config.PERCENTUAL_LIMITE)
    else:
        # Contar pessoas reais na área
        if tipo_area == 'setor':
            qtd = Empregado.query.filter(
                Empregado.sgSetor == area,
                Empregado.fkStatus == 1
            ).count()
        elif tipo_area == 'superintendencia':
            qtd = Empregado.query.filter(
                Empregado.sgSuperintendencia == area,
                Empregado.fkStatus == 1
            ).count()
        else:
            qtd = Empregado.query.filter(
                Empregado.sgDiretoria == area,
                Empregado.fkStatus == 1
            ).count()

    limite = (qtd * percentual) / 100.0
    return math.ceil(limite), qtd


def calcular_resumo_pessoas(mes_ref, area, tipo_area):
    """
    Calcula resumo de pessoas e quantidade de dias de teletrabalho no mês

    Retorna lista com:
    - Nome da pessoa
    - Quantidade de dias de teletrabalho no mês
    - Lista com as datas específicas
    """
    try:
        # Buscar todos os teletrabalhos aprovados da área no mês
        teletrabalhos = Teletrabalho.query.join(
            Usuario, Teletrabalho.USUARIO_ID == Usuario.ID
        ).filter(
            Teletrabalho.MES_REFERENCIA == mes_ref,
            Teletrabalho.AREA == area,
            Teletrabalho.TIPO_AREA == tipo_area,
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None)
        ).order_by(Usuario.NOME).all()

        # Agrupar por usuário
        resumo = {}
        for t in teletrabalhos:
            usuario_id = t.USUARIO_ID
            usuario_nome = t.usuario.NOME if t.usuario else "Desconhecido"

            if usuario_id not in resumo:
                resumo[usuario_id] = {
                    'nome': usuario_nome,
                    'quantidade': 0,
                    'datas': []
                }

            resumo[usuario_id]['quantidade'] += 1
            resumo[usuario_id]['datas'].append(t.DATA_TELETRABALHO)

        # Converter para lista ordenada por nome
        lista_resumo = sorted(resumo.values(), key=lambda x: x['nome'])

        return lista_resumo

    except Exception as e:
        print(f"[ERRO] calcular_resumo_pessoas: {str(e)}")
        import traceback
        traceback.print_exc()
        return []


def obter_areas_disponiveis():
    """Retorna lista de áreas disponíveis"""
    areas = []

    superintendencias = db.session.query(
        Empregado.sgSuperintendencia,
        func.count(Empregado.pkPessoa).label('qtd')
    ).filter(
        Empregado.sgSuperintendencia.isnot(None),
        Empregado.sgSuperintendencia != '',
        Empregado.fkStatus == 1
    ).group_by(Empregado.sgSuperintendencia).all()

    for sup, qtd in superintendencias:
        areas.append({
            'area': sup,
            'tipo': 'superintendencia',
            'qtd_pessoas': qtd
        })

    return areas


def gerar_opcoes_periodo():
    """Gera opções de ano e mês para seleção (a partir de outubro/2025)"""
    periodos = []

    # ✅ INÍCIO FIXO: Outubro de 2025
    ano_inicio = 2025
    mes_inicio = 10

    # ✅ Gerar até 2 anos a frente (até 2027)
    for ano in range(ano_inicio, ano_inicio + 3):  # 2025, 2026, 2027
        # ✅ Se for 2025, começar do mês 10 (outubro)
        # ✅ Se for 2026 ou 2027, começar do mês 1 (janeiro)
        primeiro_mes = mes_inicio if ano == ano_inicio else 1

        for mes in range(primeiro_mes, 13):
            data = date(ano, mes, 1)
            mes_ref = data.strftime('%Y%m')
            meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
            nome_mes = f"{meses_pt[mes - 1]}/{ano}"

            periodos.append({
                'valor': mes_ref,
                'nome': nome_mes,
                'ano': ano,
                'mes': mes
            })

    return periodos



@teletrabalho_bp.route('/')
@login_required
def index():
    """Página principal do sistema de teletrabalho"""
    try:
        usuario = Usuario.query.get(current_user.id)
        empregado = usuario.empregado

        if current_user.perfil == 'admin':
            if not empregado:
                areas = obter_areas_disponiveis()
                periodos = gerar_opcoes_periodo()

                return render_template('teletrabalho/index.html',
                                       area='Admin (selecione área)',
                                       tipo_area='superintendencia',
                                       limite=0,
                                       total_pessoas=0,
                                       eh_gestor=True,
                                       areas=areas,
                                       meses_disponiveis=periodos[:12],
                                       admin_sem_empregado=True)

            area = empregado.sgSuperintendencia
            tipo_area = 'superintendencia'
            limite, total_pessoas = calcular_limite_area(area, tipo_area)
            areas = obter_areas_disponiveis()
        else:
            if not empregado:
                flash('Usuário sem vínculo com empregado. Entre em contato com o RH.', 'warning')
                return redirect(url_for('main.geinc_index'))

            area = empregado.sgSuperintendencia
            tipo_area = 'superintendencia'
            limite, total_pessoas = calcular_limite_area(area, tipo_area)
            areas = None

        periodos = gerar_opcoes_periodo()

        return render_template('teletrabalho/index.html',
                               area=area,
                               tipo_area=tipo_area,
                               limite=limite,
                               total_pessoas=total_pessoas,
                               eh_gestor=eh_gestor_ou_admin(usuario),
                               areas=areas,
                               meses_disponiveis=periodos[:12],
                               admin_sem_empregado=False)

    except Exception as e:
        flash(f'Erro ao carregar página: {str(e)}', 'danger')
        return redirect(url_for('main.geinc_index'))


@teletrabalho_bp.route('/calendario/<mes_referencia>')
@login_required
def calendario(mes_referencia):
    """Calendário mensal de teletrabalho do usuário"""
    try:
        usuario = Usuario.query.get(current_user.id)
        empregado = usuario.empregado

        if not empregado:
            flash('Você precisa estar vinculado a um empregado.', 'danger')
            return redirect(url_for('teletrabalho.index'))

        area = empregado.sgSuperintendencia
        tipo_area = 'superintendencia'

        limite, total_pessoas = calcular_limite_area(area, tipo_area)

        ano = int(mes_referencia[:4])
        mes = int(mes_referencia[4:])

        meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        nome_mes = f"{meses_pt[mes - 1]} de {ano}"

        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])
        primeiro_dia_grid = (primeiro_dia.weekday() + 1) % 7

        # ====================================================================
        # ✅ CORREÇÃO: CARREGAR DIAS ÚTEIS DA TABELA PAR_TB020_CALENDARIO
        # ====================================================================
        calendario_info = Calendario.carregar_info_mes(ano, mes)

        # Buscar teletrabalhos do mês nesta área
        teletrabalhos = Teletrabalho.query.filter(
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.AREA == area,
            Teletrabalho.TIPO_AREA == tipo_area,
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        # Organizar teletrabalhos por dia
        dias_ocupados = {}
        for t in teletrabalhos:
            data_str = t.DATA_TELETRABALHO.strftime('%Y-%m-%d')
            if data_str not in dias_ocupados:
                dias_ocupados[data_str] = []
            dias_ocupados[data_str].append({
                'usuario': t.usuario.NOME if t.usuario else 'Desconhecido',
                'id': t.ID
            })

        # Buscar bloqueios do mês
        bloqueios = BloqueioDia.listar_bloqueios_mes(mes_referencia, area, tipo_area)
        dias_bloqueados = {}
        for bloq in bloqueios:
            data_str = bloq.DATA_BLOQUEIO.strftime('%Y-%m-%d')
            dias_bloqueados[data_str] = {
                'motivo': bloq.MOTIVO,
                'bloqueado_por': bloq.bloqueador.NOME if bloq.bloqueador else 'Gestor'
            }

        # Montar estrutura de dias do calendário
        dias_calendario = []
        data_atual = primeiro_dia

        while data_atual <= ultimo_dia:
            data_str = data_atual.strftime('%Y-%m-%d')

            # ✅ Verificar dia útil pela tabela PAR_TB020_CALENDARIO
            info_dia = calendario_info.get(data_atual)
            if info_dia is not None:
                eh_util = info_dia['eh_util']
            else:
                eh_util = Feriado.eh_dia_util(data_atual)

            # Buscar nome do feriado para exibição
            feriado_obj = Feriado.obter_feriado(data_atual)

            # ✅ Texto do feriado: se não é útil mas é dia de semana e sem feriado cadastrado
            feriado_texto = None
            if feriado_obj:
                feriado_texto = feriado_obj.DS_FERIADO
            elif not eh_util and data_atual.weekday() < 5:
                feriado_texto = 'Ponto Facultativo / Não Útil'

            qtd_pessoas = len(dias_ocupados.get(data_str, []))
            bloqueio_info = dias_bloqueados.get(data_str, None)

            dias_calendario.append({
                'data': data_str,
                'dia': data_atual.day,
                'eh_util': eh_util,
                'feriado': feriado_texto,
                'qtd_pessoas': qtd_pessoas,
                'limite': limite,
                'lotado': qtd_pessoas >= limite,
                'pessoas': dias_ocupados.get(data_str, []),
                'bloqueado': bloqueio_info is not None,
                'motivo_bloqueio': bloqueio_info['motivo'] if bloqueio_info else None,
                'bloqueado_por': bloqueio_info['bloqueado_por'] if bloqueio_info else None
            })

            data_atual += timedelta(days=1)

        # Buscar dias que o usuário já marcou neste mês
        meus_dias = Teletrabalho.query.filter(
            Teletrabalho.USUARIO_ID == current_user.id,
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        meus_dias_datas = [d.DATA_TELETRABALHO.strftime('%Y-%m-%d') for d in meus_dias]

        periodos = gerar_opcoes_periodo()

        return render_template('teletrabalho/calendario.html',
                               dias=dias_calendario,
                               mes_referencia=mes_referencia,
                               nome_mes=nome_mes,
                               limite=limite,
                               total_pessoas=total_pessoas,
                               area=area,
                               tipo_area=tipo_area,
                               meus_dias=meus_dias_datas,
                               eh_gestor=eh_gestor_ou_admin(usuario),
                               pode_solicitar=(empregado is not None),
                               primeiro_dia_grid=primeiro_dia_grid,
                               periodos_disponiveis=periodos)

    except Exception as e:
        flash(f'Erro ao carregar calendário: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))




@teletrabalho_bp.route('/solicitar', methods=['POST'])
@login_required
def solicitar():
    """Solicita dias de teletrabalho (usuário comum)."""
    try:
        usuario = Usuario.query.get(current_user.id)
        empregado = usuario.empregado

        if not empregado:
            flash('Você precisa estar vinculado a um empregado para solicitar teletrabalho.', 'danger')
            return redirect(url_for('teletrabalho.index'))

        datas_str = request.form.getlist('datas[]')
        mes_referencia = request.form.get('mes_referencia')
        tipo_periodo = request.form.get('tipo_periodo')
        observacao = request.form.get('observacao', '')

        area = empregado.sgSuperintendencia
        tipo_area = 'superintendencia'

        # ====================================================================
        # ✅ NOVAS REGRAS DE TELETRABALHO
        # - TRES_CONSECUTIVOS: até 2 blocos de 3 dias consecutivos (máx 6 dias/mês)
        # - CINCO_DIAS_CORRIDOS: 5 dias úteis seguidos, único marcação do mês
        # - O limite mensal geral sobe para 6 dias (2 blocos x 3 = 6)
        # ====================================================================
        LIMITE_DIAS_MES = 6

        limite, total_pessoas = calcular_limite_area(area, tipo_area)

        qtd_dias_mes = Teletrabalho.contar_dias_mes(current_user.id, mes_referencia)

        if qtd_dias_mes >= LIMITE_DIAS_MES:
            flash(f'Você já atingiu o limite de {LIMITE_DIAS_MES} dias de teletrabalho neste mês.', 'warning')
            return redirect(url_for('teletrabalho.calendario', mes_referencia=mes_referencia))

        if len(datas_str) + qtd_dias_mes > LIMITE_DIAS_MES:
            flash(
                f'Você só pode marcar mais {LIMITE_DIAS_MES - qtd_dias_mes} dia(s) neste mês.',
                'warning'
            )
            return redirect(url_for('teletrabalho.calendario', mes_referencia=mes_referencia))

        # Validar tipo de período informado
        if tipo_periodo not in ('TRES_CONSECUTIVOS', 'CINCO_DIAS_CORRIDOS'):
            flash('Selecione um tipo de período válido (3 consecutivos ou 5 corridos).', 'warning')
            return redirect(url_for('teletrabalho.calendario', mes_referencia=mes_referencia))

        datas = [datetime.strptime(d, '%Y-%m-%d').date() for d in datas_str]

        # Validação específica do tipo
        if tipo_periodo == 'CINCO_DIAS_CORRIDOS':
            if len(datas) != 5:
                flash('Para "5 dias corridos" selecione exatamente 5 dias.', 'warning')
                return redirect(url_for('teletrabalho.calendario', mes_referencia=mes_referencia))

            valido, msg = Teletrabalho.validar_cinco_dias_corridos(
                current_user.id, datas, mes_referencia
            )
            if not valido:
                flash(msg, 'warning')
                return redirect(url_for('teletrabalho.calendario', mes_referencia=mes_referencia))

        elif tipo_periodo == 'TRES_CONSECUTIVOS':
            valido, msg = Teletrabalho.validar_tres_consecutivos(
                current_user.id, datas, mes_referencia
            )
            if not valido:
                flash(msg, 'warning')
                return redirect(url_for('teletrabalho.calendario', mes_referencia=mes_referencia))

        # Validações de cada data selecionada
        for data in datas:
            # 1. Dia útil
            if not Feriado.eh_dia_util(data):
                feriado = Feriado.obter_feriado(data)
                motivo = feriado.DS_FERIADO if feriado else "fim de semana"
                flash(f'A data {data.strftime("%d/%m/%Y")} não é dia útil ({motivo}).', 'warning')
                return redirect(url_for('teletrabalho.calendario', mes_referencia=mes_referencia))

            # 2. Dia bloqueado
            bloqueado, bloqueio = BloqueioDia.dia_esta_bloqueado(data, area, tipo_area)
            if bloqueado:
                motivo_bloqueio = bloqueio.MOTIVO if bloqueio and bloqueio.MOTIVO else "bloqueado pelo gestor"
                flash(f'A data {data.strftime("%d/%m/%Y")} está bloqueada: {motivo_bloqueio}', 'danger')
                return redirect(url_for('teletrabalho.calendario', mes_referencia=mes_referencia))

            # 3. Limite de pessoas no dia (30% da área)
            qtd_atual = Teletrabalho.contar_pessoas_dia(data, area, tipo_area)
            if qtd_atual >= limite:
                flash(
                    f'Data {data.strftime("%d/%m/%Y")}: limite de {limite} pessoas atingido.',
                    'warning'
                )
                return redirect(url_for('teletrabalho.calendario', mes_referencia=mes_referencia))

        # Salvar
        for data in datas:
            teletrabalho = Teletrabalho(
                USUARIO_ID=current_user.id,
                DATA_TELETRABALHO=data,
                MES_REFERENCIA=mes_referencia,
                AREA=area,
                TIPO_AREA=tipo_area,
                STATUS='APROVADO',
                TIPO_PERIODO=tipo_periodo,
                TIPO_MARCACAO='TELETRABALHO',
                OBSERVACAO=observacao,
                APROVADO_POR=current_user.id,
                APROVADO_EM=datetime.now()
            )
            db.session.add(teletrabalho)

        registrar_log(
            acao='criar',
            entidade='teletrabalho',
            entidade_id=0,
            descricao=f'Solicitação de {len(datas)} dia(s) de teletrabalho - {tipo_periodo}'
        )

        db.session.commit()
        flash(f'{len(datas)} dia(s) de teletrabalho solicitado(s) com sucesso!', 'success')
        return redirect(url_for('teletrabalho.calendario', mes_referencia=mes_referencia))

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao solicitar teletrabalho: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))


@teletrabalho_bp.route('/cancelar/<int:id>', methods=['POST'])
@login_required
def cancelar(id):
    """Cancela um dia de teletrabalho"""
    try:
        teletrabalho = Teletrabalho.query.get_or_404(id)
        usuario = Usuario.query.get(current_user.id)

        # Verificar permissão
        if teletrabalho.USUARIO_ID != current_user.id and not eh_gestor_ou_admin(usuario):
            flash('Você não tem permissão para cancelar este teletrabalho.', 'danger')
            return redirect(url_for('teletrabalho.index'))

        # ✅ VALIDAÇÃO: Gerente só pode excluir da sua área
        nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)

        if nivel_acesso == 'gerente':
            if not usuario.empregado:
                flash('Usuário sem vínculo de empregado.', 'danger')
                return redirect(url_for('teletrabalho.index'))

            area_gerente = usuario.empregado.sgSuperintendencia

            if teletrabalho.AREA != area_gerente:
                flash('Você só pode cancelar teletrabalhos da sua área.', 'danger')
                return redirect(url_for('teletrabalho.index'))

        # Guardar informações para o log
        data_cancelada = teletrabalho.DATA_TELETRABALHO
        usuario_nome = teletrabalho.usuario.NOME if teletrabalho.usuario else 'Usuário desconhecido'

        # ✅ SOFT DELETE
        teletrabalho.DELETED_AT = datetime.utcnow()

        # ✅ COMMIT PRIMEIRO (antes do log)
        db.session.commit()

        # ✅ LOG DEPOIS DO COMMIT (para evitar rollback)
        try:
            registrar_log(
                acao='excluir',
                entidade='teletrabalho',
                entidade_id=id,
                descricao=f'{nivel_acesso.upper()} cancelou teletrabalho de {usuario_nome} em {data_cancelada.strftime("%d/%m/%Y")}'
            )
        except Exception as log_error:
            # Se falhar o log, não importa - o dado já foi salvo
            print(f"[AVISO] Erro ao registrar log (não afeta a operação): {log_error}")

        flash('Teletrabalho cancelado com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao cancelar teletrabalho: {str(e)}', 'danger')

    return redirect(request.referrer or url_for('teletrabalho.index'))


@teletrabalho_bp.route('/relatorio')
@login_required
def relatorio():
    """Relatório de teletrabalho (apenas admin)"""
    try:
        if current_user.perfil != 'admin':
            flash('Você não tem permissão para acessar esta página.', 'danger')
            return redirect(url_for('teletrabalho.index'))

        mes_ref = request.args.get('mes_ref')
        area_filtro = request.args.get('area')

        query = db.session.query(
            Teletrabalho,
            Usuario
        ).join(
            Usuario, Teletrabalho.USUARIO_ID == Usuario.ID
        ).filter(
            Teletrabalho.DELETED_AT.is_(None)
        )

        if mes_ref:
            query = query.filter(Teletrabalho.MES_REFERENCIA == mes_ref)

        if area_filtro:
            query = query.filter(Teletrabalho.AREA == area_filtro)

        resultados = query.order_by(
            Teletrabalho.DATA_TELETRABALHO.desc()
        ).all()

        areas = obter_areas_disponiveis()
        periodos = gerar_opcoes_periodo()

        total_solicitacoes = len(resultados)
        total_aprovadas = sum(1 for t, u in resultados if t.STATUS == 'APROVADO')

        return render_template('teletrabalho/relatorio.html',
                               resultados=resultados,
                               areas=areas,
                               periodos=periodos,
                               total_solicitacoes=total_solicitacoes,
                               total_aprovadas=total_aprovadas,
                               mes_ref=mes_ref,
                               area_filtro=area_filtro)

    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))


# Rotas Admin continuam as mesmas...
@teletrabalho_bp.route('/admin/cadastrar', methods=['GET'])
@login_required
def admin_cadastrar():
    """Formulário para admin/moderador/gerente cadastrar teletrabalho"""
    usuario = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario):
        flash('Acesso negado. Apenas administradores, moderadores e gerentes.', 'danger')
        return redirect(url_for('teletrabalho.index'))

    try:
        # Query base
        query = db.session.query(
            Usuario.ID,
            Usuario.NOME,
            Usuario.EMAIL,
            Empregado.sgSuperintendencia,
            Empregado.dsCargo
        ).join(
            Empregado, Usuario.FK_PESSOA == Empregado.pkPessoa
        ).filter(
            Usuario.ATIVO == True,
            Usuario.DELETED_AT.is_(None),
            Empregado.fkStatus == 1,
            Empregado.sgSuperintendencia.isnot(None)
        )

        # CORREÇÃO: Verificar nível de acesso
        # GERENTE (mesmo se moderador) vê apenas sua área
        # ADMIN e MODERADOR (não gerente) veem tudo
        if not pode_ver_todas_areas(usuario):
            # É gerente (mesmo se for moderador também)
            empregado_logado = usuario.empregado
            if empregado_logado and empregado_logado.sgSuperintendencia:
                query = query.filter(Empregado.sgSuperintendencia == empregado_logado.sgSuperintendencia)
            else:
                flash('Você não está vinculado a nenhuma superintendência.', 'warning')
                return redirect(url_for('teletrabalho.index'))

        usuarios = query.order_by(Usuario.NOME).all()

        lista_usuarios = []
        for u in usuarios:
            lista_usuarios.append({
                'id': u.ID,
                'nome': u.NOME,
                'email': u.EMAIL,
                'area': u.sgSuperintendencia,
                'cargo': u.dsCargo
            })

        periodos = gerar_opcoes_periodo()

        return render_template('teletrabalho/admin_cadastrar.html',
                               usuarios=lista_usuarios,
                               meses_disponiveis=periodos)

    except Exception as e:
        flash(f'Erro ao carregar página: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))


@teletrabalho_bp.route('/admin/obter_calendario/<int:usuario_id>/<mes_referencia>')
@login_required
def admin_obter_calendario(usuario_id, mes_referencia):
    """Retorna dados do calendário para o admin (JSON)"""
    try:
        usuario_logado = Usuario.query.get(current_user.id)
        if not eh_gestor_ou_admin(usuario_logado):
            return jsonify({'erro': 'Acesso negado'}), 403

        usuario = Usuario.query.get_or_404(usuario_id)
        empregado = usuario.empregado

        if not empregado:
            return jsonify({'erro': 'Usuário sem vínculo com empregado'}), 400

        area = empregado.sgSuperintendencia
        tipo_area = 'superintendencia'

        limite, total_pessoas = calcular_limite_area(area, tipo_area)

        ano = int(mes_referencia[:4])
        mes = int(mes_referencia[4:])

        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])
        primeiro_dia_grid = (primeiro_dia.weekday() + 1) % 7

        # ✅ CORREÇÃO: Usar PAR_TB020_CALENDARIO
        calendario_info = Calendario.carregar_info_mes(ano, mes)

        # Buscar teletrabalhos da área no mês
        teletrabalhos = Teletrabalho.query.filter(
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.AREA == area,
            Teletrabalho.TIPO_AREA == tipo_area,
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        # Ocupação por dia e nomes
        dias_ocupados = {}
        for t in teletrabalhos:
            data_str = t.DATA_TELETRABALHO.strftime('%Y-%m-%d')
            if data_str not in dias_ocupados:
                dias_ocupados[data_str] = {'qtd': 0, 'usuarios': []}
            dias_ocupados[data_str]['qtd'] += 1
            if t.usuario:
                dias_ocupados[data_str]['usuarios'].append(t.usuario.NOME)

        # Dias que o USUÁRIO SELECIONADO já marcou
        dias_usuario_marcados = set()
        for t in teletrabalhos:
            if t.USUARIO_ID == usuario_id:
                dias_usuario_marcados.add(t.DATA_TELETRABALHO.strftime('%Y-%m-%d'))

        # Montar dias do calendário
        dias_calendario = []
        data_atual = primeiro_dia

        while data_atual <= ultimo_dia:
            data_str = data_atual.strftime('%Y-%m-%d')

            # ✅ Usar PAR_TB020_CALENDARIO
            info_dia = calendario_info.get(data_atual)
            if info_dia is not None:
                eh_util = info_dia['eh_util']
            else:
                eh_util = Feriado.eh_dia_util(data_atual)

            feriado_obj = Feriado.obter_feriado(data_atual)
            feriado_texto = None
            if feriado_obj:
                feriado_texto = feriado_obj.DS_FERIADO
            elif not eh_util and data_atual.weekday() < 5:
                feriado_texto = 'Ponto Facultativo / Não Útil'

            qtd_pessoas = dias_ocupados.get(data_str, {}).get('qtd', 0)

            dias_calendario.append({
                'data': data_str,
                'dia': data_atual.day,
                'eh_util': eh_util,
                'feriado': feriado_texto,
                'qtd_pessoas': qtd_pessoas,
                'limite': limite,
                'lotado': qtd_pessoas >= limite,
                'ja_marcado': data_str in dias_usuario_marcados,
                'usuarios': dias_ocupados.get(data_str, {}).get('usuarios', [])
            })

            data_atual += timedelta(days=1)

        return jsonify({
            'dias': dias_calendario,
            'area': area,
            'limite': limite,
            'total_pessoas': total_pessoas,
            'dias_marcados': len(dias_usuario_marcados),
            'primeiro_dia_grid': primeiro_dia_grid
        })

    except Exception as e:
        print(f"[ERRO] admin_obter_calendario: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'erro': f'Erro ao carregar calendário: {str(e)}'}), 500


@teletrabalho_bp.route('/admin/cadastrar', methods=['POST'])
@login_required
def admin_cadastrar_post():
    """Admin/Moderador/Gerente cadastra teletrabalho para outro usuário."""
    usuario_logado = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario_logado):
        flash('Acesso negado. Apenas administradores, moderadores e gerentes.', 'danger')
        return redirect(url_for('teletrabalho.index'))

    try:
        usuario_id = int(request.form.get('usuario_id'))
        datas_str = request.form.getlist('datas[]')
        mes_referencia = request.form.get('mes_referencia')
        tipo_periodo = request.form.get('tipo_periodo')
        observacao = request.form.get('observacao', '')

        usuario = Usuario.query.get_or_404(usuario_id)
        empregado = usuario.empregado

        if not empregado:
            flash('Usuário sem vínculo com empregado.', 'danger')
            return redirect(url_for('teletrabalho.admin_cadastrar'))

        area = empregado.sgSuperintendencia
        tipo_area = 'superintendencia'

        limite, total_pessoas = calcular_limite_area(area, tipo_area)

        # ====================================================================
        # ✅ NOVAS REGRAS: limite mensal 6 (2 blocos de 3) ou 5 corridos
        # ====================================================================
        LIMITE_DIAS_MES = 6

        qtd_dias_mes = Teletrabalho.contar_dias_mes(usuario_id, mes_referencia)

        if qtd_dias_mes >= LIMITE_DIAS_MES:
            flash(f'{usuario.NOME} já atingiu o limite de {LIMITE_DIAS_MES} dias neste mês.', 'warning')
            return redirect(url_for('teletrabalho.admin_cadastrar'))

        if len(datas_str) + qtd_dias_mes > LIMITE_DIAS_MES:
            flash(
                f'{usuario.NOME} só pode marcar mais {LIMITE_DIAS_MES - qtd_dias_mes} dia(s) neste mês.',
                'warning'
            )
            return redirect(url_for('teletrabalho.admin_cadastrar'))

        if tipo_periodo not in ('TRES_CONSECUTIVOS', 'CINCO_DIAS_CORRIDOS'):
            flash('Selecione um tipo de período válido (3 consecutivos ou 5 corridos).', 'warning')
            return redirect(url_for('teletrabalho.admin_cadastrar'))

        datas = [datetime.strptime(d, '%Y-%m-%d').date() for d in datas_str]

        if tipo_periodo == 'CINCO_DIAS_CORRIDOS':
            if len(datas) != 5:
                flash('Para "5 dias corridos" selecione exatamente 5 dias.', 'warning')
                return redirect(url_for('teletrabalho.admin_cadastrar'))

            valido, msg = Teletrabalho.validar_cinco_dias_corridos(usuario_id, datas, mes_referencia)
            if not valido:
                flash(msg, 'warning')
                return redirect(url_for('teletrabalho.admin_cadastrar'))

        elif tipo_periodo == 'TRES_CONSECUTIVOS':
            valido, msg = Teletrabalho.validar_tres_consecutivos(usuario_id, datas, mes_referencia)
            if not valido:
                flash(msg, 'warning')
                return redirect(url_for('teletrabalho.admin_cadastrar'))

        for data in datas:
            if not Feriado.eh_dia_util(data):
                feriado = Feriado.obter_feriado(data)
                motivo = feriado.DS_FERIADO if feriado else "fim de semana"
                flash(f'{data.strftime("%d/%m/%Y")} não é dia útil ({motivo}).', 'warning')
                return redirect(url_for('teletrabalho.admin_cadastrar'))

            bloqueado, bloqueio = BloqueioDia.dia_esta_bloqueado(data, area, tipo_area)
            if bloqueado:
                motivo_bloqueio = bloqueio.MOTIVO if bloqueio and bloqueio.MOTIVO else "bloqueado pelo gestor"
                flash(f'{data.strftime("%d/%m/%Y")} está bloqueada: {motivo_bloqueio}', 'danger')
                return redirect(url_for('teletrabalho.admin_cadastrar'))

            qtd_atual = Teletrabalho.contar_pessoas_dia(data, area, tipo_area)
            if qtd_atual >= limite:
                flash(f'{data.strftime("%d/%m/%Y")}: limite de {limite} pessoas atingido.', 'warning')
                return redirect(url_for('teletrabalho.admin_cadastrar'))

        tipo_cadastrador = 'ADMIN' if usuario_logado.PERFIL == 'admin' else (
            'MODERADOR' if usuario_logado.PERFIL == 'moderador' else 'GERENTE'
        )

        for data in datas:
            teletrabalho = Teletrabalho(
                USUARIO_ID=usuario_id,
                DATA_TELETRABALHO=data,
                MES_REFERENCIA=mes_referencia,
                AREA=area,
                TIPO_AREA=tipo_area,
                STATUS='APROVADO',
                TIPO_PERIODO=tipo_periodo,
                TIPO_MARCACAO='TELETRABALHO',
                OBSERVACAO=f'[{tipo_cadastrador}] {observacao}' if observacao else f'[{tipo_cadastrador}] Cadastrado por {tipo_cadastrador.lower()}',
                APROVADO_POR=current_user.id,
                APROVADO_EM=datetime.now()
            )
            db.session.add(teletrabalho)

        registrar_log(
            acao='criar',
            entidade='teletrabalho',
            entidade_id=0,
            descricao=f'{tipo_cadastrador} cadastrou {len(datas)} dia(s) de teletrabalho para {usuario.NOME} ({tipo_periodo})'
        )

        db.session.commit()
        flash(f'{len(datas)} dia(s) cadastrado(s) para {usuario.NOME} com sucesso!', 'success')
        return redirect(url_for('teletrabalho.admin_cadastrar'))

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao cadastrar: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.admin_cadastrar'))


@teletrabalho_bp.route('/config/area', methods=['GET', 'POST'])
@login_required
def config_area():
    """Configuração da área - criar/editar"""
    usuario = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario):
        flash('Acesso negado. Apenas gestores podem configurar.', 'danger')
        return redirect(url_for('teletrabalho.index'))

    try:
        empregado = usuario.empregado

        # Admin pode escolher área, gerente só sua área
        if current_user.perfil == 'admin':
            area_param = request.args.get('area')
            tipo_area_param = request.args.get('tipo_area', 'superintendencia')

            if area_param:
                area = area_param
                tipo_area = tipo_area_param
            elif empregado:
                area = empregado.sgSuperintendencia
                tipo_area = 'superintendencia'
            else:
                # Admin sem empregado - mostrar seletor
                areas = obter_areas_disponiveis()
                return render_template('teletrabalho/config_area.html',
                                       config=None,
                                       area=None,
                                       tipo_area=None,
                                       areas_disponiveis=areas,
                                       mostrar_seletor=True)
        else:
            # Gerente/Moderador - só sua área
            if not empregado:
                flash('Usuário sem vínculo com empregado.', 'danger')
                return redirect(url_for('teletrabalho.index'))

            area = empregado.sgSuperintendencia
            tipo_area = 'superintendencia'

        if request.method == 'POST':
            qtd_pessoas = int(request.form.get('qtd_pessoas'))
            percentual = float(request.form.get('percentual', 30.0))

            if qtd_pessoas < 1:
                flash('Quantidade deve ser maior que zero.', 'warning')
                return redirect(url_for('teletrabalho.config_area', area=area, tipo_area=tipo_area))

            if percentual < 1 or percentual > 100:
                flash('Percentual deve estar entre 1 e 100.', 'warning')
                return redirect(url_for('teletrabalho.config_area', area=area, tipo_area=tipo_area))

            # Buscar config existente
            from app.models.teletrabalho import ConfigAreaTeletrabalho

            config = ConfigAreaTeletrabalho.query.filter_by(
                AREA=area,
                TIPO_AREA=tipo_area,
                DELETED_AT=None
            ).first()

            if config:
                # Atualizar
                config.QTD_TOTAL_PESSOAS = qtd_pessoas
                config.PERCENTUAL_LIMITE = percentual
                config.UPDATED_AT = datetime.utcnow()
                acao = 'atualizada'
            else:
                # Criar
                config = ConfigAreaTeletrabalho(
                    AREA=area,
                    TIPO_AREA=tipo_area,
                    QTD_TOTAL_PESSOAS=qtd_pessoas,
                    PERCENTUAL_LIMITE=percentual,
                    GESTOR_ID=current_user.id
                )
                db.session.add(config)
                acao = 'criada'

            registrar_log(
                acao='editar' if acao == 'atualizada' else 'criar',
                entidade='config_area_teletrabalho',
                entidade_id=config.ID if config.ID else 0,
                descricao=f'Configuração da área {area} {acao}: {qtd_pessoas} pessoas, {percentual}%'
            )

            db.session.commit()
            flash(f'Configuração {acao} com sucesso!', 'success')
            return redirect(url_for('teletrabalho.config_area', area=area, tipo_area=tipo_area))

        # GET - buscar config
        from app.models.teletrabalho import ConfigAreaTeletrabalho

        config = ConfigAreaTeletrabalho.query.filter_by(
            AREA=area,
            TIPO_AREA=tipo_area,
            DELETED_AT=None
        ).first()

        # Contar pessoas reais na área
        qtd_real = Empregado.query.filter(
            Empregado.sgSuperintendencia == area,
            Empregado.fkStatus == 1
        ).count()

        return render_template('teletrabalho/config_area.html',
                               config=config,
                               area=area,
                               tipo_area=tipo_area,
                               qtd_real=qtd_real,
                               areas_disponiveis=None,
                               mostrar_seletor=False)

    except Exception as e:
        db.session.rollback()
        flash(f'Erro: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))


@teletrabalho_bp.route('/config/listar')
@login_required
def config_listar():
    """Lista todas as configurações de áreas"""
    if current_user.perfil != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('teletrabalho.index'))

    try:
        from app.models.teletrabalho import ConfigAreaTeletrabalho

        configs = ConfigAreaTeletrabalho.query.filter(
            ConfigAreaTeletrabalho.DELETED_AT.is_(None)
        ).order_by(ConfigAreaTeletrabalho.AREA).all()

        # Buscar áreas sem config
        areas_com_config = [c.AREA for c in configs]

        todas_areas = db.session.query(
            Empregado.sgSuperintendencia,
            func.count(Empregado.pkPessoa).label('qtd')
        ).filter(
            Empregado.sgSuperintendencia.isnot(None),
            Empregado.sgSuperintendencia != '',
            Empregado.fkStatus == 1
        ).group_by(Empregado.sgSuperintendencia).all()

        areas_sem_config = [
            {'area': area, 'qtd': qtd}
            for area, qtd in todas_areas
            if area not in areas_com_config
        ]

        return render_template('teletrabalho/config_listar.html',
                               configs=configs,
                               areas_sem_config=areas_sem_config)

    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))


@teletrabalho_bp.route('/config/excluir/<int:id>', methods=['POST'])
@login_required
def config_excluir(id):
    """Excluir configuração de área"""
    if current_user.perfil != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('teletrabalho.index'))

    try:
        from app.models.teletrabalho import ConfigAreaTeletrabalho

        config = ConfigAreaTeletrabalho.query.get_or_404(id)
        config.DELETED_AT = datetime.utcnow()

        registrar_log(
            acao='excluir',
            entidade='config_area_teletrabalho',
            entidade_id=id,
            descricao=f'Configuração da área {config.AREA} excluída'
        )

        db.session.commit()
        flash('Configuração excluída.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro: {str(e)}', 'danger')

    return redirect(url_for('teletrabalho.config_listar'))


@teletrabalho_bp.route('/admin/sortear', methods=['GET'])
@login_required
def admin_sortear_form():
    """Formulário para sortear teletrabalho automaticamente"""
    usuario = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario):
        flash('Acesso negado. Apenas gestores e admins.', 'danger')
        return redirect(url_for('teletrabalho.index'))

    try:
        # CORREÇÃO: Gerente vê apenas sua área, Admin/Moderador não-gerente veem tudo
        if pode_ver_todas_areas(usuario):
            # Admin ou Moderador (não gerente) - vê todas
            areas = db.session.query(
                Empregado.sgSuperintendencia,
                func.count(Empregado.pkPessoa).label('qtd')
            ).filter(
                Empregado.sgSuperintendencia.isnot(None),
                Empregado.fkStatus == 1
            ).group_by(Empregado.sgSuperintendencia).all()
        else:
            # Gerente (mesmo se for moderador também) - vê apenas sua área
            if usuario.empregado and usuario.empregado.sgSuperintendencia:
                area = usuario.empregado.sgSuperintendencia
                qtd = Empregado.query.filter(
                    Empregado.sgSuperintendencia == area,
                    Empregado.fkStatus == 1
                ).count()
                areas = [(area, qtd)]
            else:
                flash('Você não está vinculado a nenhuma área.', 'warning')
                return redirect(url_for('teletrabalho.index'))

        areas_disponiveis = [{'area': a[0], 'qtd': a[1]} for a in areas]
        periodos = gerar_opcoes_periodo()

        return render_template('teletrabalho/admin_sortear.html',
                               areas=areas_disponiveis,
                               periodos=periodos)

    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))

@teletrabalho_bp.route('/admin/sortear', methods=['POST'])
@login_required
def admin_sortear_processar():
    """Processa o sorteio automático de teletrabalho"""
    usuario = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('teletrabalho.index'))

    try:
        area = request.form.get('area')
        mes_referencia = request.form.get('mes_referencia')
        qtd_dias = int(request.form.get('qtd_dias', 6))

        if not area or not mes_referencia:
            flash('Preencha todos os campos.', 'warning')
            return redirect(url_for('teletrabalho.admin_sortear_form'))

        resultado = sortear_teletrabalho_automatico(area, mes_referencia, qtd_dias, usuario.ID)

        if resultado['sucesso']:
            msg = (
                f"✅ Sorteio concluído! "
                f"{resultado['total_sorteados']}/{resultado['total_elegiveis']} "
                f"pessoas sorteadas. Total: {resultado['total_dias']} dias."
            )

            if resultado.get('removidos_anterior', 0) > 0:
                msg += f" | Anterior removido: {resultado['removidos_anterior']} registros."

            if resultado.get('excluidos_cargo', 0) > 0:
                msg += f" | Excluídos (cargo/gerente): {resultado['excluidos_cargo']}."

            flash(msg, 'success')

            for aviso in resultado.get('avisos_ferias', []):
                flash(f"⚠️ {aviso}", 'warning')

            registrar_log(
                acao='criar',
                entidade='sorteio_teletrabalho',
                entidade_id=0,
                descricao=(
                    f'Sorteio automático: {area} - {mes_referencia} - '
                    f'{resultado["total_sorteados"]} pessoas - '
                    f'{resultado["total_dias"]} dias'
                )
            )
        else:
            flash(f"❌ Erro no sorteio: {resultado['erro']}", 'danger')

        return redirect(url_for('teletrabalho.admin_sortear_form'))

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao processar sorteio: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.admin_sortear_form'))


def sortear_teletrabalho_automatico(area, mes_referencia, qtd_dias_por_pessoa, admin_id):
    """
    Sorteia teletrabalho com as regras atuais.

    REGRAS:
    - Todos os usuários elegíveis recebem 2 blocos de até 3 dias úteis
      consecutivos em semanas diferentes (total máximo: 6 dias no mês).
    - Gerentes são excluídos do sorteio (só cadastro manual).
    - Estagiários e Superintendentes são excluídos.
    - Respeita o limite diário de 30% (calculado pela config da área).
    - Usa tabela PAR_TB020_CALENDARIO como fonte definitiva de dias úteis.
    - Se repetir o sorteio, apaga o anterior antes.
    - Round-robin: 1º bloco para todos, depois 2º bloco para todos.
    - Fallback: tenta 3 dias, depois 2, depois 1 por bloco.
    """
    import random
    from datetime import datetime, timedelta

    try:
        ano = int(mes_referencia[:4])
        mes = int(mes_referencia[4:])

        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])

        # ====================================================================
        # CARREGAR DIAS ÚTEIS DA TABELA PAR_TB020_CALENDARIO
        # ====================================================================
        dias_uteis_set, dias_nao_uteis_set = Calendario.carregar_dias_uteis_mes(ano, mes)

        if not dias_uteis_set:
            print(f"[SORTEIO] AVISO: PAR_TB020_CALENDARIO sem dados para {mes}/{ano}. Usando fallback.")
            dias_uteis_set = set()
            data_check = primeiro_dia
            while data_check <= ultimo_dia:
                if Feriado.eh_dia_util(data_check):
                    dias_uteis_set.add(data_check)
                data_check += timedelta(days=1)

        print(f"[SORTEIO] Dias úteis no mês {mes}/{ano}: {len(dias_uteis_set)}")

        # Log dos feriados que caem em dia de semana
        nomes_wd = {0: 'SEG', 1: 'TER', 2: 'QUA', 3: 'QUI', 4: 'SEX', 5: 'SAB', 6: 'DOM'}
        if dias_nao_uteis_set:
            for d in sorted(dias_nao_uteis_set):
                if d.weekday() < 5:
                    print(f"[SORTEIO] FERIADO: {d.strftime('%d/%m/%Y')} ({nomes_wd.get(d.weekday())})")

        # ====================================================================
        # APAGAR SORTEIO ANTERIOR (soft delete)
        # ====================================================================
        registros_antigos = Teletrabalho.query.filter(
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.AREA == area,
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.OBSERVACAO.like('[[]SORTEIO_AUTO]%'),
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        registros_antigos_v1 = Teletrabalho.query.filter(
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.AREA == area,
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.OBSERVACAO.like('[[]SORTEIO AUTO]%'),
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        todos_antigos = registros_antigos + registros_antigos_v1
        qtd_removidos = len(todos_antigos)

        for reg in todos_antigos:
            reg.DELETED_AT = datetime.now()
            reg.STATUS = 'CANCELADO'

        if qtd_removidos > 0:
            db.session.flush()
            print(f"[SORTEIO] Removidos {qtd_removidos} registros do sorteio anterior")

        # ====================================================================
        # BUSCAR TODAS AS PESSOAS DA ÁREA
        # ====================================================================
        usuarios_query = db.session.query(Usuario, Empregado).join(
            Empregado, Usuario.FK_PESSOA == Empregado.pkPessoa
        ).filter(
            Usuario.ATIVO == True,
            Usuario.DELETED_AT.is_(None),
            Empregado.fkStatus == 1,
            Empregado.sgSuperintendencia == area
        ).all()

        # ====================================================================
        # CLASSIFICAR CADA PESSOA
        # Log detalhado para diagnóstico de quem entra e quem sai
        # ====================================================================
        usuarios_elegiveis = []
        excluidos_cargo = []

        print(f"[SORTEIO] === CLASSIFICAÇÃO DE PESSOAS ({len(usuarios_query)} encontradas na área {area}) ===")

        for usuario, empregado in usuarios_query:
            cargo = (empregado.dsCargo or '').strip().upper()
            nome = usuario.NOME

            # ✅ REGRA 1: Excluir ESTAGIÁRIO (verificação exata no início do cargo)
            if 'ESTAGIARIO' in cargo or 'ESTAGIÁRIO' in cargo:
                motivo = f"Cargo excluído: {cargo}"
                excluidos_cargo.append(f"{nome} ({cargo}) - {motivo}")
                print(f"[SORTEIO]   ❌ {nome} | Cargo: {cargo} | Motivo: {motivo}")
                continue

            # ✅ REGRA 2: Excluir SUPERINTENDENTE
            if 'SUPERINTENDENTE' in cargo:
                motivo = f"Cargo excluído: {cargo}"
                excluidos_cargo.append(f"{nome} ({cargo}) - {motivo}")
                print(f"[SORTEIO]   ❌ {nome} | Cargo: {cargo} | Motivo: {motivo}")
                continue

            # ✅ REGRA 3: Excluir GERENTE (só o cargo que COMEÇA com GERENTE)
            # Isso evita excluir "ANALISTA DA GERÊNCIA" ou "SUBGERENTE"
            # Cargos excluídos: "GERENTE", "GERENTE DE ...", "GERENTE-EXECUTIVO"
            # Cargos NÃO excluídos: "SUBGERENTE", "ANALISTA DA GERÊNCIA"
            if cargo.startswith('GERENTE'):
                motivo = "Gerente (só cadastro manual)"
                excluidos_cargo.append(f"{nome} ({cargo}) - {motivo}")
                print(f"[SORTEIO]   ⚠️ {nome} | Cargo: {cargo} | Motivo: {motivo}")
                continue

            # ✅ ELEGÍVEL
            usuarios_elegiveis.append(usuario)
            print(f"[SORTEIO]   ✅ {nome} | Cargo: {cargo} | ELEGÍVEL")

        print(f"[SORTEIO] Resultado: {len(usuarios_elegiveis)} elegíveis, {len(excluidos_cargo)} excluídos")

        if not usuarios_elegiveis:
            return {'sucesso': False, 'erro': 'Nenhum usuário elegível encontrado'}

        # ====================================================================
        # BUSCAR FÉRIAS DO MÊS
        # ====================================================================
        ferias_mes = {}
        ferias_query = Teletrabalho.query.filter(
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.AREA == area,
            Teletrabalho.TIPO_MARCACAO == 'FERIAS',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()
        for feria in ferias_query:
            ferias_mes.setdefault(feria.USUARIO_ID, []).append(feria.DATA_TELETRABALHO)

        if ferias_mes:
            print(f"[SORTEIO] Pessoas com férias no mês: {len(ferias_mes)}")

        # ====================================================================
        # BUSCAR BLOQUEIOS DO MÊS
        # ====================================================================
        bloqueios_mes = []
        bloqueios_query = BloqueioDia.query.filter(
            BloqueioDia.MES_REFERENCIA == mes_referencia,
            BloqueioDia.AREA == area,
            BloqueioDia.DELETED_AT.is_(None)
        ).all()
        for bloqueio in bloqueios_query:
            bloqueios_mes.append(bloqueio.DATA_BLOQUEIO)

        print(f"[SORTEIO] Bloqueios: {len(bloqueios_mes)}")

        # ====================================================================
        # ORGANIZAR DIAS ÚTEIS POR SEMANA ISO
        # ====================================================================
        semanas = {}
        data_atual = primeiro_dia
        while data_atual <= ultimo_dia:
            if data_atual in dias_uteis_set and data_atual not in bloqueios_mes:
                ano_iso, semana_iso, _ = data_atual.isocalendar()
                chave = (ano_iso, semana_iso)
                dia_semana = data_atual.weekday()
                semanas.setdefault(chave, {}).setdefault(dia_semana, []).append(data_atual)
            data_atual += timedelta(days=1)

        # Log das semanas
        for chave_sem, dias_sem in sorted(semanas.items()):
            dias_str = ', '.join([
                f"{nomes_wd.get(wd, '?')}({d.strftime('%d/%m')})"
                for wd, datas in sorted(dias_sem.items())
                for d in datas
            ])
            print(f"[SORTEIO] Semana ISO {chave_sem}: {dias_str}")

        # ====================================================================
        # LIMITE DIÁRIO: USA calcular_limite_area() (config da área)
        # ====================================================================
        limite_diario, total_pessoas_area = calcular_limite_area(area, 'superintendencia')
        print(f"[SORTEIO] Limite diário: {limite_diario} (config da área: {total_pessoas_area} pessoas)")

        # ====================================================================
        # OCUPAÇÃO: INICIALIZAR COM CADASTROS MANUAIS JÁ EXISTENTES
        # ====================================================================
        ocupacao_por_dia = {}
        data_atual = primeiro_dia
        while data_atual <= ultimo_dia:
            if data_atual in dias_uteis_set and data_atual not in bloqueios_mes:
                ocupacao_por_dia[data_atual] = 0
            data_atual += timedelta(days=1)

        cadastros_manuais = Teletrabalho.query.filter(
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.AREA == area,
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None),
            ~Teletrabalho.OBSERVACAO.like('[[]SORTEIO_AUTO]%'),
            ~Teletrabalho.OBSERVACAO.like('[[]SORTEIO AUTO]%')
        ).all()

        for cadastro in cadastros_manuais:
            if cadastro.DATA_TELETRABALHO in ocupacao_por_dia:
                ocupacao_por_dia[cadastro.DATA_TELETRABALHO] += 1

        dias_com_manual = {k: v for k, v in ocupacao_por_dia.items() if v > 0}
        if dias_com_manual:
            print(f"[SORTEIO] Cadastros manuais pré-existentes em {len(dias_com_manual)} dia(s):")
            for d_manual, q_manual in sorted(dias_com_manual.items()):
                print(f"  {d_manual.strftime('%d/%m/%Y')}: {q_manual}/{limite_diario}")

        # ====================================================================
        # ROUND-ROBIN: 2 PASSADAS
        # Passo 1: 1 bloco para CADA pessoa (embaralhado)
        # Passo 2: 2º bloco para CADA pessoa (re-embaralhado, semana diferente)
        # ====================================================================
        dias_por_usuario = {}
        semana_por_usuario = {}
        avisos = []

        # ---- PASSO 1: 1º bloco para cada ----
        ordem_passo1 = list(usuarios_elegiveis)
        random.shuffle(ordem_passo1)

        print(f"[SORTEIO] === PASSO 1: 1º bloco para {len(ordem_passo1)} pessoas ===")

        for usuario in ordem_passo1:
            dias_ferias_usuario = ferias_mes.get(usuario.ID, [])
            dias_excluir = list(set(dias_ferias_usuario + bloqueios_mes))

            bloco, chave_semana = sortear_um_bloco_consecutivo(
                semanas, ocupacao_por_dia, limite_diario,
                excluir_datas=dias_excluir,
                semanas_usadas=[]
            )

            if bloco:
                dias_por_usuario[usuario.ID] = list(bloco)
                semana_por_usuario[usuario.ID] = [chave_semana]
                for dia in bloco:
                    ocupacao_por_dia[dia] = ocupacao_por_dia.get(dia, 0) + 1
                dias_str = ', '.join([d.strftime('%d/%m') for d in bloco])
                print(f"[SORTEIO] P1 ✅ {usuario.NOME}: {len(bloco)} dia(s) [{dias_str}] semana {chave_semana}")
            else:
                dias_por_usuario[usuario.ID] = []
                semana_por_usuario[usuario.ID] = []
                avisos.append(f"{usuario.NOME}: Nenhum bloco disponível no 1º passo")
                print(f"[SORTEIO] P1 ❌ {usuario.NOME}: SEM BLOCO DISPONÍVEL")

        # ---- PASSO 2: 2º bloco para cada (semana diferente) ----
        ordem_passo2 = list(usuarios_elegiveis)
        random.shuffle(ordem_passo2)

        print(f"[SORTEIO] === PASSO 2: 2º bloco para {len(ordem_passo2)} pessoas ===")

        for usuario in ordem_passo2:
            dias_ferias_usuario = ferias_mes.get(usuario.ID, [])
            dias_excluir = list(set(dias_ferias_usuario + bloqueios_mes))

            semanas_ja_usadas = semana_por_usuario.get(usuario.ID, [])

            bloco, chave_semana = sortear_um_bloco_consecutivo(
                semanas, ocupacao_por_dia, limite_diario,
                excluir_datas=dias_excluir,
                semanas_usadas=semanas_ja_usadas
            )

            if bloco:
                dias_por_usuario[usuario.ID].extend(bloco)
                semana_por_usuario[usuario.ID].append(chave_semana)
                for dia in bloco:
                    ocupacao_por_dia[dia] = ocupacao_por_dia.get(dia, 0) + 1
                dias_str = ', '.join([d.strftime('%d/%m') for d in bloco])
                print(f"[SORTEIO] P2 ✅ {usuario.NOME}: +{len(bloco)} dia(s) [{dias_str}] semana {chave_semana}")
            else:
                qtd_atual = len(dias_por_usuario.get(usuario.ID, []))
                avisos.append(
                    f"{usuario.NOME}: {qtd_atual} dia(s) total (sem semana para 2º bloco)"
                )
                print(f"[SORTEIO] P2 ⚠️ {usuario.NOME}: SEM 2º BLOCO (ficou com {qtd_atual} dias)")

        # ====================================================================
        # SALVAR NO BANCO
        # ====================================================================
        total_dias_cadastrados = 0
        pessoas_com_dias = 0

        for usuario_id, dias in dias_por_usuario.items():
            if not dias:
                continue
            pessoas_com_dias += 1
            for dia in dias:
                teletrabalho = Teletrabalho(
                    USUARIO_ID=usuario_id,
                    DATA_TELETRABALHO=dia,
                    MES_REFERENCIA=mes_referencia,
                    AREA=area,
                    TIPO_AREA='superintendencia',
                    STATUS='APROVADO',
                    TIPO_PERIODO='TRES_CONSECUTIVOS',
                    TIPO_MARCACAO='TELETRABALHO',
                    OBSERVACAO=f'[SORTEIO_AUTO] {datetime.now().strftime("%d/%m/%Y %H:%M")}',
                    APROVADO_POR=admin_id,
                    APROVADO_EM=datetime.now()
                )
                db.session.add(teletrabalho)
                total_dias_cadastrados += 1

        db.session.commit()

        # ====================================================================
        # LOG FINAL: OCUPAÇÃO POR DIA E DISTRIBUIÇÃO POR PESSOA
        # ====================================================================
        print(f"[SORTEIO] === OCUPAÇÃO FINAL POR DIA ===")
        for dia_ocp in sorted(ocupacao_por_dia.keys()):
            qtd = ocupacao_por_dia[dia_ocp]
            if qtd > 0:
                wd_nome = nomes_wd.get(dia_ocp.weekday(), '?')
                status = "⚠️ LOTADO" if qtd >= limite_diario else ""
                print(f"  {dia_ocp.strftime('%d/%m/%Y')} ({wd_nome}): {qtd}/{limite_diario} {status}")

        print(f"[SORTEIO] === DISTRIBUIÇÃO POR PESSOA ===")
        for uid, dias in sorted(dias_por_usuario.items(), key=lambda x: len(x[1]), reverse=True):
            nome_user = next((u.NOME for u in usuarios_elegiveis if u.ID == uid), f"ID={uid}")
            dias_str = ', '.join([d.strftime('%d/%m') for d in sorted(dias)])
            print(f"  {nome_user}: {len(dias)} dia(s) [{dias_str}]")

        return {
            'sucesso': True,
            'total_sorteados': pessoas_com_dias,
            'total_dias': total_dias_cadastrados,
            'total_elegiveis': len(usuarios_elegiveis),
            'excluidos_cargo': len(excluidos_cargo),
            'removidos_anterior': qtd_removidos,
            'avisos_ferias': avisos
        }

    except Exception as e:
        db.session.rollback()
        print(f"[SORTEIO] ERRO: {e}")
        import traceback
        traceback.print_exc()
        return {'sucesso': False, 'erro': str(e)}


def sortear_um_bloco_consecutivo(semanas, ocupacao, limite, excluir_datas=None, semanas_usadas=None):
    """
    Sorteia UM bloco de até 3 dias úteis consecutivos em uma semana disponível.

    LÓGICA COM FALLBACK:
    1. Tenta alocar 3 dias consecutivos (SEG-TER-QUA, TER-QUA-QUI, QUA-QUI-SEX)
    2. Se não conseguir 3, tenta 2 dias consecutivos (SEG-TER, TER-QUA, QUA-QUI, QUI-SEX)
    3. Se não conseguir 2, tenta 1 dia isolado (qualquer dia útil da semana)

    :param semanas: dict {(ano_iso, semana_iso): {weekday: [datas]}}
    :param ocupacao: dict {data: qtd_pessoas_ja_alocadas}
    :param limite: int (limite diário 30%)
    :param excluir_datas: lista de datas a excluir (férias, bloqueios)
    :param semanas_usadas: lista de chaves de semana já usadas por esse usuário
    :return: (lista_de_datas_do_bloco, chave_semana) ou ([], None)
    """
    import random

    if excluir_datas is None:
        excluir_datas = []
    if semanas_usadas is None:
        semanas_usadas = []

    padroes_3 = [
        [0, 1, 2],  # SEG-TER-QUA
        [1, 2, 3],  # TER-QUA-QUI
        [2, 3, 4],  # QUA-QUI-SEX
    ]
    padroes_2 = [
        [0, 1],  # SEG-TER
        [1, 2],  # TER-QUA
        [2, 3],  # QUA-QUI
        [3, 4],  # QUI-SEX
    ]
    padroes_1 = [
        [0],  # SEG
        [1],  # TER
        [2],  # QUA
        [3],  # QUI
        [4],  # SEX
    ]

    semanas_candidatas = [s for s in semanas.keys() if s not in semanas_usadas]
    random.shuffle(semanas_candidatas)

    for chave_semana in semanas_candidatas:
        dias_da_semana = semanas[chave_semana]

        for grupo_padroes in [padroes_3, padroes_2, padroes_1]:
            padroes_embaralhados = grupo_padroes.copy()
            random.shuffle(padroes_embaralhados)

            for padrao in padroes_embaralhados:
                dias_bloco = []
                valido = True

                for wd in padrao:
                    candidatos = dias_da_semana.get(wd, [])
                    candidatos_ok = [
                        d for d in candidatos
                        if ocupacao.get(d, 0) < limite
                        and d not in excluir_datas
                    ]
                    if not candidatos_ok:
                        valido = False
                        break
                    dias_bloco.append(candidatos_ok[0])

                if valido and len(dias_bloco) == len(padrao):
                    return sorted(dias_bloco), chave_semana

    return [], None

def sortear_dois_blocos_tres_consecutivos(semanas, ocupacao, limite, excluir_datas=None):
    """
    Sorteia até 2 blocos de 3 dias úteis consecutivos em semanas ISO diferentes.

    Cada bloco segue um dos padrões de weekday:
      - [0, 1, 2] SEG-TER-QUA
      - [1, 2, 3] TER-QUA-QUI
      - [2, 3, 4] QUA-QUI-SEX

    :param semanas: dict {(ano_iso, semana_iso): {weekday: [datas]}}
    :param ocupacao: dict {data: qtd_pessoas_ja_alocadas}
    :param limite: int (limite diário de pessoas)
    :param excluir_datas: lista de datas a excluir (férias, bloqueios do usuário)
    :return: lista ordenada de até 6 datas, ou [] se não conseguir nada
    """
    import random

    if excluir_datas is None:
        excluir_datas = []

    padroes_bloco = [
        [0, 1, 2],  # SEG-TER-QUA
        [1, 2, 3],  # TER-QUA-QUI
        [2, 3, 4],  # QUA-QUI-SEX
    ]

    semanas_disponiveis = list(semanas.keys())
    random.shuffle(semanas_disponiveis)

    blocos_escolhidos = []

    for chave_semana in semanas_disponiveis:
        if len(blocos_escolhidos) >= 2:
            break

        dias_da_semana = semanas[chave_semana]

        padroes_embaralhados = padroes_bloco.copy()
        random.shuffle(padroes_embaralhados)

        for padrao in padroes_embaralhados:
            dias_bloco = []
            valido = True

            for wd in padrao:
                candidatos = dias_da_semana.get(wd, [])
                candidatos_ok = [
                    d for d in candidatos
                    if ocupacao.get(d, 0) < limite
                    and d not in excluir_datas
                ]
                if not candidatos_ok:
                    valido = False
                    break
                dias_bloco.append(candidatos_ok[0])

            if valido and len(dias_bloco) == 3:
                blocos_escolhidos.append(dias_bloco)
                break

    if not blocos_escolhidos:
        return []

    todos_dias = []
    for bloco in blocos_escolhidos:
        todos_dias.extend(bloco)

    return sorted(todos_dias)



@teletrabalho_bp.route('/visualizacao-geral')
@login_required
def visualizacao_geral():
    """Visualização consolidada de teletrabalho da área"""
    try:
        usuario = Usuario.query.get(current_user.id)

        # Pegar parâmetros
        mes_ref = request.args.get('mes_ref')
        area_filtro = request.args.get('area')

        # Se não passou mês, usa o atual
        if not mes_ref:
            hoje = date.today()
            mes_ref = f"{hoje.year}{hoje.month:02d}"

        # DEFINIR ÁREA baseado no nível de acesso
        nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)

        if nivel_acesso == 'gerente':
            if usuario.empregado and usuario.empregado.sgSuperintendencia:
                area = usuario.empregado.sgSuperintendencia
                pode_trocar_area = False
            else:
                flash('Você não está vinculado a nenhuma área.', 'warning')
                return redirect(url_for('teletrabalho.index'))

        elif nivel_acesso in ['admin', 'moderador']:
            if area_filtro:
                area = area_filtro
            else:
                primeira_area = db.session.query(Empregado.sgSuperintendencia).filter(
                    Empregado.sgSuperintendencia.isnot(None),
                    Empregado.fkStatus == 1
                ).distinct().first()
                area = primeira_area[0] if primeira_area else None
            pode_trocar_area = True

        else:
            if usuario.empregado and usuario.empregado.sgSuperintendencia:
                area = usuario.empregado.sgSuperintendencia
                pode_trocar_area = False
            else:
                flash('Você não está vinculado a nenhuma área.', 'warning')
                return redirect(url_for('teletrabalho.index'))

        if not area:
            flash('Nenhuma área disponível.', 'warning')
            return redirect(url_for('teletrabalho.index'))

        # BUSCAR DADOS DO MÊS
        ano = int(mes_ref[:4])
        mes = int(mes_ref[4:])

        # ====================================================================
        # ✅ CORREÇÃO: CARREGAR DIAS ÚTEIS DA TABELA PAR_TB020_CALENDARIO
        # Uma única query para o mês inteiro
        # ====================================================================
        calendario_info = Calendario.carregar_info_mes(ano, mes)

        teletrabalhos = Teletrabalho.query.join(
            Usuario, Teletrabalho.USUARIO_ID == Usuario.ID
        ).filter(
            Teletrabalho.MES_REFERENCIA == mes_ref,
            Teletrabalho.AREA == area,
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        # ORGANIZAR POR DIA
        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])

        dias_organizados = {}
        data_atual = primeiro_dia

        while data_atual <= ultimo_dia:
            dia_str = data_atual.strftime('%Y-%m-%d')

            # ✅ Verificar dia útil pela tabela PAR_TB020_CALENDARIO
            info_dia = calendario_info.get(data_atual)
            if info_dia is not None:
                eh_util = info_dia['eh_util']
            else:
                # Fallback se a data não existe na tabela calendario
                eh_util = Feriado.eh_dia_util(data_atual)

            # Buscar nome do feriado (para exibição) na tabela AUX_TB003_FERIADOS
            feriado = Feriado.obter_feriado(data_atual)
            dia_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo'][data_atual.weekday()]

            # Buscar pessoas neste dia
            pessoas_dia = []
            for t in teletrabalhos:
                if t.DATA_TELETRABALHO == data_atual:
                    usuario_tele = Usuario.query.get(t.USUARIO_ID)
                    if usuario_tele:
                        pessoas_dia.append({
                            'nome': usuario_tele.NOME,
                            'tipo_periodo': t.TIPO_PERIODO,
                            'observacao': t.OBSERVACAO
                        })

            # ✅ Se não é útil e não tem feriado cadastrado, mostra "Não útil"
            feriado_texto = None
            if not eh_util:
                if feriado:
                    feriado_texto = feriado.DS_FERIADO
                elif data_atual.weekday() < 5:
                    # É dia de semana mas não é útil (ponto facultativo, etc.)
                    feriado_texto = 'Ponto Facultativo / Não Útil'

            dias_organizados[dia_str] = {
                'data': data_atual,
                'dia_semana': dia_semana,
                'eh_util': eh_util,
                'feriado': feriado_texto,
                'qtd_pessoas': len(pessoas_dia),
                'pessoas': pessoas_dia
            }

            data_atual += timedelta(days=1)

        # CALCULAR ESTATÍSTICAS
        total_dias_com_pessoas = sum(1 for d in dias_organizados.values() if d['qtd_pessoas'] > 0)
        pessoas_unicas = set()
        for t in teletrabalhos:
            pessoas_unicas.add(t.USUARIO_ID)

        # BUSCAR TODAS AS ÁREAS (para filtro de admin/moderador)
        if pode_trocar_area:
            areas_disponiveis = db.session.query(
                Empregado.sgSuperintendencia,
                func.count(Empregado.pkPessoa).label('qtd')
            ).filter(
                Empregado.sgSuperintendencia.isnot(None),
                Empregado.fkStatus == 1
            ).group_by(Empregado.sgSuperintendencia).all()
        else:
            areas_disponiveis = []

        # PERÍODOS DISPONÍVEIS
        periodos = gerar_opcoes_periodo()

        # Nome do mês
        meses_nomes = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                       'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        nome_mes = f"{meses_nomes[mes - 1]} de {ano}"

        return render_template('teletrabalho/visualizacao_geral.html',
                               dias_organizados=dias_organizados,
                               nome_mes=nome_mes,
                               area=area,
                               mes_referencia=mes_ref,
                               total_dias_com_pessoas=total_dias_com_pessoas,
                               total_pessoas_unicas=len(pessoas_unicas),
                               total_agendamentos=len(teletrabalhos),
                               pode_trocar_area=pode_trocar_area,
                               areas_disponiveis=areas_disponiveis,
                               periodos=periodos)

    except Exception as e:
        flash(f'Erro ao carregar visualização: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))


@teletrabalho_bp.route('/calendario-visual')
@login_required
def calendario_visual():
    """Calendário visual com drag-and-drop para gestores"""
    try:
        usuario = Usuario.query.get(current_user.id)
        empregado = usuario.empregado

        nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)

        mes_ref = request.args.get('mes_ref')
        area_filtro = request.args.get('area')
        tipo_area_filtro = request.args.get('tipo_area', 'superintendencia')

        if not mes_ref:
            hoje = date.today()
            mes_ref = hoje.strftime('%Y%m')

        if nivel_acesso == 'gerente':
            if not empregado:
                flash('Gerente sem vínculo de empregado.', 'warning')
                return redirect(url_for('teletrabalho.index'))
            area_filtro = empregado.sgSuperintendencia
            tipo_area_filtro = 'superintendencia'
            pode_escolher_area = False

        elif nivel_acesso in ['admin', 'moderador']:
            pode_escolher_area = True
            if not area_filtro:
                if empregado:
                    area_filtro = empregado.sgSuperintendencia
                    tipo_area_filtro = 'superintendencia'
                else:
                    primeira_area = db.session.query(
                        Empregado.sgSuperintendencia
                    ).filter(
                        Empregado.sgSuperintendencia.isnot(None),
                        Empregado.fkStatus == 1
                    ).first()
                    if primeira_area:
                        area_filtro = primeira_area[0]
                        tipo_area_filtro = 'superintendencia'
        else:
            flash('Você não tem permissão para acessar esta página.', 'danger')
            return redirect(url_for('teletrabalho.index'))

        if not area_filtro:
            flash('Área não identificada.', 'warning')
            return redirect(url_for('teletrabalho.index'))

        ano = int(mes_ref[:4])
        mes = int(mes_ref[4:])

        meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        nome_mes = f"{meses_pt[mes - 1]} de {ano}"

        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])
        primeiro_dia_grid = (primeiro_dia.weekday() + 1) % 7

        # ====================================================================
        # ✅ CORREÇÃO: CARREGAR DIAS ÚTEIS DA TABELA PAR_TB020_CALENDARIO
        # ====================================================================
        calendario_info = Calendario.carregar_info_mes(ano, mes)

        # Buscar teletrabalhos da área
        teletrabalhos = Teletrabalho.query.filter(
            Teletrabalho.MES_REFERENCIA == mes_ref,
            Teletrabalho.AREA == area_filtro,
            Teletrabalho.TIPO_AREA == tipo_area_filtro,
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        # Organizar estrutura de dias do mês
        dias_organizados = {}
        data_atual = primeiro_dia

        while data_atual <= ultimo_dia:
            data_str = data_atual.strftime('%Y-%m-%d')

            # ✅ Verificar dia útil pela tabela PAR_TB020_CALENDARIO
            info_dia = calendario_info.get(data_atual)
            if info_dia is not None:
                eh_util = info_dia['eh_util']
            else:
                eh_util = Feriado.eh_dia_util(data_atual)

            # Buscar nome do feriado na tabela AUX_TB003_FERIADOS (para exibição)
            feriado_obj = Feriado.obter_feriado(data_atual)

            # ✅ Se não é útil e não tem feriado na tabela antiga, mostra texto genérico
            feriado_texto = None
            if not eh_util:
                if feriado_obj:
                    feriado_texto = feriado_obj.DS_FERIADO
                elif data_atual.weekday() < 5:
                    feriado_texto = 'Ponto Facultativo / Não Útil'

            # Buscar teletrabalhos deste dia
            pessoas_dia = []
            for t in teletrabalhos:
                if t.DATA_TELETRABALHO == data_atual:
                    usuario_nome = t.usuario.NOME if t.usuario else "Desconhecido"

                    partes = usuario_nome.split()
                    if len(partes) > 1:
                        primeiro = partes[0]
                        sobrenome = partes[-1]
                        nome_exibicao = f"{primeiro} {sobrenome}"
                    else:
                        nome_exibicao = partes[0] if partes else usuario_nome

                    pessoas_dia.append({
                        'id': t.ID,
                        'usuario': usuario_nome,
                        'nome_exibicao': nome_exibicao
                    })

            dias_organizados[data_str] = {
                'numero': data_atual.day,
                'eh_util': eh_util,
                'feriado': feriado_texto,
                'pessoas': pessoas_dia
            }

            data_atual += timedelta(days=1)

        # Calcular resumo
        resumo_pessoas = calcular_resumo_pessoas(mes_ref, area_filtro, tipo_area_filtro)

        # Limite
        limite, total_pessoas = calcular_limite_area(area_filtro, tipo_area_filtro)

        # Filtros
        areas = obter_areas_disponiveis()
        periodos = gerar_opcoes_periodo()

        eh_gestor = eh_gestor_ou_admin(usuario)

        return render_template('teletrabalho/calendario_visual.html',
                               dias_organizados=dias_organizados,
                               nome_mes=nome_mes,
                               mes_ref=mes_ref,
                               area=area_filtro,
                               tipo_area=tipo_area_filtro,
                               limite=limite,
                               total_pessoas=total_pessoas,
                               areas=areas,
                               periodos=periodos,
                               primeiro_dia_grid=primeiro_dia_grid,
                               eh_gestor=eh_gestor,
                               resumo_pessoas=resumo_pessoas,
                               pode_escolher_area=pode_escolher_area)

    except Exception as e:
        flash(f'Erro ao carregar calendário visual: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))

@teletrabalho_bp.route('/exportar-excel/<mes_ref>')
@login_required
def exportar_excel(mes_ref):
    """Exporta calendário para Excel com nomes abreviados"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from flask import send_file
        import io

        area = request.args.get('area')
        tipo_area = request.args.get('tipo_area', 'superintendencia')

        ano = int(mes_ref[:4])
        mes = int(mes_ref[4:])

        meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        nome_mes = f"{meses_pt[mes - 1]} {ano}"

        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])

        # Buscar dados
        teletrabalhos = Teletrabalho.query.filter(
            Teletrabalho.MES_REFERENCIA == mes_ref,
            Teletrabalho.AREA == area,
            Teletrabalho.TIPO_AREA == tipo_area,
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        # Organizar por dia com nomes abreviados
        dias_dados = {}
        data_atual = primeiro_dia

        while data_atual <= ultimo_dia:
            data_str = data_atual.strftime('%Y-%m-%d')
            dias_dados[data_str] = []
            data_atual += timedelta(days=1)

        # Adicionar pessoas
        for tele in teletrabalhos:
            data_str = tele.DATA_TELETRABALHO.strftime('%Y-%m-%d')
            if data_str in dias_dados:
                nome_completo = tele.usuario.NOME
                partes = nome_completo.split()

                if len(partes) >= 2:
                    nome_abreviado = f"{partes[0]} {partes[1][0]}."
                else:
                    nome_abreviado = partes[0] if partes else nome_completo

                dias_dados[data_str].append(nome_abreviado)

        # Processar nomes duplicados por dia
        for data_str, nomes in dias_dados.items():
            if len(nomes) > 1:
                # Contar primeiros nomes
                contagem = {}
                for nome in nomes:
                    primeiro = nome.split()[0]
                    contagem[primeiro] = contagem.get(primeiro, 0) + 1

                # Se houver duplicatas, já está com inicial do segundo nome
                # Então não precisa fazer nada, já vem como "João S."

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Teletrabalho {nome_mes}"

        # Estilos
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Cabeçalho principal
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Calendário de Teletrabalho - {nome_mes}"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:G2')
        ws['A2'] = f"Área: {area}"
        ws['A2'].font = Font(size=11, bold=True)
        ws['A2'].fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
        ws['A2'].alignment = center_align

        # Cabeçalho dos dias
        dias_semana = ['DOMINGO', 'SEGUNDA', 'TERÇA', 'QUARTA', 'QUINTA', 'SEXTA', 'SÁBADO']
        row = 4
        for col, dia in enumerate(dias_semana, start=1):
            cell = ws.cell(row=row, column=col, value=dia)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[chr(64 + col)].width = 18

        # Dados do calendário
        row = 5
        col = (primeiro_dia.weekday() + 1) % 7 + 1  # Posição inicial

        data_atual = primeiro_dia
        while data_atual <= ultimo_dia:
            if col > 7:
                col = 1
                row += 1

            data_str = data_atual.strftime('%Y-%m-%d')
            pessoas = dias_dados.get(data_str, [])

            # Conteúdo da célula
            conteudo = f"{data_atual.day}"
            if pessoas:
                conteudo += "\n\n" + "\n".join(pessoas)

            cell = ws.cell(row=row, column=col, value=conteudo)
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(size=10)

            # Número do dia em negrito
            if pessoas:
                cell.font = Font(size=10, bold=True)

            # Cor de fundo
            if not Feriado.eh_dia_util(data_atual):
                cell.fill = PatternFill(start_color="e0e0e0", end_color="e0e0e0", fill_type="solid")
                cell.font = Font(size=10, color="666666")
            elif pessoas:
                # Verde claro se tem pessoas
                cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                cell.font = Font(size=10, bold=True, color="155724")
            else:
                # Branco se vazio
                cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")

            col += 1
            data_atual += timedelta(days=1)

        # Ajustar altura das linhas
        for r in range(5, row + 1):
            ws.row_dimensions[r].height = 80

        # Rodapé com legenda
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "Legenda: Verde = Com teletrabalho | Cinza = Fim de semana/Feriado"
        ws[f'A{row}'].font = Font(size=9, italic=True)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")

        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Teletrabalho_{area}_{nome_mes.replace(" ", "_")}.xlsx'
        )

    except Exception as e:
        flash(f'Erro ao exportar: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.calendario_visual'))


@teletrabalho_bp.route('/mover-pessoa', methods=['POST'])
@login_required
def mover_pessoa():
    """Move pessoa de um dia para outro via drag-and-drop"""
    usuario = Usuario.query.get(current_user.id)

    # Apenas gestores podem mover
    if not eh_gestor_ou_admin(usuario):
        return jsonify({'erro': 'Sem permissão'}), 403

    try:
        data = request.get_json()
        teletrabalho_id = int(data.get('teletrabalho_id'))
        data_nova_str = data.get('data_nova')

        # Buscar teletrabalho
        teletrabalho = Teletrabalho.query.get_or_404(teletrabalho_id)

        # Validação de área para gerente
        nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)

        if nivel_acesso == 'gerente':
            if not usuario.empregado:
                return jsonify({'erro': 'Usuário sem vínculo de empregado'}), 403

            area_gerente = usuario.empregado.sgSuperintendencia

            if teletrabalho.AREA != area_gerente:
                return jsonify({'erro': 'Você só pode mover pessoas da sua área'}), 403

        data_antiga = teletrabalho.DATA_TELETRABALHO
        data_nova = datetime.strptime(data_nova_str, '%Y-%m-%d').date()

        # Validar se mudou
        if data_antiga == data_nova:
            return jsonify({'erro': 'Mesma data'}), 400

        # Validar dia útil
        if not Feriado.eh_dia_util(data_nova):
            feriado = Feriado.obter_feriado(data_nova)
            motivo = feriado.DS_FERIADO if feriado else "fim de semana"
            return jsonify({'erro': f'Data não é dia útil ({motivo})'}), 400

        # Validar mês
        mes_referencia = teletrabalho.MES_REFERENCIA
        if data_nova.strftime('%Y%m') != mes_referencia:
            return jsonify({'erro': 'Não pode mover para outro mês'}), 400

        # Validar limite de pessoas no dia
        area = teletrabalho.AREA
        tipo_area = teletrabalho.TIPO_AREA
        limite, _ = calcular_limite_area(area, tipo_area)

        qtd_no_dia = Teletrabalho.contar_pessoas_dia(data_nova, area, tipo_area)
        if qtd_no_dia >= limite:
            return jsonify({'erro': f'Dia lotado ({qtd_no_dia}/{limite})'}), 400

        # Validar 5 dias corridos (se for o tipo)
        if teletrabalho.TIPO_PERIODO == 'CINCO_DIAS_CORRIDOS':
            outros_dias = Teletrabalho.query.filter(
                Teletrabalho.USUARIO_ID == teletrabalho.USUARIO_ID,
                Teletrabalho.MES_REFERENCIA == mes_referencia,
                Teletrabalho.TIPO_PERIODO == 'CINCO_DIAS_CORRIDOS',
                Teletrabalho.STATUS == 'APROVADO',
                Teletrabalho.DELETED_AT.is_(None),
                Teletrabalho.ID != teletrabalho_id
            ).all()
            datas_mes = [d.DATA_TELETRABALHO for d in outros_dias] + [data_nova]

            valido, msg = Teletrabalho.validar_cinco_dias_corridos(
                teletrabalho.USUARIO_ID,
                datas_mes,
                mes_referencia,
                excluir_ids=[teletrabalho_id]
            )
            if not valido:
                return jsonify({'erro': msg}), 400

        elif teletrabalho.TIPO_PERIODO == 'TRES_CONSECUTIVOS':
            valido, msg = Teletrabalho.validar_tres_consecutivos(
                teletrabalho.USUARIO_ID,
                [data_nova],
                mes_referencia,
                excluir_ids=[teletrabalho_id]
            )
            if not valido:
                return jsonify({'erro': msg}), 400

        # ✅ ATUALIZAR DATA
        data_antiga_str = data_antiga.strftime('%d/%m/%Y')
        data_nova_str_br = data_nova.strftime('%d/%m/%Y')

        teletrabalho.DATA_TELETRABALHO = data_nova
        teletrabalho.UPDATED_AT = datetime.utcnow()

        # ✅ COMMIT PRIMEIRO
        db.session.commit()

        # ✅ LOG DE AUDITORIA DEPOIS DO COMMIT (para evitar rollback)
        try:
            registrar_log(
                acao='editar',
                entidade='teletrabalho',
                entidade_id=teletrabalho_id,
                descricao=f'{nivel_acesso.upper()} moveu {teletrabalho.usuario.NOME} de {data_antiga_str} para {data_nova_str_br}'
            )
        except Exception as log_error:
            # Se falhar o log, não importa - o dado já foi salvo
            print(f"[AVISO] Erro ao registrar log (não afeta a operação): {log_error}")

        return jsonify({
            'sucesso': True,
            'mensagem': f'Movido para {data_nova_str_br}'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro: {str(e)}'}), 500


@teletrabalho_bp.route('/ferias')
@login_required
def marcar_ferias():
    """Página para marcar férias"""
    try:
        usuario = Usuario.query.get(current_user.id)
        empregado = usuario.empregado

        if not empregado:
            flash('Você precisa estar vinculado a um empregado.', 'warning')
            return redirect(url_for('teletrabalho.index'))

        periodos = gerar_opcoes_periodo()

        # Buscar férias já marcadas
        ferias_marcadas = Teletrabalho.query.filter(
            Teletrabalho.USUARIO_ID == current_user.id,
            Teletrabalho.TIPO_MARCACAO == 'FERIAS',
            Teletrabalho.DELETED_AT.is_(None)
        ).order_by(Teletrabalho.DATA_TELETRABALHO).all()

        # ✅ ADICIONADO: Verificar se é gestor para passar ao template
        eh_gestor = eh_gestor_ou_admin(usuario)

        return render_template('teletrabalho/marcar_ferias.html',
                               periodos=periodos,
                               ferias_marcadas=ferias_marcadas,
                               eh_gestor=eh_gestor)  # ✅ NOVO PARÂMETRO
    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))


@teletrabalho_bp.route('/ferias/calendario/<mes_referencia>', methods=['GET'])
@login_required
def ferias_calendario(mes_referencia):
    """Retorna dados do calendário para marcação de férias via AJAX"""
    try:
        usuario = Usuario.query.get(current_user.id)
        empregado = usuario.empregado

        if not empregado:
            return jsonify({'erro': 'Sem vínculo com empregado'}), 403

        ano = int(mes_referencia[:4])
        mes = int(mes_referencia[4:])

        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])
        primeiro_dia_grid = (primeiro_dia.weekday() + 1) % 7

        # Buscar férias já marcadas
        ferias_marcadas = Teletrabalho.query.filter(
            Teletrabalho.USUARIO_ID == current_user.id,
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.TIPO_MARCACAO == 'FERIAS',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        datas_ferias = [f.DATA_TELETRABALHO.strftime('%Y-%m-%d') for f in ferias_marcadas]

        # Buscar teletrabalhos já marcados
        teletrabalhos = Teletrabalho.query.filter(
            Teletrabalho.USUARIO_ID == current_user.id,
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        datas_teletrabalho = [t.DATA_TELETRABALHO.strftime('%Y-%m-%d') for t in teletrabalhos]

        # Montar lista de dias
        dias = []
        data_atual = primeiro_dia

        while data_atual <= ultimo_dia:
            data_str = data_atual.strftime('%Y-%m-%d')
            eh_util = Feriado.eh_dia_util(data_atual)
            feriado = Feriado.obter_feriado(data_atual)

            dias.append({
                'data': data_str,
                'dia': data_atual.day,
                'eh_util': eh_util,
                'feriado': feriado.DS_FERIADO if feriado else None,
                'ja_marcado': data_str in datas_ferias,
                'tem_teletrabalho': data_str in datas_teletrabalho
            })

            data_atual += timedelta(days=1)

        return jsonify({
            'dias': dias,
            'primeiro_dia_grid': primeiro_dia_grid
        })

    except Exception as e:
        print(f"[ERRO] ferias_calendario: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@teletrabalho_bp.route('/ferias/salvar', methods=['POST'])
@login_required
def salvar_ferias():
    """Salva marcação de férias"""
    try:
        usuario = Usuario.query.get(current_user.id)
        empregado = usuario.empregado

        if not empregado:
            return jsonify({'erro': 'Sem vínculo com empregado'}), 403

        data = request.get_json()
        datas_str = data.get('datas', [])

        if not datas_str:
            return jsonify({'erro': 'Selecione pelo menos um dia'}), 400

        area = empregado.sgSuperintendencia
        tipo_area = 'superintendencia'

        datas = [datetime.strptime(d, '%Y-%m-%d').date() for d in datas_str]

        # Validar se já tem teletrabalho marcado
        teletrabalhos_conflito = Teletrabalho.query.filter(
            Teletrabalho.USUARIO_ID == current_user.id,
            Teletrabalho.DATA_TELETRABALHO.in_(datas),
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        if teletrabalhos_conflito:
            datas_conflito = [t.DATA_TELETRABALHO.strftime('%d/%m/%Y') for t in teletrabalhos_conflito]
            return jsonify({
                'erro': f'Você já tem teletrabalho marcado em: {", ".join(datas_conflito)}'
            }), 400

        # Salvar férias
        for data_ferias in datas:
            # Verificar se já existe
            existe = Teletrabalho.query.filter(
                Teletrabalho.USUARIO_ID == current_user.id,
                Teletrabalho.DATA_TELETRABALHO == data_ferias,
                Teletrabalho.TIPO_MARCACAO == 'FERIAS',
                Teletrabalho.DELETED_AT.is_(None)
            ).first()

            if not existe:
                ferias = Teletrabalho(
                    USUARIO_ID=current_user.id,
                    DATA_TELETRABALHO=data_ferias,
                    MES_REFERENCIA=data_ferias.strftime('%Y%m'),
                    AREA=area,
                    TIPO_AREA=tipo_area,
                    STATUS='APROVADO',
                    TIPO_PERIODO='FERIAS',
                    TIPO_MARCACAO='FERIAS',  # ✅ IMPORTANTE
                    OBSERVACAO='Férias marcadas pelo usuário',
                    APROVADO_POR=current_user.id,
                    APROVADO_EM=datetime.now()
                )
                db.session.add(ferias)

        db.session.commit()

        return jsonify({
            'sucesso': True,
            'mensagem': f'{len(datas)} dia(s) de férias marcado(s)!'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 500


@teletrabalho_bp.route('/ferias/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_ferias(id):
    """Exclui marcação de férias"""
    try:
        ferias = Teletrabalho.query.get_or_404(id)

        if ferias.USUARIO_ID != current_user.id:
            return jsonify({'erro': 'Sem permissão'}), 403

        if ferias.TIPO_MARCACAO != 'FERIAS':
            return jsonify({'erro': 'Não é registro de férias'}), 400

        ferias.DELETED_AT = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'sucesso': True,
            'mensagem': 'Férias removidas!'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 500


@teletrabalho_bp.route('/admin/marcar-ferias', methods=['GET'])
@login_required
def admin_marcar_ferias():
    """Página para gestor marcar férias de usuários"""
    usuario = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('teletrabalho.index'))

    try:
        # Buscar usuários da área (similar ao admin_cadastrar)
        nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)

        if nivel_acesso in ['admin', 'moderador']:
            # Admin/Moderador vê todos
            usuarios_query = db.session.query(
                Usuario.ID,
                Usuario.NOME,
                Empregado.sgSuperintendencia,
                Empregado.dsCargo
            ).join(
                Empregado, Usuario.FK_PESSOA == Empregado.pkPessoa
            ).filter(
                Usuario.ATIVO == True,
                Usuario.DELETED_AT.is_(None),
                Empregado.fkStatus == 1,
                Empregado.sgSuperintendencia.isnot(None)
            ).order_by(Usuario.NOME).all()

        else:  # Gerente
            if not usuario.empregado:
                flash('Você não está vinculado a nenhuma área.', 'warning')
                return redirect(url_for('teletrabalho.index'))

            area_gerente = usuario.empregado.sgSuperintendencia

            usuarios_query = db.session.query(
                Usuario.ID,
                Usuario.NOME,
                Empregado.sgSuperintendencia,
                Empregado.dsCargo
            ).join(
                Empregado, Usuario.FK_PESSOA == Empregado.pkPessoa
            ).filter(
                Usuario.ATIVO == True,
                Usuario.DELETED_AT.is_(None),
                Empregado.fkStatus == 1,
                Empregado.sgSuperintendencia == area_gerente
            ).order_by(Usuario.NOME).all()

        usuarios = [
            {
                'id': u.ID,
                'nome': u.NOME,
                'area': u.sgSuperintendencia,
                'cargo': u.dsCargo
            }
            for u in usuarios_query
        ]

        periodos = gerar_opcoes_periodo()

        return render_template('teletrabalho/admin_marcar_ferias.html',
                               usuarios=usuarios,
                               periodos=periodos)

    except Exception as e:
        flash(f'Erro: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))


@teletrabalho_bp.route('/admin/marcar-ferias/calendario/<int:usuario_id>/<mes_referencia>', methods=['GET'])
@login_required
def admin_ferias_calendario(usuario_id, mes_referencia):
    """Retorna dados do calendário para gestor marcar férias de usuário"""
    usuario_logado = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario_logado):
        return jsonify({'erro': 'Sem permissão'}), 403

    try:
        # Buscar usuário alvo
        usuario_alvo = Usuario.query.get_or_404(usuario_id)
        empregado = usuario_alvo.empregado

        if not empregado:
            return jsonify({'erro': 'Usuário sem vínculo com empregado'}), 404

        # Validar permissão de área (gerente só pode marcar da sua área)
        nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario_logado)

        if nivel_acesso == 'gerente':
            if not usuario_logado.empregado:
                return jsonify({'erro': 'Gerente sem vínculo de empregado'}), 403

            area_gerente = usuario_logado.empregado.sgSuperintendencia

            if empregado.sgSuperintendencia != area_gerente:
                return jsonify({'erro': 'Você só pode marcar férias de usuários da sua área'}), 403

        ano = int(mes_referencia[:4])
        mes = int(mes_referencia[4:])

        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])
        primeiro_dia_grid = (primeiro_dia.weekday() + 1) % 7

        # Buscar férias já marcadas
        ferias_marcadas = Teletrabalho.query.filter(
            Teletrabalho.USUARIO_ID == usuario_id,
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.TIPO_MARCACAO == 'FERIAS',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        datas_ferias = [f.DATA_TELETRABALHO.strftime('%Y-%m-%d') for f in ferias_marcadas]

        # Buscar teletrabalhos já marcados
        teletrabalhos_marcados = Teletrabalho.query.filter(
            Teletrabalho.USUARIO_ID == usuario_id,
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        datas_teletrabalho = [t.DATA_TELETRABALHO.strftime('%Y-%m-%d') for t in teletrabalhos_marcados]

        # Montar lista de dias
        dias = []
        data_atual = primeiro_dia

        while data_atual <= ultimo_dia:
            data_str = data_atual.strftime('%Y-%m-%d')
            eh_util = Feriado.eh_dia_util(data_atual)
            feriado = Feriado.obter_feriado(data_atual)

            dias.append({
                'data': data_str,
                'dia': data_atual.day,
                'eh_util': eh_util,
                'feriado': feriado.DS_FERIADO if feriado else None,
                'ja_marcado_ferias': data_str in datas_ferias,
                'ja_marcado_teletrabalho': data_str in datas_teletrabalho
            })

            data_atual += timedelta(days=1)

        # ✅ IMPORTANTE: RETORNAR JSON COM SUCESSO
        return jsonify({
            'sucesso': True,
            'dias': dias,
            'primeiro_dia_grid': primeiro_dia_grid
        })

    except Exception as e:
        print(f"[ERRO] admin_ferias_calendario: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@teletrabalho_bp.route('/admin/marcar-ferias/salvar', methods=['POST'])
@login_required
def admin_salvar_ferias():
    """Gestor salva férias de usuário"""
    usuario_logado = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario_logado):
        return jsonify({'erro': 'Sem permissão'}), 403

    try:
        data = request.get_json()
        usuario_id = int(data.get('usuario_id'))
        datas_str = data.get('datas', [])

        if not datas_str:
            return jsonify({'erro': 'Selecione pelo menos um dia'}), 400

        # Buscar usuário alvo
        usuario_alvo = Usuario.query.get_or_404(usuario_id)
        empregado = usuario_alvo.empregado

        if not empregado:
            return jsonify({'erro': 'Usuário sem vínculo com empregado'}), 404

        # Validar permissão de área
        nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario_logado)

        if nivel_acesso == 'gerente':
            if not usuario_logado.empregado:
                return jsonify({'erro': 'Sem vínculo com empregado'}), 403

            area_gerente = usuario_logado.empregado.sgSuperintendencia

            if empregado.sgSuperintendencia != area_gerente:
                return jsonify({'erro': 'Você só pode marcar férias de usuários da sua área'}), 403

        area = empregado.sgSuperintendencia
        tipo_area = 'superintendencia'

        datas = [datetime.strptime(d, '%Y-%m-%d').date() for d in datas_str]

        # Validar conflito com teletrabalho
        teletrabalhos_conflito = Teletrabalho.query.filter(
            Teletrabalho.USUARIO_ID == usuario_id,
            Teletrabalho.DATA_TELETRABALHO.in_(datas),
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        if teletrabalhos_conflito:
            datas_conflito = [t.DATA_TELETRABALHO.strftime('%d/%m/%Y') for t in teletrabalhos_conflito]
            return jsonify({
                'erro': f'Usuário já tem teletrabalho marcado em: {", ".join(datas_conflito)}'
            }), 400

        # Salvar férias
        for data_ferias in datas:
            existe = Teletrabalho.query.filter(
                Teletrabalho.USUARIO_ID == usuario_id,
                Teletrabalho.DATA_TELETRABALHO == data_ferias,
                Teletrabalho.TIPO_MARCACAO == 'FERIAS',
                Teletrabalho.DELETED_AT.is_(None)
            ).first()

            if not existe:
                ferias = Teletrabalho(
                    USUARIO_ID=usuario_id,
                    DATA_TELETRABALHO=data_ferias,
                    MES_REFERENCIA=data_ferias.strftime('%Y%m'),
                    AREA=area,
                    TIPO_AREA=tipo_area,
                    STATUS='APROVADO',
                    TIPO_PERIODO='FERIAS',
                    TIPO_MARCACAO='FERIAS',
                    OBSERVACAO=f'[{nivel_acesso.upper()}] Férias marcadas por {usuario_logado.NOME}',
                    APROVADO_POR=current_user.id,
                    APROVADO_EM=datetime.now()
                )
                db.session.add(ferias)

        db.session.commit()

        try:
            registrar_log(
                acao='criar',
                entidade='ferias',
                entidade_id=0,
                descricao=f'{nivel_acesso.upper()} marcou {len(datas)} dia(s) de férias para {usuario_alvo.NOME}'
            )
        except:
            pass

        return jsonify({
            'sucesso': True,
            'mensagem': f'{len(datas)} dia(s) de férias marcado(s) para {usuario_alvo.NOME}!'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 500


# ==================== BLOQUEIO DE DIAS ====================

@teletrabalho_bp.route('/bloqueios')
@login_required
def listar_bloqueios():
    """Lista todos os bloqueios de dias"""
    usuario = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario):
        flash('Acesso negado. Apenas gestores e admins.', 'danger')
        return redirect(url_for('teletrabalho.index'))

    try:
        mes_ref = request.args.get('mes_ref')
        area_filtro = request.args.get('area')

        # Determinar área do filtro
        if not area_filtro:
            nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)
            if nivel_acesso == 'gerente':
                if usuario.empregado:
                    area_filtro = usuario.empregado.sgSuperintendencia
                else:
                    flash('Você não está vinculado a nenhuma área.', 'warning')
                    return redirect(url_for('teletrabalho.index'))

        # ✅ NOVA LÓGICA: Query na nova tabela
        query = BloqueioDia.query.filter(
            BloqueioDia.DELETED_AT.is_(None)
        )

        # Filtrar por área se for gerente
        if not pode_ver_todas_areas(usuario) and area_filtro:
            query = query.filter(BloqueioDia.AREA == area_filtro)
        elif area_filtro:
            query = query.filter(BloqueioDia.AREA == area_filtro)

        # Filtrar por mês se especificado
        if mes_ref:
            query = query.filter(BloqueioDia.MES_REFERENCIA == mes_ref)

        bloqueios = query.order_by(BloqueioDia.DATA_BLOQUEIO.desc()).all()

        # Buscar áreas disponíveis
        if pode_ver_todas_areas(usuario):
            areas = db.session.query(
                Empregado.sgSuperintendencia
            ).filter(
                Empregado.sgSuperintendencia.isnot(None),
                Empregado.fkStatus == 1
            ).group_by(Empregado.sgSuperintendencia).all()
            areas_disponiveis = [a[0] for a in areas]
        else:
            areas_disponiveis = [area_filtro] if area_filtro else []

        periodos = gerar_opcoes_periodo()

        return render_template('teletrabalho/bloqueios.html',
                               bloqueios=bloqueios,
                               areas_disponiveis=areas_disponiveis,
                               periodos=periodos,
                               mes_ref=mes_ref,
                               area_filtro=area_filtro)

    except Exception as e:
        flash(f'Erro ao carregar bloqueios: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))


@teletrabalho_bp.route('/bloqueios/adicionar', methods=['GET', 'POST'])
@login_required
def adicionar_bloqueio():
    """Adiciona bloqueio de dia"""
    usuario = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario):
        flash('Acesso negado. Apenas gestores e admins.', 'danger')
        return redirect(url_for('teletrabalho.index'))

    if request.method == 'POST':
        try:
            datas_str = request.form.getlist('datas[]')
            motivo = request.form.get('motivo', '').strip()
            area = request.form.get('area')
            tipo_area = request.form.get('tipo_area', 'superintendencia')
            mes_referencia = request.form.get('mes_referencia')

            if not datas_str:
                flash('Selecione pelo menos um dia para bloquear.', 'warning')
                return redirect(url_for('teletrabalho.adicionar_bloqueio'))

            # Determinar área
            nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)

            if nivel_acesso == 'gerente':
                if not usuario.empregado:
                    flash('Você não está vinculado a nenhuma área.', 'warning')
                    return redirect(url_for('teletrabalho.adicionar_bloqueio'))

                area = usuario.empregado.sgSuperintendencia
                tipo_area = 'superintendencia'
            elif not area:
                flash('Selecione uma área.', 'warning')
                return redirect(url_for('teletrabalho.adicionar_bloqueio'))

            # Converter strings para datas
            datas = [datetime.strptime(d, '%Y-%m-%d').date() for d in datas_str]

            bloqueios_adicionados = 0
            bloqueios_existentes = 0

            for data_bloqueio in datas:
                # ✅ NOVA LÓGICA: Verificar se já existe bloqueio na nova tabela
                ja_bloqueado, _ = BloqueioDia.dia_esta_bloqueado(data_bloqueio, area, tipo_area)

                if ja_bloqueado:
                    bloqueios_existentes += 1
                    continue

                # ✅ NOVA LÓGICA: Criar bloqueio na tabela específica
                bloqueio = BloqueioDia(
                    DATA_BLOQUEIO=data_bloqueio,
                    MES_REFERENCIA=data_bloqueio.strftime('%Y%m'),
                    AREA=area,
                    TIPO_AREA=tipo_area,
                    MOTIVO=motivo if motivo else 'Dia bloqueado pelo gestor',
                    BLOQUEADO_POR=current_user.id,
                    BLOQUEADO_EM=datetime.utcnow()
                )
                db.session.add(bloqueio)
                bloqueios_adicionados += 1

            db.session.commit()

            registrar_log(
                acao='criar',
                entidade='bloqueio_dia',
                entidade_id=0,
                descricao=f'{nivel_acesso.upper()} bloqueou {bloqueios_adicionados} dia(s) na área {area}'
            )

            mensagem = f'{bloqueios_adicionados} dia(s) bloqueado(s) com sucesso!'
            if bloqueios_existentes > 0:
                mensagem += f' ({bloqueios_existentes} dia(s) já estavam bloqueados)'

            flash(mensagem, 'success')
            return redirect(url_for('teletrabalho.listar_bloqueios'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao bloquear dias: {str(e)}', 'danger')
            return redirect(url_for('teletrabalho.adicionar_bloqueio'))

    # GET - Exibir formulário (continua igual)
    try:
        nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)

        if nivel_acesso == 'gerente':
            if not usuario.empregado:
                flash('Você não está vinculado a nenhuma área.', 'warning')
                return redirect(url_for('teletrabalho.index'))

            area = usuario.empregado.sgSuperintendencia
            areas_disponiveis = [area]
        else:
            # Admin ou Moderador
            areas = db.session.query(
                Empregado.sgSuperintendencia
            ).filter(
                Empregado.sgSuperintendencia.isnot(None),
                Empregado.fkStatus == 1
            ).group_by(Empregado.sgSuperintendencia).all()
            areas_disponiveis = [a[0] for a in areas]

        periodos = gerar_opcoes_periodo()

        return render_template('teletrabalho/adicionar_bloqueio.html',
                               areas_disponiveis=areas_disponiveis,
                               periodos=periodos)

    except Exception as e:
        flash(f'Erro ao carregar formulário: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.index'))


@teletrabalho_bp.route('/bloqueios/calendario/<mes_referencia>/<area>/<tipo_area>')
@login_required
def bloqueios_calendario(mes_referencia, area, tipo_area):
    """Retorna calendário para tela de bloqueios (JSON)"""
    try:
        usuario = Usuario.query.get(current_user.id)
        if not eh_gestor_ou_admin(usuario):
            return jsonify({'erro': 'Acesso negado'}), 403

        nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)

        if nivel_acesso == 'gerente':
            if not usuario.empregado:
                return jsonify({'erro': 'Sem vínculo com empregado'}), 403
            if area != usuario.empregado.sgSuperintendencia:
                return jsonify({'erro': 'Você só pode bloquear dias da sua área'}), 403

        ano = int(mes_referencia[:4])
        mes = int(mes_referencia[4:])

        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])

        # ✅ CORREÇÃO: Usar PAR_TB020_CALENDARIO
        calendario_info = Calendario.carregar_info_mes(ano, mes)

        # Buscar bloqueios
        bloqueios = BloqueioDia.listar_bloqueios_mes(mes_referencia, area, tipo_area)
        datas_bloqueadas = {b.DATA_BLOQUEIO.strftime('%Y-%m-%d'): b.MOTIVO for b in bloqueios}

        # Buscar teletrabalhos já agendados
        teletrabalhos = Teletrabalho.query.filter(
            Teletrabalho.MES_REFERENCIA == mes_referencia,
            Teletrabalho.AREA == area,
            Teletrabalho.TIPO_AREA == tipo_area,
            Teletrabalho.TIPO_MARCACAO == 'TELETRABALHO',
            Teletrabalho.STATUS == 'APROVADO',
            Teletrabalho.DELETED_AT.is_(None)
        ).all()

        teletrabalhos_por_dia = {}
        for t in teletrabalhos:
            data_str = t.DATA_TELETRABALHO.strftime('%Y-%m-%d')
            if data_str not in teletrabalhos_por_dia:
                teletrabalhos_por_dia[data_str] = 0
            teletrabalhos_por_dia[data_str] += 1

        # Montar dias
        dias = []
        data_atual = primeiro_dia

        while data_atual <= ultimo_dia:
            data_str = data_atual.strftime('%Y-%m-%d')

            # ✅ Usar PAR_TB020_CALENDARIO
            info_dia = calendario_info.get(data_atual)
            if info_dia is not None:
                eh_util = info_dia['eh_util']
            else:
                eh_util = Feriado.eh_dia_util(data_atual)

            feriado_obj = Feriado.obter_feriado(data_atual)
            feriado_texto = None
            if feriado_obj:
                feriado_texto = feriado_obj.DS_FERIADO
            elif not eh_util and data_atual.weekday() < 5:
                feriado_texto = 'Ponto Facultativo / Não Útil'

            dias.append({
                'data': data_str,
                'dia': data_atual.day,
                'eh_util': eh_util,
                'ja_bloqueado': data_str in datas_bloqueadas,
                'motivo_bloqueio': datas_bloqueadas.get(data_str, ''),
                'qtd_teletrabalhos': teletrabalhos_por_dia.get(data_str, 0),
                'feriado': feriado_texto
            })

            data_atual += timedelta(days=1)

        return jsonify({
            'sucesso': True,
            'dias': dias,
            'primeiro_dia_grid': (primeiro_dia.weekday() + 1) % 7
        })

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@teletrabalho_bp.route('/bloqueios/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_bloqueio(id):
    """Remove bloqueio de dia (soft delete)"""
    usuario = Usuario.query.get(current_user.id)

    if not eh_gestor_ou_admin(usuario):
        flash('Acesso negado. Apenas gestores e admins.', 'danger')
        return redirect(url_for('teletrabalho.listar_bloqueios'))

    try:
        # ✅ NOVA LÓGICA: Buscar na nova tabela
        bloqueio = BloqueioDia.query.get_or_404(id)

        # Validar permissão para gerentes
        nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)

        if nivel_acesso == 'gerente':
            if not usuario.empregado:
                flash('Sem vínculo com empregado.', 'danger')
                return redirect(url_for('teletrabalho.listar_bloqueios'))

            if bloqueio.AREA != usuario.empregado.sgSuperintendencia:
                flash('Você só pode desbloquear dias da sua área.', 'danger')
                return redirect(url_for('teletrabalho.listar_bloqueios'))

        # Soft delete
        data_bloqueada = bloqueio.DATA_BLOQUEIO
        area_bloqueada = bloqueio.AREA

        bloqueio.DELETED_AT = datetime.utcnow()
        db.session.commit()

        registrar_log(
            acao='excluir',
            entidade='bloqueio_dia',
            entidade_id=id,
            descricao=f'{nivel_acesso.upper()} desbloqueou dia {data_bloqueada.strftime("%d/%m/%Y")} na área {area_bloqueada}'
        )

        flash('Bloqueio removido com sucesso!', 'success')
        return redirect(url_for('teletrabalho.listar_bloqueios'))

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover bloqueio: {str(e)}', 'danger')
        return redirect(url_for('teletrabalho.listar_bloqueios'))

@teletrabalho_bp.route('/exportar-pdf/<mes_referencia>')
@login_required
def exportar_pdf(mes_referencia):
    """Exporta a visualização geral para PDF"""
    usuario = Usuario.query.get(current_user.id)
    empregado = usuario.empregado

    # Determinar área (similar à visualizacao_geral)
    nivel_acesso = verificar_nivel_acesso_teletrabalho(usuario)

    if nivel_acesso == 'gerente':
        if usuario.empregado and usuario.empregado.sgSuperintendencia:
            area = usuario.empregado.sgSuperintendencia
        else:
            flash('Você não está vinculado a nenhuma área.', 'warning')
            return redirect(url_for('teletrabalho.index'))
    else:
        area = request.args.get('area', empregado.sgSuperintendencia if empregado else None)
        if not area:
            flash('Área não identificada.', 'warning')
            return redirect(url_for('teletrabalho.index'))

    # Buscar dados (reutilizar lógica da visualizacao_geral)
    ano = int(mes_referencia[:4])
    mes = int(mes_referencia[4:])
    primeiro_dia = date(ano, mes, 1)
    ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])
    meses_pt = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    nome_mes = f"{meses_pt[mes - 1]} de {ano}"

    teletrabalhos = Teletrabalho.query.join(
        Usuario, Teletrabalho.USUARIO_ID == Usuario.ID
    ).filter(
        Teletrabalho.MES_REFERENCIA == mes_referencia,
        Teletrabalho.AREA == area,
        Teletrabalho.STATUS == 'APROVADO',
        Teletrabalho.DELETED_AT.is_(None)
    ).all()

    dias_organizados = {}
    data_atual = primeiro_dia
    while data_atual <= ultimo_dia:
        data_str = data_atual.strftime('%Y-%m-%d')
        dias_organizados[data_str] = {
            'data': data_atual,
            'dia_semana': ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo'][data_atual.weekday()],
            'pessoas': []
        }
        data_atual += timedelta(days=1)

    for tele in teletrabalhos:
        data_str = tele.DATA_TELETRABALHO.strftime('%Y-%m-%d')
        if data_str in dias_organizados:
            nome_completo = tele.usuario.NOME
            partes = nome_completo.split()
            nome_abreviado = f"{partes[0]} {partes[1][0]}." if len(partes) >= 2 else partes[0]
            dias_organizados[data_str]['pessoas'].append({
                'nome': nome_abreviado,
                'tipo_periodo': tele.TIPO_PERIODO
            })

    # Gerar PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=inch, leftMargin=inch, topMargin=inch, bottomMargin=inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=18, spaceAfter=30, alignment=1)
    normal_style = styles['Normal']

    elements = []

    # Cabeçalho
    elements.append(Paragraph(f"<b>Calendário de Teletrabalho - {nome_mes}</b>", title_style))
    elements.append(Paragraph(f"<b>Área:</b> {area}", normal_style))
    elements.append(Spacer(1, 0.5 * inch))

    # Tabela com dias e pessoas
    data = [['Data', 'Dia da Semana', 'Pessoas em Teletrabalho']]
    for data_str, info in sorted(dias_organizados.items()):
        pessoas = ', '.join([p['nome'] for p in info['pessoas']]) if info['pessoas'] else 'Nenhuma'
        data.append([info['data'].strftime('%d/%m/%Y'), info['dia_semana'], pessoas])

    table = Table(data, colWidths=[1.5*inch, 1.5*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)

    # Rodapé
    elements.append(Spacer(1, 0.5 * inch))
    elements.append(Paragraph("Gerado automaticamente pelo sistema de teletrabalho.", normal_style))

    doc.build(elements)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name=f'visualizacao_geral_{area}_{mes_referencia}.pdf', mimetype='application/pdf')
