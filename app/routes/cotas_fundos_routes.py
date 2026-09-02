from datetime import timedelta
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

from flask import Blueprint, render_template, request, jsonify
from flask_login import login_required

from app import db
from app.models.cotas_fundos import (
    FUNDOS, obter_ultimo_registro, obter_proxima_data,
    e_dia_util, calcular_ind_cota,
)
from app.utils.audit import registrar_log
import io
from flask import Response
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime


cotas_fundos_bp = Blueprint(
    'cotas_fundos', __name__, url_prefix='/cotas-fundos'
)

# Em qual tipo de dia o preenchimento é AUTOMÁTICO (repete o dia anterior e
# zera o IND_COTA). 'nao_util' = fim de semana/feriado (padrão, pois a cota
# não se move e o índice dá zero). Troque para 'util' se a regra for outra.
AUTO_PREENCHER_EM = 'nao_util'

_MES_ABREV = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']

def _to_decimal(valor):
    """Converte string do formulário em Decimal. None se vazio/inválido."""
    if valor is None:
        return None
    txt = str(valor).strip().replace('.', '').replace(',', '.') \
        if ',' in str(valor) else str(valor).strip()
    if txt == '':
        return None
    try:
        return Decimal(txt)
    except (InvalidOperation, ValueError):
        return None


def _auto_preencher(data):
    """
    True somente quando a próxima DATA NÃO é dia útil (fim de semana/feriado).
    Nesse caso os valores são copiados do dia anterior e o IND_COTA vai zerado,
    pois a cota não se move. Em dia útil o usuário digita o VR_COTA.
    Data fora do calendário: exige digitação (retorna False).
    """
    util = e_dia_util(data)
    if util is None:
        return False
    return not util


def _montar_contexto_fundo(chave, cfg):
    """Monta o estado atual de um fundo para a tela."""
    model = cfg['model']
    ultimo = obter_ultimo_registro(model)
    proxima = obter_proxima_data(model)

    dados_anteriores = {}
    if ultimo is not None:
        for (attr, _label) in cfg['campos']:
            valor = getattr(ultimo, attr, None)
            dados_anteriores[attr] = float(valor) if valor is not None else None

    util = e_dia_util(proxima) if proxima else None
    return {
        'chave': chave,
        'label': cfg['label'],
        'tabela': cfg['tabela'],
        'campos': cfg['campos'],
        'proxima_data': proxima,
        'data_anterior': ultimo.DATA if ultimo else None,
        'vr_cota_anterior': float(ultimo.VR_COTA) if (ultimo and ultimo.VR_COTA is not None) else None,
        'dados_anteriores': dados_anteriores,
        'e_dia_util': util,
        'auto': _auto_preencher(proxima) if proxima else False,
        'vazia': ultimo is None,
    }

def _fmt_vr(v):
    """Formata valor no padrão BR: 1234567.89 -> 'R$ 1.234.567,89'. None/vazio -> 'R$ 0,00'."""
    if v is None or v == '':
        v = 0
    try:
        d = Decimal(str(v)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        d = Decimal('0.00')
    # separador de milhar e vírgula decimal
    s = f'{d:,.2f}'                       # 1,234,567.89
    s = s.replace(',', 'X').replace('.', ',').replace('X', '.')  # 1.234.567,89
    return f'R$ {s}'


@cotas_fundos_bp.route('/')
@login_required
def index():
    """Tela de entrada de dados das cotas dos fundos (FIN_TB026/027/028)."""
    fundos = [_montar_contexto_fundo(chave, cfg) for chave, cfg in FUNDOS.items()]
    return render_template('cotas_fundos/index.html', fundos=fundos)


@cotas_fundos_bp.route('/salvar/<chave>', methods=['POST'])
@login_required
def salvar(chave):
    """Grava a linha da próxima DATA do fundo informado."""
    cfg = FUNDOS.get(chave)
    if not cfg:
        return jsonify({'success': False, 'message': 'Fundo inválido.'}), 400

    model = cfg['model']
    ultimo = obter_ultimo_registro(model)
    proxima = obter_proxima_data(model)

    if proxima is None:
        return jsonify({
            'success': False,
            'message': (f'A tabela {cfg["tabela"]} está vazia. Cadastre a primeira '
                        f'linha (data inicial) diretamente no banco para iniciar a série.')
        }), 400

    try:
        if _auto_preencher(proxima):
            # Dia de repetição: copia tudo do dia anterior e zera o IND_COTA.
            registro = model(DATA=proxima)
            registro.VR_COTA = ultimo.VR_COTA
            registro.IND_COTA = Decimal('0.00000000')
            for (attr, _label) in cfg['campos']:
                setattr(registro, attr, getattr(ultimo, attr, None))
            origem = 'automático (repetiu o dia anterior)'
        else:
            vr_cota = _to_decimal(request.form.get('VR_COTA'))
            if vr_cota is None:
                return jsonify({'success': False,
                                'message': 'Informe o VR_COTA.'}), 400
            registro = model(DATA=proxima)
            registro.VR_COTA = vr_cota
            registro.IND_COTA = calcular_ind_cota(
                vr_cota, ultimo.VR_COTA if ultimo else None
            )
            for (attr, _label) in cfg['campos']:
                setattr(registro, attr, _to_decimal(request.form.get(attr)))
            origem = 'manual'

        db.session.add(registro)
        db.session.commit()

        registrar_log(
            acao='inclusao',
            entidade='cotas_fundos',
            entidade_id=None,
            descricao=f'Cotas {cfg["label"]} — {proxima.strftime("%d/%m/%Y")} ({origem})',
            dados_novos={
                'tabela': cfg['tabela'],
                'DATA': proxima.strftime('%Y-%m-%d'),
                'VR_COTA': str(registro.VR_COTA),
                'IND_COTA': str(registro.IND_COTA),
            },
        )

        return jsonify({
            'success': True,
            'message': (f'{cfg["label"]}: {proxima.strftime("%d/%m/%Y")} gravado '
                        f'({origem}). IND_COTA = {registro.IND_COTA}.'),
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro ao gravar: {str(e)}'}), 500


@cotas_fundos_bp.route('/preencher-automaticos/<chave>', methods=['POST'])
@login_required
def preencher_automaticos(chave):
    """
    Avança automaticamente todas as datas seguidas que são de repetição
    (não úteis), parando na primeira que exige digitação.
    """
    cfg = FUNDOS.get(chave)
    if not cfg:
        return jsonify({'success': False, 'message': 'Fundo inválido.'}), 400

    model = cfg['model']
    criados = []
    try:
        for _ in range(60):  # trava de segurança
            ultimo = obter_ultimo_registro(model)
            proxima = obter_proxima_data(model)
            if proxima is None or not _auto_preencher(proxima):
                break
            registro = model(DATA=proxima)
            registro.VR_COTA = ultimo.VR_COTA
            registro.IND_COTA = Decimal('0.00000000')
            for (attr, _label) in cfg['campos']:
                setattr(registro, attr, getattr(ultimo, attr, None))
            db.session.add(registro)
            db.session.flush()
            criados.append(proxima.strftime('%d/%m/%Y'))
        db.session.commit()

        if not criados:
            return jsonify({'success': True,
                            'message': 'Nenhum dia automático pendente.'})

        registrar_log(
            acao='inclusao',
            entidade='cotas_fundos',
            entidade_id=None,
            descricao=f'Cotas {cfg["label"]} — {len(criados)} dia(s) automático(s)',
            dados_novos={'tabela': cfg['tabela'], 'datas': criados},
        )
        return jsonify({
            'success': True,
            'message': f'{len(criados)} dia(s) preenchido(s): ' + ', '.join(criados) + '.',
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@cotas_fundos_bp.route('/performance/excel')
@login_required
def performance_excel():
    """Performance Diária (FIN_VW034): ano mais recente com detalhe mensal
    (dias agrupados, só o mês atual aberto) + ACUMULADO; anos anteriores
    aparecem apenas como uma linha de ACUMULADO cada. Totais com 4 casas."""
    from collections import OrderedDict

    rows = db.session.execute(text("""
        SELECT ANO, DT_ATUALIZACAO, FAE_COTA, FAE_IND_COTA, BB_COTA, BB_IND_COTA,
               CX_COTA, CX_IND_COTA, SELIC, ANBIMA,
               PC_IRFM_FAE, PC_SELIC_FAE, PC_IRFM_BB, PC_SELIC_BB, PC_IRFM_CX, PC_SELIC_CX
        FROM [BDG].[FIN_VW034_PERFORMANCE_DIARIA_FUNDOS]
        ORDER BY ANO, DT_ATUALIZACAO
    """)).fetchall()

    if not rows:
        return Response('Sem dados na FIN_VW034.', mimetype='text/plain')

    # índices dentro de cada linha (com ANO na frente)
    I_DT, I_FAEC, I_FAEI, I_BBC, I_BBI, I_CXC, I_CXI, I_SELIC, I_ANB = 1, 2, 3, 4, 5, 6, 7, 8, 9
    I_PC = 10  # começam os 6 PC_* (10..15)

    _n = lambda v: float(v) if v is not None else None

    # agrupa por ano
    por_ano = OrderedDict()
    for r in rows:
        por_ano.setdefault(int(r[0]), []).append(r)
    anos = sorted(por_ano.keys())
    ano = max(anos)                       # ano mais recente = detalhado
    outros = sorted([a for a in anos if a != ano], reverse=True)

    # último registro de cada ano (p/ base do ano seguinte)
    last_rec = {a: recs[-1] for a, recs in por_ano.items()}
    def base_cotas(a):
        p = a - 1
        if p in last_rec:
            lr = last_rec[p]
            return (_n(lr[I_FAEC]), _n(lr[I_BBC]), _n(lr[I_CXC]))
        return (None, None, None)

    # ---------- helper: acumulado de um ano inteiro (valores) ----------
    def acum_ano(recs, base):
        first, last = recs[0], recs[-1]
        def var(idx, b):
            c1 = _n(last[idx]); c0 = b if b is not None else _n(first[idx])
            return (c1 / c0 - 1) if (c0 not in (None, 0) and c1 is not None) else None
        fae = var(I_FAEC, base[0]); bb = var(I_BBC, base[1]); cx = var(I_CXC, base[2])
        ps = pa = 1.0
        for r in recs:
            s = _n(r[I_SELIC]) or 0.0; a_ = _n(r[I_ANB]) or 0.0
            ps *= (1 + s / 100.0); pa *= (1 + a_ / 100.0)
        selic, anb = ps - 1, pa - 1
        rt = lambda f, i: (f / i) if (f is not None and i not in (None, 0)) else None
        return {'C': fae, 'E': bb, 'G': cx, 'H': selic, 'I': anb,
                'J': rt(fae, anb), 'K': rt(fae, selic), 'L': rt(bb, anb),
                'M': rt(bb, selic), 'N': rt(cx, anb), 'O': rt(cx, selic)}

    # ================= monta a planilha =================
    wb = Workbook(); ws = wb.active; ws.title = str(ano)
    ws.sheet_properties.outlinePr.summaryBelow = True

    thin = Side(style='thin'); borda = Border(thin, thin, thin, thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    azul = PatternFill('solid', fgColor='1F4E79'); fbranco = Font(bold=True, color='FFFFFF')
    realce = PatternFill('solid', fgColor='D9E1F2'); realce_mes = PatternFill('solid', fgColor='EDF1F7')
    realce_ano = PatternFill('solid', fgColor='FDE9D9')
    F_COTA = '0.000000000'; F_PCT6 = '0.000000%'; F_TOT = '0.0000%'

    ws['A1'] = 'Diretoria Contábil e Financeira - Difin'
    ws['A2'] = 'Superintendência Financeira - Sufin'
    ws['A3'] = 'Gerência de Finanças - Gefin'
    ws['A5'] = 'Performance dos Fundos de Investimentos'
    for c in ('A1', 'A2', 'A3'): ws[c].font = Font(bold=True, size=10)
    ws['A5'].font = Font(bold=True, size=13, color='1F4E79')

    L1, L2 = 7, 8
    ws.merge_cells(start_row=L1, start_column=1, end_row=L2, end_column=1); ws.cell(L1, 1, 'DATA')
    def grp(ci, tit, span=2):
        ws.merge_cells(start_row=L1, start_column=ci, end_row=L1, end_column=ci+span-1); ws.cell(L1, ci, tit)
    grp(2, 'Extramercado FAE 2'); grp(4, 'BB RF Exclusivo Emgea'); grp(6, 'Caixa RF Exclusivo XXI')
    ws.merge_cells(start_row=L1, start_column=8, end_row=L1, end_column=9); ws.cell(L1, 8, 'Benchmarks')
    ws.merge_cells(start_row=L1, start_column=10, end_row=L1, end_column=15); ws.cell(L1, 10, 'Performance (%)')
    sub = [None, 'Cota', 'Variação', 'Cota', 'Variação', 'Cota', 'Variação', 'Selic', 'Anbima IRFM 1',
           '% IRFM FAE2', '% SELIC FAE2', '% IRFM BB', '% SELIC BB', '% IRFM CX', '% SELIC CX']
    for i, t in enumerate(sub, start=1):
        if t: ws.cell(L2, i, t)
    for r in (L1, L2):
        for c in range(1, 16):
            cell = ws.cell(r, c); cell.font = fbranco; cell.fill = azul; cell.alignment = center; cell.border = borda

    # ---------- ano mais recente: detalhe mensal ----------
    registros_ano = por_ano[ano]
    por_mes = OrderedDict()
    for r in registros_ano:
        por_mes.setdefault(r[I_DT].month, []).append(r)
    meses = list(por_mes.keys())
    mes_mais_recente = meses[-1] if meses else None

    linhas_mes = []
    prev_ud = None
    AUX_L, AUX_M = 21, 22
    lin = 9

    for m in meses:
        registros = por_mes[m]
        pi = lin
        for r in registros:
            ws.cell(lin, 1, r[I_DT]).number_format = 'dd/mm/yyyy'; ws.cell(lin, 1).alignment = center
            ws.cell(lin, 2, _n(r[I_FAEC])).number_format = F_COTA
            ws.cell(lin, 4, _n(r[I_BBC])).number_format = F_COTA
            ws.cell(lin, 6, _n(r[I_CXC])).number_format = F_COTA
            for col, idx in [(3, I_FAEI), (5, I_BBI), (7, I_CXI), (8, I_SELIC), (9, I_ANB)]:
                v = _n(r[idx]); ws.cell(lin, col, (v/100.0) if v is not None else None).number_format = F_PCT6
            # 6 colunas PC_* vindas da view (já em %)
            for col, idx in [(10, I_PC), (11, I_PC+1), (12, I_PC+2), (13, I_PC+3), (14, I_PC+4), (15, I_PC+5)]:
                v = _n(r[idx]); ws.cell(lin, col, (v/100.0) if v is not None else None).number_format = F_PCT6
            ws.cell(lin, AUX_L, f"=1+H{lin}")
            ws.cell(lin, AUX_M, f"=1+I{lin}")
            ws.row_dimensions[lin].outline_level = 1
            ws.row_dimensions[lin].hidden = (m != mes_mais_recente)
            lin += 1
        ud = lin - 1

        auxL = get_column_letter(AUX_L); auxM = get_column_letter(AUX_M)
        ws.cell(lin, 1, f"{_MES_ABREV[m-1]}/{ano}").font = Font(bold=True)
        base = prev_ud if prev_ud else pi
        ws.cell(lin, 3, f"=(B{ud}/B{base})-1").number_format = F_TOT
        ws.cell(lin, 5, f"=(D{ud}/D{base})-1").number_format = F_TOT
        ws.cell(lin, 7, f"=(F{ud}/F{base})-1").number_format = F_TOT
        ws.cell(lin, 8, f"=PRODUCT({auxL}{pi}:{auxL}{ud})-1").number_format = F_TOT
        ws.cell(lin, 9, f"=PRODUCT({auxM}{pi}:{auxM}{ud})-1").number_format = F_TOT
        ws.cell(lin, 10, f"=C{lin}/I{lin}").number_format = F_TOT
        ws.cell(lin, 11, f"=C{lin}/H{lin}").number_format = F_TOT
        ws.cell(lin, 12, f"=E{lin}/I{lin}").number_format = F_TOT
        ws.cell(lin, 13, f"=E{lin}/H{lin}").number_format = F_TOT
        ws.cell(lin, 14, f"=G{lin}/I{lin}").number_format = F_TOT
        ws.cell(lin, 15, f"=G{lin}/H{lin}").number_format = F_TOT
        for c in range(1, 16):
            ws.cell(lin, c).fill = realce_mes; ws.cell(lin, c).border = borda
        linhas_mes.append(lin)
        prev_ud = ud
        lin += 1

    ws.column_dimensions[get_column_letter(AUX_L)].hidden = True
    ws.column_dimensions[get_column_letter(AUX_M)].hidden = True

    # ACUMULADO do ano mais recente (produtório das linhas mensais)
    lac = lin + 1
    ws.cell(lac, 1, f'ACUMULADO {ano}').font = Font(bold=True)
    def prod_meses(colL):
        return "=" + "*".join([f"(1+{colL}{lmr})" for lmr in linhas_mes]) + "-1"
    for colL, colnum in [('C', 3), ('E', 5), ('G', 7), ('H', 8), ('I', 9)]:
        ws.cell(lac, colnum, prod_meses(colL)).number_format = F_TOT
    for col, (num, den) in {10:('C','I'),11:('C','H'),12:('E','I'),13:('E','H'),14:('G','I'),15:('G','H')}.items():
        ws.cell(lac, col, f"={num}{lac}/{den}{lac}").number_format = F_TOT
    for c in range(1, 16):
        ws.cell(lac, c).fill = realce; ws.cell(lac, c).font = Font(bold=True); ws.cell(lac, c).border = borda

    # ---------- outros anos: só ACUMULADO (valores) ----------
    lin = lac + 2
    for a in outros:
        vals = acum_ano(por_ano[a], base_cotas(a))
        ws.cell(lin, 1, f'ACUMULADO {a}').font = Font(bold=True)
        for colL, colnum in [('C',3),('E',5),('G',7),('H',8),('I',9),
                             ('J',10),('K',11),('L',12),('M',13),('N',14),('O',15)]:
            v = vals[colL]
            cc = ws.cell(lin, colnum, v if v is not None else None)
            cc.number_format = F_TOT
        for c in range(1, 16):
            ws.cell(lin, c).fill = realce_ano; ws.cell(lin, c).font = Font(bold=True); ws.cell(lin, c).border = borda
        lin += 1

    larg = {'A':12,'B':13,'C':12,'D':13,'E':12,'F':13,'G':12,'H':11,'I':13,
            'J':12,'K':12,'L':12,'M':12,'N':12,'O':12}
    for col, w in larg.items(): ws.column_dimensions[col].width = w

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return Response(bio.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="PERFORMANCE_FUNDOS_{ano}.xlsx"'})

@cotas_fundos_bp.route('/rentabilidade-calculada')
@login_required
def rentabilidade_calculada():
    """Página: escolhe uma data e vê a rentabilidade calculada dos 3 fundos."""
    dia_str = (request.args.get('dia') or '').strip()
    dia = None
    if dia_str:
        try:
            dia = datetime.strptime(dia_str, '%Y-%m-%d').date()
        except ValueError:
            dia = None

    resultados = []
    if dia:
        for chave, cfg in FUNDOS.items():
            resultados.append(_rentabilidade_fundo(chave, cfg, dia))

    return render_template(
        'cotas_fundos/rentabilidade_calculada.html',
        dia=dia_str,
        dia_fmt=dia.strftime('%d/%m/%Y') if dia else '',
        resultados=resultados,
        tem_data=bool(dia),
    )


def _rentabilidade_fundo(chave, cfg, dia):
    """Calcula a rentabilidade do fundo na data:
       (SD_BRUTO_atual - SD_BRUTO_anterior) + VR_RESGATE + VR_IR - VR_APLICACAO + VR_IOF.
       Campos ausentes contam como 0. 'anterior' = registro imediatamente anterior."""
    model = cfg['model']
    sd_attr = cfg['sd_bruto']

    def _d(v):
        return Decimal(str(v)) if v is not None else Decimal('0')

    # registro da data escolhida
    atual = db.session.query(model).filter(model.DATA == dia).first()
    if atual is None:
        return {'label': cfg['label'], 'ok': False,
                'msg': f'Não há lançamento em {dia.strftime("%d/%m/%Y")} para este fundo.'}

    # registro imediatamente anterior (a data anterior mais recente)
    anterior = (db.session.query(model)
                .filter(model.DATA < dia)
                .order_by(model.DATA.desc())
                .first())
    if anterior is None:
        return {'label': cfg['label'], 'ok': False,
                'msg': 'Não há registro anterior para comparar.'}

    sd_atual = _d(getattr(atual, sd_attr, None))
    sd_anterior = _d(getattr(anterior, sd_attr, None))
    aplic = _d(getattr(atual, 'VR_APLICACAO', None))
    resg = _d(getattr(atual, 'VR_RESGATE', None))
    ir = _d(getattr(atual, 'VR_IR', None))
    iof = _d(getattr(atual, 'VR_IOF', None))

    rent = (sd_atual - sd_anterior) + resg + ir - aplic + iof

    return {
        'label': cfg['label'], 'ok': True,
        'data_atual': atual.DATA.strftime('%d/%m/%Y'),
        'data_anterior': anterior.DATA.strftime('%d/%m/%Y'),
        'sd_atual': _fmt_vr(sd_atual), 'sd_anterior': _fmt_vr(sd_anterior),
        'aplicacao': _fmt_vr(aplic), 'resgate': _fmt_vr(resg),
        'ir': _fmt_vr(ir), 'iof': _fmt_vr(iof),
        'rentabilidade': _fmt_vr(rent),
        'positiva': rent >= 0,
    }