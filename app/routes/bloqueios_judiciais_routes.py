from datetime import datetime
from decimal import Decimal, InvalidOperation

from flask import Blueprint, render_template, request, jsonify
from flask_login import login_required
from sqlalchemy import text

from app import db
from app.utils.audit import registrar_log
from decimal import Decimal
from datetime import datetime
from sqlalchemy import text
from app import db  # ajuste se o import do db for diferente no seu arquivo
import csv
import io
from flask import Response

from datetime import datetime
from decimal import Decimal
from flask import Response
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

bloqueios_judiciais_bp = Blueprint(
    'bloqueios_judiciais', __name__, url_prefix='/bloqueios-judiciais'
)

_TB = '[BDG].[FIN_TB022_BLOQUEIOS_JUDICIAIS]'


def _parse_data(s):
    s = (s or '').strip()
    if not s:
        return None
    # aceita AAAA-MM-DD, AAAAMMDD, DD/MM/AAAA
    for fmt in ('%Y-%m-%d', '%Y%m%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def _parse_decimal(s):
    s = (s or '').strip()
    if not s:
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _fmt_vr(v):
    if v is None:
        return ''
    inteiro, _, dec = f"{abs(Decimal(str(v))):.2f}".partition('.')
    import re
    inteiro = re.sub(r'(?<=\d)(?=(?:\d{3})+$)', '.', inteiro)
    sinal = '-' if Decimal(str(v)) < 0 else ''
    return f"{sinal}R$ {inteiro},{dec}"


def _carregar_contas():
    sql = text("""
        SELECT ID_CONTA, DSC_CONTA
        FROM [BDG].[PAR_TB027_CONTAS_EMGEA]
        ORDER BY DSC_CONTA
    """)
    return [{'id': r[0], 'dsc': (r[1] or '').strip()}
            for r in db.session.execute(sql).fetchall()]

def _parse_bit(s):
    """'1'/'0'/'sim'/'nao' -> 1/0 ; vazio -> None."""
    s = (s or '').strip().lower()
    if s in ('1', 'sim', 'true', 's'):
        return 1
    if s in ('0', 'nao', 'não', 'false', 'n'):
        return 0
    return None


def _parse_evento(s):
    """Aceita apenas 'D', 'T', 'D/T'. Caso contrário, None."""
    s = (s or '').strip().upper()
    return s if s in ('D', 'T', 'D/T') else None

_MEMO_SUFIXO = '/Gefin/Sufin/Difin'


def _montar_memo_sei(valor):
    """
    Monta o Memo SEI completo pra gravar no banco.
    O usuário digita só a parte variável (ex.: '715/2026') e o sistema anexa o
    sufixo fixo '/Gefin/Sufin/Difin', gravando o valor COMPLETO
    (ex.: '715/2026/Gefin/Sufin/Difin').

    Idempotente: se o valor já vier com o sufixo (ex.: numa reedição), ele é
    removido antes de reanexar — nunca duplica. Tolera espaços e caixa.
    Retorna None quando o usuário não informa nada.
    Compatível com Python 3.9 e 3.12.
    """
    import re
    base = (valor or '').strip()
    if not base:
        return None
    # remove o sufixo fixo (uma ou mais vezes) do fim, se já estiver presente
    base = re.sub(r'(\s*/\s*gefin\s*/\s*sufin\s*/\s*difin\s*)+$', '',
                  base, flags=re.IGNORECASE).strip()
    base = base.rstrip('/').strip()
    if not base:
        return None
    return base + _MEMO_SUFIXO


@bloqueios_judiciais_bp.route('/')
@login_required
def index():
    """Entrada e edição dos Bloqueios Judiciais (FIN_TB022)."""
    f_processo = (request.args.get('processo') or '').strip()
    f_autor = (request.args.get('autor') or '').strip()
    f_conta = (request.args.get('conta') or '').strip()
    f_situacao = (request.args.get('situacao') or 'todos').strip()
    f_memo = (request.args.get('memo_sei') or '').strip()
    f_memo_status = (request.args.get('memo_status') or 'todos').strip()
    f_vr_exato = _parse_decimal(request.args.get('vr_exato'))
    f_dt_dep = _parse_data(request.args.get('dt_deposito'))
    f_dt_desb = _parse_data(request.args.get('dt_desbloqueio'))

    condicoes, params = [], {}
    if f_processo:
        condicoes.append("PROCESSO LIKE :p_proc"); params['p_proc'] = f'%{f_processo}%'
    if f_autor:
        condicoes.append("AUTOR LIKE :p_autor"); params['p_autor'] = f'%{f_autor}%'
    if f_conta:
        condicoes.append("CONTA = :p_conta"); params['p_conta'] = f_conta
    if f_memo:
        condicoes.append("MEMO_SEI LIKE :p_memo"); params['p_memo'] = f'%{f_memo}%'
    if f_memo_status == 'preenchido':
        condicoes.append("MEMO_SEI IS NOT NULL AND LTRIM(RTRIM(MEMO_SEI)) <> ''")
    elif f_memo_status == 'nao_preenchido':
        condicoes.append("(MEMO_SEI IS NULL OR LTRIM(RTRIM(MEMO_SEI)) = '')")
    if f_dt_dep:
        condicoes.append("CONVERT(varchar(8), DT_DEPOSITO, 112) = :p_dtdep")
        params['p_dtdep'] = f_dt_dep.strftime('%Y%m%d')
    if f_dt_desb:
        condicoes.append("CONVERT(varchar(8), DT_DESBLOQUEIO, 112) = :p_dtdesb")
        params['p_dtdesb'] = f_dt_desb.strftime('%Y%m%d')
    if f_situacao == 'bloqueado':
        condicoes.append("DT_DESBLOQUEIO IS NULL")
    elif f_situacao == 'desbloqueado':
        condicoes.append("DT_DESBLOQUEIO IS NOT NULL")
    if f_vr_exato is not None:
        condicoes.append("VR_BLOQUEADO = :p_vrex")
        params['p_vrex'] = f_vr_exato
    where = ("WHERE " + " AND ".join(condicoes)) if condicoes else ""

    sql = text(f"""
        SELECT DT_DEPOSITO, VR_BLOQUEADO, CONTA, PROCESSO, VARA, AUTOR, DT_DESBLOQUEIO,
               EXCE, EVENTO, MEMO_SEI
        FROM {_TB} {where}
        ORDER BY DT_DEPOSITO DESC, PROCESSO
    """)
    rows = db.session.execute(sql, params).fetchall()

    lista = []
    for r in rows:
        exce = r[7]
        lista.append({
            'dt_deposito': r[0], 'dt_deposito_iso': r[0].strftime('%Y-%m-%d') if r[0] else '',
            'vr': r[1], 'vr_str': ('' if r[1] is None else str(r[1])), 'vr_fmt': _fmt_vr(r[1]),
            'conta': (r[2] or ''), 'processo': (r[3] or ''),
            'vara': (r[4] or ''), 'autor': (r[5] or ''),
            'dt_desbloqueio': r[6], 'dt_desbloqueio_iso': r[6].strftime('%Y-%m-%d') if r[6] else '',
            'exce': exce,
            'exce_str': ('' if exce is None else ('1' if exce else '0')),
            'exce_lbl': ('Sim' if exce == 1 else ('Não' if exce == 0 else '—')),
            'evento': (r[8] or ''),
            'memo_sei': (r[9] or ''),
        })

    # ===== Totais sensíveis aos filtros (calculados sobre a lista já filtrada) =====
    qtd = len(lista)
    total_valor = Decimal('0')
    qtd_bloq = 0
    total_bloq = Decimal('0')
    qtd_desb = 0
    total_desb = Decimal('0')
    for l in lista:
        vr = Decimal(str(l['vr'])) if l['vr'] is not None else Decimal('0')
        total_valor += vr
        if l['dt_desbloqueio']:
            qtd_desb += 1; total_desb += vr
        else:
            qtd_bloq += 1; total_bloq += vr

    totais = {
        'qtd': qtd,
        'valor': _fmt_vr(total_valor),
        'qtd_bloq': qtd_bloq, 'valor_bloq': _fmt_vr(total_bloq),
        'qtd_desb': qtd_desb, 'valor_desb': _fmt_vr(total_desb),
    }

    return render_template(
        'bloqueios_judiciais/index.html',
        contas=_carregar_contas(), lista=lista, totais=totais,
        filtros={'processo': f_processo, 'autor': f_autor, 'conta': f_conta,
                 'situacao': f_situacao, 'memo_sei': f_memo, 'memo_status': f_memo_status,
                 'vr_exato': request.args.get('vr_exato', ''),
                 'dt_deposito': request.args.get('dt_deposito', ''),
                 'dt_desbloqueio': request.args.get('dt_desbloqueio', '')},
    )


@bloqueios_judiciais_bp.route('/incluir', methods=['POST'])
@login_required
def incluir():
    dt_dep = _parse_data(request.form.get('DT_DEPOSITO'))
    vr = _parse_decimal(request.form.get('VR_BLOQUEADO'))
    conta = (request.form.get('CONTA') or '').strip()
    processo = (request.form.get('PROCESSO') or '').strip()
    vara = (request.form.get('VARA') or '').strip()
    autor = (request.form.get('AUTOR') or '').strip()
    exce = _parse_bit(request.form.get('EXCE'))
    memo_sei = _montar_memo_sei(request.form.get('MEMO_SEI'))  # monta o valor completo

    if not dt_dep:
        return jsonify({'success': False, 'message': 'Informe a data do depósito.'}), 400
    if vr is None:
        return jsonify({'success': False, 'message': 'Informe o valor bloqueado.'}), 400
    if not conta:
        return jsonify({'success': False, 'message': 'Selecione a conta.'}), 400
    if not processo:
        return jsonify({'success': False, 'message': 'Informe o processo.'}), 400

    try:
        db.session.execute(text(f"""
            INSERT INTO {_TB}
                (DT_DEPOSITO, VR_BLOQUEADO, CONTA, PROCESSO, VARA, AUTOR,
                 DT_DESBLOQUEIO, EXCE, EVENTO, MEMO_SEI)
            VALUES (:dt_dep, :vr, :conta, :processo, :vara, :autor,
                    NULL, :exce, NULL, :memo)
        """), {'dt_dep': dt_dep, 'vr': vr, 'conta': conta,
               'processo': processo, 'vara': vara, 'autor': autor,
               'exce': exce, 'memo': memo_sei})
        db.session.commit()

        registrar_log(
            acao='inclusao', entidade='bloqueios_judiciais', entidade_id=None,
            descricao=f'Novo bloqueio — processo {processo}',
            dados_novos={'DT_DEPOSITO': dt_dep.strftime('%Y-%m-%d'), 'VR_BLOQUEADO': str(vr),
                         'CONTA': conta, 'PROCESSO': processo, 'VARA': vara, 'AUTOR': autor,
                         'EXCE': exce, 'MEMO_SEI': memo_sei},
        )
        return jsonify({'success': True, 'message': f'Bloqueio do processo {processo} incluído.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro ao incluir: {str(e)}'}), 500


@bloqueios_judiciais_bp.route('/editar', methods=['POST'])
@login_required
def editar():
    dt_dep = _parse_data(request.form.get('DT_DEPOSITO'))
    vr = _parse_decimal(request.form.get('VR_BLOQUEADO'))
    conta = (request.form.get('CONTA') or '').strip()
    processo = (request.form.get('PROCESSO') or '').strip()
    vara = (request.form.get('VARA') or '').strip()
    autor = (request.form.get('AUTOR') or '').strip()
    dt_desb = _parse_data(request.form.get('DT_DESBLOQUEIO'))  # pode ser None
    exce = _parse_bit(request.form.get('EXCE'))
    evento = _parse_evento(request.form.get('EVENTO')) if dt_desb else None  # só se desbloqueado
    memo_sei = _montar_memo_sei(request.form.get('MEMO_SEI'))  # monta o valor completo (idempotente)

    if not dt_dep or vr is None or not conta or not processo:
        return jsonify({'success': False,
                        'message': 'Data do depósito, valor, conta e processo são obrigatórios.'}), 400

    o_dep = _parse_data(request.form.get('o_DT_DEPOSITO'))
    o_vr = _parse_decimal(request.form.get('o_VR_BLOQUEADO'))
    o_desb = _parse_data(request.form.get('o_DT_DESBLOQUEIO'))
    params = {
        'dt_dep': dt_dep, 'vr': vr, 'conta': conta, 'processo': processo,
        'vara': vara, 'autor': autor, 'dt_desb': dt_desb, 'exce': exce, 'evento': evento,
        'memo': (memo_sei or None),
        'o_dep': o_dep.strftime('%Y%m%d') if o_dep else '',
        'o_vr': o_vr,
        'o_conta': (request.form.get('o_CONTA') or '').strip(),
        'o_processo': (request.form.get('o_PROCESSO') or '').strip(),
        'o_vara': (request.form.get('o_VARA') or '').strip(),
        'o_autor': (request.form.get('o_AUTOR') or '').strip(),
        'o_desb': o_desb.strftime('%Y%m%d') if o_desb else '',
    }

    try:
        result = db.session.execute(text(f"""
            UPDATE {_TB}
            SET DT_DEPOSITO = :dt_dep, VR_BLOQUEADO = :vr, CONTA = :conta,
                PROCESSO = :processo, VARA = :vara, AUTOR = :autor,
                DT_DESBLOQUEIO = :dt_desb, EXCE = :exce, EVENTO = :evento,
                MEMO_SEI = :memo
            WHERE CONVERT(varchar(8), DT_DEPOSITO, 112) = :o_dep
              AND ROUND(VR_BLOQUEADO, 2) = ROUND(:o_vr, 2)
              AND RTRIM(ISNULL(CONTA, '')) = RTRIM(:o_conta)
              AND RTRIM(ISNULL(PROCESSO, '')) = RTRIM(:o_processo)
              AND RTRIM(ISNULL(VARA, '')) = RTRIM(:o_vara)
              AND RTRIM(ISNULL(AUTOR, '')) = RTRIM(:o_autor)
        """), params)
        db.session.commit()

        if result.rowcount == 0:
            return jsonify({'success': False,
                            'message': 'Nenhuma linha correspondente encontrada (a linha pode ter sido alterada).'}), 404

        registrar_log(
            acao='atualizacao', entidade='bloqueios_judiciais', entidade_id=None,
            descricao=f'Edição de bloqueio — processo {processo}',
            dados_novos={'DT_DEPOSITO': dt_dep.strftime('%Y-%m-%d'), 'VR_BLOQUEADO': str(vr),
                         'CONTA': conta, 'PROCESSO': processo, 'VARA': vara, 'AUTOR': autor,
                         'DT_DESBLOQUEIO': dt_desb.strftime('%Y-%m-%d') if dt_desb else None,
                         'EXCE': exce, 'EVENTO': evento, 'MEMO_SEI': memo_sei},
        )
        aviso = '' if result.rowcount == 1 else f' ({result.rowcount} linhas idênticas atualizadas)'
        return jsonify({'success': True, 'message': f'Bloqueio atualizado.{aviso}'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro ao editar: {str(e)}'}), 500

def _bj_moeda(v):
    if v is None:
        return '-'
    s = f"{Decimal(str(v)):,.2f}"           # 1,234.56
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')  # 1.234,56


def _bj_data(d):
    if d is None:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime('%d/%m/%Y')
    s = str(d)[:10]
    try:
        y, m, dd = s.split('-')
        return f"{dd}/{m}/{y}"
    except Exception:
        return s


@bloqueios_judiciais_bp.route('/planilha')
@login_required
def planilha():
    """Planilha de Bloqueios Judiciais em Vigor (FIN_VW033), agrupada por conta."""
    rows = db.session.execute(text("""
        SELECT ORDEM, CONTA, DT_DEPOSITO, VR_BLOQUEADO, PROCESSO, VARA, AUTOR
        FROM [BDG].[FIN_VW033_BLOQUEIOS_JUDICIAIS_PLANILHA]
        ORDER BY ORDEM, DT_DEPOSITO DESC
    """)).fetchall()

    grupos = []            # [{conta, linhas[], total}]
    idx_por_conta = {}
    total_geral = Decimal('0')

    for r in rows:
        conta = (r[1] or '').strip()
        vr = Decimal(str(r[3])) if r[3] is not None else Decimal('0')
        total_geral += vr

        if conta not in idx_por_conta:
            idx_por_conta[conta] = len(grupos)
            grupos.append({'conta': conta, 'linhas': [], 'total': Decimal('0')})

        g = grupos[idx_por_conta[conta]]
        g['linhas'].append({
            'data': _bj_data(r[2]),
            'valor': _bj_moeda(vr),
            'processo': (r[4] or '').strip(),
            'vara': (r[5] or '').strip(),
            'autor': (r[6] or '').strip(),
        })
        g['total'] += vr

    # formata os totais por grupo
    for g in grupos:
        g['total_fmt'] = _bj_moeda(g['total'])

    return render_template(
        'bloqueios_judiciais/planilha.html',
        grupos=grupos,
        total_geral=_bj_moeda(total_geral),
        qtd_contas=len(grupos),
        data_posicao=datetime.now().strftime('%d/%m/%Y'),
        sem_dados=(len(grupos) == 0),
    )


@bloqueios_judiciais_bp.route('/planilha/excel')
@login_required
def planilha_excel():
    """Exporta Bloqueios Judiciais em Vigor (FIN_VW033) em .xlsx, no layout do modelo."""
    rows = db.session.execute(text("""
        SELECT ORDEM, CONTA, DT_DEPOSITO, VR_BLOQUEADO, PROCESSO, VARA, AUTOR
        FROM [BDG].[FIN_VW033_BLOQUEIOS_JUDICIAIS_PLANILHA]
        ORDER BY ORDEM, DT_DEPOSITO DESC
    """)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = 'BLOQ VIGOR'

    # Larguras (iguais ao modelo)
    larguras = {'A': 25.7, 'B': 20.7, 'C': 29.3, 'D': 101.3, 'E': 70.7}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    thin = Side(style='thin')
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right')
    FMT_DATA = 'mm-dd-yy'
    FMT_MOEDA = '"R$"\\ #,##0.00'

    # --- Título (A1:B1 mesclado, tam.14, centralizado) + data (B2, d.m.yyyy) ---
    ws.merge_cells('A1:B1')
    ws['A1'] = 'Bloqueios Judiciais em Vigor'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center
    ws['B2'] = datetime.now()
    ws['B2'].font = Font(bold=True, size=11)
    ws['B2'].alignment = right
    ws['B2'].number_format = 'd\\.m\\.yyyy'

    linha = 3
    conta_atual = None
    ini_dados = None       # primeira linha de dados do bloco (p/ o SUM)
    ultima_dado = None
    total_geral = 0.0

    def escreve_total(l_total, l_ini, l_fim):
        ws.cell(l_total, 1, 'TOTAL').font = Font(bold=True, size=12)
        ws.cell(l_total, 1).alignment = center
        c = ws.cell(l_total, 2, f'=SUM(B{l_ini}:B{l_fim})')
        c.font = Font(bold=True, size=12)
        c.alignment = right
        c.number_format = FMT_MOEDA

    for r in rows:
        conta = (r[1] or '').strip()

        if conta != conta_atual:
            # fecha total do bloco anterior
            if conta_atual is not None:
                escreve_total(linha, ini_dados, ultima_dado)
                linha += 2  # total + linha em branco

            # cabeçalho do novo bloco (conta em A:B + PROCESSO/VARA/AUTOR), negrito, borda
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=2)
            ca = ws.cell(linha, 1, conta)
            ca.font = Font(bold=True, size=12); ca.alignment = center; ca.border = borda
            ws.cell(linha, 2).border = borda
            for col, txt in [(3, 'PROCESSO'), (4, 'VARA'), (5, 'AUTOR')]:
                cc = ws.cell(linha, col, txt)
                cc.font = Font(bold=True, size=12); cc.alignment = center; cc.border = borda
            conta_atual = conta
            linha += 1
            ini_dados = linha

        # linha de dados
        cA = ws.cell(linha, 1, r[2])                     # data (datetime)
        cA.alignment = center; cA.font = Font(size=12); cA.number_format = FMT_DATA
        vr = float(r[3]) if r[3] is not None else 0.0
        total_geral += vr
        cB = ws.cell(linha, 2, vr)                       # valor
        cB.alignment = right; cB.font = Font(size=12); cB.number_format = FMT_MOEDA
        ws.cell(linha, 3, (r[4] or '').strip()).font = Font(size=10)
        ws.cell(linha, 3).alignment = center
        ws.cell(linha, 4, (r[5] or '').strip()).font = Font(size=12)
        ws.cell(linha, 4).alignment = center
        ws.cell(linha, 5, (r[6] or '').strip()).font = Font(size=12)
        ws.cell(linha, 5).alignment = center

        ultima_dado = linha
        linha += 1

    # total do último bloco + TOTAL GERAL na última linha
    if conta_atual is not None:
        escreve_total(linha, ini_dados, ultima_dado)
        linha += 2  # pula o total + 1 linha em branco

        cg = ws.cell(linha, 1, 'TOTAL GERAL')
        cg.font = Font(bold=True, size=12); cg.alignment = center
        cgv = ws.cell(linha, 2, total_geral)
        cgv.font = Font(bold=True, size=12); cgv.alignment = right
        cgv.number_format = FMT_MOEDA

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    nome = f"BLOQUEIOS_JUDICIAIS_EM_VIGOR_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        bio.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nome}"'}
    )

    def escreve_total(l_total, l_ini, l_fim):
        ws.cell(l_total, 1, 'TOTAL').font = Font(bold=True, size=12)
        ws.cell(l_total, 1).alignment = center
        c = ws.cell(l_total, 2, f'=SUM(B{l_ini}:B{l_fim})')
        c.font = Font(bold=True, size=12)
        c.alignment = right
        c.number_format = FMT_MOEDA

    ultima_dado = None
    for r in rows:
        conta = (r[1] or '').strip()

        if conta != conta_atual:
            # fecha total do bloco anterior
            if conta_atual is not None:
                l_total = linha
                escreve_total(l_total, ini_dados, ultima_dado)
                linha += 2  # total + linha em branco

            # cabeçalho do novo bloco (conta em A:B + PROCESSO/VARA/AUTOR), negrito, borda
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=2)
            ca = ws.cell(linha, 1, conta)
            ca.font = Font(bold=True, size=12); ca.alignment = center; ca.border = borda
            ws.cell(linha, 2).border = borda
            for col, txt in [(3, 'PROCESSO'), (4, 'VARA'), (5, 'AUTOR')]:
                cc = ws.cell(linha, col, txt)
                cc.font = Font(bold=True, size=12); cc.alignment = center; cc.border = borda
            conta_atual = conta
            linha += 1
            ini_dados = linha

        # linha de dados
        cA = ws.cell(linha, 1, r[2])                    # data (datetime)
        cA.alignment = center; cA.font = Font(size=12); cA.number_format = FMT_DATA
        vr = float(r[3]) if r[3] is not None else 0.0
        cB = ws.cell(linha, 2, vr)                       # valor
        cB.alignment = right; cB.font = Font(size=12); cB.number_format = FMT_MOEDA
        ws.cell(linha, 3, (r[4] or '').strip()).font = Font(size=10)
        ws.cell(linha, 3).alignment = center
        ws.cell(linha, 4, (r[5] or '').strip()).font = Font(size=12)
        ws.cell(linha, 4).alignment = center
        ws.cell(linha, 5, (r[6] or '').strip()).font = Font(size=12)
        ws.cell(linha, 5).alignment = center

        ultima_dado = linha
        linha += 1

    # total do último bloco
    if conta_atual is not None:
        escreve_total(linha, ini_dados, ultima_dado)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    nome = f"BLOQUEIOS_JUDICIAIS_EM_VIGOR_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        bio.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nome}"'}
    )

@bloqueios_judiciais_bp.route('/liberar-parcial', methods=['POST'])
@login_required
def liberar_parcial():
    """Libera PARTE do valor de um bloqueio:
       - cria um novo registro (idêntico ao original) com o valor liberado e a
         DT_DESBLOQUEIO informada, mantendo a DT_DEPOSITO original;
       - subtrai o valor liberado do registro original, que continua bloqueado.
       A linha original é identificada pelos valores originais (o_*)."""
    valor_liberado = _parse_decimal(request.form.get('VR_LIBERADO'))
    dt_desb = _parse_data(request.form.get('DT_DESBLOQUEIO'))

    # Valores ORIGINAIS (identificam a linha) — mesmo esquema do editar
    o_dep = _parse_data(request.form.get('o_DT_DEPOSITO'))
    o_vr = _parse_decimal(request.form.get('o_VR_BLOQUEADO'))
    o_conta = (request.form.get('o_CONTA') or '').strip()
    o_processo = (request.form.get('o_PROCESSO') or '').strip()
    o_vara = (request.form.get('o_VARA') or '').strip()
    o_autor = (request.form.get('o_AUTOR') or '').strip()
    evento = _parse_evento(request.form.get('EVENTO'))

    if not o_dep or o_vr is None or not o_processo:
        return jsonify({'success': False, 'message': 'Registro original inválido.'}), 400
    if valor_liberado is None or valor_liberado <= 0:
        return jsonify({'success': False, 'message': 'Informe um valor liberado maior que zero.'}), 400
    if valor_liberado >= o_vr:
        return jsonify({'success': False,
                        'message': f'O valor liberado deve ser menor que o valor bloqueado '
                                   f'({_fmt_vr(o_vr)}). Para liberar tudo, use a edição e informe a data de desbloqueio.'}), 400
    if not dt_desb:
        return jsonify({'success': False, 'message': 'Informe a data do desbloqueio (liberação).'}), 400

    novo_valor_original = o_vr - valor_liberado

    params = {
        'lib': valor_liberado, 'novo': novo_valor_original, 'dt_desb': dt_desb,
        'o_dep': o_dep.strftime('%Y%m%d'), 'o_vr': o_vr,
        'o_conta': o_conta, 'o_processo': o_processo, 'o_vara': o_vara, 'o_autor': o_autor,
    }

    try:
        # 1) Subtrai a parte liberada do registro original (continua bloqueado)
        upd = db.session.execute(text(f"""
            UPDATE {_TB}
            SET VR_BLOQUEADO = :novo
            WHERE CONVERT(varchar(8), DT_DEPOSITO, 112) = :o_dep
              AND VR_BLOQUEADO = :o_vr
              AND ISNULL(CONTA, '') = :o_conta
              AND ISNULL(PROCESSO, '') = :o_processo
              AND ISNULL(VARA, '') = :o_vara
              AND ISNULL(AUTOR, '') = :o_autor
              AND DT_DESBLOQUEIO IS NULL
        """), params)

        if upd.rowcount == 0:
            db.session.rollback()
            return jsonify({'success': False,
                            'message': 'Registro original não encontrado (pode ter sido alterado).'}), 404

        # 2) Cria o novo registro (parte liberada), igual ao original,
        #    com a mesma DT_DEPOSITO e a DT_DESBLOQUEIO informada
        db.session.execute(text(f"""
            INSERT INTO {_TB}
                (DT_DEPOSITO, VR_BLOQUEADO, CONTA, PROCESSO, VARA, AUTOR, DT_DESBLOQUEIO, EXCE, EVENTO)
            VALUES (:dt_dep, :lib, :conta, :processo, :vara, :autor, :dt_desb, :exce, :evento)
        """), {'dt_dep': o_dep, 'lib': valor_liberado, 'conta': o_conta,
               'processo': o_processo, 'vara': o_vara, 'autor': o_autor,
               'dt_desb': dt_desb, 'exce': _parse_bit(request.form.get('EXCE')), 'evento': evento})

        db.session.commit()

        registrar_log(
            acao='liberacao_parcial', entidade='bloqueios_judiciais', entidade_id=None,
            descricao=f'Liberação parcial — processo {o_processo}: liberado {_fmt_vr(valor_liberado)}, '
                      f'restante bloqueado {_fmt_vr(novo_valor_original)}',
            dados_novos={'DT_DEPOSITO': o_dep.strftime('%Y-%m-%d'),
                         'VR_LIBERADO': str(valor_liberado),
                         'VR_RESTANTE': str(novo_valor_original),
                         'DT_DESBLOQUEIO': dt_desb.strftime('%Y-%m-%d'),
                         'CONTA': o_conta, 'PROCESSO': o_processo,
                         'VARA': o_vara, 'AUTOR': o_autor},
        )
        return jsonify({'success': True,
                        'message': f'Liberado {_fmt_vr(valor_liberado)} do processo {o_processo}. '
                                   f'Restam {_fmt_vr(novo_valor_original)} bloqueados.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro na liberação parcial: {str(e)}'}), 500


def _fmt_data_ponto(d):
    """DD.M.AAAA (dia com zero à esquerda não; mês sem zero, como no modelo)."""
    if not d:
        return ''
    return f"{d.day}.{d.month}.{d.year}"


def _dados_bloqueio_diario(dia):
    """Registros com DEPÓSITO = dia OU DESBLOQUEIO = dia, AGRUPADOS por
    (processo, vara, autor, memo_sei, data do depósito). Cada grupo = um cartão
    com várias contas e o total somado. Situação aparece quando há desbloqueio na data."""
    from collections import OrderedDict
    dia_112 = dia.strftime('%Y%m%d')
    rows = db.session.execute(text(f"""
        SELECT DT_DEPOSITO, VR_BLOQUEADO, CONTA, PROCESSO, VARA, AUTOR,
               DT_DESBLOQUEIO, MEMO_SEI
        FROM {_TB}
        WHERE CONVERT(varchar(8), DT_DEPOSITO, 112) = :d
           OR CONVERT(varchar(8), DT_DESBLOQUEIO, 112) = :d
        ORDER BY PROCESSO, DT_DEPOSITO
    """), {'d': dia_112}).fetchall()

    grupos = OrderedDict()
    for r in rows:
        dep, vr, conta, proc, vara, autor, desb, memo = r
        dep_d = dep.date() if hasattr(dep, 'date') else dep
        desb_d = desb.date() if (desb and hasattr(desb, 'date')) else desb

        chave = (
            (proc or '').strip(), (vara or '').strip(),
            (autor or '').strip(), (memo or '').strip(),
            dep_d.strftime('%Y%m%d') if dep_d else '',
        )

        if chave not in grupos:
            grupos[chave] = {
                'dt_bloqueio': _fmt_data_ponto(dep_d),
                'processo': (proc or ''), 'vara': (vara or ''),
                'autor': (autor or ''), 'memo_sei': (memo or ''),
                'contas': [], 'total': Decimal('0'),
                'desb_na_data': False, 'dt_transf': '',
            }

        g = grupos[chave]
        vrd = Decimal(str(vr)) if vr is not None else Decimal('0')
        g['contas'].append({'conta': (conta or ''), 'vr_fmt': _fmt_vr(vrd), 'valor': vrd})
        g['total'] += vrd

        if desb_d and desb_d.strftime('%Y%m%d') == dia_112:
            g['desb_na_data'] = True
            g['dt_transf'] = _fmt_data_ponto(desb_d)

    cartoes = []
    for g in grupos.values():
        situacao = ''
        if g['desb_na_data']:
            situacao = f"Desbloqueados e Transferidos em {g['dt_transf']}"
        cartoes.append({
            'dt_bloqueio': g['dt_bloqueio'], 'dt_transf': g['dt_transf'],
            'processo': g['processo'], 'vara': g['vara'],
            'autor': g['autor'], 'memo_sei': g['memo_sei'],
            'contas': g['contas'],
            'total_fmt': _fmt_vr(g['total']), 'total_valor': g['total'],
            'situacao': situacao,
        })
    return cartoes


@bloqueios_judiciais_bp.route('/bloqueio-diario')
@login_required
def bloqueio_diario():
    """Página do Bloqueio Diário: escolhe uma data e vê os cartões."""
    dia_str = (request.args.get('dia') or '').strip()
    dia = _parse_data(dia_str)
    cartoes = _dados_bloqueio_diario(dia) if dia else []
    return render_template(
        'bloqueios_judiciais/bloqueio_diario.html',
        dia=dia_str, dia_fmt=_fmt_data_ponto(dia) if dia else '',
        cartoes=cartoes, tem_data=bool(dia),
    )


@bloqueios_judiciais_bp.route('/bloqueio-diario/excel')
@login_required
def bloqueio_diario_excel():
    """Excel do dia (uma aba = a data), agrupado por processo/vara/autor/memo/data.
    Várias contas somando o total; Situação vertical à direita."""
    from openpyxl.styles import PatternFill
    dia = _parse_data((request.args.get('dia') or '').strip())
    if not dia:
        return Response('Informe a data.', mimetype='text/plain')
    cartoes = _dados_bloqueio_diario(dia)

    wb = Workbook(); ws = wb.active
    ws.title = _fmt_data_ponto(dia)[:31]
    ws.sheet_view.showGridLines = False

    fina = Side(style='thin', color='D6DEEA')
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    azul_txt = Font(bold=True, color='1F3A5F')
    azul_hdr = PatternFill('solid', fgColor='1F4E79')
    branco_b = Font(bold=True, color='FFFFFF')
    azul_claro = PatternFill('solid', fgColor='EAF1FB')
    cinza = PatternFill('solid', fgColor='F5F7FA')
    verm = Font(bold=True, color='C0392B')
    F_MOEDA = '"R$"\\ #,##0.00'

    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 34

    ws.merge_cells('A1:C1')
    t = ws.cell(1, 1, f"Bloqueios do dia {_fmt_data_ponto(dia)}")
    t.font = Font(bold=True, size=15, color='1F3A5F'); t.alignment = center
    ws.row_dimensions[1].height = 26

    lin = 3
    for c in cartoes:
        tem_desb = bool(c.get('situacao'))
        n_datas = 2 if tem_desb else 1

        # Topo: datas (A:B) em faixa azul
        ws.cell(lin, 1, f"Data do Bloqueio: {c['dt_bloqueio']}")
        ws.cell(lin, 1).font = branco_b; ws.cell(lin, 1).fill = azul_hdr; ws.cell(lin, 1).alignment = left
        ws.cell(lin, 2).fill = azul_hdr
        if tem_desb:
            ws.cell(lin + 1, 1, f"Data dos Desbloqueios/Transferências: {c['dt_transf']}")
            ws.cell(lin + 1, 1).font = branco_b; ws.cell(lin + 1, 1).fill = azul_hdr; ws.cell(lin + 1, 1).alignment = left
            ws.cell(lin + 1, 2).fill = azul_hdr

        ini_bloco = lin
        base = lin + n_datas

        # Processo/Vara/Autor/Memo SEI (memo ao lado do autor)
        ws.merge_cells(start_row=base, start_column=1, end_row=base, end_column=2)
        linha_proc = (f"Processo: {c['processo']}"
                      + (f"   ·   Vara: {c['vara']}" if c['vara'] else "")
                      + (f"   ·   Autor: {c['autor']}" if c['autor'] else "")
                      + (f"   ·   Memo SEI: {c['memo_sei']}" if c['memo_sei'] else ""))
        ws.cell(base, 1, linha_proc).alignment = left

        # Valor da ordem = total
        ws.cell(base + 1, 1, 'Valor da ordem:').font = azul_txt
        vo = ws.cell(base + 1, 2, float(c['total_valor']))
        vo.number_format = F_MOEDA; vo.font = azul_txt; vo.alignment = right

        # Cabeçalho da tabela conta
        ws.cell(base + 2, 1, 'Conta corrente/Investimento').font = azul_txt
        ws.cell(base + 2, 1).fill = azul_claro; ws.cell(base + 2, 1).border = borda
        h2 = ws.cell(base + 2, 2, 'Valor Bloqueado'); h2.font = azul_txt; h2.fill = azul_claro
        h2.alignment = right; h2.border = borda

        # Linhas das contas
        rr = base + 3
        for ct in c['contas']:
            ws.cell(rr, 1, ct['conta']).border = borda
            vb = ws.cell(rr, 2, float(ct['valor']))
            vb.number_format = F_MOEDA; vb.alignment = right; vb.border = borda
            rr += 1

        # Total
        ws.cell(rr, 1, 'Total Bloqueado').font = azul_txt
        ws.cell(rr, 1).fill = cinza; ws.cell(rr, 1).border = borda
        vt = ws.cell(rr, 2, float(c['total_valor']))
        vt.number_format = F_MOEDA; vt.font = azul_txt; vt.alignment = right; vt.fill = cinza; vt.border = borda

        fim_bloco = rr

        # Situação: título na faixa azul + texto no corpo (coluna C)
        ws.merge_cells(start_row=ini_bloco, start_column=3, end_row=ini_bloco + n_datas - 1, end_column=3)
        stit = ws.cell(ini_bloco, 3, 'Situação')
        stit.font = branco_b; stit.fill = azul_hdr; stit.alignment = center

        ws.merge_cells(start_row=base, start_column=3, end_row=fim_bloco, end_column=3)
        sc = ws.cell(base, 3, c['situacao'] if tem_desb else '')
        sc.alignment = center; sc.font = verm; sc.border = borda

        lin = fim_bloco + 2

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    nome = f"BLOQUEIO_DIARIO_{dia.strftime('%Y%m%d')}.xlsx"
    return Response(
        bio.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nome}"'})