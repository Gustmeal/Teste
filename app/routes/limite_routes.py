from flask import Blueprint, render_template, redirect, url_for, request, flash, send_file, jsonify
from app.models.limite_distribuicao import LimiteDistribuicao
from app.models.edital import Edital
from app.models.periodo import PeriodoAvaliacao
from app.models.empresa_participante import EmpresaParticipante
from app.models.criterio_selecao import CriterioSelecao
from app.utils.distribuir_contratos import obter_resultados_finais_distribuicao
from app.utils.redistribuir_contratos import processar_redistribuicao_contratos
from app import db
from datetime import datetime
from flask_login import login_required, current_user
from app.utils.audit import registrar_log
from sqlalchemy import or_, func, text
import math
import random
import logging

limite_bp = Blueprint('limite', __name__, url_prefix='/credenciamento')


@limite_bp.context_processor
def inject_current_year():
    return {'current_year': datetime.utcnow().year}

# Função auxiliar para truncar valores decimais
def truncate_decimal(valor, casas=2):
    """
    Trunca um valor decimal para o número de casas especificado,
    sem realizar nenhum arredondamento.
    """
    fator = 10 ** casas
    return int(valor * fator) / fator

@limite_bp.app_template_filter('truncate_2dec')
def truncate_2dec_filter(value):
    """Filtro Jinja2 para truncar valores com duas casas decimais."""
    if isinstance(value, (int, float)):
        return "{0:.2f}".format(int(value * 100) / 100)
    return value

@limite_bp.route('/limites')
@login_required
def lista_limites():
    from sqlalchemy.orm import joinedload

    periodo_id = request.args.get('periodo_id', type=int)
    criterio_id = request.args.get('criterio_id', type=int)

    query = db.session.query(
        LimiteDistribuicao,
        CriterioSelecao.DS_CRITERIO_SELECAO
    ).outerjoin(
        CriterioSelecao,
        LimiteDistribuicao.COD_CRITERIO_SELECAO == CriterioSelecao.COD
    ).options(
        joinedload(LimiteDistribuicao.periodo),
        joinedload(LimiteDistribuicao.edital)
    ).filter(
        LimiteDistribuicao.DELETED_AT == None,
        or_(CriterioSelecao.DELETED_AT == None, CriterioSelecao.DELETED_AT.is_(None))
    )

    if periodo_id:
        query = query.filter(LimiteDistribuicao.ID_PERIODO == periodo_id)

    if criterio_id:
        query = query.filter(LimiteDistribuicao.COD_CRITERIO_SELECAO == criterio_id)

    results = query.all()

    limites = []
    for limite, ds_criterio in results:
        # Obter nome da empresa
        empresas = EmpresaParticipante.query.filter_by(ID_EMPRESA=limite.ID_EMPRESA).all()
        empresa_nome = None
        empresa_nome_abreviado = None
        if empresas:
            for empresa in empresas:
                if empresa.NO_EMPRESA:
                    empresa_nome = empresa.NO_EMPRESA
                    empresa_nome_abreviado = empresa.NO_EMPRESA_ABREVIADA
                    break

        if not empresa_nome:
            try:
                from app.models.empresa_responsavel import EmpresaResponsavel
                empresa_resp = EmpresaResponsavel.query.filter_by(
                    pkEmpresaResponsavelCobranca=limite.ID_EMPRESA).first()
                if empresa_resp:
                    empresa_nome = empresa_resp.nmEmpresaResponsavelCobranca
                    empresa_nome_abreviado = empresa_resp.NO_ABREVIADO_EMPRESA
            except:
                empresa_nome = f"Empresa ID {limite.ID_EMPRESA}"
                empresa_nome_abreviado = f"ID {limite.ID_EMPRESA}"

        limite.empresa_nome = empresa_nome
        limite.empresa_nome_abreviado = empresa_nome_abreviado
        limite.criterio_descricao = ds_criterio if ds_criterio else f"Critério {limite.COD_CRITERIO_SELECAO}"

        limites.append(limite)

    periodos = PeriodoAvaliacao.query.filter(PeriodoAvaliacao.DELETED_AT == None).all()
    criterios = CriterioSelecao.query.filter(CriterioSelecao.DELETED_AT == None).all()

    return render_template(
        'credenciamento/lista_limites.html',
        limites=limites,
        periodos=periodos,
        criterios=criterios,
        filtro_periodo_id=periodo_id,
        filtro_criterio_id=criterio_id
    )


def selecionar_contratos():
    """
    VERSÃO OTIMIZADA - Permite contratos com suspensão judicial
    """
    try:
        with db.engine.connect() as connection:
            try:
                logging.info("Limpando tabelas...")
                connection.execute(text("TRUNCATE TABLE [BDG].[DCA_TB006_DISTRIBUIVEIS]"))
                connection.execute(text("TRUNCATE TABLE [BDG].[DCA_TB007_ARRASTAVEIS]"))

                insert_sql = text("""
                    -- ================================================================
                    -- PREPARAÇÃO: Criar tabelas temporárias para otimização
                    -- ================================================================

                    -- Último edital e período (calcula UMA VEZ)
                    DECLARE @UltimoEdital INT = (
                        SELECT TOP 1 ID 
                        FROM [BDG].[DCA_TB008_EDITAIS] 
                        WHERE DELETED_AT IS NULL 
                        ORDER BY ID DESC
                    );

                    DECLARE @UltimoPeriodo INT = (
                        SELECT TOP 1 ID_PERIODO 
                        FROM [BDG].[DCA_TB001_PERIODO_AVALIACAO] 
                        WHERE DELETED_AT IS NULL 
                        ORDER BY ID_PERIODO DESC
                    );

                    PRINT 'Edital: ' + CAST(@UltimoEdital AS VARCHAR(10)) + ' | Período: ' + CAST(@UltimoPeriodo AS VARCHAR(10));

                    -- CPFs que têm contrato no SERASA (para exclusão)
                    IF OBJECT_ID('tempdb..#CPFsNoSerasa') IS NOT NULL
                        DROP TABLE #CPFsNoSerasa;

                    SELECT DISTINCT NR_CPF_CNPJ
                    INTO #CPFsNoSerasa
                    FROM [BDDASHBOARDBI].[BDG].[TEMP_DISTRIBUICAO_SERASA_ASSESSORIA]
                    WHERE ONDE = 'SERASA';

                    CREATE INDEX IX_CPF ON #CPFsNoSerasa(NR_CPF_CNPJ);

                    DECLARE @QtdeCPFsSerasa INT;
                    SELECT @QtdeCPFsSerasa = COUNT(*) FROM #CPFsNoSerasa;
                    PRINT 'CPFs no SERASA (para exclusão): ' + CAST(@QtdeCPFsSerasa AS VARCHAR(10));

                    -- ================================================================
                    -- ETAPA 1: Contratos ASSESSORIA (EXCLUINDO CPFs no SERASA)
                    -- PERMITE contratos com suspensão judicial
                    -- ================================================================
                    INSERT INTO [BDG].[DCA_TB006_DISTRIBUIVEIS]
                    SELECT 
                        ECA.fkContratoSISCTR,
                        CON.NR_CPF_CNPJ,
                        SIT.VR_SD_DEVEDOR,
                        GETDATE(),
                        NULL,
                        NULL
                    FROM
                        [BDG].[COM_TB011_EMPRESA_COBRANCA_ATUAL] AS ECA
                        INNER JOIN [BDG].[COM_TB001_CONTRATO] AS CON
                            ON ECA.fkContratoSISCTR = CON.fkContratoSISCTR
                        INNER JOIN [BDG].[COM_TB007_SITUACAO_CONTRATOS] AS SIT
                            ON ECA.fkContratoSISCTR = SIT.fkContratoSISCTR
                        INNER JOIN [BDDASHBOARDBI].[BDG].[TEMP_DISTRIBUICAO_SERASA_ASSESSORIA] AS TEMP
                            ON ECA.fkContratoSISCTR = TEMP.fkContratoSISCTR
                        LEFT JOIN #CPFsNoSerasa AS SERASA
                            ON CON.NR_CPF_CNPJ = SERASA.NR_CPF_CNPJ
                    WHERE
                        SIT.[fkSituacaoCredito] = 1
                        AND TEMP.ONDE = 'ASSESSORIA'
                        AND TEMP.SIT_ESPECIAL IS NULL
                        AND SERASA.NR_CPF_CNPJ IS NULL;

                    DECLARE @QtdeEtapa1 INT = @@ROWCOUNT;
                    PRINT 'ETAPA 1 - Contratos ASSESSORIA: ' + CAST(@QtdeEtapa1 AS VARCHAR(10));

                    -- ================================================================
                    -- ETAPA 2: CPFs com acordo em empresas que PERMANECEM
                    -- (dos contratos que NÃO passaram na TEMP)
                    -- PERMITE contratos com suspensão judicial
                    -- ================================================================
                    IF OBJECT_ID('tempdb..#CPFsComAcordo') IS NOT NULL
                        DROP TABLE #CPFsComAcordo;

                    SELECT DISTINCT CON.NR_CPF_CNPJ
                    INTO #CPFsComAcordo
                    FROM [BDG].[COM_TB011_EMPRESA_COBRANCA_ATUAL] AS ECA
                        INNER JOIN [BDG].[COM_TB001_CONTRATO] AS CON
                            ON ECA.fkContratoSISCTR = CON.fkContratoSISCTR
                        INNER JOIN [BDG].[COM_TB007_SITUACAO_CONTRATOS] AS SIT
                            ON ECA.fkContratoSISCTR = SIT.fkContratoSISCTR
                        INNER JOIN [BDG].[COM_TB009_ACORDOS_LIQUIDADOS_VIGENTES] AS ALV
                            ON ECA.fkContratoSISCTR = ALV.fkContratoSISCTR
                        INNER JOIN [BDG].[DCA_TB002_EMPRESAS_PARTICIPANTES] AS EMP
                            ON ECA.COD_EMPRESA_COBRANCA = EMP.ID_EMPRESA
                            AND EMP.ID_EDITAL = @UltimoEdital
                            AND EMP.ID_PERIODO = @UltimoPeriodo
                            AND EMP.DS_CONDICAO = 'PERMANECE'
                        LEFT JOIN [BDDASHBOARDBI].[BDG].[TEMP_DISTRIBUICAO_SERASA_ASSESSORIA] AS TEMP
                            ON ECA.fkContratoSISCTR = TEMP.fkContratoSISCTR
                            AND TEMP.ONDE = 'ASSESSORIA'
                            AND TEMP.SIT_ESPECIAL IS NULL
                    WHERE
                        SIT.[fkSituacaoCredito] = 1
                        AND ALV.fkEstadoAcordo = 1
                        AND TEMP.fkContratoSISCTR IS NULL;

                    CREATE INDEX IX_CPF ON #CPFsComAcordo(NR_CPF_CNPJ);

                    DECLARE @QtdeCPFsComAcordo INT;
                    SELECT @QtdeCPFsComAcordo = COUNT(*) FROM #CPFsComAcordo;
                    PRINT 'ETAPA 2 - CPFs com acordo: ' + CAST(@QtdeCPFsComAcordo AS VARCHAR(10));

                    -- ================================================================
                    -- ETAPA 3: Adicionar TODOS os contratos dos CPFs com acordo
                    -- PERMITE contratos com suspensão judicial
                    -- ================================================================
                    INSERT INTO [BDG].[DCA_TB006_DISTRIBUIVEIS]
                    SELECT 
                        ECA.fkContratoSISCTR,
                        CON.NR_CPF_CNPJ,
                        SIT.VR_SD_DEVEDOR,
                        GETDATE(),
                        NULL,
                        NULL
                    FROM
                        [BDG].[COM_TB011_EMPRESA_COBRANCA_ATUAL] AS ECA
                        INNER JOIN [BDG].[COM_TB001_CONTRATO] AS CON
                            ON ECA.fkContratoSISCTR = CON.fkContratoSISCTR
                        INNER JOIN [BDG].[COM_TB007_SITUACAO_CONTRATOS] AS SIT
                            ON ECA.fkContratoSISCTR = SIT.fkContratoSISCTR
                        INNER JOIN #CPFsComAcordo AS CPF_ACORDO
                            ON CON.NR_CPF_CNPJ = CPF_ACORDO.NR_CPF_CNPJ
                        LEFT JOIN [BDG].[DCA_TB006_DISTRIBUIVEIS] AS JA_INS
                            ON ECA.fkContratoSISCTR = JA_INS.FkContratoSISCTR
                    WHERE
                        SIT.[fkSituacaoCredito] = 1
                        AND JA_INS.FkContratoSISCTR IS NULL;

                    DECLARE @QtdeEtapa3 INT = @@ROWCOUNT;
                    PRINT 'ETAPA 3 - Contratos arrastados: ' + CAST(@QtdeEtapa3 AS VARCHAR(10));

                    -- Limpar tabelas temporárias
                    DROP TABLE #CPFsNoSerasa;
                    DROP TABLE #CPFsComAcordo;
                """)

                connection.execute(insert_sql)
                logging.info("Inserção concluída")

                count_sql = text("SELECT COUNT(*) FROM [BDG].[DCA_TB006_DISTRIBUIVEIS]")
                num_contratos = connection.execute(count_sql).scalar()

                logging.info(f"Total: {num_contratos}")
                return num_contratos

            except Exception as e:
                logging.error(f"Erro: {str(e)}")
                import traceback
                logging.error(traceback.format_exc())
                raise

    except Exception as e:
        logging.error(f"Erro: {str(e)}")
        return 0


def ajustar_percentuais_tabela1(empresas, include_total=False):
    """
    Ajusta os percentuais na tabela de participação na arrecadação (Tabela 1)

    Considera empresas com DS_CONDICAO 'PERMANECE' e 'DESCREDENCIADA';
    Distribui os ajustes de 0,01% por empresa em ordem decrescente de arrecadação.
    Uma empresa só recebe o segundo incremento após todas receberem o primeiro.

    Args:
        empresas: Lista de objetos EmpresaParticipante
        include_total: Se True, inclui linha de TOTAL nos resultados
    """
    # Converter objetos para dicionários, filtrando NOVAS
    dados_empresas = []
    idx = 1
    for empresa in empresas:
        if empresa.DS_CONDICAO == 'NOVA':
            continue
        sit = 'PERMANECE' if empresa.DS_CONDICAO == 'PERMANECE' else 'DESCREDENCIADA'
        dados_empresas.append({
            'idx': idx,
            'id_empresa': empresa.ID_EMPRESA,
            'empresa': empresa.NO_EMPRESA_ABREVIADA or empresa.NO_EMPRESA,
            'situacao': sit,
            'arrecadacao': getattr(empresa, 'arrecadacao', 0.0),
        })
        idx += 1

    # Calcular total de arrecadação
    total_arrecadacao = sum(e['arrecadacao'] for e in dados_empresas)

    # Calcular percentual de arrecadação para cada empresa
    for e in dados_empresas:
        e['pct_arrecadacao'] = (e['arrecadacao'] / total_arrecadacao * 100) if total_arrecadacao > 0 else 0.0

    # Separar PERMANECEM e DESCREDENCIADAS
    empresas_permanece = [e for e in dados_empresas if e['situacao'] == 'PERMANECE']
    empresas_descred = [e for e in dados_empresas if e['situacao'] == 'DESCREDENCIADA']

    # Redistribuir percentual das DESCREDENCIADAS entre as PERMANECEM
    pct_total_saem = sum(e['pct_arrecadacao'] for e in empresas_descred)
    pct_redistrib = (pct_total_saem / len(empresas_permanece)) if empresas_permanece else 0.0

    # Inicializar campos para todas as empresas
    for e in dados_empresas:
        if e['situacao'] == 'PERMANECE':
            e['pct_redistribuido'] = pct_redistrib
            e['pct_novo'] = e['pct_arrecadacao'] + pct_redistrib
            # Trunca valor em duas casas decimais (sem arredondamento)
            e['pct_novo_truncado'] = int(e['pct_novo'] * 100) / 100
            e['ajuste'] = 0.0
            e['pct_novofinal'] = 0.0
        else:
            e['pct_redistribuido'] = 0.0
            e['pct_novo'] = 0.0
            e['pct_novo_truncado'] = 0.0
            e['ajuste'] = 0.0
            e['pct_novofinal'] = 0.0

    # Soma truncada e diferença para 100%
    soma_trunc = sum(e['pct_novo_truncado'] for e in empresas_permanece)
    diferenca = 100.0 - soma_trunc

    # IMPORTANTE: Nova implementação da distribuição de ajustes
    # Ordenar empresas por arrecadação decrescente
    empresas_ordenadas = sorted(empresas_permanece, key=lambda x: x['arrecadacao'], reverse=True)

    # Inicializar contadores de ajustes recebidos
    ajustes_recebidos = {e['id_empresa']: 0 for e in empresas_ordenadas}

    # Calcular quantos incrementos de 0.01% são necessários
    incrementos = int(diferenca * 100)  # Cada 0.01% = 1 incremento

    # Distribuir incrementos um a um
    for _ in range(incrementos):
        # Encontrar a empresa que tem menos ajustes
        min_ajustes = min(ajustes_recebidos.values())

        # Empresas candidatas (com o menor número de ajustes)
        candidatas = [e for e in empresas_ordenadas if ajustes_recebidos[e['id_empresa']] == min_ajustes]

        # Em caso de empate, escolher a empresa com maior arrecadação
        # (já está ordenado, então é a primeira da lista)
        empresa_escolhida = candidatas[0]

        # Incrementar o contador de ajustes
        ajustes_recebidos[empresa_escolhida['id_empresa']] += 1

        # Adicionar o ajuste à empresa
        empresa_escolhida['ajuste'] += 0.01

    # Calcular pct_novofinal para cada empresa
    for e in empresas_permanece:
        e['pct_novofinal'] = e['pct_novo_truncado'] + e['ajuste']

    # Verificar o total final (deve ser exatamente 100%)
    soma_final = sum(e['pct_novofinal'] for e in empresas_permanece)

    # Ajuste fino se necessário (embora não deveria ser com a implementação acima)
    if abs(soma_final - 100.0) > 0.0001 and empresas_permanece:
        diferenca_final = 100.0 - soma_final
        # Encontrar a empresa com menos ajustes para fazer o ajuste final
        min_ajustes = min(ajustes_recebidos.values())
        candidatas = [e for e in empresas_ordenadas if ajustes_recebidos[e['id_empresa']] == min_ajustes]
        candidatas[0]['ajuste'] += diferenca_final
        candidatas[0]['pct_novofinal'] = candidatas[0]['pct_novo_truncado'] + candidatas[0]['ajuste']

    # Adicionar linha de TOTAL se solicitado
    if include_total:
        total_row = {
            'idx': 'TOTAL',
            'empresa': 'TOTAL',
            'situacao': '',
            'arrecadacao': sum(e['arrecadacao'] for e in dados_empresas),
            'pct_arrecadacao': sum(e['pct_arrecadacao'] for e in dados_empresas),
            'pct_redistribuido': sum(e['pct_redistribuido'] for e in empresas_permanece),
            'pct_novo': sum(e['pct_novo'] for e in empresas_permanece),
            'pct_novo_truncado': sum(e['pct_novo_truncado'] for e in empresas_permanece),
            'ajuste': sum(e['ajuste'] for e in empresas_permanece),
            'pct_novofinal': sum(e.get('pct_novofinal', 0) for e in empresas_permanece)
        }
        dados_empresas.append(total_row)

    return dados_empresas

def distribuir_contratos_tabela3(empresas_permanece, total_contratos_permanece):
    """
    Distribui contratos para empresas que permanecem com base no percentual final,
    respeitando a ordem de faturamento para ajustes.

    Args:
        empresas_permanece: Lista de objetos com empresas que permanecem
        total_contratos_permanece: Total de contratos a distribuir (da Tabela 2)

    Returns:
        Lista atualizada de empresas com contratos, ajustes e totais calculados
    """
    # Ordenar empresas por faturamento (decrescente)
    empresas_ordenadas = sorted(empresas_permanece, key=lambda x: x.get('arrecadacao', 0), reverse=True)

    # Inicializar campos para cada empresa
    for empresa in empresas_ordenadas:
        # Calcular quantidade base de contratos (truncado, sem arredondamento)
        pct_novofinal = empresa.get('pct_novofinal', 0)
        empresa['contratos'] = int(total_contratos_permanece * pct_novofinal / 100)  # Truncar para inteiro
        empresa['ajuste_contratos'] = 0  # Inicializa ajustes

    # Calcular contratos distribuídos e sobra
    contratos_distribuidos = sum(empresa['contratos'] for empresa in empresas_ordenadas)
    sobra_contratos = total_contratos_permanece - contratos_distribuidos

    # Distribuir sobras uma a uma, em ciclos, por ordem de faturamento
    if sobra_contratos > 0:
        # Inicializar contador de ajustes recebidos por cada empresa
        ajustes_recebidos = [0] * len(empresas_ordenadas)

        # Distribuir cada contrato da sobra
        for _ in range(sobra_contratos):
            # Encontrar empresas com menor número de ajustes
            min_ajustes = min(ajustes_recebidos)
            candidatos = [i for i, v in enumerate(ajustes_recebidos) if v == min_ajustes]

            # Selecionar a primeira empresa candidata (mantém ordem de faturamento)
            idx = candidatos[0]

            # Aplicar ajuste
            empresas_ordenadas[idx]['contratos'] += 1
            empresas_ordenadas[idx]['ajuste_contratos'] += 1
            ajustes_recebidos[idx] += 1

    # Verificar se o total está correto (soma dos contratos = total_contratos_permanece)
    total_final = sum(empresa['contratos'] for empresa in empresas_ordenadas)
    assert total_final == total_contratos_permanece, "Erro na distribuição dos contratos"

    return empresas_ordenadas


def distribuir_restantes_tabela4(empresas_novas, total_contratos_novas):
    """
    Redistribui total_contratos_novas igualitariamente entre empresas_novas,
    truncando para inteiro e aplicando ajustes 1 a 1 em ciclos aleatórios até esgotar a sobra.
    """
    import random

    n = len(empresas_novas)
    if n == 0:
        return []  # sem empresas novas

    # 1) Base truncada
    base = total_contratos_novas // n
    for emp in empresas_novas:
        emp['contratos'] = base
        emp['ajuste_contratos'] = 0

    # 2) Calcular sobra
    restante = total_contratos_novas - base * n

    # 3) Ciclo aleatório para distribuir a sobra
    if restante > 0:
        indices = list(range(n))
        random.shuffle(indices)
        for i in range(restante):
            idx = indices[i % n]
            empresas_novas[idx]['contratos'] += 1
            empresas_novas[idx]['ajuste_contratos'] += 1

    # 4) Verificação final
    total_dist = sum(e['contratos'] for e in empresas_novas)
    assert total_dist == total_contratos_novas, (
        f"Soma incorreta: {total_dist} != {total_contratos_novas}"
    )

    return empresas_novas


def ajustar_percentuais_tabela5(empresas_todas, total_contratos):
    # 1) pct_base e init ajuste_pct
    for e in empresas_todas:
        if total_contratos > 0:
            e['pct_base'] = int((e['contratos'] / total_contratos * 100) * 100) / 100
        else:
            e['pct_base'] = 0.0
        e['ajuste_pct'] = 0.0

    # 2) calcula faltante
    soma_base = sum(e['pct_base'] for e in empresas_todas)
    faltante = int(round((100.00 - soma_base) * 100))

    # 3) ciclo distribuindo
    perm = sorted(
        [e for e in empresas_todas if e['situacao'] == 'PERMANECE'],
        key=lambda x: x.get('arrecadacao', 0),
        reverse=True
    )
    nov = [e for e in empresas_todas if e['situacao'] == 'NOVA']
    ciclo = perm + nov
    ciclo_len = len(ciclo)

    if faltante > 0 and ciclo_len > 0:
        for i in range(faltante):
            pos = i % ciclo_len
            if pos < len(perm):
                alvo = perm[pos]
            else:
                # se não houver NOVA, volta a distribuir entre as PERMANECEM
                alvo = random.choice(nov) if nov else perm[pos % len(perm)]
            alvo['ajuste_pct'] += 0.01

    # 4) recalcula pct_final
    for e in empresas_todas:
        e['pct_final'] = e['pct_base'] + e['ajuste_pct']

    # 5) ajuste fino
    total_final = sum(e['pct_final'] for e in empresas_todas)
    delta = round((100.00 - total_final) * 100) / 100
    if abs(delta) >= 0.01 and empresas_todas:
        # prefere primeira PERMANECEM, senão a primeira da lista completa
        alvo = perm[0] if perm else empresas_todas[0]
        alvo['ajuste_pct'] += delta
        alvo['pct_final'] = alvo['pct_base'] + alvo['ajuste_pct']

    return empresas_todas

def modulo_empresas_mistas(ultimo_edital, ultimo_periodo, periodo_anterior, empresas, num_contratos):
    """
    Função principal para o cálculo de distribuição com empresas mistas

    Args:
        ultimo_edital: Objeto do último edital
        ultimo_periodo: Objeto do último período
        periodo_anterior: Objeto do período anterior
        empresas: Lista de objetos EmpresaParticipante
        num_contratos: Número total de contratos a distribuir

    Returns:
        Dicionário com resultados para cada tabela
    """
    # Separar empresas por situação
    empresas_permanece = [e for e in empresas if e.DS_CONDICAO == 'PERMANECE']
    empresas_novas = [e for e in empresas if e.DS_CONDICAO == 'NOVA']

    # 1. Calcular percentuais e ajustes para Tabela 1
    resultados_tabela1 = ajustar_percentuais_tabela1(empresas)

    # 2. Distribuir contratos por situação (Tabela 2)
    qtde_empresas_permanece = len(empresas_permanece)
    qtde_empresas_novas = len(empresas_novas)
    qtde_total_empresas = qtde_empresas_permanece + qtde_empresas_novas

    # Calcular contratos para cada tipo
    qtde_contratos_permanece = int(
        (qtde_empresas_permanece / qtde_total_empresas) * num_contratos) if qtde_total_empresas > 0 else 0
    qtde_contratos_novas = num_contratos - qtde_contratos_permanece

    # 3. Distribuir contratos para empresas que permanecem (Tabela 3)
    dados_permanece = [e for e in resultados_tabela1 if e['situacao'] == 'PERMANECE']
    resultados_tabela3 = distribuir_contratos_tabela3(dados_permanece, qtde_contratos_permanece)

    # 4. Distribuir contratos para empresas novas (Tabela 4)
    dados_novas = [e for e in resultados_tabela1 if e['situacao'] == 'NOVA']
    resultados_tabela4 = distribuir_restantes_tabela4(dados_novas, qtde_contratos_novas)

    # 5. Calcular percentuais finais (Tabela 5)
    dados_combinados = resultados_tabela3 + resultados_tabela4
    resultados_tabela5 = ajustar_percentuais_tabela5(dados_combinados, num_contratos)

    # Adicionar empresas descredenciadas aos resultados finais
    dados_descredenciadas = [e for e in resultados_tabela1 if e['situacao'] == 'DESCREDENCIADA']

    # Montar o resultado final incluindo todas as empresas
    resultados_finais = resultados_tabela5 + dados_descredenciadas

    # Adicionar linha de total
    # Calcular totais
    total_arrecadacao = sum(e.get('arrecadacao', 0) for e in resultados_finais)
    total_pct_arrecadacao = sum(e.get('pct_arrecadacao', 0) for e in resultados_finais)
    total_pct_redistribuido = sum(e.get('pct_redistribuido', 0) for e in resultados_finais)
    total_pct_novo = sum(e.get('pct_novo_truncado', 0) for e in resultados_finais)
    total_ajuste = sum(e.get('ajuste', 0) for e in resultados_finais)
    total_pct_final = sum(e.get('pct_final', 0) for e in resultados_finais)
    total_contratos = sum(e.get('contratos', 0) for e in resultados_finais)

    # Adicionar linha de total
    linha_total = {
        'idx': 'TOTAL',
        'empresa': 'TOTAL',
        'situacao': '',
        'arrecadacao': total_arrecadacao,
        'pct_arrecadacao': total_pct_arrecadacao,
        'pct_redistribuido': total_pct_redistribuido,
        'pct_novo': total_pct_novo,
        'pct_novo_truncado': total_pct_novo,
        'ajuste': total_ajuste,
        'pct_final': total_pct_final,
        'contratos': total_contratos,
        'ajuste_contratos': sum(e.get('ajuste_contratos', 0) for e in resultados_finais),
        'total_contratos': sum(e.get('total_contratos', 0) for e in resultados_finais),
        'pct_base': sum(e.get('pct_base', 0) for e in resultados_finais),
        'ajuste_pct': sum(e.get('ajuste_pct', 0) for e in resultados_finais),
        'pct_novofinal': sum(e.get('pct_novofinal', 0) for e in resultados_finais)
    }

    resultados_finais.append(linha_total)

    # Retornar resultados consolidados
    return {
        'resultados': resultados_finais,  # Resultado final com todos os dados
        'todas_empresas_anteriores': {e['id_empresa']: e for e in resultados_tabela1},
        'empresas_que_saem': {e['id_empresa']: e for e in resultados_tabela1 if e['situacao'] == 'DESCREDENCIADA'},
        'num_contratos': num_contratos,
        'tipo_calculo': 'mistas',
        'qtde_empresas_permanece': qtde_empresas_permanece,
        'qtde_empresas_novas': qtde_empresas_novas,
        'qtde_contratos_permanece': qtde_contratos_permanece,
        'qtde_contratos_novas': qtde_contratos_novas,
        'total_pct_que_saem': sum(
            e['pct_arrecadacao'] for e in resultados_tabela1 if e['situacao'] == 'DESCREDENCIADA'),
        'valor_redistribuicao': sum(
            e['pct_redistribuido'] for e in resultados_tabela1 if e['situacao'] == 'PERMANECE') / len(
            empresas_permanece) if empresas_permanece else 0
    }

def calcular_limites_empresas_permanece(ultimo_edital, ultimo_periodo, periodo_anterior):
    """
    Realiza o cálculo dos limites de distribuição quando todas as empresas permanecem.

    Args:
        ultimo_edital: Objeto Edital com o último edital
        ultimo_periodo: Objeto PeriodoAvaliacao com o último período
        periodo_anterior: Objeto PeriodoAvaliacao com o período anterior

    Returns:
        dict: Dicionário com os resultados do cálculo e metadados
    """
    try:
        # Obter os contratos distribuíveis
        num_contratos = selecionar_contratos()
        if num_contratos <= 0:
            return None

        # Buscar empresas participantes
        empresas = EmpresaParticipante.query.filter(
            EmpresaParticipante.ID_EDITAL == ultimo_edital.ID,
            EmpresaParticipante.ID_PERIODO == ultimo_periodo.ID_PERIODO,
            EmpresaParticipante.DELETED_AT == None
        ).all()

        if not empresas:
            return None

        # Buscar dados de arrecadação do período anterior
        with db.engine.connect() as connection:
            # Query para obter arrecadação das empresas no período anterior
            sql = text("""
            SELECT 
                EP.ID_EMPRESA,
                EP.NO_EMPRESA,
                EP.NO_EMPRESA_ABREVIADA,
                EP.DS_CONDICAO,
                -- Obter dados de arrecadação da tabela real
                COALESCE(REE.VR_ARRECADACAO, 0) AS VR_ARRECADACAO
            FROM 
                BDG.DCA_TB002_EMPRESAS_PARTICIPANTES AS EP
            LEFT JOIN (
                SELECT 
                    REE.CO_EMPRESA_COBRANCA,
                    SUM(REE.VR_ARRECADACAO_TOTAL) AS VR_ARRECADACAO
                FROM 
                    BDG.COM_TB062_REMUNERACAO_ESTIMADA AS REE
                WHERE 
                    REE.DT_ARRECADACAO BETWEEN :data_inicio AND :data_fim
                GROUP BY 
                    REE.CO_EMPRESA_COBRANCA
            ) AS REE ON EP.ID_EMPRESA = REE.CO_EMPRESA_COBRANCA
            WHERE 
                EP.ID_EDITAL = :id_edital 
                AND EP.ID_PERIODO = :id_periodo 
                AND EP.DS_CONDICAO = 'PERMANECE'
                AND EP.DELETED_AT IS NULL
            ORDER BY
                VR_ARRECADACAO DESC
            """).bindparams(
                id_edital=ultimo_edital.ID,
                id_periodo=ultimo_periodo.ID_PERIODO,
                data_inicio=periodo_anterior.DT_INICIO,
                data_fim=periodo_anterior.DT_FIM
            )

            # Executar a consulta
            result = connection.execute(sql)
            rows = result.fetchall()

            if not rows:
                return None

            # Processar dados de arrecadação reais
            dados_arrecadacao = []
            for row in rows:
                dados_arrecadacao.append({
                    'id_empresa': row[0],
                    'nome': row[1],
                    'nome_abreviado': row[2] or (row[1][:3] if row[1] else ''),
                    'situacao': row[3],
                    'arrecadacao': float(row[4]) if row[4] else 0.0
                })

            # Ordenar por arrecadação (maior para menor)
            dados_arrecadacao.sort(key=lambda x: x['arrecadacao'], reverse=True)

            # Calcular arrecadação total
            total_arrecadacao = sum(item['arrecadacao'] for item in dados_arrecadacao)

            # Calcular percentuais (truncados, sem arredondamento)
            dados_processados = []
            for idx, item in enumerate(dados_arrecadacao):
                # Calcular percentual bruto e truncar para duas casas decimais
                pct_arrecadacao = truncate_decimal((item['arrecadacao'] / total_arrecadacao) * 100)

                # Adicionar dados processados
                dados_processados.append({
                    'idx': idx + 1,
                    'id_empresa': item['id_empresa'],
                    'empresa': item['nome_abreviado'],
                    'situacao': item['situacao'],
                    'arrecadacao': item['arrecadacao'],
                    'pct_arrecadacao': pct_arrecadacao,
                    'ajuste': 0.00,  # Será atualizado posteriormente
                    'pct_final': pct_arrecadacao  # Será atualizado posteriormente
                })

            # Calcular a soma dos percentuais truncados
            soma_percentuais = truncate_decimal(sum(item['pct_arrecadacao'] for item in dados_processados))

            # Calcular quanto falta para chegar a 100%
            diferenca = truncate_decimal(100.00 - soma_percentuais)

            # Aplicar ajuste de 0,01% às empresas com maior arrecadação até chegar exatamente a 100%
            if diferenca > 0:
                # Quantas empresas precisam ser ajustadas inicialmente (um "ciclo" completo de empresas)
                ajuste_por_empresa = 0.01
                num_ciclos_completos = int(diferenca / (ajuste_por_empresa * len(dados_processados)))
                ajustes_restantes = int(
                    (diferenca % (ajuste_por_empresa * len(dados_processados))) / ajuste_por_empresa)

                # Aplicar ajustes para ciclos completos (todas as empresas recebem ajuste)
                if num_ciclos_completos > 0:
                    for i in range(len(dados_processados)):
                        dados_processados[i]['ajuste'] = truncate_decimal(num_ciclos_completos * ajuste_por_empresa)
                        dados_processados[i]['pct_final'] = truncate_decimal(
                            dados_processados[i]['pct_arrecadacao'] + dados_processados[i]['ajuste'])

                # Aplicar ajustes restantes (apenas algumas empresas recebem ajuste adicional)
                for i in range(ajustes_restantes):
                    # O índice sempre será menor que o número de empresas
                    indice = i % len(dados_processados)
                    dados_processados[indice]['ajuste'] = truncate_decimal(
                        dados_processados[indice]['ajuste'] + ajuste_por_empresa)
                    dados_processados[indice]['pct_final'] = truncate_decimal(
                        dados_processados[indice]['pct_arrecadacao'] + dados_processados[indice]['ajuste'])

                # Recalcular totais
                total_pct_arrecadacao = truncate_decimal(sum(item['pct_arrecadacao'] for item in dados_processados))
                total_ajuste = truncate_decimal(sum(item['ajuste'] for item in dados_processados))
                total_pct_final = truncate_decimal(sum(item['pct_final'] for item in dados_processados))

                # Verificar se o total final é exatamente 100%
                if total_pct_final != 100.00:
                    # Se ainda não for exatamente 100%, ajustar a primeira empresa
                    diferenca_restante = truncate_decimal(100.00 - total_pct_final)
                    dados_processados[0]['pct_final'] = truncate_decimal(
                        dados_processados[0]['pct_final'] + diferenca_restante)
                    dados_processados[0]['ajuste'] = truncate_decimal(
                        dados_processados[0]['ajuste'] + diferenca_restante)

                    # Recalcular totais finais
                    total_ajuste = truncate_decimal(sum(item['ajuste'] for item in dados_processados))
                    total_pct_final = truncate_decimal(sum(item['pct_final'] for item in dados_processados))

            # Adicionar linha de total
            dados_processados.append({
                'idx': 'TOTAL',
                'id_empresa': '',
                'empresa': 'TOTAL',
                'situacao': '',
                'arrecadacao': total_arrecadacao,
                'pct_arrecadacao': soma_percentuais,
                'ajuste': total_ajuste if 'total_ajuste' in locals() else 0.00,
                'pct_final': total_pct_final if 'total_pct_final' in locals() else soma_percentuais
            })

            # Retornar resultados e metadados
            return {
                'resultados': dados_processados,
                'num_contratos': num_contratos,
                'tipo_calculo': 'permanece'
            }

    except Exception as e:
        print(f"Erro no cálculo para empresas permanece: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return None


def calcular_limites_empresas_novas(ultimo_edital, ultimo_periodo, empresas):
    """
    Realiza o cálculo dos limites de distribuição quando todas as empresas são novas.
    Neste caso, os contratos são distribuídos igualitariamente entre as empresas.

    Args:
        ultimo_edital: Objeto Edital com o último edital
        ultimo_periodo: Objeto PeriodoAvaliacao com o último período
        empresas: Lista de objetos EmpresaParticipante das empresas

    Returns:
        dict: Dicionário com os resultados do cálculo e metadados
    """
    try:
        # Obter o número de contratos distribuíveis
        num_contratos = selecionar_contratos()
        if num_contratos <= 0:
            return None

        # Número de empresas participantes
        qtde_empresas = len(empresas)
        if qtde_empresas <= 0:
            return None

        # Calcular o percentual base para cada empresa (truncado para duas casas sem arredondamento)
        percentual_base = truncate_decimal(100.00 / qtde_empresas)

        # Calcular a diferença que precisa ser distribuída
        total_percentual_base = truncate_decimal(percentual_base * qtde_empresas)
        diferenca = truncate_decimal(100.00 - total_percentual_base)

        # Preparar dados para retorno
        resultados = []

        # Primeiro, criar a lista de empresas com percentual base
        for idx, empresa in enumerate(empresas):
            resultados.append({
                'idx': idx + 1,
                'id_empresa': empresa.ID_EMPRESA,
                'empresa': empresa.NO_EMPRESA_ABREVIADA or empresa.NO_EMPRESA,
                'situacao': empresa.DS_CONDICAO,
                'contratos': num_contratos // qtde_empresas,  # Divisão inteira
                'pct_distribuicao': percentual_base,
                'ajuste': 0.00,
                'pct_final': percentual_base
            })

        # Nova abordagem: manter um contador de quantos incrementos cada empresa recebeu
        # e distribuir os incrementos de forma equilibrada
        contador_ajustes = [0] * len(resultados)
        incremento = 0.01
        incrementos_totais_necessarios = int(diferenca / incremento)

        # Usar um algoritmo estrito de nivelamento
        for _ in range(incrementos_totais_necessarios):
            # Encontrar o menor valor atual no contador
            menor_contador = min(contador_ajustes)

            # Coletar todos os índices com este valor
            indices_candidatos = []
            for i, valor in enumerate(contador_ajustes):
                if valor == menor_contador:
                    indices_candidatos.append(i)

            # Escolher aleatoriamente um dos índices candidatos
            from random import choice
            indice_escolhido = choice(indices_candidatos)

            # Aplicar o incremento
            contador_ajustes[indice_escolhido] += 1
            resultados[indice_escolhido]['ajuste'] = truncate_decimal(contador_ajustes[indice_escolhido] * incremento)
            resultados[indice_escolhido]['pct_final'] = truncate_decimal(
                percentual_base + resultados[indice_escolhido]['ajuste'])

        # Verificação CRÍTICA: Garantir que a soma seja EXATAMENTE 100.00%
        soma_pct_final = truncate_decimal(sum(item['pct_final'] for item in resultados))

        # Se não for exatamente 100.00%, fazer ajuste adicional
        if soma_pct_final != 100.00:
            # Calcular diferença exata
            diferenca_final = truncate_decimal(100.00 - soma_pct_final)

            # Arredondar para exatamente 0.01 se estiver próximo
            if abs(diferenca_final - 0.01) < 0.001:
                diferenca_final = 0.01

            # Encontrar a empresa com menos incrementos para adicionar a diferença
            menor_contador_final = min(contador_ajustes)
            indices_menor_final = [i for i, v in enumerate(contador_ajustes) if v == menor_contador_final]
            indice_ajuste_final = choice(indices_menor_final)

            # Aplicar o ajuste final
            resultados[indice_ajuste_final]['ajuste'] = truncate_decimal(
                resultados[indice_ajuste_final]['ajuste'] + diferenca_final)
            resultados[indice_ajuste_final]['pct_final'] = truncate_decimal(
                percentual_base + resultados[indice_ajuste_final]['ajuste']
            )

        # Calcular totais finais
        total_pct_distribuicao = truncate_decimal(sum(item['pct_distribuicao'] for item in resultados))
        total_ajuste = truncate_decimal(sum(item['ajuste'] for item in resultados))
        total_pct_final = truncate_decimal(sum(item['pct_final'] for item in resultados))

        # Adicionar linha de total
        resultados.append({
            'idx': 'TOTAL',
            'id_empresa': '',
            'empresa': 'TOTAL',
            'situacao': '',
            'contratos': num_contratos,
            'pct_distribuicao': total_pct_distribuicao,
            'ajuste': total_ajuste,
            'pct_final': total_pct_final
        })

        # Retornar resultados e metadados
        return {
            'resultados': resultados,
            'num_contratos': num_contratos,
            'tipo_calculo': 'novas'
        }

    except Exception as e:
        print(f"Erro no cálculo para empresas novas: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return None


def calcular_limites_empresas_mistas(ultimo_edital, ultimo_periodo, periodo_anterior, empresas):
    """
    Realiza o cálculo dos limites de distribuição quando há empresas que permanecem e empresas novas,
    incluindo Tabela 1 (pct), Tabela 2 (qtde por situação), Tabela 3 (distribuição para permanece)
    e Tabela 4 (distribuição para novas).
    """
    try:
        # 1) Obter total de contratos distribuíveis (Tabela 2)
        num_contratos = selecionar_contratos()
        if num_contratos <= 0:
            return None

        # 2) Buscar dados de arrecadação do período anterior
        with db.engine.connect() as connection:
            sql = text("""
                SELECT 
                    EP.ID_EMPRESA,
                    COALESCE(REE.VR_ARRECADACAO, 0) AS VR_ARRECADACAO
                FROM BDG.DCA_TB002_EMPRESAS_PARTICIPANTES AS EP
                LEFT JOIN (
                    SELECT 
                        CO_EMPRESA_COBRANCA, 
                        SUM(VR_ARRECADACAO_TOTAL) AS VR_ARRECADACAO
                    FROM BDG.COM_TB062_REMUNERACAO_ESTIMADA
                    WHERE DT_ARRECADACAO BETWEEN :data_inicio AND :data_fim
                    GROUP BY CO_EMPRESA_COBRANCA
                ) REE ON EP.ID_EMPRESA = REE.CO_EMPRESA_COBRANCA
                WHERE EP.ID_EDITAL = :id_edital
                  AND EP.ID_PERIODO = :id_periodo_anterior
                  AND EP.DELETED_AT IS NULL
            """).bindparams(
                id_edital=ultimo_edital.ID,
                id_periodo_anterior=periodo_anterior.ID_PERIODO,
                data_inicio=periodo_anterior.DT_INICIO,
                data_fim=periodo_anterior.DT_FIM
            )
            rows = connection.execute(sql).fetchall()

        # 3) Preencher arrecadação em cada empresa
        for emp in empresas:
            emp.arrecadacao = 0.0
            for row in rows:
                if row[0] == emp.ID_EMPRESA:
                    emp.arrecadacao = float(row[1]) or 0.0
                    break

        # 4) Tabela 1: ajustar percentuais (SEM incluir linha TOTAL)
        resultados_tabela1 = ajustar_percentuais_tabela1(empresas, include_total=False)

        # 5) Tabela 2: contratos por situação
        empresas_permanece = [e for e in empresas if e.DS_CONDICAO == 'PERMANECE']
        empresas_novas = [e for e in empresas if e.DS_CONDICAO == 'NOVA']
        total_empresas = len(empresas_permanece) + len(empresas_novas)
        qtde_contratos_permanece = (
            int((len(empresas_permanece) / total_empresas) * num_contratos)
            if total_empresas > 0 else 0
        )
        qtde_contratos_novas = num_contratos - qtde_contratos_permanece

        # 6) Tabela 3: distribuir contratos para quem permanece
        dados_perm = [e for e in resultados_tabela1 if e['situacao'] == 'PERMANECE']
        res3 = distribuir_contratos_tabela3(dados_perm, qtde_contratos_permanece)

        # Transferir resultados de Tabela 3 para os dados consolidados
        mapa3 = {e['id_empresa']: e for e in res3}
        for e in resultados_tabela1:
            if e['situacao'] == 'PERMANECE':
                m = mapa3.get(e['id_empresa'], {})
                e['contratos'] = m.get('contratos', 0)
                e['ajuste_contratos'] = m.get('ajuste_contratos', 0)
                e['total'] = m.get('total', e['contratos'])

        # 7) Tabela 4: Distribuição para empresas NOVAS
        # Montar lista de dicionários apenas com as NOVAS
        dados_nov_raw = []
        for emp in empresas:
            if emp.DS_CONDICAO == 'NOVA':
                dados_nov_raw.append({
                    'idx': None,
                    'id_empresa': emp.ID_EMPRESA,
                    'empresa': emp.NO_EMPRESA_ABREVIADA or emp.NO_EMPRESA,
                    'situacao': 'NOVA',
                    'contratos': 0,
                    'ajuste_contratos': 0,
                })

        resultados_novas = distribuir_restantes_tabela4(dados_nov_raw, qtde_contratos_novas)

        # 8) Combinar resultados para Tabela 5
        # Limpar entradas NOVA existentes e mesclar as novas
        resultados_sem_novas = [e for e in resultados_tabela1 if e.get('situacao') != 'NOVA']

        # Garantir que apenas empresas relevantes sejam passadas para ajuste de percentuais
        empresas_para_tabela5 = []

        # Adicionar empresas que permanecem
        for e in res3:
            # Copiar apenas os dados necessários para a Tabela 5
            empresas_para_tabela5.append({
                'idx': e.get('idx'),
                'id_empresa': e.get('id_empresa'),
                'empresa': e.get('empresa'),
                'situacao': 'PERMANECE',
                'contratos': e.get('contratos', 0),
                'arrecadacao': e.get('arrecadacao', 0),
                'pct_distribuicao': 0.0,  # Será calculado na Tabela 5
            })

        # Adicionar empresas novas
        for e in resultados_novas:
            empresas_para_tabela5.append({
                'idx': e.get('idx'),
                'id_empresa': e.get('id_empresa'),
                'empresa': e.get('empresa'),
                'situacao': 'NOVA',
                'contratos': e.get('contratos', 0),
                'arrecadacao': 0.0,  # Empresas novas não têm arrecadação
                'pct_distribuicao': 0.0,  # Será calculado na Tabela 5
            })

        # 9) Calcular percentuais finais (Tabela 5)
        resultados_tabela5 = ajustar_percentuais_tabela5(empresas_para_tabela5, num_contratos)

        # 10) Integrar resultados da Tabela 5 de volta ao resultado consolidado
        mapa5 = {e['id_empresa']: e for e in resultados_tabela5 if e.get('idx') != 'TOTAL'}

        # Atualizar dados nas empresas que permanecem
        for e in resultados_sem_novas:
            if e.get('idx') != 'TOTAL' and e['situacao'] == 'PERMANECE':
                m = mapa5.get(e['id_empresa'], {})
                e['pct_distribuicao'] = m.get('pct_base', 0.0)
                e['ajuste_pct'] = m.get('ajuste_pct', 0.0)  # Usar ajuste da Tabela 5
                e['pct_final'] = m.get('pct_final', 0.0)

        # Adicionar empresas novas com percentuais calculados
        for e in resultados_novas:
            m = mapa5.get(e['id_empresa'], {})
            e['pct_distribuicao'] = m.get('pct_base', 0.0)
            e['ajuste_pct'] = m.get('ajuste_pct', 0.0)  # Usar ajuste da Tabela 5
            e['pct_final'] = m.get('pct_final', 0.0)
            resultados_sem_novas.append(e)

        # Agora adicionar a linha TOTAL única para todas as tabelas
        total_arrecadacao = sum(e.get('arrecadacao', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL')
        total_pct_arrecadacao = sum(e.get('pct_arrecadacao', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL')
        total_pct_redistribuido = sum(e.get('pct_redistribuido', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL')
        total_pct_novo = sum(e.get('pct_novo_truncado', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL')
        total_ajuste = sum(e.get('ajuste', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL')
        total_pct_final = sum(e.get('pct_final', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL')
        total_contratos = sum(e.get('contratos', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL')
        total_pct_novofinal = sum(e.get('pct_novofinal', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL' and e.get('situacao') == 'PERMANECE')

        linha_total = {
            'idx': 'TOTAL',
            'empresa': 'TOTAL',
            'situacao': '',
            'arrecadacao': total_arrecadacao,
            'pct_arrecadacao': total_pct_arrecadacao,
            'pct_redistribuido': total_pct_redistribuido,
            'pct_novo': total_pct_novo,
            'pct_novo_truncado': total_pct_novo,
            'ajuste': total_ajuste,
            'pct_novofinal': total_pct_novofinal,
            'pct_final': total_pct_final,
            'contratos': total_contratos,
            'ajuste_contratos': sum(e.get('ajuste_contratos', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL'),
            'total_contratos': total_contratos,
            'pct_base': sum(e.get('pct_distribuicao', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL'),
            'pct_distribuicao': sum(e.get('pct_distribuicao', 0) for e in resultados_sem_novas if e.get('idx') != 'TOTAL'),
        }

        resultados_sem_novas.append(linha_total)

        # 11) Montar retorno
        return {
            'resultados': resultados_sem_novas,
            'todas_empresas_anteriores': {
                e['id_empresa']: e for e in resultados_tabela1 if e.get('idx') != 'TOTAL'
            },
            'empresas_que_saem': {
                e['id_empresa']: e for e in resultados_tabela1
                if e['situacao'] == 'DESCREDENCIADA' and e.get('idx') != 'TOTAL'
            },
            'num_contratos': num_contratos,
            'tipo_calculo': 'mistas',
            'qtde_empresas_permanece': len(empresas_permanece),
            'qtde_empresas_novas': len(empresas_novas),
            'qtde_contratos_permanece': qtde_contratos_permanece,
            'qtde_contratos_novas': qtde_contratos_novas,
            'total_pct_que_saem': sum(
                e['pct_arrecadacao'] for e in resultados_tabela1
                if e['situacao'] == 'DESCREDENCIADA' and e.get('idx') != 'TOTAL'
            ),
            'valor_redistribuicao': (
                sum(
                    e['pct_redistribuido'] for e in resultados_tabela1
                    if e['situacao'] == 'PERMANECE' and e.get('idx') != 'TOTAL'
                ) / max(1, len(empresas_permanece))
            )
        }

    except Exception as e:
        print(f"Erro no cálculo para empresas mistas: {e}")
        import traceback
        traceback.print_exc()
        return None



def ajustar_percentuais_finais(empresas):
    total_pct = truncate_decimal(sum(e['pct_distribuicao'] for e in empresas))
    diferenca = truncate_decimal(100.00 - total_pct)

    if diferenca <= 0:
        for e in empresas:
            e['pct_final'] = truncate_decimal(e['pct_distribuicao'])
            e['ajuste'] = truncate_decimal(0.00)
        return empresas

    # Organiza em ciclos: permanece (por arrecadação), depois novas
    permanece = sorted(
        [e for e in empresas if e['situacao'] == 'PERMANECE'],
        key=lambda x: x.get('arrecadacao', 0), reverse=True
    )
    novas = [e for e in empresas if e['situacao'] == 'NOVA']
    ciclo = permanece + novas

    # Distribui 0.01% por vez para cada empresa, começando pelas que mais arrecadam
    idx = 0
    while truncate_decimal(diferenca, 4) > 0 and ciclo:
        empresa = ciclo[idx % len(ciclo)]
        empresa['ajuste'] = truncate_decimal(empresa.get('ajuste', 0.00) + 0.01)
        diferenca = truncate_decimal(diferenca - 0.01, 4)
        idx += 1

    # Calcula os percentuais finais
    for e in empresas:
        e['ajuste'] = truncate_decimal(e.get('ajuste', 0.00))
        e['pct_final'] = truncate_decimal(e['pct_distribuicao'] + e['ajuste'])

    # Verifica se o total ainda é 100%
    total_final = truncate_decimal(sum(e['pct_final'] for e in empresas))
    if total_final != 100.00 and empresas:
        # Ajustar a última discrepância na primeira empresa
        diferenca_final = truncate_decimal(100.00 - total_final)
        empresas[0]['pct_final'] = truncate_decimal(empresas[0]['pct_final'] + diferenca_final)
        empresas[0]['ajuste'] = truncate_decimal(empresas[0]['ajuste'] + diferenca_final)

    return empresas


@limite_bp.route('/limites/analise')
@login_required
def analise_limites():
    try:
        # Obter o edital mais recente
        ultimo_edital = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).first()

        if not ultimo_edital:
            flash('Não foram encontrados editais cadastrados.', 'warning')
            return redirect(url_for('edital.lista_editais'))

        # Obter o período mais recente
        ultimo_periodo = PeriodoAvaliacao.query.filter(
            PeriodoAvaliacao.DELETED_AT == None,
            PeriodoAvaliacao.ID_EDITAL == ultimo_edital.ID
        ).order_by(PeriodoAvaliacao.ID_PERIODO.desc()).first()

        if not ultimo_periodo:
            flash('Não foram encontrados períodos para o edital mais recente.', 'warning')
            return redirect(url_for('periodo.lista_periodos'))

        # Buscar empresas participantes do último período
        empresas = EmpresaParticipante.query.filter(
            EmpresaParticipante.ID_EDITAL == ultimo_edital.ID,
            EmpresaParticipante.ID_PERIODO == ultimo_periodo.ID_PERIODO,
            EmpresaParticipante.DELETED_AT == None
        ).all()

        if not empresas:
            flash('Não foram encontradas empresas participantes para o período atual.', 'warning')
            return redirect(url_for('empresa.lista_empresas', periodo_id=ultimo_periodo.ID))

        # Analisar condições das empresas
        todas_permanece = all(empresa.DS_CONDICAO == 'PERMANECE' for empresa in empresas)
        todas_novas = all(empresa.DS_CONDICAO == 'NOVA' for empresa in empresas)
        alguma_permanece = any(empresa.DS_CONDICAO == 'PERMANECE' for empresa in empresas)

        # Obter período anterior para os cálculos
        periodo_anterior = None
        if not todas_novas:  # Se todas são novas, não precisamos do período anterior
            periodo_anterior = PeriodoAvaliacao.query.filter(
                PeriodoAvaliacao.ID_EDITAL == ultimo_edital.ID,
                PeriodoAvaliacao.ID_PERIODO < ultimo_periodo.ID_PERIODO,
                PeriodoAvaliacao.DELETED_AT == None
            ).order_by(PeriodoAvaliacao.ID_PERIODO.desc()).first()

            if not periodo_anterior and not todas_novas:
                flash('Não foi encontrado período anterior para realizar os cálculos.', 'warning')
                return redirect(url_for('limite.lista_limites'))

        # Realizar cálculos conforme a condição das empresas
        resultado_calculo = None
        num_contratos = selecionar_contratos()

        if todas_permanece:
            if periodo_anterior:
                resultado_calculo = calcular_limites_empresas_permanece(
                    ultimo_edital, ultimo_periodo, periodo_anterior)
            else:
                flash('Não foi encontrado período anterior para realizar os cálculos.', 'warning')
                return redirect(url_for('limite.lista_limites'))
        elif todas_novas:
            resultado_calculo = calcular_limites_empresas_novas(
                ultimo_edital, ultimo_periodo, empresas)
        elif alguma_permanece:
            # Empresas mistas (PERMANECE + NOVAS)
            if periodo_anterior:
                resultado_calculo = calcular_limites_empresas_mistas(
                    ultimo_edital, ultimo_periodo, periodo_anterior, empresas)
            else:
                flash('Não foi encontrado período anterior para realizar os cálculos.', 'warning')
                return redirect(url_for('limite.lista_limites'))

        if not resultado_calculo:
            # Se não for possível calcular, redirecionar com mensagem
            flash('Não foi possível realizar o cálculo dos limites. Verifique a configuração das empresas e períodos.',
                  'warning')
            return redirect(url_for('limite.lista_limites'))

        # Verificar o tipo de resultado e preparar dados para o template
        if isinstance(resultado_calculo, dict):
            resultados_calculo = resultado_calculo.get('resultados', [])
            tipo_calculo = resultado_calculo.get('tipo_calculo', '')
            metadados = resultado_calculo.get('metadados', {})
            num_contratos = resultado_calculo.get('num_contratos', 0)

            # Pega a variável diretamente se ela não vier no metadados
            todas_empresas_anteriores = resultado_calculo.get('todas_empresas_anteriores', {})

            dados_template = {
                'edital': ultimo_edital,
                'periodo': ultimo_periodo,
                'periodo_anterior': periodo_anterior,
                'empresas': empresas,
                'todas_permanece': todas_permanece,
                'todas_novas': todas_novas,
                'alguma_permanece': alguma_permanece,
                'resultados_calculo': resultados_calculo,
                'num_contratos': num_contratos,
                'tipo_calculo': tipo_calculo,
                'todas_empresas_anteriores': todas_empresas_anteriores  # ✅ adiciona aqui
            }

            # Adiciona outras variáveis extras (mistas)
            if tipo_calculo == 'mistas':
                for chave, valor in resultado_calculo.items():
                    if chave != 'resultados':
                        dados_template[chave] = valor

            return render_template('credenciamento/analise_limite.html', **dados_template)
        else:
            flash('Erro no formato dos resultados do cálculo.', 'danger')
            return redirect(url_for('limite.lista_limites'))

    except Exception as e:
        flash(f'Erro durante a análise: {str(e)}', 'danger')
        import traceback
        print(f"Erro detalhado na análise de limites: {traceback.format_exc()}")
        return redirect(url_for('limite.lista_limites'))


# Modificação no arquivo app/routes/limite_routes.py
# Na função salvar_limites()

@limite_bp.route('/limites/salvar', methods=['POST'])
@login_required
def salvar_limites():
    try:
        edital_id = request.form.get('edital_id', type=int)
        periodo_id = request.form.get('periodo_id', type=int)

        # Buscar DT_APURACAO da tabela de situação de contratos
        dt_apuracao = None
        with db.engine.connect() as connection:
            try:
                sql = text("""
                        SELECT TOP 1 DT_REFERENCIA 
                        FROM [BDG].[COM_TB007_SITUACAO_CONTRATOS]
                        ORDER BY DT_REFERENCIA DESC
                    """)
                result = connection.execute(sql)
                row = result.fetchone()
                dt_apuracao = row[0].date() if row and isinstance(row[0], datetime) else datetime.now().date()
            except Exception as e:
                print(f"Erro ao buscar DT_APURACAO: {str(e)}")
                dt_apuracao = datetime.now().date()

        # Obter dados do formulário
        empresas_data = request.form.getlist('empresa_id[]')
        percentuais = request.form.getlist('percentual_final[]')
        situacoes = request.form.getlist('situacao[]')

        if len(empresas_data) != len(percentuais) or len(empresas_data) != len(situacoes):
            flash('Erro: Dados inconsistentes.', 'danger')
            return redirect(url_for('limite.analise_limites'))

        if not empresas_data:
            flash('Erro: Nenhuma empresa informada.', 'danger')
            return redirect(url_for('limite.analise_limites'))

        cod_criterio = 7  # critério fixo

        # Excluir limites anteriores para o mesmo edital/período/critério
        db.session.query(LimiteDistribuicao).filter(
            LimiteDistribuicao.ID_EDITAL == edital_id,
            LimiteDistribuicao.ID_PERIODO == periodo_id,
            LimiteDistribuicao.COD_CRITERIO_SELECAO == cod_criterio,
            LimiteDistribuicao.DELETED_AT == None
        ).delete(synchronize_session=False)
        db.session.commit()

        # Inserir os novos limites
        limites_criados = 0

        # Coletar todos os valores percentuais para validação
        percentuais_validados = []
        for i in range(len(empresas_data)):
            if situacoes[i].upper() == 'DESCREDENCIADA':
                continue

            # Tratar o valor percentual diretamente como float (evitar conversões intermediárias)
            try:
                percentual_valor = float(percentuais[i]) if percentuais[i] else 0.0
                # Arredondar para 2 casas decimais para manter consistência
                percentual_valor = round(percentual_valor, 2)
                percentuais_validados.append((i, percentual_valor))
            except ValueError:
                percentuais_validados.append((i, 0.0))

        # Verificar se a soma é exatamente 100%
        soma_percentuais = sum(pct for _, pct in percentuais_validados)
        if abs(soma_percentuais - 100.0) > 0.01:
            # Se a soma não for 100%, ajustar o maior valor para compensar
            if percentuais_validados:
                maior_idx, _ = max(percentuais_validados, key=lambda x: x[1])
                ajuste = 100.0 - soma_percentuais
                percentuais_validados = [(i, pct + ajuste if i == maior_idx else pct) for i, pct in
                                         percentuais_validados]

        # Inserir os limites com valores validados
        for idx, percentual_valor in percentuais_validados:
            i = idx  # Índice original

            # Criar novo limite apenas para empresas que não são DESCREDENCIADAS
            if situacoes[i].upper() != 'DESCREDENCIADA':
                novo_limite = LimiteDistribuicao(
                    ID_EDITAL=edital_id,
                    ID_PERIODO=periodo_id,
                    ID_EMPRESA=int(empresas_data[i]),
                    COD_CRITERIO_SELECAO=cod_criterio,
                    DT_APURACAO=dt_apuracao,
                    QTDE_MAXIMA=None,
                    VALOR_MAXIMO=None,
                    PERCENTUAL_FINAL=percentual_valor  # Valor já tratado
                )
                db.session.add(novo_limite)
                limites_criados += 1

        db.session.commit()

        # Registrar log de auditoria
        registrar_log(
            acao='criar',
            entidade='limite',
            entidade_id=0,  # Não temos um ID único para o conjunto
            descricao=f'Criação de {limites_criados} limites de distribuição a partir da análise',
            dados_novos={'total_limites': limites_criados}
        )

        flash(f'Limites salvos com sucesso. Total: {limites_criados}', 'success')
        return redirect(url_for('limite.lista_limites'))

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao salvar limites: {str(e)}', 'danger')
        print(f"Erro detalhado: {e}")
        return redirect(url_for('limite.analise_limites'))


@limite_bp.route('/limites/novo', methods=['GET', 'POST'])
@login_required
def novo_limite():
    editais = Edital.query.filter(Edital.DELETED_AT == None).all()
    periodos = PeriodoAvaliacao.query.filter(PeriodoAvaliacao.DELETED_AT == None).all()
    empresas = EmpresaParticipante.query.filter(EmpresaParticipante.DELETED_AT == None).all()
    criterios = CriterioSelecao.query.filter(CriterioSelecao.DELETED_AT == None).all()

    # Verificar se há editais e períodos cadastrados
    if not editais:
        flash('Não há editais cadastrados. Cadastre um edital primeiro.', 'warning')
        return redirect(url_for('edital.lista_editais'))

    if not periodos:
        flash('Não há períodos cadastrados. Cadastre um período primeiro.', 'warning')
        return redirect(url_for('periodo.lista_periodos'))

    if not empresas:
        flash('Não há empresas cadastradas. Cadastre uma empresa primeiro.', 'warning')
        return redirect(url_for('periodo.lista_periodos'))

    if request.method == 'POST':
        try:
            edital_id = int(request.form['edital_id'])
            periodo_id = int(request.form['periodo_id'])
            empresa_id = int(request.form['empresa_id'])
            cod_criterio = int(request.form['cod_criterio'])

            # Valores opcionais
            qtde_maxima = request.form.get('qtde_maxima')
            if qtde_maxima:
                qtde_maxima = int(qtde_maxima)

            valor_maximo = request.form.get('valor_maximo')
            if valor_maximo:
                valor_maximo = truncate_decimal(float(valor_maximo))

            percentual_final = request.form.get('percentual_final')
            if percentual_final:
                percentual_final = truncate_decimal(float(percentual_final))

            # Verificar se já existe limite para esta combinação
            limite_existente = LimiteDistribuicao.query.filter_by(
                ID_EDITAL=edital_id,
                ID_PERIODO=periodo_id,
                ID_EMPRESA=empresa_id,
                COD_CRITERIO_SELECAO=cod_criterio,
                DELETED_AT=None
            ).first()

            if limite_existente:
                flash('Já existe um limite cadastrado com estes critérios.', 'danger')
                return render_template('credenciamento/form_limite.html', editais=editais, periodos=periodos,
                                       empresas=empresas, criterios=criterios)

            novo_limite = LimiteDistribuicao(
                ID_EDITAL=edital_id,
                ID_PERIODO=periodo_id,
                ID_EMPRESA=empresa_id,
                COD_CRITERIO_SELECAO=cod_criterio,
                QTDE_MAXIMA=qtde_maxima,
                VALOR_MAXIMO=valor_maximo,
                PERCENTUAL_FINAL=percentual_final
            )

            db.session.add(novo_limite)
            db.session.commit()

            # Registrar log de auditoria
            dados_novos = {
                'id_edital': edital_id,
                'id_periodo': periodo_id,
                'id_empresa': empresa_id,
                'cod_criterio': cod_criterio,
                'qtde_maxima': qtde_maxima,
                'valor_maximo': valor_maximo,
                'percentual_final': percentual_final
            }

            registrar_log(
                acao='criar',
                entidade='limite',
                entidade_id=novo_limite.ID,
                descricao=f'Cadastro de limite de distribuição',
                dados_novos=dados_novos
            )

            flash('Limite de distribuição cadastrado com sucesso!', 'success')
            return redirect(url_for('limite.lista_limites'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro: {str(e)}', 'danger')

    return render_template('credenciamento/form_limite.html', editais=editais, periodos=periodos, empresas=empresas,
                           criterios=criterios)


@limite_bp.route('/limites/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_limite(id):
    limite = LimiteDistribuicao.query.get_or_404(id)
    editais = Edital.query.filter(Edital.DELETED_AT == None).all()
    periodos = PeriodoAvaliacao.query.filter(PeriodoAvaliacao.DELETED_AT == None).all()
    empresas = EmpresaParticipante.query.filter(EmpresaParticipante.DELETED_AT == None).all()
    criterios = CriterioSelecao.query.filter(CriterioSelecao.DELETED_AT == None).all()

    if request.method == 'POST':
        try:
            # Capturar dados antigos para auditoria
            dados_antigos = {
                'id_edital': limite.ID_EDITAL,
                'id_periodo': limite.ID_PERIODO,
                'id_empresa': limite.ID_EMPRESA,
                'cod_criterio': limite.COD_CRITERIO_SELECAO,
                'qtde_maxima': limite.QTDE_MAXIMA,
                'valor_maximo': limite.VALOR_MAXIMO,
                'percentual_final': limite.PERCENTUAL_FINAL
            }

            # Atualizar dados
            limite.ID_EDITAL = int(request.form['edital_id'])
            limite.ID_PERIODO = int(request.form['periodo_id'])
            limite.ID_EMPRESA = int(request.form['empresa_id'])
            limite.COD_CRITERIO_SELECAO = int(request.form['cod_criterio'])

            # Valores opcionais
            qtde_maxima = request.form.get('qtde_maxima')
            if qtde_maxima:
                limite.QTDE_MAXIMA = int(qtde_maxima)
            else:
                limite.QTDE_MAXIMA = None

            valor_maximo = request.form.get('valor_maximo')
            if valor_maximo:
                limite.VALOR_MAXIMO = truncate_decimal(float(valor_maximo))
            else:
                limite.VALOR_MAXIMO = None

            percentual_final = request.form.get('percentual_final')
            if percentual_final:
                limite.PERCENTUAL_FINAL = truncate_decimal(float(percentual_final))
            else:
                limite.PERCENTUAL_FINAL = None

            limite.UPDATED_AT = datetime.utcnow()

            # Verificar se já existe limite para esta combinação (excluindo o próprio registro)
            limite_existente = LimiteDistribuicao.query.filter(
                LimiteDistribuicao.ID_EDITAL == limite.ID_EDITAL,
                LimiteDistribuicao.ID_PERIODO == limite.ID_PERIODO,
                LimiteDistribuicao.ID_EMPRESA == limite.ID_EMPRESA,
                LimiteDistribuicao.COD_CRITERIO_SELECAO == limite.COD_CRITERIO_SELECAO,
                LimiteDistribuicao.DELETED_AT == None,
                LimiteDistribuicao.ID != id
            ).first()

            if limite_existente:
                flash('Já existe um limite cadastrado com estes critérios.', 'danger')
                return render_template(
                    'credenciamento/form_limite.html',
                    limite=limite,
                    editais=editais,
                    periodos=periodos,
                    empresas=empresas,
                    criterios=criterios
                )

            # Dados novos para auditoria
            dados_novos = {
                'id_edital': limite.ID_EDITAL,
                'id_periodo': limite.ID_PERIODO,
                'id_empresa': limite.ID_EMPRESA,
                'cod_criterio': limite.COD_CRITERIO_SELECAO,
                'qtde_maxima': limite.QTDE_MAXIMA,
                'valor_maximo': limite.VALOR_MAXIMO,
                'percentual_final': limite.PERCENTUAL_FINAL
            }

            db.session.commit()

            # Registrar log de auditoria
            registrar_log(
                acao='editar',
                entidade='limite',
                entidade_id=limite.ID,
                descricao=f'Atualização de limite de distribuição',
                dados_antigos=dados_antigos,
                dados_novos=dados_novos
            )

            flash('Limite de distribuição atualizado com sucesso!', 'success')
            return redirect(url_for('limite.lista_limites'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro: {str(e)}', 'danger')

    return render_template(
        'credenciamento/form_limite.html',
        limite=limite,
        editais=editais,
        periodos=periodos,
        empresas=empresas,
        criterios=criterios)


@limite_bp.route('/limites/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_limite(id):
    limite = LimiteDistribuicao.query.get_or_404(id)

    try:
        # Capturar dados antigos para auditoria
        dados_antigos = {
            'id_edital': limite.ID_EDITAL,
            'id_periodo': limite.ID_PERIODO,
            'id_empresa': limite.ID_EMPRESA,
            'cod_criterio': limite.COD_CRITERIO_SELECAO,
            'qtde_maxima': limite.QTDE_MAXIMA,
            'valor_maximo': limite.VALOR_MAXIMO,
            'percentual_final': limite.PERCENTUAL_FINAL
        }

        # Soft delete - apenas marca como excluído
        limite.DELETED_AT = datetime.utcnow()
        db.session.commit()

        # Registrar log de auditoria
        registrar_log(
            acao='excluir',
            entidade='limite',
            entidade_id=limite.ID,
            descricao=f'Exclusão de limite de distribuição',
            dados_antigos=dados_antigos
        )

        flash('Limite de distribuição excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir limite de distribuição: {str(e)}', 'danger')

    return redirect(url_for('limite.lista_limites'))


@limite_bp.route('/limites/filtrar', methods=['GET'])
@login_required
def filtrar_limites():
    periodo_id = request.args.get('periodo_id', type=int)
    criterio_id = request.args.get('criterio_id', type=int)

    # Redirecionar para a lista com os parâmetros de filtro
    return redirect(url_for('limite.lista_limites', periodo_id=periodo_id, criterio_id=criterio_id))


@limite_bp.route('/limites/detalhe/<int:id>')
@login_required
def detalhe_limite(id):
    # Consulta para obter o limite e a descrição do critério
    resultado = db.session.query(
        LimiteDistribuicao,
        CriterioSelecao.DS_CRITERIO_SELECAO,
        Edital.NU_EDITAL,
        Edital.ANO,
        PeriodoAvaliacao.DT_INICIO,
        PeriodoAvaliacao.DT_FIM,
        EmpresaParticipante.NO_EMPRESA
    ).outerjoin(
        CriterioSelecao,
        LimiteDistribuicao.COD_CRITERIO_SELECAO == CriterioSelecao.COD
    ).outerjoin(
        Edital,
        LimiteDistribuicao.ID_EDITAL == Edital.ID
    ).outerjoin(
        PeriodoAvaliacao,
        LimiteDistribuicao.ID_PERIODO == PeriodoAvaliacao.ID
    ).outerjoin(
        EmpresaParticipante,
        LimiteDistribuicao.ID_EMPRESA == EmpresaParticipante.ID_EMPRESA
    ).filter(
        LimiteDistribuicao.ID == id,
        LimiteDistribuicao.DELETED_AT == None
    ).first_or_404()

    limite, ds_criterio, nu_edital, nu_ano, dt_inicio, dt_fim, no_empresa = resultado

    return render_template(
        'credenciamento/detalhe_limite.html',
        limite=limite,
        ds_criterio=ds_criterio,
        nu_edital=nu_edital,
        nu_ano=nu_ano,
        dt_inicio=dt_inicio,
        dt_fim=dt_fim,
        no_empresa=no_empresa
    )


# SUBSTITUA A FUNÇÃO INTEIRA NO SEU ARQUIVO limite_routes.py

@limite_bp.route('/limites/distribuir-contratos', methods=['GET', 'POST'])
@login_required
def distribuir_contratos():
    """
    Página para distribuição de contratos conforme limites cadastrados.
    """
    try:
        # Encontrar automaticamente o edital mais recente
        ultimo_edital = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).first()

        if not ultimo_edital:
            flash('Não foram encontrados editais cadastrados.', 'warning')
            return redirect(url_for('edital.lista_editais'))

        # Encontrar automaticamente o período mais recente do último edital
        ultimo_periodo = PeriodoAvaliacao.query.filter(
            PeriodoAvaliacao.ID_EDITAL == ultimo_edital.ID,
            PeriodoAvaliacao.DELETED_AT == None
        ).order_by(PeriodoAvaliacao.ID_PERIODO.desc()).first()

        if not ultimo_periodo:
            flash('Não foram encontrados períodos para o edital mais recente.', 'warning')
            return redirect(url_for('periodo.lista_periodos'))

        # Buscar empresas descredenciadas para o formulário
        empresas_descredenciadas = EmpresaParticipante.query.filter(
            EmpresaParticipante.ID_EDITAL == ultimo_edital.ID,
            EmpresaParticipante.ID_PERIODO == ultimo_periodo.ID_PERIODO,
            EmpresaParticipante.DS_CONDICAO == 'DESCREDENCIADA',
            EmpresaParticipante.DELETED_AT == None
        ).all()

        logging.info(f"Distribuição de contratos - Edital: {ultimo_edital.NU_EDITAL}")
        logging.info(f"Distribuição de contratos - Período: {ultimo_periodo.ID_PERIODO}")
        logging.info(f"Empresas descredenciadas encontradas: {len(empresas_descredenciadas)}")

        resultados = None

        if request.method == 'POST':
            try:
                edital_id = ultimo_edital.ID
                periodo_id = ultimo_periodo.ID_PERIODO

                # Importar as funções necessárias do seu arquivo de utils
                from app.utils.distribuir_contratos import selecionar_contratos_distribuiveis, \
                    processar_distribuicao_completa

                modo_execucao = request.form.get('modo_execucao', 'selecao')
                logging.info(f"Modo de execução selecionado: {modo_execucao}")

                if modo_execucao == 'completo':
                    # Verificar se deve usar distribuição igualitária
                    usar_distribuicao_igualitaria = request.form.get('usar_distribuicao_igualitaria') == '1'
                    empresa_descredenciada_id = None
                    data_fim_periodo_anterior = None  # ADICIONADO: Inicializar a variável

                    if usar_distribuicao_igualitaria:
                        empresa_descredenciada_id = request.form.get('empresa_descredenciada_id', type=int)

                        # ADICIONADO: Obter a data do formulário
                        data_fim_periodo_anterior = request.form.get('data_fim_periodo_anterior')

                        if not empresa_descredenciada_id or not data_fim_periodo_anterior:
                            flash(
                                'Para distribuição igualitária, selecione a empresa e a data de fim do período anterior.',
                                'warning')
                            return redirect(url_for('limite.distribuir_contratos'))

                        logging.info(
                            f"Distribuição igualitária ativada para empresa {empresa_descredenciada_id} com data {data_fim_periodo_anterior}")

                    # Executar o processo completo de distribuição
                    # MODIFICADO: A chamada agora inclui o parâmetro 'data_fim_periodo_anterior'
                    resultados = processar_distribuicao_completa(
                        edital_id,
                        periodo_id,
                        usar_distribuicao_igualitaria=usar_distribuicao_igualitaria,
                        empresa_descredenciada_id=empresa_descredenciada_id,
                        data_fim_periodo_anterior=data_fim_periodo_anterior
                    )

                    logging.info(f"Resultados obtidos: {resultados}")

                    if resultados.get('contratos_distribuiveis', 0) > 0:
                        mensagem = f'Processo de distribuição concluído. {resultados.get("total_distribuido", 0)} contratos distribuídos.'

                        if usar_distribuicao_igualitaria:
                            mensagem += f' A distribuição igualitária foi aplicada para a empresa selecionada.'

                        flash(mensagem, 'success')

                        # Registrar log de auditoria
                        registrar_log(
                            acao='distribuir',
                            entidade='distribuicao',
                            entidade_id=periodo_id,
                            descricao=f'Distribuição de contratos - Edital {ultimo_edital.NU_EDITAL}/{ultimo_edital.ANO}',
                            dados_novos={
                                'modo': 'completo',
                                'usou_distribuicao_igualitaria': usar_distribuicao_igualitaria,
                                'empresa_descredenciada_id': empresa_descredenciada_id,
                                'data_fim_periodo_anterior': data_fim_periodo_anterior,
                                'resultados': resultados
                            }
                        )
                    else:
                        flash('Nenhum contrato disponível para distribuição.', 'warning')

                else:  # modo == 'selecao'
                    logging.info("Iniciando seleção de contratos distribuíveis...")
                    num_contratos = selecionar_contratos_distribuiveis()
                    logging.info(f"Contratos selecionados: {num_contratos}")

                    if num_contratos > 0:
                        flash(f'Seleção concluída. {num_contratos} contratos disponíveis.', 'success')
                        resultados = {'contratos_distribuiveis': num_contratos}
                    else:
                        flash('Nenhum contrato disponível para distribuição.', 'warning')

            except Exception as e:
                logging.error(f"Erro ao processar a distribuição: {str(e)}")
                import traceback
                logging.error(traceback.format_exc())
                flash(f'Erro ao processar a distribuição: {str(e)}', 'danger')

        return render_template(
            'credenciamento/distribuir_contratos.html',
            ultimo_edital=ultimo_edital,
            ultimo_periodo=ultimo_periodo,
            resultados=resultados,
            empresas_descredenciadas=empresas_descredenciadas
        )

    except Exception as e:
        logging.error(f"Erro na página de distribuição: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        flash(f'Erro na página de distribuição: {str(e)}', 'danger')
        return redirect(url_for('limite.lista_limites'))


@limite_bp.route('/limites/homologar-distribuicao', methods=['POST'])
@login_required
def homologar_distribuicao():
    """
    Homologa a distribuição de contratos e opcionalmente faz download do arquivo TXT de homologação.
    """
    try:
        edital_id = request.form.get('edital_id', type=int)
        periodo_id = request.form.get('periodo_id', type=int)
        download_arquivo = request.form.get('download_arquivo') == '1'

        # Buscar informações do edital e período para o log
        edital = Edital.query.get_or_404(edital_id)
        periodo = PeriodoAvaliacao.query.filter_by(ID_PERIODO=periodo_id).first_or_404()

        # Obter informações da distribuição
        resultados = obter_resultados_finais_distribuicao(edital_id, periodo_id)

        # Registrar log de homologação
        registrar_log(
            acao='homologar',
            entidade='distribuicao',
            entidade_id=periodo_id,  # Usando ID do período como identificador da distribuição
            descricao=f'Homologação da distribuição do Edital {edital.NU_EDITAL}/{edital.ANO} - Período {periodo.ID_PERIODO}',
            dados_novos={
                'homologado_por': current_user.nome,
                'data_homologacao': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'qtde_contratos': resultados.get('total_qtde', 0),
                'valor_total': resultados.get('total_saldo', 0)
            }
        )

        # Redirecionar ou fazer download direto
        if download_arquivo:
            # Gerar o arquivo TXT para download
            return gerar_arquivo_homologacao(edital_id, periodo_id)
        else:
            flash(f'Distribuição homologada com sucesso por {current_user.nome}.', 'success')
            return redirect(url_for('limite.distribuir_contratos'))

    except Exception as e:
        flash(f'Erro ao homologar distribuição: {str(e)}', 'danger')
        return redirect(url_for('limite.distribuir_contratos'))

def gerar_arquivo_homologacao(edital_id, periodo_id):
    """
    Gera um arquivo TXT com a homologação da distribuição.

    Args:
        edital_id: ID do edital
        periodo_id: ID do período

    Returns:
        Response: Download do arquivo TXT
    """
    try:
        # Consultar dados da distribuição
        query = text("""
            SELECT 
                [fkContratoSISCTR],
                [COD_EMPRESA_COBRANCA] AS NOVA_EMPRESA
            FROM [BDG].[DCA_TB005_DISTRIBUICAO]
            WHERE ID_EDITAL = :edital_id
            AND ID_PERIODO = :periodo_id
            AND DELETED_AT IS NULL
            ORDER BY [fkContratoSISCTR]
        """)

        resultados = db.session.execute(query, {"edital_id": edital_id, "periodo_id": periodo_id}).fetchall()

        # Gerar conteúdo do arquivo
        conteudo = "fkContratoSISCTR;NOVA_EMPRESA\n"
        for resultado in resultados:
            conteudo += f"{resultado[0]};{resultado[1]}\n"

        # Preparar resposta para download
        from flask import Response
        from datetime import datetime

        # Nome do arquivo com data atual
        data_atual = datetime.now().strftime('%Y%m%d')
        filename = f"DISTRIBUICAO_COBRANCA_{data_atual}_TI.TXT"

        # Configurar resposta para forçar download pelo navegador
        response = Response(
            conteudo,
            mimetype='text/plain; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'text/plain; charset=utf-8',
                # Desabilitar cache para garantir download fresco
                'Cache-Control': 'no-cache, no-store, must-revalidate',
                'Pragma': 'no-cache',
                'Expires': '0'
            }
        )

        return response

    except Exception as e:
        # Em caso de erro, log e redirecionar
        import traceback
        print(f"Erro ao gerar arquivo de homologação: {e}")
        print(traceback.format_exc())

        flash(f'Erro ao gerar arquivo de homologação: {str(e)}', 'danger')
        return redirect(url_for('limite.distribuir_contratos'))


# Adicionar no início do arquivo, junto com os outros imports
from app.utils.redistribuir_contratos import processar_redistribuicao_contratos


# Adicionar essa nova rota ao arquivo
@limite_bp.route('/limites/redistribuir', methods=['GET', 'POST'])
@login_required
def redistribuir_contratos():
    try:
        # Encontrar automaticamente o edital mais recente
        ultimo_edital = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).first()

        if not ultimo_edital:
            flash('Não foram encontrados editais cadastrados.', 'warning')
            return redirect(url_for('edital.lista_editais'))

        # Encontrar automaticamente o período mais recente do último edital
        ultimo_periodo = PeriodoAvaliacao.query.filter(
            PeriodoAvaliacao.ID_EDITAL == ultimo_edital.ID,
            PeriodoAvaliacao.DELETED_AT == None
        ).order_by(PeriodoAvaliacao.ID_PERIODO.desc()).first()

        if not ultimo_periodo:
            flash('Não foram encontrados períodos para o edital mais recente.', 'warning')
            return redirect(url_for('periodo.lista_periodos'))

        # CORREÇÃO: Buscar APENAS empresas DESCREDENCIADA NO PERÍODO
        empresas_descredenciadas = EmpresaParticipante.query.filter(
            EmpresaParticipante.ID_EDITAL == ultimo_edital.ID,
            EmpresaParticipante.ID_PERIODO == ultimo_periodo.ID_PERIODO,
            EmpresaParticipante.DS_CONDICAO == 'DESCREDENCIADA NO PERÍODO',  # CORREÇÃO AQUI
            EmpresaParticipante.DELETED_AT == None
        ).all()

        # Buscar o total de empresas que vão receber a redistribuição (NOVA + PERMANECE)
        empresas_receptoras = EmpresaParticipante.query.filter(
            EmpresaParticipante.ID_EDITAL == ultimo_edital.ID,
            EmpresaParticipante.ID_PERIODO == ultimo_periodo.ID_PERIODO,
            EmpresaParticipante.DS_CONDICAO.in_(['NOVA', 'PERMANECE']),
            EmpresaParticipante.DELETED_AT == None
        ).all()

        if not empresas_descredenciadas:
            flash('Não foram encontradas empresas descredenciadas no período atual para redistribuição.', 'warning')
            return redirect(url_for('limite.lista_limites'))

        if not empresas_receptoras:
            flash('Não foram encontradas empresas para receber a redistribuição.', 'warning')
            return redirect(url_for('limite.lista_limites'))

        # Log para debug
        logging.info(f"Empresas descredenciadas no período encontradas: {len(empresas_descredenciadas)}")
        logging.info(f"Empresas receptoras encontradas: {len(empresas_receptoras)}")

        # Usar empresas_descredenciadas no formulário para escolher qual está saindo
        empresas = empresas_descredenciadas

        # Buscar critérios de seleção para uso na redistribuição
        criterios = CriterioSelecao.query.filter(
            CriterioSelecao.DELETED_AT == None
        ).all()

        # Status da execução
        resultados = None

        # Se for POST, processar a redistribuição
        if request.method == 'POST':
            try:
                empresa_id = request.form.get('empresa_id', type=int)
                criterio_id = request.form.get('criterio_id', type=int)

                # ADIÇÃO: Verificar se é um novo critério a ser cadastrado
                novo_criterio = request.form.get('novo_criterio_flag') == '1'
                if novo_criterio:
                    # Extrair dados do novo critério
                    cod_criterio = request.form.get('novo_criterio_cod', type=int)
                    ds_criterio = request.form.get('novo_criterio_descricao')

                    # Validar dados
                    if not cod_criterio or not ds_criterio:
                        flash('Código e descrição do critério são obrigatórios.', 'warning')
                        return render_template(
                            'credenciamento/redistribuir_contratos.html',
                            ultimo_edital=ultimo_edital,
                            ultimo_periodo=ultimo_periodo,
                            empresas=empresas,
                            criterios=criterios,
                            resultados=resultados
                        )

                    # Verificar se o código já existe
                    criterio_existente = CriterioSelecao.query.filter_by(COD=cod_criterio, DELETED_AT=None).first()
                    if criterio_existente:
                        flash(f'Já existe um critério com o código {cod_criterio}.', 'warning')
                        return render_template(
                            'credenciamento/redistribuir_contratos.html',
                            ultimo_edital=ultimo_edital,
                            ultimo_periodo=ultimo_periodo,
                            empresas=empresas,
                            criterios=criterios,
                            resultados=resultados
                        )

                    # Criar novo critério
                    novo_criterio = CriterioSelecao(
                        COD=cod_criterio,
                        DS_CRITERIO_SELECAO=ds_criterio
                    )
                    db.session.add(novo_criterio)
                    db.session.commit()

                    # Usar o novo critério
                    criterio_id = cod_criterio
                    flash(f'Critério "{ds_criterio}" cadastrado com sucesso.', 'success')

                # Validar seleção da empresa
                if not empresa_id:
                    flash('Selecione a empresa que está saindo.', 'warning')
                    return render_template(
                        'credenciamento/redistribuir_contratos.html',
                        ultimo_edital=ultimo_edital,
                        ultimo_periodo=ultimo_periodo,
                        empresas=empresas,
                        criterios=criterios,
                        resultados=resultados
                    )

                # Validar seleção do critério
                if not criterio_id:
                    flash('Selecione o critério de redistribuição.', 'warning')
                    return render_template(
                        'credenciamento/redistribuir_contratos.html',
                        ultimo_edital=ultimo_edital,
                        ultimo_periodo=ultimo_periodo,
                        empresas=empresas,
                        criterios=criterios,
                        resultados=resultados
                    )

                # Log informações importantes
                logging.info(f"Iniciando redistribuição - Empresa saindo: {empresa_id}, Empresas receptoras: {len(empresas_receptoras)}")

                # Executar a redistribuição
                resultados = processar_redistribuicao_contratos(
                    ultimo_edital.ID,
                    ultimo_periodo.ID_PERIODO,
                    empresa_id,
                    criterio_id
                )

                if resultados["success"]:
                    flash(
                        f'Redistribuição concluída. {resultados["contratos_redistribuidos"]} contratos redistribuídos entre {resultados["total_empresas"]} empresas.',
                        'success')
                else:
                    flash('Falha na redistribuição de contratos. Verifique os logs.', 'danger')

            except Exception as e:
                flash(f'Erro ao processar redistribuição: {str(e)}', 'danger')
                import traceback
                logging.error(traceback.format_exc())

        # Renderizar o formulário inicial
        return render_template(
            'credenciamento/redistribuir_contratos.html',
            ultimo_edital=ultimo_edital,
            ultimo_periodo=ultimo_periodo,
            empresas=empresas,
            criterios=criterios,
            resultados=resultados,
            total_empresas_receptoras=len(empresas_receptoras)  # Passar informação adicional
        )
    except Exception as e:
        flash(f'Erro na página de redistribuição: {str(e)}', 'danger')
        import traceback
        logging.error(traceback.format_exc())
        return redirect(url_for('limite.lista_limites'))


def obter_resultados_finais_redistribuicao(edital_id, periodo_id, criterio_id):
    """
    Obter os resultados finais da redistribuição por empresa
    """
    try:
        query = text("""
        SELECT 
            LD.ID_EMPRESA AS cod_empresa,
            EP.NO_EMPRESA_ABREVIADA AS empresa_abrev,
            COUNT(D.fkContratoSISCTR) AS qtde,
            COALESCE(SUM(D.VR_SD_DEVEDOR), 0) AS saldo,
            LD.VR_ARRECADACAO
        FROM [BDG].[DCA_TB003_LIMITES_DISTRIBUICAO] LD
        LEFT JOIN [BDG].[DCA_TB005_DISTRIBUICAO] D
            ON LD.ID_EMPRESA = D.COD_EMPRESA_COBRANCA
            AND D.ID_EDITAL = :edital_id
            AND D.ID_PERIODO = :periodo_id
            AND D.COD_CRITERIO_SELECAO = :criterio_id
        LEFT JOIN [BDG].[DCA_TB002_EMPRESAS_PARTICIPANTES] EP
            ON LD.ID_EMPRESA = EP.ID_EMPRESA
            AND EP.ID_EDITAL = :edital_id
            AND EP.ID_PERIODO = :periodo_id
        WHERE LD.ID_EDITAL = :edital_id
            AND LD.ID_PERIODO = :periodo_id
            AND LD.COD_CRITERIO_SELECAO = :criterio_id
            AND LD.DELETED_AT IS NULL
        GROUP BY LD.ID_EMPRESA, EP.NO_EMPRESA_ABREVIADA, LD.VR_ARRECADACAO
        ORDER BY LD.ID_EMPRESA
        """)

        resultados = db.session.execute(query, {
            "edital_id": edital_id,
            "periodo_id": periodo_id,
            "criterio_id": criterio_id
        }).fetchall()

        resultados_formatados = []
        total_qtde = 0
        total_saldo = 0
        total_arrecadacao = 0

        for row in resultados:
            resultado = {
                "cod_empresa": row.cod_empresa,
                "empresa_abrev": row.empresa_abrev or f"Empresa {row.cod_empresa}",
                "qtde": int(row.qtde) if row.qtde else 0,
                "saldo": float(row.saldo) if row.saldo else 0.0,
                "arrecadacao": float(row.VR_ARRECADACAO) if row.VR_ARRECADACAO else 0.0
            }

            resultados_formatados.append(resultado)
            total_qtde += resultado["qtde"]
            total_saldo += resultado["saldo"]
            total_arrecadacao += resultado["arrecadacao"]

        # Calcular percentuais
        for resultado in resultados_formatados:
            resultado["pct_qtde"] = (resultado["qtde"] / total_qtde * 100) if total_qtde > 0 else 0
            resultado["pct_saldo"] = (resultado["saldo"] / total_saldo * 100) if total_saldo > 0 else 0
            resultado["pct_arrecadacao"] = (
                        resultado["arrecadacao"] / total_arrecadacao * 100) if total_arrecadacao > 0 else 0

        return {
            "resultados": resultados_formatados,
            "total_qtde": total_qtde,
            "total_saldo": total_saldo,
            "total_arrecadacao": total_arrecadacao
        }

    except Exception as e:
        logging.error(f"Erro ao obter resultados finais da redistribuição: {str(e)}")
        return None


@limite_bp.route('/limites/homologar-redistribuicao', methods=['POST'])
@login_required
def homologar_redistribuicao():
    """
    Homologa a redistribuição de contratos e opcionalmente faz download do arquivo TXT.
    """
    try:
        edital_id = request.form.get('edital_id', type=int)
        periodo_id = request.form.get('periodo_id', type=int)
        criterio_id = request.form.get('criterio_id', type=int)
        empresa_id = request.form.get('empresa_id', type=int)
        download_arquivo = request.form.get('download_arquivo') == '1'

        # Buscar informações do edital e período para o log
        edital = Edital.query.get_or_404(edital_id)
        periodo = PeriodoAvaliacao.query.filter_by(ID_PERIODO=periodo_id).first_or_404()

        # Buscar informações da empresa
        empresa = EmpresaParticipante.query.filter_by(
            ID_EDITAL=edital_id,
            ID_PERIODO=periodo_id,
            ID_EMPRESA=empresa_id
        ).first()

        empresa_nome = empresa.NO_EMPRESA_ABREVIADA if empresa else f"Empresa ID {empresa_id}"

        # Registrar log de homologação
        registrar_log(
            acao='homologar',
            entidade='redistribuicao',
            entidade_id=periodo_id,  # Usando ID do período como identificador
            descricao=f'Homologação da redistribuição do Edital {edital.NU_EDITAL}/{edital.ANO} - Período {periodo.ID_PERIODO} - Empresa {empresa_nome}',
            dados_novos={
                'homologado_por': current_user.nome,
                'data_homologacao': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'criterio_id': criterio_id,
                'empresa_id': empresa_id
            }
        )

        # Redirecionar ou fazer download direto
        if download_arquivo:
            # Gerar o arquivo TXT para download
            return gerar_arquivo_redistribuicao(edital_id, periodo_id, criterio_id, empresa_id)
        else:
            flash(f'Redistribuição homologada com sucesso por {current_user.nome}.', 'success')
            return redirect(url_for('limite.redistribuir_contratos'))

    except Exception as e:
        flash(f'Erro ao homologar redistribuição: {str(e)}', 'danger')
        return redirect(url_for('limite.redistribuir_contratos'))


def gerar_arquivo_redistribuicao(edital_id, periodo_id, criterio_id, empresa_id):
    """
    Gera um arquivo TXT com a homologação da redistribuição.

    Args:
        edital_id: ID do edital
        periodo_id: ID do período
        criterio_id: ID do critério
        empresa_id: ID da empresa redistribuída

    Returns:
        Response: Download do arquivo TXT
    """
    try:
        # Consultar dados da redistribuição com o critério específico
        query = text("""
            SELECT 
                [fkContratoSISCTR],
                [COD_EMPRESA_COBRANCA] AS ID_EMPRESA
            FROM [BDG].[DCA_TB005_DISTRIBUICAO]
            WHERE ID_EDITAL = :edital_id
            AND ID_PERIODO = :periodo_id
            AND COD_CRITERIO_SELECAO = :criterio_id
            ORDER BY [fkContratoSISCTR]
        """)

        resultados = db.session.execute(query, {
            "edital_id": edital_id,
            "periodo_id": periodo_id,
            "criterio_id": criterio_id
        }).fetchall()

        # Gerar conteúdo do arquivo
        conteudo = "fkContratoSISCTR;ID_EMPRESA\n"
        for resultado in resultados:
            conteudo += f"{resultado[0]};{resultado[1]}\n"

        # Preparar resposta para download
        from flask import Response
        from datetime import datetime

        # Nome do arquivo com data atual e identificação da empresa
        data_atual = datetime.now().strftime('%Y%m%d')
        filename = f"REDISTRIBUICAO_COBRANCA_{empresa_id}_{data_atual}_TI.TXT"

        # Configurar resposta para forçar download pelo navegador
        response = Response(
            conteudo,
            mimetype='text/plain; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'text/plain; charset=utf-8',
                # Desabilitar cache para garantir download fresco
                'Cache-Control': 'no-cache, no-store, must-revalidate',
                'Pragma': 'no-cache',
                'Expires': '0'
            }
        )

        return response

    except Exception as e:
        # Em caso de erro, log e redirecionar
        import traceback
        print(f"Erro ao gerar arquivo de homologação da redistribuição: {e}")
        print(traceback.format_exc())

        flash(f'Erro ao gerar arquivo de homologação: {str(e)}', 'danger')
        return redirect(url_for('limite.redistribuir_contratos'))


@limite_bp.route('/limites/analitico-distribuicao', methods=['GET', 'POST'])
@login_required
def analitico_distribuicao():
    """
    Página para gerar relatório analítico de distribuição por empresa.
    """
    try:
        # Buscar automaticamente o maior edital
        ultimo_edital = Edital.query.filter(
            Edital.DELETED_AT == None
        ).order_by(Edital.ID.desc()).first()

        if not ultimo_edital:
            flash('Não foram encontrados editais cadastrados.', 'warning')
            return redirect(url_for('limite.lista_limites'))

        # Buscar automaticamente o maior período do edital
        ultimo_periodo = PeriodoAvaliacao.query.filter(
            PeriodoAvaliacao.ID_EDITAL == ultimo_edital.ID,
            PeriodoAvaliacao.DELETED_AT == None
        ).order_by(PeriodoAvaliacao.ID_PERIODO.desc()).first()

        if not ultimo_periodo:
            flash('Não foram encontrados períodos para o edital.', 'warning')
            return redirect(url_for('limite.lista_limites'))

        # Buscar empresas que participaram da distribuição
        with db.engine.connect() as connection:
            sql = text("""
                SELECT DISTINCT 
                    D.COD_EMPRESA_COBRANCA,
                    E.nmEmpresaResponsavelCobranca,
                    E.NO_ABREVIADO_EMPRESA
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] D
                INNER JOIN [BDG].[PAR_TB002_EMPRESA_RESPONSAVEL_COBRANCA] E
                    ON D.COD_EMPRESA_COBRANCA = E.pkEmpresaResponsavelCobranca
                WHERE D.ID_EDITAL = :edital_id
                    AND D.ID_PERIODO = :periodo_id
                    AND D.DELETED_AT IS NULL
                ORDER BY E.NO_ABREVIADO_EMPRESA
            """)

            result = connection.execute(sql, {
                'edital_id': ultimo_edital.ID,
                'periodo_id': ultimo_periodo.ID_PERIODO
            })

            empresas = []
            for row in result:
                empresas.append({
                    'id': row[0],
                    'nome': row[1],
                    'nome_abreviado': row[2] or row[1][:10]  # Se não tiver abreviado, pega os 10 primeiros caracteres
                })

        if not empresas:
            flash('Não foram encontradas empresas com distribuição para este período.', 'warning')
            return redirect(url_for('limite.lista_limites'))

        # Se for POST, gerar o relatório
        if request.method == 'POST':
            empresa_id = request.form.get('empresa_id', type=int)

            if not empresa_id:
                flash('Selecione uma empresa.', 'warning')
                return render_template(
                    'credenciamento/analitico_distribuicao.html',
                    edital=ultimo_edital,
                    periodo=ultimo_periodo,
                    empresas=empresas
                )

            # Redirecionar para gerar o arquivo
            return redirect(url_for(
                'limite.gerar_analitico_distribuicao',
                edital_id=ultimo_edital.ID,
                periodo_id=ultimo_periodo.ID_PERIODO,
                empresa_id=empresa_id
            ))

        return render_template(
            'credenciamento/analitico_distribuicao.html',
            edital=ultimo_edital,
            periodo=ultimo_periodo,
            empresas=empresas
        )

    except Exception as e:
        flash(f'Erro ao acessar página: {str(e)}', 'danger')
        logging.error(f"Erro em analitico_distribuicao: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return redirect(url_for('limite.lista_limites'))


@limite_bp.route('/limites/gerar-analitico/<int:edital_id>/<int:periodo_id>/<int:empresa_id>')
@login_required
def gerar_analitico_distribuicao(edital_id, periodo_id, empresa_id):
    """
    Gera o arquivo TXT com o relatório analítico de distribuição para a empresa selecionada.
    Formato de colunas fixas conforme especificado.
    """
    try:
        from flask import Response
        from datetime import datetime
        from io import StringIO

        # Buscar dados usando o SQL fornecido
        with db.engine.connect() as connection:
            sql = text("""
                SELECT 
                    PR.NO_ABREVIADO_PRODUTO,
                    EM.NO_ABREVIADO_EMPRESA,
                    DIS.COD_EMPRESA_COBRANCA,
                    CTR.NR_CONTRATO,
                    DIS.NR_CPF_CNPJ,
                    DIS.VR_SD_DEVEDOR,
                    CR.[DS_CRITERIO_SELECAO],
                    SIT.QT_DIAS_ATRASO
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS WITH (NOLOCK)
                INNER JOIN BDG.PAR_TB002_EMPRESA_RESPONSAVEL_COBRANCA EM WITH (NOLOCK)
                    ON DIS.COD_EMPRESA_COBRANCA = EM.pkEmpresaResponsavelCobranca
                INNER JOIN BDG.COM_TB001_CONTRATO CTR WITH (NOLOCK)
                    ON CTR.fkContratoSISCTR = DIS.fkContratoSISCTR
                INNER JOIN BDG.PAR_TB001_PRODUTOS PR WITH (NOLOCK)
                    ON PR.pkSistemaOriginario = CTR.COD_PRODUTO
                INNER JOIN [BDG].[DCA_TB004_CRITERIO_SELECAO] CR WITH (NOLOCK)
                    ON CR.COD = DIS.COD_CRITERIO_SELECAO
                INNER JOIN BDG.COM_TB007_SITUACAO_CONTRATOS SIT WITH (NOLOCK)
                    ON SIT.fkContratoSISCTR = DIS.fkContratoSISCTR
                WHERE DIS.ID_PERIODO = :periodo_id 
                    AND DIS.ID_EDITAL = :edital_id
                    AND DIS.COD_EMPRESA_COBRANCA = :empresa_id
                ORDER BY CTR.NR_CONTRATO
            """)

            # Usar StringIO para construir o arquivo em memória
            output = StringIO()

            # Definir larguras das colunas baseado no exemplo fornecido
            col_widths = {
                'NO_ABREVIADO_PRODUTO': 20,
                'NO_ABREVIADO_EMPRESA': 20,
                'COD_EMPRESA_COBRANCA': 20,
                'NR_CONTRATO': 39,
                'NR_CPF_CNPJ': 20,
                'VR_SD_DEVEDOR': 39,
                'DS_CRITERIO_SELECAO': 100,
                'QT_DIAS_ATRASO': 14
            }

            # Escrever cabeçalho
            header = []
            header.append('NO_ABREVIADO_PRODUTO'.ljust(col_widths['NO_ABREVIADO_PRODUTO']))
            header.append('NO_ABREVIADO_EMPRESA'.ljust(col_widths['NO_ABREVIADO_EMPRESA']))
            header.append('COD_EMPRESA_COBRANCA'.ljust(col_widths['COD_EMPRESA_COBRANCA']))
            header.append('NR_CONTRATO'.ljust(col_widths['NR_CONTRATO']))
            header.append('NR_CPF_CNPJ'.ljust(col_widths['NR_CPF_CNPJ']))
            header.append('VR_SD_DEVEDOR'.ljust(col_widths['VR_SD_DEVEDOR']))
            header.append('DS_CRITERIO_SELECAO'.ljust(col_widths['DS_CRITERIO_SELECAO']))
            header.append('QT_DIAS_ATRASO')

            output.write(' '.join(header) + '\n')

            # Escrever linha de separação
            separator = []
            separator.append('-' * col_widths['NO_ABREVIADO_PRODUTO'])
            separator.append('-' * col_widths['NO_ABREVIADO_EMPRESA'])
            separator.append('-' * col_widths['COD_EMPRESA_COBRANCA'])
            separator.append('-' * col_widths['NR_CONTRATO'])
            separator.append('-' * col_widths['NR_CPF_CNPJ'])
            separator.append('-' * col_widths['VR_SD_DEVEDOR'])
            separator.append('-' * col_widths['DS_CRITERIO_SELECAO'])
            separator.append('-' * col_widths['QT_DIAS_ATRASO'])

            output.write(' '.join(separator) + '\n')

            # Processar dados em chunks para melhor performance
            chunk_size = 10000
            offset = 0
            total_registros = 0
            total_saldo = 0.0
            nome_empresa_abreviado = ""

            while True:
                # SQL com paginação
                sql_paginated = text("""
                    SELECT 
                        PR.NO_ABREVIADO_PRODUTO,
                        EM.NO_ABREVIADO_EMPRESA,
                        DIS.COD_EMPRESA_COBRANCA,
                        CTR.NR_CONTRATO,
                        DIS.NR_CPF_CNPJ,
                        DIS.VR_SD_DEVEDOR,
                        CR.[DS_CRITERIO_SELECAO],
                        SIT.QT_DIAS_ATRASO
                    FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS WITH (NOLOCK)
                    INNER JOIN BDG.PAR_TB002_EMPRESA_RESPONSAVEL_COBRANCA EM WITH (NOLOCK)
                        ON DIS.COD_EMPRESA_COBRANCA = EM.pkEmpresaResponsavelCobranca
                    INNER JOIN BDG.COM_TB001_CONTRATO CTR WITH (NOLOCK)
                        ON CTR.fkContratoSISCTR = DIS.fkContratoSISCTR
                    INNER JOIN BDG.PAR_TB001_PRODUTOS PR WITH (NOLOCK)
                        ON PR.pkSistemaOriginario = CTR.COD_PRODUTO
                    INNER JOIN [BDG].[DCA_TB004_CRITERIO_SELECAO] CR WITH (NOLOCK)
                        ON CR.COD = DIS.COD_CRITERIO_SELECAO
                    INNER JOIN BDG.COM_TB007_SITUACAO_CONTRATOS SIT WITH (NOLOCK)
                        ON SIT.fkContratoSISCTR = DIS.fkContratoSISCTR
                    WHERE DIS.ID_PERIODO = :periodo_id 
                        AND DIS.ID_EDITAL = :edital_id
                        AND DIS.COD_EMPRESA_COBRANCA = :empresa_id
                    ORDER BY CTR.NR_CONTRATO
                    OFFSET :offset ROWS
                    FETCH NEXT :chunk_size ROWS ONLY
                """)

                result = connection.execute(sql_paginated, {
                    'edital_id': edital_id,
                    'periodo_id': periodo_id,
                    'empresa_id': empresa_id,
                    'offset': offset,
                    'chunk_size': chunk_size
                })

                rows = result.fetchall()
                if not rows:
                    break

                # Processar cada linha
                for row in rows:
                    # Guardar nome da empresa do primeiro registro
                    if not nome_empresa_abreviado and row[1]:
                        nome_empresa_abreviado = row[1]

                    # Formatar cada campo com a largura apropriada
                    linha = []
                    linha.append((str(row[0]) if row[0] else '').ljust(col_widths['NO_ABREVIADO_PRODUTO']))
                    linha.append((str(row[1]) if row[1] else '').ljust(col_widths['NO_ABREVIADO_EMPRESA']))
                    linha.append(str(row[2]).ljust(col_widths['COD_EMPRESA_COBRANCA']))
                    linha.append((str(row[3]) if row[3] else '').ljust(col_widths['NR_CONTRATO']))
                    linha.append((str(row[4]) if row[4] else '').ljust(col_widths['NR_CPF_CNPJ']))

                    # Formatar valor com 2 casas decimais
                    valor = float(row[5]) if row[5] else 0.0
                    total_saldo += valor
                    linha.append(f"{valor:.2f}".ljust(col_widths['VR_SD_DEVEDOR']))

                    linha.append((str(row[6]) if row[6] else '').ljust(col_widths['DS_CRITERIO_SELECAO']))
                    linha.append(str(row[7]) if row[7] else '')

                    output.write(' '.join(linha) + '\n')
                    total_registros += 1

                offset += chunk_size

                # Log de progresso para grandes volumes
                if total_registros % 50000 == 0:
                    logging.info(f"Processados {total_registros} registros...")

            if total_registros == 0:
                flash('Não foram encontrados dados para a empresa selecionada.', 'warning')
                return redirect(url_for('limite.analitico_distribuicao'))

            # Adicionar linha de total no final
            output.write('\n')
            output.write(' ' * (col_widths['NO_ABREVIADO_PRODUTO'] + 1 +
                                col_widths['NO_ABREVIADO_EMPRESA'] + 1 +
                                col_widths['COD_EMPRESA_COBRANCA'] + 1))
            output.write(f"TOTAL: {total_registros} contratos".ljust(col_widths['NR_CONTRATO']))
            output.write(' ' * (col_widths['NR_CPF_CNPJ'] + 1))
            output.write(f"{total_saldo:.2f}".ljust(col_widths['VR_SD_DEVEDOR']))
            output.write('\n')

            # Gerar nome do arquivo conforme especificado
            data_atual = datetime.now().strftime('%d%m%Y')
            nome_arquivo = f"{nome_empresa_abreviado.upper()}_DISTRIBUICAO_{data_atual}_{edital_id}_{periodo_id}.txt"

            # Preparar conteúdo para download
            content = output.getvalue()
            output.close()

            # Registrar log
            registrar_log(
                acao='exportar',
                entidade='analitico_distribuicao',
                entidade_id=empresa_id,
                descricao=f'Geração de relatório analítico de distribuição - Empresa {nome_empresa_abreviado}',
                dados_novos={
                    'edital_id': edital_id,
                    'periodo_id': periodo_id,
                    'empresa_id': empresa_id,
                    'total_contratos': total_registros,
                    'valor_total': float(total_saldo)
                }
            )

            # Retornar arquivo para download
            response = Response(
                content,
                mimetype='text/plain; charset=utf-8',
                headers={
                    'Content-Disposition': f'attachment; filename={nome_arquivo}',
                    'Content-Type': 'text/plain; charset=utf-8',
                    'Cache-Control': 'no-cache, no-store, must-revalidate',
                    'Pragma': 'no-cache',
                    'Expires': '0'
                }
            )

            # Adicionar cookie para indicar que o download começou
            response.set_cookie('downloadStarted', '1', max_age=60)

            return response

    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        logging.error(f"Erro em gerar_analitico_distribuicao: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return redirect(url_for('limite.analitico_distribuicao'))


@limite_bp.route('/limites/gerar-arquivos-analiticos')
@login_required
def gerar_arquivos_analiticos():
    """
    Página para seleção de tipo de relatório analítico a ser gerado.
    Oferece duas opções:
    1. PDF com resumo da distribuição
    2. Excel individualizado por empresa
    """
    try:
        # Buscar automaticamente o maior edital
        ultimo_edital = Edital.query.filter(
            Edital.DELETED_AT == None
        ).order_by(Edital.ID.desc()).first()

        if not ultimo_edital:
            flash('Não foram encontrados editais cadastrados.', 'warning')
            return redirect(url_for('limite.lista_limites'))

        # Buscar automaticamente o maior período do edital
        ultimo_periodo = PeriodoAvaliacao.query.filter(
            PeriodoAvaliacao.ID_EDITAL == ultimo_edital.ID,
            PeriodoAvaliacao.DELETED_AT == None
        ).order_by(PeriodoAvaliacao.ID_PERIODO.desc()).first()

        if not ultimo_periodo:
            flash('Não foram encontrados períodos para o edital.', 'warning')
            return redirect(url_for('limite.lista_limites'))

        # Contar empresas com distribuição no período
        with db.engine.connect() as connection:
            sql = text("""
                SELECT COUNT(DISTINCT COD_EMPRESA_COBRANCA)
                FROM [BDG].[DCA_TB005_DISTRIBUICAO]
                WHERE ID_EDITAL = :edital_id
                    AND ID_PERIODO = :periodo_id
                    AND DELETED_AT IS NULL
            """)

            result = connection.execute(sql, {
                'edital_id': ultimo_edital.ID,
                'periodo_id': ultimo_periodo.ID_PERIODO
            })

            total_empresas = result.scalar() or 0

        return render_template(
            'credenciamento/gerar_arquivos_analiticos.html',
            edital=ultimo_edital,
            periodo=ultimo_periodo,
            total_empresas=total_empresas
        )

    except Exception as e:
        flash(f'Erro ao acessar página: {str(e)}', 'danger')
        logging.error(f"Erro em gerar_arquivos_analiticos: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return redirect(url_for('limite.lista_limites'))


@limite_bp.route('/limites/gerar-pdf-resumo/<int:edital_id>/<int:periodo_id>', methods=['POST'])
@login_required
def gerar_pdf_resumo(edital_id, periodo_id):
    """
    Gera PDF com resumo da distribuição por assessoria e produto.
    Formatação aprimorada com merge de células para assessorias.
    """
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from datetime import datetime

        # Buscar dados usando o SQL (resumo antes da distribuição no SISGEA)
        with db.engine.connect() as connection:
            sql = text("""
                SELECT 
                    EM.NO_ABREVIADO_EMPRESA AS Assessoria,
                    PR.NO_ABREVIADO_PRODUTO AS Produto,
                    COUNT(DIS.[fkContratoSISCTR]) AS Qtde,
                    SUM([VR_SD_DEVEDOR]) AS Saldo
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS
                INNER JOIN [BDG].[COM_TB001_CONTRATO] CTR
                    ON DIS.[fkContratoSISCTR] = CTR.fkContratoSISCTR
                INNER JOIN BDG.PAR_TB001_PRODUTOS PR
                    ON PR.pkSistemaOriginario = CTR.COD_PRODUTO
                INNER JOIN BDG.PAR_TB002_EMPRESA_RESPONSAVEL_COBRANCA EM
                    ON DIS.COD_EMPRESA_COBRANCA = EM.pkEmpresaResponsavelCobranca
                WHERE DIS.[ID_PERIODO] = :periodo_id
                    AND DIS.[ID_EDITAL] = :edital_id
                    AND DIS.DELETED_AT IS NULL
                GROUP BY
                    EM.NO_ABREVIADO_EMPRESA,
                    PR.NO_ABREVIADO_PRODUTO
                ORDER BY
                    EM.NO_ABREVIADO_EMPRESA,
                    PR.NO_ABREVIADO_PRODUTO
            """)

            result = connection.execute(sql, {
                'periodo_id': periodo_id,
                'edital_id': edital_id
            })

            dados = result.fetchall()

        if not dados:
            flash('Não há dados de distribuição para este período.', 'warning')
            return redirect(url_for('limite.gerar_arquivos_analiticos'))

        # Agrupar dados por assessoria
        assessorias = {}
        for row in dados:
            assessoria = row.Assessoria
            if assessoria not in assessorias:
                assessorias[assessoria] = []
            assessorias[assessoria].append({
                'produto': row.Produto,
                'qtde': int(row.Qtde),
                'saldo': float(row.Saldo)
            })

        # Criar PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm
        )

        story = []
        styles = getSampleStyleSheet()

        # Título principal
        titulo_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2c3e50'),
            alignment=TA_CENTER,
            spaceAfter=3 * mm,
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph(f"Distribuição Período {periodo_id}", titulo_style))

        # Data de geração
        data_style = ParagraphStyle(
            'DataStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#7f8c8d'),
            alignment=TA_CENTER,
            spaceAfter=8 * mm
        )
        data_geracao = datetime.now().strftime('%d/%m/%Y às %H:%M')
        story.append(Paragraph(f"Gerado em: {data_geracao}", data_style))

        # Preparar dados da tabela com merge de células para assessorias
        table_data = []

        # Cabeçalho
        table_data.append(['Assessoria', 'Produto', 'Qtde', 'Saldo'])

        total_qtde_geral = 0
        total_saldo_geral = 0

        # Processar cada assessoria
        for assessoria, produtos in sorted(assessorias.items()):
            num_produtos = len(produtos)

            # Calcular totais da assessoria
            total_qtde_assessoria = sum(p['qtde'] for p in produtos)
            total_saldo_assessoria = sum(p['saldo'] for p in produtos)

            total_qtde_geral += total_qtde_assessoria
            total_saldo_geral += total_saldo_assessoria

            # Adicionar produtos (primeira linha tem o nome da assessoria)
            for i, produto_info in enumerate(produtos):
                qtde_formatada = f"{produto_info['qtde']:,}".replace(',', '.')
                saldo_formatado = f"{produto_info['saldo']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

                if i == 0:
                    # Primeira linha: mostrar nome da assessoria
                    table_data.append([
                        assessoria,
                        produto_info['produto'],
                        qtde_formatada,
                        saldo_formatado
                    ])
                else:
                    # Demais linhas: célula vazia para merge posterior
                    table_data.append([
                        '',  # Será mesclada com a primeira célula
                        produto_info['produto'],
                        qtde_formatada,
                        saldo_formatado
                    ])

            # Linha de total da assessoria
            qtde_total_fmt = f"{total_qtde_assessoria:,}".replace(',', '.')
            saldo_total_fmt = f"{total_saldo_assessoria:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

            table_data.append([
                f'Total {assessoria}',
                '',
                qtde_total_fmt,
                saldo_total_fmt
            ])

        # Linha de total geral
        qtde_geral_fmt = f"{total_qtde_geral:,}".replace(',', '.')
        saldo_geral_fmt = f"{total_saldo_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

        table_data.append([
            'Total Distribuído',
            '',
            qtde_geral_fmt,
            saldo_geral_fmt
        ])

        # Criar tabela
        table = Table(table_data, colWidths=[55 * mm, 60 * mm, 30 * mm, 40 * mm])

        # Estilo base da tabela
        table_style = [
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),

            # Corpo - estilos gerais
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#34495e')),

            # Linha de total geral
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('TOPPADDING', (0, -1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 7),
        ]

        # Aplicar merges e estilos específicos para cada assessoria
        current_row = 1  # Começa após o cabeçalho

        for assessoria, produtos in sorted(assessorias.items()):
            num_produtos = len(produtos)

            if num_produtos > 1:
                # Merge da célula da assessoria
                table_style.append(('SPAN', (0, current_row), (0, current_row + num_produtos - 1)))

            # Background alternado para área da assessoria
            cor_fundo = colors.HexColor('#ecf0f1')
            table_style.append(('BACKGROUND', (0, current_row), (-1, current_row + num_produtos - 1), cor_fundo))

            # Nome da assessoria em negrito e centralizado verticalmente
            table_style.append(('FONTNAME', (0, current_row), (0, current_row + num_produtos - 1), 'Helvetica-Bold'))
            table_style.append(('FONTSIZE', (0, current_row), (0, current_row + num_produtos - 1), 11))
            table_style.append(('VALIGN', (0, current_row), (0, current_row + num_produtos - 1), 'MIDDLE'))

            current_row += num_produtos

            # Linha de total da assessoria
            table_style.append(('BACKGROUND', (0, current_row), (-1, current_row), colors.HexColor('#bdc3c7')))
            table_style.append(('FONTNAME', (0, current_row), (-1, current_row), 'Helvetica-Bold'))
            table_style.append(('SPAN', (0, current_row), (1, current_row)))  # Merge das duas primeiras células

            current_row += 1

        # Aplicar todos os estilos
        table.setStyle(TableStyle(table_style))
        story.append(table)

        # Rodapé com data
        story.append(Spacer(1, 10 * mm))
        rodape_style = ParagraphStyle(
            'Rodape',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#7f8c8d'),
            alignment=TA_LEFT
        )
        story.append(Paragraph(f"Posição: {datetime.now().strftime('%d/%m/%Y')}", rodape_style))

        # Gerar PDF
        doc.build(story)
        buffer.seek(0)

        # Registrar log
        registrar_log(
            acao='gerar',
            entidade='relatorio_pdf_resumo',
            entidade_id=periodo_id,
            descricao=f'PDF de resumo gerado para período {periodo_id}'
        )

        nome_arquivo = f'resumo_distribuicao_periodo_{periodo_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'danger')
        logging.error(f"Erro ao gerar PDF resumo: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return redirect(url_for('limite.gerar_arquivos_analiticos'))


@limite_bp.route('/limites/gerar-excel-empresas/<int:edital_id>/<int:periodo_id>', methods=['POST'])
@login_required
def gerar_excel_empresas(edital_id, periodo_id):
    """
    Gera múltiplos arquivos Excel, um para cada empresa do período.
    OTIMIZADO:
    - Removida coluna "Faixa Propensão" (FX_PROPENSAO_ATUAL_CREDITO)
    - Mantém JOIN com COM_TB007_SITUACAO_CONTRATOS para buscar QT_DIAS_ATRASO e VR_SD_DEVEDOR
    - Formatação de números longos para evitar notação científica
    """
    try:
        from io import BytesIO
        from zipfile import ZipFile
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime

        # Buscar empresas que participaram da distribuição e não estão descredenciadas
        with db.engine.connect() as connection:
            sql_empresas = text("""
                SELECT DISTINCT 
                    DIS.COD_EMPRESA_COBRANCA,
                    EM.NO_ABREVIADO_EMPRESA,
                    EM.nmEmpresaResponsavelCobranca
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS
                INNER JOIN BDG.PAR_TB002_EMPRESA_RESPONSAVEL_COBRANCA EM
                    ON DIS.COD_EMPRESA_COBRANCA = EM.pkEmpresaResponsavelCobranca
                INNER JOIN [BDG].[DCA_TB002_EMPRESAS_PARTICIPANTES] EP
                    ON DIS.COD_EMPRESA_COBRANCA = EP.ID_EMPRESA
                    AND EP.ID_EDITAL = :edital_id
                    AND EP.ID_PERIODO = :periodo_id
                WHERE DIS.ID_EDITAL = :edital_id
                    AND DIS.ID_PERIODO = :periodo_id
                    AND DIS.DELETED_AT IS NULL
                    AND EP.DS_CONDICAO <> 'DESCREDENCIADA'
                ORDER BY EM.NO_ABREVIADO_EMPRESA
            """)

            empresas = connection.execute(sql_empresas, {
                'edital_id': edital_id,
                'periodo_id': periodo_id
            }).fetchall()

        if not empresas:
            flash('Não há empresas com distribuição para este período.', 'warning')
            return redirect(url_for('limite.gerar_arquivos_analiticos'))

        # Criar arquivo ZIP para armazenar todos os Excel
        zip_buffer = BytesIO()

        with ZipFile(zip_buffer, 'w') as zip_file:
            for empresa in empresas:
                cod_empresa = empresa.COD_EMPRESA_COBRANCA
                nome_abreviado = empresa.NO_ABREVIADO_EMPRESA or empresa.nmEmpresaResponsavelCobranca[:20]

                # ✅ QUERY CORRIGIDA: Mantém JOIN com COM_TB007_SITUACAO_CONTRATOS
                # Mas REMOVE a coluna FX_PROPENSAO_ATUAL_CREDITO
                with db.engine.connect() as connection:
                    sql_analitico = text("""
                        SELECT 
                            PRO.NO_ABREVIADO_PRODUTO AS Produto,
                            DIS.[COD_EMPRESA_COBRANCA] AS COD_EMPRESA,
                            ASS.[NO_ABREVIADO_EMPRESA] AS Assessoria,
                            CTR.NR_CONTRATO AS Nr_Contrato,
                            CTR.[NR_CPF_CNPJ] AS CPF,
                            SIT.QT_DIAS_ATRASO,
                            SIT.[VR_SD_DEVEDOR],
                            CRI.DS_CRITERIO_SELECAO
                        FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS 
                        INNER JOIN [BDG].[COM_TB001_CONTRATO] CTR
                            ON DIS.fkContratoSISCTR = CTR.fkContratoSISCTR
                        INNER JOIN [BDG].[PAR_TB001_PRODUTOS] PRO
                            ON CTR.COD_PRODUTO = PRO.pkSistemaOriginario
                        INNER JOIN [BDG].[PAR_TB002_EMPRESA_RESPONSAVEL_COBRANCA] ASS
                            ON ASS.pkEmpresaResponsavelCobranca = DIS.[COD_EMPRESA_COBRANCA]
                        INNER JOIN [BDG].[COM_TB007_SITUACAO_CONTRATOS] SIT
                            ON DIS.fkContratoSISCTR = SIT.fkContratoSISCTR
                        INNER JOIN [BDG].[DCA_TB004_CRITERIO_SELECAO] CRI
                            ON CRI.[COD] = DIS.COD_CRITERIO_SELECAO
                        WHERE SIT.[fkSituacaoCredito] = 1
                            AND DIS.[COD_EMPRESA_COBRANCA] = :cod_empresa
                            AND DIS.[ID_PERIODO] = :periodo_id
                            AND DIS.[ID_EDITAL] = :edital_id
                            AND DIS.DELETED_AT IS NULL
                        ORDER BY PRO.NO_ABREVIADO_PRODUTO, CTR.NR_CONTRATO
                    """)

                    dados_empresa = connection.execute(sql_analitico, {
                        'cod_empresa': cod_empresa,
                        'periodo_id': periodo_id,
                        'edital_id': edital_id
                    }).fetchall()

                if not dados_empresa:
                    continue

                # Criar arquivo Excel para a empresa
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Período {periodo_id}"

                # Estilos
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="5893D4", end_color="5893D4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Título
                ws.merge_cells('A1:H1')  # ✅ 8 colunas (sem Faixa Propensão)
                cell_titulo = ws['A1']
                cell_titulo.value = f"Distribuição Analítica - {nome_abreviado} - Período {periodo_id}"
                cell_titulo.font = Font(bold=True, size=14)
                cell_titulo.alignment = Alignment(horizontal="center")

                # ✅ Cabeçalhos SEM "Faixa Propensão"
                headers = ['Produto', 'Cód. Empresa', 'Assessoria', 'Nº Contrato', 'CPF/CNPJ',
                           'Dias Atraso', 'Saldo Devedor', 'Critério Seleção']

                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=3, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border_style

                # ✅ Formatar colunas como texto ANTES de inserir dados
                max_rows = len(dados_empresa) + 10

                for row in range(4, max_rows + 4):
                    ws.cell(row=row, column=4).number_format = '@'  # Coluna D (Nº Contrato)
                    ws.cell(row=row, column=5).number_format = '@'  # Coluna E (CPF/CNPJ)

                # Dados
                row_idx = 4
                total_saldo = 0

                for row_data in dados_empresa:
                    ws.cell(row=row_idx, column=1, value=row_data.Produto)
                    ws.cell(row=row_idx, column=2, value=row_data.COD_EMPRESA)
                    ws.cell(row=row_idx, column=3, value=row_data.Assessoria)

                    # Inserir como string para garantir formato texto
                    ws.cell(row=row_idx, column=4, value=str(row_data.Nr_Contrato))
                    ws.cell(row=row_idx, column=5, value=str(row_data.CPF))

                    ws.cell(row=row_idx, column=6, value=int(row_data.QT_DIAS_ATRASO or 0))

                    # Saldo Devedor formatado
                    saldo = float(row_data.VR_SD_DEVEDOR or 0)
                    total_saldo += saldo
                    cell_saldo = ws.cell(row=row_idx, column=7, value=saldo)
                    cell_saldo.number_format = 'R$ #,##0.00'

                    # ✅ Coluna H = Critério Seleção (sem Faixa Propensão)
                    ws.cell(row=row_idx, column=8, value=row_data.DS_CRITERIO_SELECAO)

                    # Aplicar bordas
                    for col in range(1, 9):  # ✅ 8 colunas
                        ws.cell(row=row_idx, column=col).border = border_style
                        ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="center")

                    row_idx += 1

                # Linha de total
                ws.cell(row=row_idx, column=6, value="TOTAL:")
                ws.cell(row=row_idx, column=6).font = Font(bold=True)
                cell_total = ws.cell(row=row_idx, column=7, value=total_saldo)
                cell_total.font = Font(bold=True)
                cell_total.number_format = 'R$ #,##0.00'
                cell_total.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")

                # ✅ Ajustar largura das colunas (8 colunas - sem Faixa Propensão)
                column_widths = {
                    'A': 15,  # Produto
                    'B': 12,  # Cód. Empresa
                    'C': 20,  # Assessoria
                    'D': 18,  # Nº Contrato
                    'E': 18,  # CPF/CNPJ
                    'F': 12,  # Dias Atraso
                    'G': 18,  # Saldo Devedor
                    'H': 25   # Critério Seleção
                }

                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width

                # Salvar Excel em buffer
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)

                # Adicionar ao ZIP
                nome_arquivo_excel = f"analitico_{nome_abreviado.replace(' ', '_')}_periodo_{periodo_id}.xlsx"
                zip_file.writestr(nome_arquivo_excel, excel_buffer.getvalue())

        # Preparar ZIP para download
        zip_buffer.seek(0)

        # Registrar log
        registrar_log(
            acao='gerar',
            entidade='relatorio_excel_empresas',
            entidade_id=periodo_id,
            descricao=f'Arquivos Excel gerados para {len(empresas)} empresas do período {periodo_id} (otimizado)'
        )

        nome_arquivo_zip = f'analiticos_empresas_periodo_{periodo_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'

        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=nome_arquivo_zip
        )

    except Exception as e:
        flash(f'Erro ao gerar arquivos Excel: {str(e)}', 'danger')
        logging.error(f"Erro ao gerar Excel empresas: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return redirect(url_for('limite.gerar_arquivos_analiticos'))


@limite_bp.route('/limites/gerar-pdf-analitico-cpf/<int:edital_id>/<int:periodo_id>', methods=['POST'])
@login_required
def gerar_pdf_analitico_cpf(edital_id, periodo_id):
    """
    Gera PDF com análise agregada por empresa de cobrança.
    Mostra: Quantidade de contratos, quantidade de CPFs únicos e saldo devedor total.
    Baseado no SQL fornecido pelo usuário.
    """
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from datetime import datetime

        # ✅ SQL CORRIGIDO: Buscar VR_SD_DEVEDOR da tabela COM_TB007_SITUACAO_CONTRATOS
        with db.engine.connect() as connection:
            sql = text("""
                SELECT 
                    [COD_EMPRESA_COBRANCA],
                    COUNT(DIS.[fkContratoSISCTR]) AS QTDE,
                    COUNT(DISTINCT(DIS.[NR_CPF_CNPJ])) AS QTDE_CPF,
                    SUM(SIT.[VR_SD_DEVEDOR]) AS SD
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS
                INNER JOIN [BDG].[COM_TB001_CONTRATO] CTR 
                    ON DIS.[fkContratoSISCTR] = CTR.fkContratoSISCTR
                INNER JOIN BDG.PAR_TB001_PRODUTOS PR 
                    ON PR.pkSistemaOriginario = CTR.COD_PRODUTO
                INNER JOIN [BDG].[COM_TB007_SITUACAO_CONTRATOS] SIT
                    ON DIS.fkContratoSISCTR = SIT.fkContratoSISCTR
                WHERE [ID_PERIODO] = :periodo_id
                    AND [ID_EDITAL] = :edital_id
                    AND DIS.DELETED_AT IS NULL
                    AND SIT.[fkSituacaoCredito] = 1
                GROUP BY [COD_EMPRESA_COBRANCA]
                ORDER BY [COD_EMPRESA_COBRANCA]
            """)

            result = connection.execute(sql, {
                'periodo_id': periodo_id,
                'edital_id': edital_id
            })

            dados = result.fetchall()

        if not dados:
            flash('Não há dados de distribuição para este período.', 'warning')
            return redirect(url_for('limite.gerar_arquivos_analiticos'))

        # Buscar nome das empresas
        empresas_nomes = {}
        with db.engine.connect() as connection:
            sql_empresas = text("""
                SELECT DISTINCT
                    DIS.COD_EMPRESA_COBRANCA,
                    EM.NO_ABREVIADO_EMPRESA,
                    EM.nmEmpresaResponsavelCobranca
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS
                INNER JOIN BDG.PAR_TB002_EMPRESA_RESPONSAVEL_COBRANCA EM
                    ON DIS.COD_EMPRESA_COBRANCA = EM.pkEmpresaResponsavelCobranca
                WHERE DIS.ID_EDITAL = :edital_id
                    AND DIS.ID_PERIODO = :periodo_id
                    AND DIS.DELETED_AT IS NULL
            """)

            empresas = connection.execute(sql_empresas, {
                'edital_id': edital_id,
                'periodo_id': periodo_id
            }).fetchall()

            for emp in empresas:
                empresas_nomes[emp.COD_EMPRESA_COBRANCA] = emp.NO_ABREVIADO_EMPRESA or emp.nmEmpresaResponsavelCobranca

        # Buscar informações do edital e período
        edital = Edital.query.get_or_404(edital_id)
        periodo = PeriodoAvaliacao.query.filter_by(ID_PERIODO=periodo_id).first_or_404()

        # Criar buffer para o PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm,
                                topMargin=15 * mm, bottomMargin=15 * mm)

        story = []
        styles = getSampleStyleSheet()

        # Título principal (MESMAS CORES DO PDF RESUMO NORMAL)
        titulo_style = ParagraphStyle(
            'TituloPrincipal',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2c3e50'),  # ✅ Cinza escuro (igual ao PDF resumo)
            alignment=TA_CENTER,
            spaceAfter=3 * mm,
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph("Resumo da Distribuição por CPF/CNPJ", titulo_style))

        # Subtítulo com informações do período
        subtitulo_style = ParagraphStyle(
            'Subtitulo',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#7f8c8d'),  # ✅ Cinza médio (igual ao PDF resumo)
            alignment=TA_CENTER,
            spaceAfter=8 * mm
        )
        subtitulo_texto = f"Edital: {edital.NU_EDITAL}/{edital.ANO} | Período: {periodo.ID_PERIODO} | " \
                          f"{periodo.DT_INICIO.strftime('%d/%m/%Y')} a {periodo.DT_FIM.strftime('%d/%m/%Y')}"
        story.append(Paragraph(subtitulo_texto, subtitulo_style))

        # Preparar dados da tabela
        table_data = [['Cód. Empresa', 'Assessoria', 'Qtde Contratos', 'Qtde CPFs', 'Saldo Devedor (R$)']]

        total_qtde = 0
        total_cpfs = 0
        total_saldo = 0.0

        for row in dados:
            cod_empresa = row.COD_EMPRESA_COBRANCA
            qtde = row.QTDE
            qtde_cpf = row.QTDE_CPF
            saldo = float(row.SD or 0)

            total_qtde += qtde
            total_cpfs += qtde_cpf
            total_saldo += saldo

            nome_empresa = empresas_nomes.get(cod_empresa, f"Empresa {cod_empresa}")

            table_data.append([
                str(cod_empresa),
                nome_empresa,
                f"{qtde:,}".replace(',', '.'),
                f"{qtde_cpf:,}".replace(',', '.'),
                f"R$ {saldo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])

        # Linha de total
        table_data.append([
            '',
            'TOTAL',
            f"{total_qtde:,}".replace(',', '.'),
            f"{total_cpfs:,}".replace(',', '.'),
            f"R$ {total_saldo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        ])

        # Criar tabela
        table = Table(table_data, colWidths=[25 * mm, 60 * mm, 30 * mm, 25 * mm, 40 * mm])

        # Estilo da tabela (MESMAS CORES DO PDF RESUMO NORMAL)
        table_style = [
            # Cabeçalho (igual ao PDF resumo)
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),  # ✅ Cinza escuro
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8 * mm),
            ('TOPPADDING', (0, 0), (-1, 0), 8 * mm),

            # Corpo da tabela
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#ffffff')),
            ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Código empresa centralizado
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Nome empresa à esquerda
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),  # Números à direita
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 10),
            ('TOPPADDING', (0, 1), (-1, -2), 5 * mm),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 5 * mm),
            ('LEFTPADDING', (0, 0), (-1, -1), 6 * mm),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6 * mm),

            # Linha de total (igual ao PDF resumo - azul)
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#3498db')),  # ✅ Azul
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('TOPPADDING', (0, -1), (-1, -1), 5 * mm),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 5 * mm),

            # Bordas (igual ao PDF resumo)
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),  # ✅ Cinza claro
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#34495e')),  # ✅ Cinza escuro
        ]

        # Zebrar linhas (alternar cores - igual ao PDF resumo)
        for i in range(1, len(table_data) - 1):
            if i % 2 == 0:
                table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8f9fa')))

        table.setStyle(TableStyle(table_style))
        story.append(table)

        # Rodapé com data
        story.append(Spacer(1, 10 * mm))
        rodape_style = ParagraphStyle(
            'Rodape',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#7f8c8d'),
            alignment=TA_LEFT
        )
        story.append(Paragraph(f"Posição: {datetime.now().strftime('%d/%m/%Y')}", rodape_style))

        # Gerar PDF
        doc.build(story)
        buffer.seek(0)

        # Registrar log
        registrar_log(
            acao='gerar',
            entidade='relatorio_pdf_analitico_cpf',
            entidade_id=periodo_id,
            descricao=f'PDF analítico por CPF gerado para período {periodo_id}'
        )

        nome_arquivo = f'analitico_cpf_periodo_{periodo_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'danger')
        logging.error(f"Erro ao gerar PDF analítico por CPF: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return redirect(url_for('limite.gerar_arquivos_analiticos'))


@limite_bp.route('/limites/dashboard')
@login_required
def dashboard_distribuicao():
    """
    Dashboard principal de distribuição de contratos.
    Exibe KPIs, gráficos e informações consolidadas.
    """
    try:
        # Buscar último edital e período
        ultimo_edital = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).first()

        if not ultimo_edital:
            flash('Não foram encontrados editais cadastrados.', 'warning')
            return redirect(url_for('edital.lista_editais'))

        ultimo_periodo = PeriodoAvaliacao.query.filter(
            PeriodoAvaliacao.ID_EDITAL == ultimo_edital.ID,
            PeriodoAvaliacao.DELETED_AT == None
        ).order_by(PeriodoAvaliacao.ID_PERIODO.desc()).first()

        # Buscar todos os editais e períodos para filtros
        editais = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).all()
        periodos = PeriodoAvaliacao.query.filter(PeriodoAvaliacao.DELETED_AT == None).order_by(
            PeriodoAvaliacao.ID_PERIODO.desc()).all()

        # NOVO: Buscar empresas participantes para o filtro
        with db.engine.connect() as connection:
            sql_empresas = text("""
                SELECT DISTINCT
                    EM.pkEmpresaResponsavelCobranca,
                    EM.NO_ABREVIADO_EMPRESA,
                    EM.nmEmpresaResponsavelCobranca
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS
                INNER JOIN [BDG].[PAR_TB002_EMPRESA_RESPONSAVEL_COBRANCA] EM
                    ON DIS.COD_EMPRESA_COBRANCA = EM.pkEmpresaResponsavelCobranca
                WHERE DIS.DELETED_AT IS NULL
                ORDER BY EM.NO_ABREVIADO_EMPRESA
            """)

            resultado_empresas = connection.execute(sql_empresas).fetchall()
            empresas = [
                {
                    'id': row[0],
                    'nome_abreviado': row[1],
                    'nome_completo': row[2]
                }
                for row in resultado_empresas
            ]

        return render_template(
            'credenciamento/dashboard_distribuicao.html',
            ultimo_edital=ultimo_edital,
            ultimo_periodo=ultimo_periodo,
            editais=editais,
            periodos=periodos,
            empresas=empresas  # NOVO
        )

    except Exception as e:
        flash(f'Erro ao carregar dashboard: {str(e)}', 'danger')
        logging.error(f"Erro em dashboard_distribuicao: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return redirect(url_for('credenciamento.index'))


@limite_bp.route('/limites/api/dashboard-data')
@login_required
def api_dashboard_data():
    """
    API para buscar dados do dashboard de forma assíncrona.
    Retorna JSON com KPIs e dados para gráficos.
    """
    try:
        edital_id = request.args.get('edital_id', type=int)
        periodo_id = request.args.get('periodo_id', type=int)
        empresa_id = request.args.get('empresa_id', type=int)  # NOVO FILTRO

        # Se não informado, buscar o mais recente
        if not edital_id or not periodo_id:
            ultimo_edital = Edital.query.filter(Edital.DELETED_AT == None).order_by(Edital.ID.desc()).first()
            if ultimo_edital:
                edital_id = ultimo_edital.ID
                ultimo_periodo = PeriodoAvaliacao.query.filter(
                    PeriodoAvaliacao.ID_EDITAL == edital_id,
                    PeriodoAvaliacao.DELETED_AT == None
                ).order_by(PeriodoAvaliacao.ID_PERIODO.desc()).first()
                if ultimo_periodo:
                    periodo_id = ultimo_periodo.ID_PERIODO

        with db.engine.connect() as connection:
            # Construir WHERE clause com filtro de empresa
            where_empresa = ""
            params = {
                'edital_id': edital_id,
                'periodo_id': periodo_id
            }

            if empresa_id:
                where_empresa = "AND DIS.COD_EMPRESA_COBRANCA = :empresa_id"
                params['empresa_id'] = empresa_id

            # KPIs Gerais
            sql_kpis = text(f"""
                SELECT 
                    COUNT(DISTINCT DIS.fkContratoSISCTR) AS total_contratos,
                    COUNT(DISTINCT DIS.NR_CPF_CNPJ) AS total_cpfs,
                    COUNT(DISTINCT DIS.COD_EMPRESA_COBRANCA) AS total_empresas,
                    SUM(DIS.VR_SD_DEVEDOR) AS valor_total,
                    COUNT(DISTINCT DIS.COD_CRITERIO_SELECAO) AS total_criterios
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS
                WHERE DIS.ID_EDITAL = :edital_id
                    AND DIS.ID_PERIODO = :periodo_id
                    {where_empresa}
                    AND DIS.DELETED_AT IS NULL
            """)

            kpis = connection.execute(sql_kpis, params).fetchone()

            # Distribuição por Empresa
            sql_empresas = text(f"""
                SELECT 
                    EM.NO_ABREVIADO_EMPRESA AS empresa,
                    COUNT(DISTINCT DIS.fkContratoSISCTR) AS qtde_contratos,
                    SUM(DIS.VR_SD_DEVEDOR) AS valor_total
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS
                INNER JOIN [BDG].[PAR_TB002_EMPRESA_RESPONSAVEL_COBRANCA] EM
                    ON DIS.COD_EMPRESA_COBRANCA = EM.pkEmpresaResponsavelCobranca
                WHERE DIS.ID_EDITAL = :edital_id
                    AND DIS.ID_PERIODO = :periodo_id
                    {where_empresa}
                    AND DIS.DELETED_AT IS NULL
                GROUP BY EM.NO_ABREVIADO_EMPRESA
                ORDER BY qtde_contratos DESC
            """)

            empresas = connection.execute(sql_empresas, params).fetchall()

            # Distribuição por Critério
            sql_criterios = text(f"""
                SELECT 
                    CR.DS_CRITERIO_SELECAO AS criterio,
                    COUNT(DISTINCT DIS.fkContratoSISCTR) AS qtde_contratos,
                    SUM(DIS.VR_SD_DEVEDOR) AS valor_total
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS
                INNER JOIN [BDG].[DCA_TB004_CRITERIO_SELECAO] CR
                    ON DIS.COD_CRITERIO_SELECAO = CR.COD
                WHERE DIS.ID_EDITAL = :edital_id
                    AND DIS.ID_PERIODO = :periodo_id
                    {where_empresa}
                    AND DIS.DELETED_AT IS NULL
                GROUP BY CR.DS_CRITERIO_SELECAO
                ORDER BY qtde_contratos DESC
            """)

            criterios = connection.execute(sql_criterios, params).fetchall()

            # Distribuição por Produto
            sql_produtos = text(f"""
                SELECT 
                    PR.NO_ABREVIADO_PRODUTO AS produto,
                    COUNT(DISTINCT DIS.fkContratoSISCTR) AS qtde_contratos,
                    SUM(DIS.VR_SD_DEVEDOR) AS valor_total
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS
                INNER JOIN [BDG].[COM_TB001_CONTRATO] CTR
                    ON DIS.fkContratoSISCTR = CTR.fkContratoSISCTR
                INNER JOIN [BDG].[PAR_TB001_PRODUTOS] PR
                    ON CTR.COD_PRODUTO = PR.pkSistemaOriginario
                WHERE DIS.ID_EDITAL = :edital_id
                    AND DIS.ID_PERIODO = :periodo_id
                    {where_empresa}
                    AND DIS.DELETED_AT IS NULL
                GROUP BY PR.NO_ABREVIADO_PRODUTO
                ORDER BY qtde_contratos DESC
            """)

            produtos = connection.execute(sql_produtos, params).fetchall()

        # Montar resposta JSON
        response_data = {
            'kpis': {
                'total_contratos': kpis[0] if kpis else 0,
                'total_cpfs': kpis[1] if kpis else 0,
                'total_empresas': kpis[2] if kpis else 0,
                'valor_total': float(kpis[3]) if kpis and kpis[3] else 0,
                'total_criterios': kpis[4] if kpis else 0
            },
            'empresas': [
                {
                    'nome': row[0],
                    'qtde_contratos': row[1],
                    'valor_total': float(row[2]) if row[2] else 0
                }
                for row in empresas
            ],
            'criterios': [
                {
                    'nome': row[0],
                    'qtde_contratos': row[1],
                    'valor_total': float(row[2]) if row[2] else 0
                }
                for row in criterios
            ],
            'produtos': [
                {
                    'nome': row[0],
                    'qtde_contratos': row[1],
                    'valor_total': float(row[2]) if row[2] else 0
                }
                for row in produtos
            ]
        }

        return jsonify(response_data)

    except Exception as e:
        logging.error(f"Erro na API dashboard: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@limite_bp.route('/limites/api/buscar-contrato')
@login_required
def api_buscar_contrato():
    """
    API para buscar informações detalhadas de um contrato específico.
    Aceita tanto fkContratoSISCTR quanto NR_CONTRATO
    """
    try:
        contrato_busca = request.args.get('contrato', '').strip()

        if not contrato_busca:
            return jsonify({'error': 'Informe o número do contrato'}), 400

        with db.engine.connect() as connection:
            # Tentar buscar por fkContratoSISCTR (número) ou NR_CONTRATO (string)
            sql = text("""
                SELECT 
                    CTR.NR_CONTRATO,
                    DIS.NR_CPF_CNPJ,
                    EM.NO_ABREVIADO_EMPRESA AS empresa,
                    CR.DS_CRITERIO_SELECAO AS criterio,
                    PR.NO_ABREVIADO_PRODUTO AS produto,
                    DIS.VR_SD_DEVEDOR,
                    SIT.QT_DIAS_ATRASO,
                    DIS.DT_REFERENCIA,
                    ED.NU_EDITAL,
                    ED.ANO AS ANO_EDITAL,
                    DIS.ID_PERIODO,
                    DIS.fkContratoSISCTR
                FROM [BDG].[DCA_TB005_DISTRIBUICAO] DIS
                INNER JOIN [BDG].[COM_TB001_CONTRATO] CTR
                    ON DIS.fkContratoSISCTR = CTR.fkContratoSISCTR
                INNER JOIN [BDG].[PAR_TB002_EMPRESA_RESPONSAVEL_COBRANCA] EM
                    ON DIS.COD_EMPRESA_COBRANCA = EM.pkEmpresaResponsavelCobranca
                INNER JOIN [BDG].[DCA_TB004_CRITERIO_SELECAO] CR
                    ON DIS.COD_CRITERIO_SELECAO = CR.COD
                INNER JOIN [BDG].[PAR_TB001_PRODUTOS] PR
                    ON CTR.COD_PRODUTO = PR.pkSistemaOriginario
                INNER JOIN [BDG].[COM_TB007_SITUACAO_CONTRATOS] SIT
                    ON DIS.fkContratoSISCTR = SIT.fkContratoSISCTR
                INNER JOIN [BDG].[DCA_TB001_EDITAIS] ED
                    ON DIS.ID_EDITAL = ED.ID
                WHERE DIS.DELETED_AT IS NULL
                    AND SIT.fkSituacaoCredito = 1
                    AND (
                        CAST(DIS.fkContratoSISCTR AS VARCHAR) = :contrato
                        OR CTR.NR_CONTRATO LIKE '%' + :contrato + '%'
                    )
            """)

            resultado = connection.execute(sql, {'contrato': contrato_busca}).fetchone()

            if not resultado:
                return jsonify({'error': 'Contrato não encontrado na distribuição atual'}), 404

            return jsonify({
                'nr_contrato': resultado[0],
                'cpf_cnpj': resultado[1],
                'empresa': resultado[2],
                'criterio': resultado[3],
                'produto': resultado[4],
                'saldo_devedor': float(resultado[5]) if resultado[5] else 0,
                'dias_atraso': resultado[6] if resultado[6] else 0,
                'data_referencia': resultado[7].strftime('%d/%m/%Y') if resultado[7] else None,
                'edital': f"{resultado[8]}/{resultado[9]}",
                'periodo': resultado[10],
                'fk_contrato': resultado[11]
            })

    except Exception as e:
        logging.error(f"Erro ao buscar contrato: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': f'Erro ao buscar contrato: {str(e)}'}), 500