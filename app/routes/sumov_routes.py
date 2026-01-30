from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify
from flask_login import login_required, current_user
from app import db
from app.models.relacao_imovel_contrato import RelacaoImovelContratoParcelamento
from app.utils.audit import registrar_log
from datetime import datetime
from sqlalchemy import text
from app.models.despesas_analitico import DespesasAnalitico, OcorrenciasMovItemServico
from decimal import Decimal
from app.models.evidencias_sumov import EvidenciasSumov
from datetime import datetime, date
from app.models.deliberacao_pagamento import DeliberacaoPagamento
from flask import send_file
from dateutil.relativedelta import relativedelta
from datetime import timedelta
from flask import request, render_template, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy import text, func, distinct
from app import db
from app.models.deliberacao_pagamento import DeliberacaoPagamento
from app.utils.audit import registrar_log
from datetime import datetime
from decimal import Decimal

sumov_bp = Blueprint('sumov', __name__, url_prefix='/sumov')

@sumov_bp.context_processor
def inject_current_year():
    from datetime import datetime
    return {
        'current_year': datetime.utcnow().year,
        'data_hoje': datetime.now().strftime('%d/%m/%Y')  # ADICIONE ESTA LINHA
    }

@sumov_bp.context_processor
def inject_datetime():
    """Injeta datetime e ano atual no contexto de todos os templates do SUMOV"""
    return {
        'datetime': datetime,
        'current_year': datetime.utcnow().year
    }
@sumov_bp.route('/')

@login_required
def index():
    """Dashboard principal do SUMOV"""
    # Estatísticas
    total_vinculacoes = RelacaoImovelContratoParcelamento.query.filter_by(
        DELETED_AT=None
    ).count()

    # Vinculações do mês atual
    inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0)
    vinculacoes_mes = RelacaoImovelContratoParcelamento.query.filter(
        RelacaoImovelContratoParcelamento.CREATED_AT >= inicio_mes,
        RelacaoImovelContratoParcelamento.DELETED_AT == None
    ).count()

    # Últimas vinculações
    ultimas_vinculacoes = RelacaoImovelContratoParcelamento.query.filter_by(
        DELETED_AT=None
    ).order_by(
        RelacaoImovelContratoParcelamento.CREATED_AT.desc()
    ).limit(5).all()

    return render_template('sumov/index.html',
                           total_vinculacoes=total_vinculacoes,
                           vinculacoes_mes=vinculacoes_mes,
                           ultimas_vinculacoes=ultimas_vinculacoes)


@sumov_bp.route('/relacao-imovel-contrato')
@login_required
def relacao_imovel_contrato():
    """Dashboard do sistema de relação imóvel/contrato"""
    # Estatísticas específicas
    total_contratos = db.session.query(
        RelacaoImovelContratoParcelamento.NR_CONTRATO
    ).filter_by(DELETED_AT=None).distinct().count()

    total_imoveis = db.session.query(
        RelacaoImovelContratoParcelamento.NR_IMOVEL
    ).filter_by(DELETED_AT=None).distinct().count()

    total_vinculacoes = RelacaoImovelContratoParcelamento.query.filter_by(
        DELETED_AT=None
    ).count()

    return render_template('sumov/relacao_imovel_contrato/index.html',
                           total_contratos=total_contratos,
                           total_imoveis=total_imoveis,
                           total_vinculacoes=total_vinculacoes)


@sumov_bp.route('/relacao-imovel-contrato/lista')
@login_required
def lista_vinculacoes():
    """Lista todas as vinculações"""
    vinculacoes = RelacaoImovelContratoParcelamento.listar_vinculacoes_ativas()

    return render_template('sumov/relacao_imovel_contrato/lista.html',
                           vinculacoes=vinculacoes)


@sumov_bp.route('/relacao-imovel-contrato/nova', methods=['GET', 'POST'])
@login_required
def nova_vinculacao():
    """Criar nova vinculação entre contrato e imóvel"""
    if request.method == 'POST':
        try:
            nr_contrato = request.form.get('nr_contrato', '').strip()
            nr_imovel = request.form.get('nr_imovel', '').strip()

            # Validações
            if not nr_contrato or not nr_imovel:
                flash('Por favor, preencha todos os campos.', 'danger')
                return redirect(url_for('sumov.nova_vinculacao'))

            # Validar formato (12 dígitos)
            if len(nr_contrato) != 12 or not nr_contrato.isdigit():
                flash('Número do contrato deve ter 12 dígitos.', 'danger')
                return redirect(url_for('sumov.nova_vinculacao'))

            if len(nr_imovel) != 12 or not nr_imovel.isdigit():
                flash('Número do imóvel deve ter 12 dígitos.', 'danger')
                return redirect(url_for('sumov.nova_vinculacao'))

            # Verificar se já existe vinculação
            if RelacaoImovelContratoParcelamento.verificar_vinculacao_existente(nr_contrato, nr_imovel):
                flash('Esta vinculação já existe no sistema.', 'warning')
                return redirect(url_for('sumov.nova_vinculacao'))

            # Criar nova vinculação
            nova_relacao = RelacaoImovelContratoParcelamento(
                NR_CONTRATO=nr_contrato,
                NR_IMOVEL=nr_imovel
            )

            db.session.add(nova_relacao)
            db.session.commit()

            # Registrar log de auditoria com informação do usuário
            registrar_log(
                acao='criar',
                entidade='relacao_imovel_contrato',
                entidade_id=f"{nr_contrato}-{nr_imovel}",
                descricao=f'Nova vinculação criada: Contrato {nr_contrato} - Imóvel {nr_imovel}'
            )

            flash('Vinculação criada com sucesso!', 'success')
            return redirect(url_for('sumov.lista_vinculacoes'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar vinculação: {str(e)}', 'danger')
            return redirect(url_for('sumov.nova_vinculacao'))

    return render_template('sumov/relacao_imovel_contrato/nova.html')


@sumov_bp.route('/relacao-imovel-contrato/excluir/<contrato>/<imovel>', methods=['POST'])
@login_required
def excluir_vinculacao(contrato, imovel):
    """Excluir vinculação (soft delete)"""
    try:
        vinculacao = RelacaoImovelContratoParcelamento.query.filter_by(
            NR_CONTRATO=contrato,
            NR_IMOVEL=imovel,
            DELETED_AT=None
        ).first()

        if not vinculacao:
            flash('Vinculação não encontrada.', 'warning')
            return redirect(url_for('sumov.lista_vinculacoes'))

        # Soft delete - apenas marca a data de exclusão
        vinculacao.DELETED_AT = datetime.utcnow()

        db.session.commit()

        # Registrar log de auditoria com informação do usuário
        registrar_log(
            acao='excluir',
            entidade='relacao_imovel_contrato',
            entidade_id=f"{contrato}-{imovel}",
            descricao=f'Vinculação excluída: Contrato {contrato} - Imóvel {imovel}'
        )

        flash('Vinculação excluída com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir vinculação: {str(e)}', 'danger')

    return redirect(url_for('sumov.lista_vinculacoes'))


@sumov_bp.route('/relacao-imovel-contrato/buscar-vinculacoes', methods=['GET'])
@login_required
def buscar_vinculacoes():
    """Busca vinculações por contrato ou imóvel"""
    termo = request.args.get('termo', '').strip()
    tipo = request.args.get('tipo', 'contrato')  # 'contrato' ou 'imovel'

    if not termo:
        return jsonify({'vinculacoes': []})

    query = RelacaoImovelContratoParcelamento.query.filter_by(DELETED_AT=None)

    if tipo == 'contrato':
        query = query.filter(RelacaoImovelContratoParcelamento.NR_CONTRATO.like(f'%{termo}%'))
    else:
        query = query.filter(RelacaoImovelContratoParcelamento.NR_IMOVEL.like(f'%{termo}%'))

    vinculacoes = query.limit(10).all()

    resultado = []
    for v in vinculacoes:
        resultado.append({
            'nr_contrato': v.NR_CONTRATO,
            'nr_imovel': v.NR_IMOVEL,
            'data_criacao': v.CREATED_AT.strftime('%d/%m/%Y %H:%M') if v.CREATED_AT else ''
        })

    return jsonify({'vinculacoes': resultado})


@sumov_bp.route('/despesas-pagamentos')
@login_required
def despesas_pagamentos():
    """Dashboard do sistema de despesas e pagamentos"""
    total_registros = DespesasAnalitico.query.filter_by(
        NO_ORIGEM_REGISTRO='SUMOV'
    ).count()

    # Total do mês atual
    inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0)
    registros_mes = DespesasAnalitico.query.filter(
        DespesasAnalitico.DT_REFERENCIA >= inicio_mes,
        DespesasAnalitico.NO_ORIGEM_REGISTRO == 'SUMOV'
    ).count()

    # Valor total de despesas
    valor_total = db.session.query(
        db.func.sum(DespesasAnalitico.VR_DESPESA)
    ).filter_by(NO_ORIGEM_REGISTRO='SUMOV').scalar() or 0

    return render_template('sumov/despesas_pagamentos/index.html',
                           total_registros=total_registros,
                           registros_mes=registros_mes,
                           valor_total=valor_total)


@sumov_bp.route('/despesas-pagamentos/lista')
@login_required
def lista_despesas():
    """Lista todas as despesas registradas pelo SUMOV"""
    despesas = DespesasAnalitico.listar_despesas_sumov()
    return render_template('sumov/despesas_pagamentos/lista.html',
                           despesas=despesas)


@sumov_bp.route('/despesas-pagamentos/verificar-duplicata', methods=['POST'])
@login_required
def verificar_duplicata_despesa():
    """Verifica se já existe registro com os mesmos dados"""
    try:
        data = request.get_json()

        nr_contrato = data.get('nr_contrato', '').strip()
        id_item_servico = int(data.get('id_item_servico'))
        dt_lancamento = data.get('dt_lancamento')
        vr_despesa = data.get('vr_despesa', '0').replace(',', '.')

        origem = DespesasAnalitico.verificar_registro_existente(
            nr_contrato,
            id_item_servico,
            dt_lancamento,
            vr_despesa
        )

        return jsonify({
            'existe': origem is not None,
            'origem': origem
        })

    except Exception as e:
        return jsonify({
            'erro': str(e)
        }), 400


@sumov_bp.route('/despesas-pagamentos/nova', methods=['GET', 'POST'])
@login_required
def nova_despesa():
    """Criar novo registro de pagamento de despesa"""
    if request.method == 'POST':
        try:
            # Captura os dados do formulário
            nr_contrato = request.form.get('nr_contrato', '').strip()
            id_item_servico = int(request.form.get('id_item_servico'))
            dt_lancamento = datetime.strptime(request.form.get('dt_lancamento_pagamento'), '%Y-%m-%d')
            vr_despesa = Decimal(request.form.get('vr_despesa', '0').replace(',', '.'))
            dsc_tipo_forma_pgto = request.form.get('dsc_tipo_forma_pgto')
            confirmar_sisgea = request.form.get('confirmar_sisgea') == 'true'

            # Validações
            if not nr_contrato:
                flash('Por favor, informe o número do contrato.', 'danger')
                return redirect(url_for('sumov.nova_despesa'))

            # Verificar duplicata ANTES de prosseguir
            origem_existente = DespesasAnalitico.verificar_registro_existente(
                nr_contrato,
                id_item_servico,
                dt_lancamento,
                vr_despesa
            )

            # Se já existe registro SUMOV, bloqueia totalmente
            if origem_existente == 'SUMOV':
                flash(
                    'Este registro já foi inserido anteriormente pelo SUMOV com os mesmos dados (Contrato, Item de Serviço, Data de Pagamento e Valor).',
                    'danger')
                return redirect(url_for('sumov.nova_despesa'))

            # Se existe registro SISGEA e usuário não confirmou, retorna para confirmação
            if origem_existente == 'SISGEA' and not confirmar_sisgea:
                # Isso nunca deve acontecer porque o JavaScript trata isso
                # Mas deixamos como segurança
                flash('Este contrato já possui registro no SISGEA. Por favor, confirme se deseja continuar.', 'warning')
                return redirect(url_for('sumov.nova_despesa'))

            # Busca o item de serviço selecionado
            item_servico = OcorrenciasMovItemServico.query.get(id_item_servico)
            if not item_servico:
                flash('Item de serviço inválido.', 'danger')
                return redirect(url_for('sumov.nova_despesa'))

            # Gera o número de ocorrência
            nr_ocorrencia = DespesasAnalitico.obter_proximo_numero_ocorrencia(nr_contrato)

            # NOVA LÓGICA: Determina o ID_ITEM_SISCOR baseado no ID_ITEM_SERVICO
            if id_item_servico in [32, 33, 66]:
                id_item_siscor = 1491
            elif id_item_servico in [35, 36, 69]:
                id_item_siscor = 1490
            else:
                id_item_siscor = 1159

            # Cria o novo registro
            nova_despesa = DespesasAnalitico(
                DT_REFERENCIA=datetime.now().date(),
                NR_CONTRATO=nr_contrato,
                NR_OCORRENCIA=nr_ocorrencia,
                DSC_ITEM_SERVICO=item_servico.DSC_ITEM_SERVICO,
                DT_LANCAMENTO_PAGAMENTO=dt_lancamento.date(),
                VR_DESPESA=vr_despesa,
                estadoLancamento='Pagamento Efetuado',
                ID_ITEM_SISCOR=id_item_siscor,  # Agora usa a variável determinada pela lógica
                ID_ITEM_SERVICO=id_item_servico,
                NR_CTR_ORIGINAL_TABELA=nr_contrato,
                DSC_TIPO_FORMA_PGTO=dsc_tipo_forma_pgto,
                NO_ORIGEM_REGISTRO='SUMOV'
            )

            db.session.add(nova_despesa)
            db.session.commit()

            # Registrar log
            msg_log = f'Novo pagamento registrado: Contrato {nr_contrato} - Ocorrência {nr_ocorrencia} - Valor R$ {vr_despesa} - ID_ITEM_SISCOR: {id_item_siscor}'
            if origem_existente == 'SISGEA':
                msg_log += ' (Confirmado apesar de existir no SISGEA)'

            registrar_log(
                acao='criar',
                entidade='despesa_pagamento',
                entidade_id=nr_ocorrencia,
                descricao=msg_log
            )

            flash('Pagamento registrado com sucesso!', 'success')
            return redirect(url_for('sumov.lista_despesas'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao registrar pagamento: {str(e)}', 'danger')
            return redirect(url_for('sumov.nova_despesa'))

    # GET - Carrega os itens de serviço permitidos
    itens_servico = OcorrenciasMovItemServico.listar_itens_permitidos()

    return render_template('sumov/despesas_pagamentos/nova.html',
                           itens_servico=itens_servico)

@sumov_bp.route('/evidencias')
@login_required
def evidencias_index():
    """Dashboard do sistema de Evidências SUMOV"""
    # Estatísticas diretas do banco
    total_evidencias = EvidenciasSumov.query.count()

    # Evidências do mês atual - usando DATE
    from datetime import datetime
    hoje = datetime.now()
    primeiro_dia_mes = date(hoje.year, hoje.month, 1)

    evidencias_mes = EvidenciasSumov.query.filter(
        db.func.year(EvidenciasSumov.MESANO) == hoje.year,
        db.func.month(EvidenciasSumov.MESANO) == hoje.month
    ).count()

    # Valor total das evidências
    valor_total = db.session.query(
        db.func.sum(EvidenciasSumov.VALOR)
    ).scalar() or 0

    # Últimas 5 evidências cadastradas (por ID decrescente)
    ultimas_evidencias = EvidenciasSumov.query.order_by(
        EvidenciasSumov.ID.desc()
    ).limit(5).all()

    return render_template('sumov/evidencias/index.html',
                           total_evidencias=total_evidencias,
                           evidencias_mes=evidencias_mes,
                           valor_total=valor_total,
                           ultimas_evidencias=ultimas_evidencias)


@sumov_bp.route('/evidencias/lista')
@login_required
def evidencias_lista():
    """Lista todas as evidências cadastradas no banco"""
    # Buscar todas as evidências diretamente do banco
    evidencias = EvidenciasSumov.listar_todas()

    return render_template('sumov/evidencias/lista.html',
                           evidencias=evidencias)


@sumov_bp.route('/evidencias/nova', methods=['GET', 'POST'])
@login_required
def evidencias_nova():
    """Criar nova evidência no banco"""
    if request.method == 'POST':
        try:
            # Captura dados do formulário
            nr_contrato = request.form.get('nr_contrato', '').strip()
            mesano = request.form.get('mesano', '').strip()
            valor_str = request.form.get('valor', '0').strip()
            descricao = request.form.get('descricao', '').strip()
            id_item = request.form.get('id_item', '').strip()

            # Validações dos campos obrigatórios
            if not mesano:
                flash('Por favor, informe o mês/ano de referência.', 'danger')
                return redirect(url_for('sumov.evidencias_nova'))

            if not descricao:
                flash('Por favor, informe a descrição.', 'danger')
                return redirect(url_for('sumov.evidencias_nova'))

            # Converter valor - remover formatação brasileira
            try:
                # Remove pontos de milhar e troca vírgula por ponto
                valor_str = valor_str.replace('.', '').replace(',', '.')
                valor = Decimal(valor_str) if valor_str else Decimal('0')

                # MODIFICAÇÃO: Aceita valores negativos, apenas impede zero
                if valor == 0:
                    flash('Por favor, informe um valor diferente de zero.', 'danger')
                    return redirect(url_for('sumov.evidencias_nova'))
            except:
                flash('Valor inválido. Use o formato: 1.234,56', 'danger')
                return redirect(url_for('sumov.evidencias_nova'))

            # Converter MESANO de MM/YYYY para date
            data_mesano = None
            if '/' in mesano:
                partes = mesano.split('/')
                if len(partes) == 2:
                    try:
                        mes = int(partes[0])
                        ano = int(partes[1])
                        # Criar data com primeiro dia do mês
                        data_mesano = date(ano, mes, 1)
                    except:
                        flash('Mês/Ano inválido. Use o formato MM/YYYY', 'danger')
                        return redirect(url_for('sumov.evidencias_nova'))

            # Converter NR_CONTRATO para Decimal ou None
            nr_contrato_decimal = None
            if nr_contrato and nr_contrato.replace(' ', '').isdigit():
                nr_contrato_decimal = Decimal(nr_contrato.replace(' ', ''))

            # Converter ID_ITEM para int ou None
            id_item_int = None
            if id_item and id_item.isdigit():
                id_item_int = int(id_item)

            # Criar nova evidência diretamente no banco
            nova_evidencia = EvidenciasSumov(
                NR_CONTRATO=nr_contrato_decimal,
                MESANO=data_mesano,
                VALOR=valor,
                DESCRICAO=descricao[:150],  # Limitar a 150 caracteres
                ID_ITEM=id_item_int
            )

            db.session.add(nova_evidencia)
            db.session.commit()

            # Registrar log de auditoria
            registrar_log(
                acao='criar',
                entidade='evidencia_sumov',
                entidade_id=nova_evidencia.ID,
                descricao=f'Nova evidência criada: {mesano} - Valor R$ {valor}'
            )

            flash('Evidência cadastrada com sucesso!', 'success')
            return redirect(url_for('sumov.evidencias_lista'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar evidência: {str(e)}', 'danger')
            return redirect(url_for('sumov.evidencias_nova'))

    # GET - Exibe o formulário
    return render_template('sumov/evidencias/nova.html')


@sumov_bp.route('/evidencias/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def evidencias_editar(id):
    """Editar evidência existente no banco"""
    # Buscar evidência diretamente do banco
    evidencia = EvidenciasSumov.buscar_por_id(id)
    if not evidencia:
        flash('Evidência não encontrada.', 'danger')
        return redirect(url_for('sumov.evidencias_lista'))

    if request.method == 'POST':
        try:
            # Capturar dados antigos para auditoria
            dados_antigos = {
                'nr_contrato': str(int(evidencia.NR_CONTRATO)) if evidencia.NR_CONTRATO else None,
                'mesano': evidencia.formatar_mesano(),
                'valor': float(evidencia.VALOR) if evidencia.VALOR else 0,
                'descricao': evidencia.DESCRICAO,
                'id_item': evidencia.ID_ITEM
            }

            # Capturar novos dados
            nr_contrato = request.form.get('nr_contrato', '').strip()
            mesano = request.form.get('mesano', '').strip()
            valor_str = request.form.get('valor', '0').strip()
            descricao_nova = request.form.get('descricao', '').strip()
            id_item = request.form.get('id_item', '').strip()

            # Validações
            if not mesano:
                flash('Por favor, informe o mês/ano de referência.', 'danger')
                return render_template('sumov/evidencias/editar.html', evidencia=evidencia)

            if not descricao_nova:
                flash('Por favor, informe a descrição.', 'danger')
                return render_template('sumov/evidencias/editar.html', evidencia=evidencia)

            # Converter valor - remover formatação brasileira
            try:
                # Remove pontos de milhar e troca vírgula por ponto
                valor_str = valor_str.replace('.', '').replace(',', '.')
                valor_novo = Decimal(valor_str) if valor_str else Decimal('0')

                # MODIFICAÇÃO: Aceita valores negativos, apenas impede zero
                if valor_novo == 0:
                    flash('Por favor, informe um valor diferente de zero.', 'danger')
                    return render_template('sumov/evidencias/editar.html', evidencia=evidencia)
            except:
                flash('Valor inválido. Use o formato: 1.234,56', 'danger')
                return render_template('sumov/evidencias/editar.html', evidencia=evidencia)

            # Converter MESANO de MM/YYYY para date
            if '/' in mesano:
                partes = mesano.split('/')
                if len(partes) == 2:
                    try:
                        mes = int(partes[0])
                        ano = int(partes[1])
                        evidencia.MESANO = date(ano, mes, 1)
                    except:
                        flash('Mês/Ano inválido. Use o formato MM/YYYY', 'danger')
                        return render_template('sumov/evidencias/editar.html', evidencia=evidencia)

            # Converter NR_CONTRATO para Decimal ou None
            if nr_contrato and nr_contrato.replace(' ', '').isdigit():
                evidencia.NR_CONTRATO = Decimal(nr_contrato.replace(' ', ''))
            else:
                evidencia.NR_CONTRATO = None

            # Converter ID_ITEM para int ou None
            if id_item and id_item.isdigit():
                evidencia.ID_ITEM = int(id_item)
            else:
                evidencia.ID_ITEM = None

            # Atualizar outros campos
            evidencia.VALOR = valor_novo
            evidencia.DESCRICAO = descricao_nova[:150]  # Limitar a 150 caracteres

            # Salvar alterações no banco
            db.session.commit()

            # Registrar log de auditoria
            dados_novos = {
                'nr_contrato': str(int(evidencia.NR_CONTRATO)) if evidencia.NR_CONTRATO else None,
                'mesano': evidencia.formatar_mesano(),
                'valor': float(evidencia.VALOR) if evidencia.VALOR else 0,
                'descricao': evidencia.DESCRICAO,
                'id_item': evidencia.ID_ITEM
            }

            registrar_log(
                acao='editar',
                entidade='evidencia_sumov',
                entidade_id=evidencia.ID,
                descricao=f'Evidência editada: ID {evidencia.ID}',
                dados_antigos=dados_antigos,
                dados_novos=dados_novos
            )

            flash('Evidência atualizada com sucesso!', 'success')
            return redirect(url_for('sumov.evidencias_lista'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar evidência: {str(e)}', 'danger')
            return render_template('sumov/evidencias/editar.html', evidencia=evidencia)

    return render_template('sumov/evidencias/editar.html', evidencia=evidencia)

@sumov_bp.route('/evidencias/excluir/<int:id>')
@login_required
def evidencias_excluir(id):
    """Excluir evidência do banco (delete real)"""
    try:
        # Buscar evidência no banco
        evidencia = EvidenciasSumov.buscar_por_id(id)
        if not evidencia:
            flash('Evidência não encontrada.', 'danger')
            return redirect(url_for('sumov.evidencias_lista'))

        # Guardar informações para o log antes de deletar
        info_evidencia = f"ID {evidencia.ID} - {evidencia.formatar_mesano()} - R$ {evidencia.VALOR}"

        # Deletar do banco (delete real)
        db.session.delete(evidencia)
        db.session.commit()

        # Registrar log
        registrar_log(
            acao='excluir',
            entidade='evidencia_sumov',
            entidade_id=id,
            descricao=f'Evidência excluída: {info_evidencia}'
        )

        flash('Evidência excluída com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir evidência: {str(e)}', 'danger')

    return redirect(url_for('sumov.evidencias_lista'))


@sumov_bp.route('/despesas-pagamentos/analise', methods=['GET'])
@login_required
def analise_pagamentos():
    """Página de análise de pagamentos por mês/ano e item de serviço"""
    # Buscar todos os itens de serviço permitidos
    itens_servico = OcorrenciasMovItemServico.listar_itens_permitidos()

    # Buscar todas as datas distintas de DT_LANCAMENTO_PAGAMENTO para montar o select de mês/ano
    datas_disponiveis = db.session.query(
        db.func.year(DespesasAnalitico.DT_LANCAMENTO_PAGAMENTO).label('ano'),
        db.func.month(DespesasAnalitico.DT_LANCAMENTO_PAGAMENTO).label('mes')
    ).filter(
        DespesasAnalitico.DT_LANCAMENTO_PAGAMENTO.isnot(None)
    ).distinct().order_by(
        db.text('ano DESC, mes DESC')
    ).all()

    # Formatar datas para exibição
    meses_anos = []
    meses_nomes = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }

    for data in datas_disponiveis:
        meses_anos.append({
            'ano': data.ano,
            'mes': data.mes,
            'label': f"{meses_nomes[data.mes]} de {data.ano}"
        })

    return render_template('sumov/despesas_pagamentos/analise_pagamentos.html',
                           itens_servico=itens_servico,
                           meses_anos=meses_anos)


@sumov_bp.route('/despesas-pagamentos/analise/buscar', methods=['POST'])
@login_required
def buscar_analise_pagamentos():
    """Busca dados de pagamentos filtrados por mês/ano e itens de serviço"""
    try:
        data = request.get_json()

        mes = int(data.get('mes'))
        ano = int(data.get('ano'))
        itens_selecionados = data.get('itens', [])

        if not itens_selecionados:
            return jsonify({'erro': 'Selecione pelo menos um item de serviço'}), 400

        # Converter itens para inteiros
        itens_selecionados = [int(item) for item in itens_selecionados]

        # Buscar registros SUMOV usando DT_LANCAMENTO_PAGAMENTO
        registros_sumov = DespesasAnalitico.query.filter(
            db.func.year(DespesasAnalitico.DT_LANCAMENTO_PAGAMENTO) == ano,
            db.func.month(DespesasAnalitico.DT_LANCAMENTO_PAGAMENTO) == mes,
            DespesasAnalitico.ID_ITEM_SERVICO.in_(itens_selecionados),
            DespesasAnalitico.NO_ORIGEM_REGISTRO == 'SUMOV'
        ).all()

        # Buscar registros SISGEA usando DT_LANCAMENTO_PAGAMENTO
        registros_sisgea = DespesasAnalitico.query.filter(
            db.func.year(DespesasAnalitico.DT_LANCAMENTO_PAGAMENTO) == ano,
            db.func.month(DespesasAnalitico.DT_LANCAMENTO_PAGAMENTO) == mes,
            DespesasAnalitico.ID_ITEM_SERVICO.in_(itens_selecionados),
            DespesasAnalitico.NO_ORIGEM_REGISTRO == 'SISGEA'
        ).all()

        # Calcular estatísticas SUMOV
        total_sumov = Decimal('0')
        contratos_unicos_sumov = set()
        itens_agrupados_sumov = {}

        for registro in registros_sumov:
            valor = registro.VR_DESPESA if registro.VR_DESPESA else Decimal('0')
            total_sumov += valor
            contratos_unicos_sumov.add(registro.NR_CONTRATO)

            # Agrupar por item de serviço
            item = registro.DSC_ITEM_SERVICO or 'Sem descrição'
            if item not in itens_agrupados_sumov:
                itens_agrupados_sumov[item] = {'quantidade': 0, 'valor_total': Decimal('0')}
            itens_agrupados_sumov[item]['quantidade'] += 1
            itens_agrupados_sumov[item]['valor_total'] += valor

        # Calcular estatísticas SISGEA
        total_sisgea = Decimal('0')
        contratos_unicos_sisgea = set()
        itens_agrupados_sisgea = {}

        for registro in registros_sisgea:
            valor = registro.VR_DESPESA if registro.VR_DESPESA else Decimal('0')
            total_sisgea += valor
            contratos_unicos_sisgea.add(registro.NR_CONTRATO)

            # Agrupar por item de serviço
            item = registro.DSC_ITEM_SERVICO or 'Sem descrição'
            if item not in itens_agrupados_sisgea:
                itens_agrupados_sisgea[item] = {'quantidade': 0, 'valor_total': Decimal('0')}
            itens_agrupados_sisgea[item]['quantidade'] += 1
            itens_agrupados_sisgea[item]['valor_total'] += valor

        # Formatar itens agrupados SUMOV
        itens_formatados_sumov = []
        for item, dados in itens_agrupados_sumov.items():
            itens_formatados_sumov.append({
                'item': item,
                'quantidade': dados['quantidade'],
                'valor_total': float(dados['valor_total']),
                'valor_total_formatado': f"R$ {dados['valor_total']:,.2f}".replace(",", "X").replace(".", ",").replace(
                    "X", ".")
            })

        # Formatar itens agrupados SISGEA
        itens_formatados_sisgea = []
        for item, dados in itens_agrupados_sisgea.items():
            itens_formatados_sisgea.append({
                'item': item,
                'quantidade': dados['quantidade'],
                'valor_total': float(dados['valor_total']),
                'valor_total_formatado': f"R$ {dados['valor_total']:,.2f}".replace(",", "X").replace(".", ",").replace(
                    "X", ".")
            })

        return jsonify({
            'sumov': {
                'total': float(total_sumov),
                'total_formatado': f"R$ {total_sumov:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                'quantidade_registros': len(registros_sumov),
                'quantidade_contratos': len(contratos_unicos_sumov),
                'itens': itens_formatados_sumov
            },
            'sisgea': {
                'total': float(total_sisgea),
                'total_formatado': f"R$ {total_sisgea:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                'quantidade_registros': len(registros_sisgea),
                'quantidade_contratos': len(contratos_unicos_sisgea),
                'itens': itens_formatados_sisgea
            }
        })

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# ==================== DELIBERAÇÃO DE PAGAMENTO ====================

@sumov_bp.route('/deliberacao-pagamento')
@login_required
def deliberacao_pagamento():
    """Dashboard do sistema de Deliberação de Pagamento com lista de deliberações e filtros"""
    try:
        from datetime import datetime
        from app.models.usuario import Empregado  # Importar o modelo Empregado

        # Capturar parâmetros de filtro
        filtro_contrato = request.args.get('filtro_contrato', '').strip()
        filtro_analista = request.args.get('filtro_analista', '').strip()
        filtro_status = request.args.get('filtro_status', '').strip()

        # Construir query base
        query = DeliberacaoPagamento.query.filter(
            DeliberacaoPagamento.DELETED_AT.is_(None)
        )

        # Aplicar filtros
        if filtro_contrato:
            query = query.filter(DeliberacaoPagamento.NU_CONTRATO.like(f'%{filtro_contrato}%'))

        if filtro_analista:
            query = query.filter(DeliberacaoPagamento.COLABORADOR_ANALISOU == filtro_analista)

        if filtro_status:
            query = query.filter(DeliberacaoPagamento.STATUS_DOCUMENTO == filtro_status)

        # Ordenar e buscar
        deliberacoes = query.order_by(DeliberacaoPagamento.CREATED_AT.desc()).all()

        # ===== NOVO: Buscar TODOS os empregados da GEADI =====
        # Busca todos que têm sgSetor = 'GEADI' e estão ativos (fkStatus = 1)
        empregados_geadi = Empregado.query.filter(
            Empregado.sgSetor == 'GEADI',
            Empregado.fkStatus == 1  # Apenas ativos
        ).order_by(Empregado.nmPessoa).all()

        # Lista de nomes dos analistas da GEADI
        lista_analistas = [emp.nmPessoa for emp in empregados_geadi if emp.nmPessoa]
        # Remove duplicados e ordena (pode haver registros duplicados por data de referência)
        lista_analistas = sorted(list(set(lista_analistas)))

        # Buscar valores distintos para status
        status_distintos = db.session.query(
            distinct(DeliberacaoPagamento.STATUS_DOCUMENTO)
        ).filter(
            DeliberacaoPagamento.DELETED_AT.is_(None),
            DeliberacaoPagamento.STATUS_DOCUMENTO.isnot(None)
        ).order_by(DeliberacaoPagamento.STATUS_DOCUMENTO).all()

        # Transformar em lista simples
        lista_status = [s[0] for s in status_distintos if s[0]]

        return render_template('sumov/deliberacao_pagamento/index.html',
                               deliberacoes=deliberacoes,
                               data_hoje=datetime.now().strftime('%d/%m/%Y'),
                               lista_analistas=lista_analistas,
                               lista_status=lista_status,
                               filtro_contrato=filtro_contrato,
                               filtro_analista=filtro_analista,
                               filtro_status=filtro_status)
    except Exception as e:
        from datetime import datetime
        flash(f'Erro ao carregar deliberações: {str(e)}', 'danger')
        return render_template('sumov/deliberacao_pagamento/index.html',
                               deliberacoes=[],
                               data_hoje=datetime.now().strftime('%d/%m/%Y'),
                               lista_analistas=[],
                               lista_status=[],
                               filtro_contrato='',
                               filtro_analista='',
                               filtro_status='')


@sumov_bp.route('/deliberacao-pagamento/nova', methods=['GET', 'POST'])
@login_required
def deliberacao_pagamento_nova():
    """Criar nova Deliberação de Pagamento seguindo estrutura do Word"""
    if request.method == 'POST':
        try:
            from decimal import Decimal, InvalidOperation

            # ===== CAPTURA DADOS DO FORMULÁRIO =====
            contrato = request.form.get('contrato', '').strip()
            matricula = request.form.get('matricula', '').strip() or None
            dt_arrematacao_str = request.form.get('dt_arrematacao', '').strip()
            indice_selecionado_idx = request.form.get('indice_economico', '').strip()
            dt_registro_str = request.form.get('dt_registro', '').strip()

            # NOVOS CAMPOS
            vr_divida_condominio_2_str = request.form.get('vr_divida_condominio_2', '').strip()
            dt_periodo_inicio_str = request.form.get('dt_periodo_cobranca_inicio', '').strip()
            dt_periodo_fim_str = request.form.get('dt_periodo_cobranca_fim', '').strip()

            # Campos da parte final
            gravame_matricula = request.form.get('gravame_matricula', '').strip() or None
            acoes_negociais_adm = request.form.get('acoes_negociais_administrativas', '').strip() or None
            nr_processos = request.form.get('nr_processos_judiciais', '').strip() or None
            vara_processo = request.form.get('vara_processo', '').strip() or None
            fase_processo = request.form.get('fase_processo', '').strip() or None
            relatorio_assessoria = request.form.get('relatorio_assessoria_juridica', '').strip() or None
            penalidade_ans = request.form.get('penalidade_ans_caixa', '').strip() or None
            prejuizo_financeiro = request.form.get('prejuizo_financeiro_caixa', '').strip() or None
            consideracoes_analista = request.form.get('consideracoes_analista', '').strip() or None
            consideracoes_gestor = request.form.get('consideracoes_gestor', '').strip() or None
            consideracoes_gestor_sumov = request.form.get('consideracoes_gestor_sumov', '').strip() or None
            tipo_pagamento_venda = request.form.get('tipo_pagamento_venda', '').strip() or None

            # ===== VALIDAÇÕES OBRIGATÓRIAS =====
            if not contrato:
                flash('Por favor, informe o Contrato/Imóvel.', 'danger')
                return redirect(url_for('sumov.deliberacao_pagamento_nova'))

            if not indice_selecionado_idx:
                flash('Por favor, selecione o Índice Econômico.', 'danger')
                return redirect(url_for('sumov.deliberacao_pagamento_nova'))

            # ===== CONVERSÃO DE DATAS =====

            # Converter dt_arrematacao (OPCIONAL)
            dt_arrematacao = None
            if dt_arrematacao_str:
                try:
                    dt_arrematacao = datetime.strptime(dt_arrematacao_str, '%Y-%m-%d').date()
                except ValueError:
                    flash('Data de Arrematação inválida.', 'danger')
                    return redirect(url_for('sumov.deliberacao_pagamento_nova'))

            # ✅ Converter dt_registro (OPCIONAL - PODE SER VAZIO AGORA)
            dt_registro = None
            if dt_registro_str:
                try:
                    dt_registro = datetime.strptime(dt_registro_str, '%Y-%m-%d').date()
                except ValueError:
                    flash('Data do Registro inválida.', 'danger')
                    return redirect(url_for('sumov.deliberacao_pagamento_nova'))

            # Converter datas do período de cobrança (OPCIONAL)
            dt_periodo_cobranca_inicio = None
            if dt_periodo_inicio_str:
                try:
                    dt_periodo_cobranca_inicio = datetime.strptime(dt_periodo_inicio_str, '%Y-%m-%d').date()
                except ValueError:
                    flash('Data de início do período de cobrança inválida.', 'danger')
                    return redirect(url_for('sumov.deliberacao_pagamento_nova'))

            dt_periodo_cobranca_fim = None
            if dt_periodo_fim_str:
                try:
                    dt_periodo_cobranca_fim = datetime.strptime(dt_periodo_fim_str, '%Y-%m-%d').date()
                except ValueError:
                    flash('Data de fim do período de cobrança inválida.', 'danger')
                    return redirect(url_for('sumov.deliberacao_pagamento_nova'))

            # ===== CONVERSÃO DE VALORES =====

            # Converter valor manual (OPCIONAL)
            vr_divida_condominio_2 = None
            if vr_divida_condominio_2_str:
                try:
                    # Remover formatação brasileira e converter
                    valor_limpo = vr_divida_condominio_2_str.replace('R$', '').replace('.', '').replace(',',
                                                                                                        '.').strip()
                    if valor_limpo:
                        vr_divida_condominio_2 = Decimal(valor_limpo)
                except (ValueError, InvalidOperation):
                    flash('Valor da Dívida (Manual) inválido.', 'danger')
                    return redirect(url_for('sumov.deliberacao_pagamento_nova'))

            # ===== BUSCA 1: Data de Entrada no Estoque =====
            sql_estoque = text("""
                SELECT TOP 1 
                    [DT_ENTRADA_ESTOQUE]
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB012_IMOVEIS_NAO_USO_ESTOQUE]
                WHERE [NR_CONTRATO] = :contrato
            """)
            result_estoque = db.session.execute(sql_estoque, {'contrato': contrato}).fetchone()
            dt_entrada_estoque = result_estoque[0] if result_estoque else None

            # ===== BUSCA 2: Período de Prescrição =====
            sql_prescricao = text("""
                SELECT TOP 1
                    [PERIODO_PRESCRICAO]
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB032_SISCALCULO_PRESCRICOES]
                WHERE [IMOVEL] = :contrato
                ORDER BY [DT_VENCIMENTO] DESC
            """)
            result_prescricao = db.session.execute(sql_prescricao, {'contrato': contrato}).fetchone()
            periodo_prescrito = result_prescricao[0] if result_prescricao else None

            # ===== BUSCA 3: Valor Inicial (DA ÚLTIMA ATUALIZAÇÃO) =====
            sql_valor_inicial = text("""
                SELECT 
                    SUM([VR_COTA]) as VALOR_INICIAL,
                    MAX([NOME_CONDOMINIO]) as NOME_CONDOMINIO
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB030_SISCALCULO_DADOS]
                WHERE [IMOVEL] = :contrato
                AND [DT_ATUALIZACAO] = (
                    SELECT MAX([DT_ATUALIZACAO])
                    FROM [BDDASHBOARDBI].[BDG].[MOV_TB030_SISCALCULO_DADOS]
                    WHERE [IMOVEL] = :contrato
                )
            """)
            result_valor = db.session.execute(sql_valor_inicial, {'contrato': contrato}).fetchone()

            vr_divida_inicial = result_valor[0] if result_valor and result_valor[0] else Decimal('0')
            nome_condominio = result_valor[1] if result_valor and result_valor[1] else None

            # ===== BUSCA 4: Valor de Débito EXCLUÍDOS as Cotas Prescritas =====
            sql_valor_prescrito = text("""
                SELECT 
                    SUM([VR_COTA]) as VALOR_PRESCRITO
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB032_SISCALCULO_PRESCRICOES]
                WHERE [IMOVEL] = :contrato
                AND [ID_INDICE_ECONOMICO] = (
                    SELECT MIN([ID_INDICE_ECONOMICO])
                    FROM [BDDASHBOARDBI].[BDG].[MOV_TB032_SISCALCULO_PRESCRICOES]
                    WHERE [IMOVEL] = :contrato
                    AND [DT_ATUALIZACAO] = (
                        SELECT MAX([DT_ATUALIZACAO])
                        FROM [BDDASHBOARDBI].[BDG].[MOV_TB032_SISCALCULO_PRESCRICOES]
                        WHERE [IMOVEL] = :contrato
                    )
                )
                AND [DT_ATUALIZACAO] = (
                    SELECT MAX([DT_ATUALIZACAO])
                    FROM [BDDASHBOARDBI].[BDG].[MOV_TB032_SISCALCULO_PRESCRICOES]
                    WHERE [IMOVEL] = :contrato
                )
            """)
            result_prescrito = db.session.execute(sql_valor_prescrito, {'contrato': contrato}).fetchone()
            vr_debito_excluido_prescritas = result_prescrito[0] if result_prescrito and result_prescrito[
                0] else Decimal('0')

            # ===== BUSCA 5: Cálculos do SISCalculo =====
            sql_calculos = text("""
                SELECT 
                    ID_INDICE_ECONOMICO,
                    SUM(VR_TOTAL) AS VR_TOTAL_SEM_HONORARIOS,
                    MAX(PERC_HONORARIOS) AS PERC_HONORARIOS,
                    SUM(VR_TOTAL) * MAX(PERC_HONORARIOS) / 100.0 AS VR_HONORARIOS,
                    SUM(VR_TOTAL) + (SUM(VR_TOTAL) * MAX(PERC_HONORARIOS) / 100.0) AS VR_TOTAL_COM_HONORARIOS
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB031_SISCALCULO_CALCULOS]
                WHERE [IMOVEL] = :contrato
                AND [DT_ATUALIZACAO] = (
                    SELECT MAX([DT_ATUALIZACAO])
                    FROM [BDDASHBOARDBI].[BDG].[MOV_TB031_SISCALCULO_CALCULOS]
                    WHERE [IMOVEL] = :contrato
                )
                GROUP BY ID_INDICE_ECONOMICO
                ORDER BY ID_INDICE_ECONOMICO
            """)
            result_calculos = db.session.execute(sql_calculos, {'contrato': contrato}).fetchall()

            if not result_calculos:
                flash('Nenhum cálculo encontrado no SISCalculo para este contrato. Processe os cálculos primeiro.',
                      'warning')
                return redirect(url_for('sumov.deliberacao_pagamento_nova'))

            # Pegar o índice selecionado
            try:
                idx = int(indice_selecionado_idx)
                calculo = result_calculos[idx]
            except (ValueError, IndexError):
                flash('Índice econômico inválido.', 'danger')
                return redirect(url_for('sumov.deliberacao_pagamento_nova'))

            id_indice = calculo[0]
            vr_divida_calculada = calculo[1]
            perc_honorarios_emgea = calculo[2]
            vr_honorarios = calculo[3]
            vr_debito_calculado_emgea = calculo[4]

            # Buscar nome do índice
            from app.models.siscalculo import ParamIndicesEconomicos
            indice_obj = ParamIndicesEconomicos.query.get(id_indice)
            indice_debito_emgea = indice_obj.DSC_INDICE_ECONOMICO if indice_obj else f"Índice {id_indice}"

            # ===== BUSCA 6: Valor de Avaliação e Data do Laudo =====
            sql_avaliacao = text("""
                SELECT TOP 1
                    [VR_LAUDO_AVALIACAO],
                    [DT_LAUDO]
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB001_IMOVEIS_NAO_USO_STATUS]
                WHERE [NR_CONTRATO] = :contrato
                ORDER BY [DT_REFERENCIA] DESC
            """)
            result_avaliacao = db.session.execute(sql_avaliacao, {'contrato': contrato}).fetchone()

            vr_avaliacao = result_avaliacao[0] if result_avaliacao and result_avaliacao[0] else None
            dt_laudo = result_avaliacao[1] if result_avaliacao and result_avaliacao[1] else None

            # ===== BUSCA 7: Status do Imóvel =====
            sql_status = text("""
                SELECT TOP 1
                    [DSC_SITUACAO]
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB020_IMOVEIS_RM_TOTVS]
                WHERE [IMOVEL] = :contrato
            """)
            result_status = db.session.execute(sql_status, {'contrato': contrato}).fetchone()
            status_imovel = result_status[0] if result_status else None

            # ===== BUSCA 8: Dados de Venda =====
            sql_venda = text("""
                SELECT TOP 1
                    [VR_VENDA],
                    [DT_VENDA],
                    [NO_COMPRADOR]
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB023_VENDA_IMOVEIS_RM_TOTVS]
                WHERE [NU_IMOVEL] = :contrato
                ORDER BY [DT_VENDA] DESC
            """)
            result_venda = db.session.execute(sql_venda, {'contrato': contrato}).fetchone()

            vr_venda = result_venda[0] if result_venda and result_venda[0] else None
            dt_venda = result_venda[1] if result_venda and result_venda[1] else None
            nome_comprador = result_venda[2] if result_venda and result_venda[2] else None

            # ===== BUSCA 9: Débitos Pagos (SISDEX + SISGEA + SISINC) =====
            sql_sisdex = text("""
                SELECT SUM([VR_PAGO]) as TOTAL_SISDEX
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB017_DESPESAS_MANUTENCAO]
                WHERE [NU_CONTRATO] = :contrato
            """)
            result_sisdex = db.session.execute(sql_sisdex, {'contrato': contrato}).fetchone()
            vr_debitos_sisdex = result_sisdex[0] if result_sisdex and result_sisdex[0] else Decimal('0')

            sql_sisgea = text("""
                SELECT SUM([VR_DESPESA]) as TOTAL_SISGEA
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB004_DESPESAS_ANALITICO]
                WHERE [NR_CONTRATO] = :contrato
            """)
            result_sisgea = db.session.execute(sql_sisgea, {'contrato': contrato}).fetchone()
            vr_debitos_sisgea = result_sisgea[0] if result_sisgea and result_sisgea[0] else Decimal('0')

            # ===== NOVO: BUSCA SISINC =====
            sql_sisinc = text("""
                SELECT SUM([VR_LANCAMENTO]) as TOTAL_SISINC
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB014_DESPESAS_EXECUCAO_SISINC]
                WHERE [NU_CONTRATO] = :contrato
            """)
            result_sisinc = db.session.execute(sql_sisinc, {'contrato': contrato}).fetchone()
            vr_debitos_sisinc = result_sisinc[0] if result_sisinc and result_sisinc[0] else Decimal('0')

            # TOTAL agora inclui SISINC
            vr_debitos_total = vr_debitos_sisdex + vr_debitos_sisgea + vr_debitos_sisinc

            # ===== VERIFICAR SE JÁ EXISTE REGISTRO =====
            deliberacao_existente = DeliberacaoPagamento.buscar_por_contrato(contrato)

            # Determinar status e usuários que deliberaram
            status_documento = 'RASCUNHO'
            usuario_deliberou = None
            usuario_gestor_geadi = None
            usuario_gestor_sumov = None
            dt_deliberacao_geadi = None
            dt_deliberacao_sumov = None

            # Se preencheu considerações do Gestor GEADI
            if consideracoes_gestor:
                status_documento = 'DELIBERADO'
                usuario_gestor_geadi = current_user.nome
                dt_deliberacao_geadi = datetime.utcnow()
                if not usuario_deliberou:  # Primeiro a deliberar
                    usuario_deliberou = current_user.nome

            # Se preencheu considerações do Gestor SUMOV
            if consideracoes_gestor_sumov:
                status_documento = 'DELIBERADO'
                usuario_gestor_sumov = current_user.nome
                dt_deliberacao_sumov = datetime.utcnow()
                if not usuario_deliberou:  # Primeiro a deliberar
                    usuario_deliberou = current_user.nome

            if deliberacao_existente:
                # ATUALIZAR registro existente
                deliberacao_existente.COLABORADOR_ANALISOU = current_user.nome
                deliberacao_existente.DT_ANALISE = datetime.now().date()
                deliberacao_existente.MATRICULA_CAIXA_EMGEA = matricula
                deliberacao_existente.DT_ARREMATACAO_AQUISICAO = dt_arrematacao
                deliberacao_existente.DT_ENTRADA_ESTOQUE = dt_entrada_estoque
                deliberacao_existente.PERIODO_PRESCRITO = periodo_prescrito
                deliberacao_existente.DT_PERIODO_COBRANCA_INICIO = dt_periodo_cobranca_inicio
                deliberacao_existente.DT_PERIODO_COBRANCA_FIM = dt_periodo_cobranca_fim
                deliberacao_existente.VR_DIVIDA_CONDOMINIO_1 = vr_divida_inicial
                deliberacao_existente.VR_DIVIDA_CONDOMINIO_2 = vr_divida_condominio_2
                deliberacao_existente.VR_DEBITO_EXCLUIDO_PRESCRITAS = vr_debito_excluido_prescritas
                deliberacao_existente.VR_DEBITO_CALCULADO_EMGEA = vr_debito_calculado_emgea
                deliberacao_existente.INDICE_DEBITO_EMGEA = indice_debito_emgea
                deliberacao_existente.PERC_HONORARIOS_EMGEA = perc_honorarios_emgea
                deliberacao_existente.VR_HONORARIOS_EMGEA = vr_honorarios
                deliberacao_existente.DT_CALCULO_EMGEA = datetime.now().date()
                deliberacao_existente.VR_AVALIACAO = vr_avaliacao
                deliberacao_existente.DT_LAUDO = dt_laudo
                deliberacao_existente.STATUS_IMOVEL = status_imovel
                deliberacao_existente.VR_VENDA = vr_venda
                deliberacao_existente.DT_VENDA = dt_venda
                deliberacao_existente.NOME_COMPRADOR = nome_comprador
                deliberacao_existente.DT_REGISTRO = dt_registro
                deliberacao_existente.VR_DEBITOS_SISDEX = vr_debitos_sisdex
                deliberacao_existente.VR_DEBITOS_SISGEA = vr_debitos_sisgea
                deliberacao_existente.VR_DEBITOS_SISINC = vr_debitos_sisinc  # NOVO
                deliberacao_existente.VR_DEBITOS_TOTAL = vr_debitos_total
                deliberacao_existente.GRAVAME_MATRICULA = gravame_matricula
                deliberacao_existente.ACOES_NEGOCIAIS_ADMINISTRATIVAS = acoes_negociais_adm
                deliberacao_existente.NR_PROCESSOS_JUDICIAIS = nr_processos
                deliberacao_existente.VARA_PROCESSO = vara_processo
                deliberacao_existente.FASE_PROCESSO = fase_processo
                deliberacao_existente.RELATORIO_ASSESSORIA_JURIDICA = relatorio_assessoria
                deliberacao_existente.PENALIDADE_ANS_CAIXA = penalidade_ans
                deliberacao_existente.PREJUIZO_FINANCEIRO_CAIXA = prejuizo_financeiro
                deliberacao_existente.CONSIDERACOES_ANALISTA_GEADI = consideracoes_analista
                deliberacao_existente.CONSIDERACOES_GESTOR_GEADI = consideracoes_gestor
                deliberacao_existente.CONSIDERACOES_GESTOR_SUMOV = consideracoes_gestor_sumov
                deliberacao_existente.TIPO_PAGAMENTO_VENDA = tipo_pagamento_venda
                deliberacao_existente.STATUS_DOCUMENTO = status_documento
                # Atualizar usuário geral (mantém o primeiro)
                if status_documento == 'DELIBERADO' and not deliberacao_existente.USUARIO_DELIBEROU:
                    deliberacao_existente.USUARIO_DELIBEROU = usuario_deliberou
                # Atualizar usuário GEADI (sempre atualiza se tiver considerações)
                if consideracoes_gestor:
                    deliberacao_existente.USUARIO_GESTOR_GEADI = usuario_gestor_geadi
                    deliberacao_existente.DT_DELIBERACAO_GEADI = dt_deliberacao_geadi
                # Atualizar usuário SUMOV (sempre atualiza se tiver considerações)
                if consideracoes_gestor_sumov:
                    deliberacao_existente.USUARIO_GESTOR_SUMOV = usuario_gestor_sumov
                    deliberacao_existente.DT_DELIBERACAO_SUMOV = dt_deliberacao_sumov
                deliberacao_existente.USUARIO_ATUALIZACAO = current_user.nome
                deliberacao_existente.UPDATED_AT = datetime.utcnow()
                deliberacao = deliberacao_existente
                acao_log = 'atualizar'
                mensagem = 'Deliberação de Pagamento atualizada com sucesso!'
            else:
                # CRIAR novo registro
                deliberacao = DeliberacaoPagamento(
                    NU_CONTRATO=contrato,
                    COLABORADOR_ANALISOU=current_user.nome,
                    DT_ANALISE=datetime.now().date(),
                    MATRICULA_CAIXA_EMGEA=matricula,
                    DT_ARREMATACAO_AQUISICAO=dt_arrematacao,
                    DT_ENTRADA_ESTOQUE=dt_entrada_estoque,
                    PERIODO_PRESCRITO=periodo_prescrito,
                    DT_PERIODO_COBRANCA_INICIO=dt_periodo_cobranca_inicio,
                    DT_PERIODO_COBRANCA_FIM=dt_periodo_cobranca_fim,
                    VR_DIVIDA_CONDOMINIO_1=vr_divida_inicial,
                    VR_DIVIDA_CONDOMINIO_2=vr_divida_condominio_2,
                    VR_DEBITO_EXCLUIDO_PRESCRITAS=vr_debito_excluido_prescritas,
                    VR_DEBITO_CALCULADO_EMGEA=vr_debito_calculado_emgea,
                    INDICE_DEBITO_EMGEA=indice_debito_emgea,
                    PERC_HONORARIOS_EMGEA=perc_honorarios_emgea,
                    VR_HONORARIOS_EMGEA=vr_honorarios,
                    DT_CALCULO_EMGEA=datetime.now().date(),
                    VR_AVALIACAO=vr_avaliacao,
                    DT_LAUDO=dt_laudo,
                    STATUS_IMOVEL=status_imovel,
                    VR_VENDA=vr_venda,
                    DT_VENDA=dt_venda,
                    NOME_COMPRADOR=nome_comprador,
                    DT_REGISTRO=dt_registro,
                    VR_DEBITOS_SISDEX=vr_debitos_sisdex,
                    VR_DEBITOS_SISGEA=vr_debitos_sisgea,
                    VR_DEBITOS_SISINC=vr_debitos_sisinc,  # NOVO
                    VR_DEBITOS_TOTAL=vr_debitos_total,
                    GRAVAME_MATRICULA=gravame_matricula,
                    ACOES_NEGOCIAIS_ADMINISTRATIVAS=acoes_negociais_adm,
                    NR_PROCESSOS_JUDICIAIS=nr_processos,
                    VARA_PROCESSO=vara_processo,
                    FASE_PROCESSO=fase_processo,
                    RELATORIO_ASSESSORIA_JURIDICA=relatorio_assessoria,
                    PENALIDADE_ANS_CAIXA=penalidade_ans,
                    PREJUIZO_FINANCEIRO_CAIXA=prejuizo_financeiro,
                    CONSIDERACOES_ANALISTA_GEADI=consideracoes_analista,
                    CONSIDERACOES_GESTOR_GEADI=consideracoes_gestor,
                    CONSIDERACOES_GESTOR_SUMOV=consideracoes_gestor_sumov,
                    TIPO_PAGAMENTO_VENDA=tipo_pagamento_venda,
                    STATUS_DOCUMENTO=status_documento,
                    USUARIO_GESTOR_GEADI=usuario_gestor_geadi,  # ← NOVO
                    USUARIO_GESTOR_SUMOV=usuario_gestor_sumov,  # ← NOVO
                    DT_DELIBERACAO_GEADI=dt_deliberacao_geadi,  # ← NOVO
                    DT_DELIBERACAO_SUMOV=dt_deliberacao_sumov,  # ← NOVO
                    USUARIO_DELIBEROU=usuario_deliberou,
                    USUARIO_CRIACAO=current_user.nome,
                    CREATED_AT=datetime.utcnow()
                )

                acao_log = 'criar'
                mensagem = 'Deliberação de Pagamento criada com sucesso!'

            # ===== SALVAR NO BANCO =====
            if deliberacao.salvar():
                registrar_log(
                    acao=acao_log,
                    entidade='deliberacao_pagamento',
                    entidade_id=contrato,
                    descricao=f'Deliberação de Pagamento para contrato {contrato}: {indice_debito_emgea}'
                )

                flash(mensagem, 'success')
                return redirect(url_for('sumov.deliberacao_pagamento'))
            else:
                flash('Erro ao salvar deliberação no banco de dados.', 'danger')
                return redirect(url_for('sumov.deliberacao_pagamento_nova'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao processar deliberação: {str(e)}', 'danger')
            import traceback
            traceback.print_exc()
            return redirect(url_for('sumov.deliberacao_pagamento_nova'))

    # GET - Carregar página do formulário
    return render_template('sumov/deliberacao_pagamento/nova.html')

@sumov_bp.route('/deliberacao-pagamento/buscar-dados-contrato', methods=['POST'])
@login_required
def buscar_dados_contrato():
    """Busca dados do contrato nas tabelas do banco incluindo SISINC"""
    try:
        data = request.get_json()
        contrato = data.get('contrato', '').strip()

        if not contrato:
            return jsonify({'success': False, 'message': 'Contrato não informado'})

        # ===== BUSCA 1: Data de Entrada no Estoque =====
        sql_estoque = text("""
            SELECT TOP 1 
                [DT_ENTRADA_ESTOQUE]
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB012_IMOVEIS_NAO_USO_ESTOQUE]
            WHERE [NR_CONTRATO] = :contrato
        """)
        result_estoque = db.session.execute(sql_estoque, {'contrato': contrato}).fetchone()
        dt_entrada_estoque = result_estoque[0].strftime('%Y-%m-%d') if result_estoque and result_estoque[0] else None

        # ===== BUSCA 2: Período de Prescrição =====
        sql_prescricao = text("""
            SELECT TOP 1
                [PERIODO_PRESCRICAO]
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB032_SISCALCULO_PRESCRICOES]
            WHERE [IMOVEL] = :contrato
            ORDER BY [DT_VENCIMENTO] DESC
        """)
        result_prescricao = db.session.execute(sql_prescricao, {'contrato': contrato}).fetchone()
        periodo_prescricao = result_prescricao[0] if result_prescricao else None

        # ===== BUSCA 2.5: Período da Cobrança dos Débitos (MIN e MAX DT_VENCIMENTO) =====
        sql_periodo_cobranca = text("""
            SELECT 
                MIN([DT_VENCIMENTO]) as DT_INICIO,
                MAX([DT_VENCIMENTO]) as DT_FIM
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB030_SISCALCULO_DADOS]
            WHERE [IMOVEL] = :contrato
            AND [DT_ATUALIZACAO] = (
                SELECT MAX([DT_ATUALIZACAO])
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB030_SISCALCULO_DADOS]
                WHERE [IMOVEL] = :contrato
            )
        """)
        result_periodo_cobranca = db.session.execute(sql_periodo_cobranca, {'contrato': contrato}).fetchone()

        dt_periodo_cobranca_inicio = None
        dt_periodo_cobranca_fim = None
        periodo_cobranca_encontrado = False

        if result_periodo_cobranca and result_periodo_cobranca[0] and result_periodo_cobranca[1]:
            dt_periodo_cobranca_inicio = result_periodo_cobranca[0].strftime('%Y-%m-%d')
            dt_periodo_cobranca_fim = result_periodo_cobranca[1].strftime('%Y-%m-%d')
            periodo_cobranca_encontrado = True

        # ===== BUSCA 3: Valor Inicial (Soma dos VR_COTA - DA ÚLTIMA ATUALIZAÇÃO) =====
        sql_valor_inicial = text("""
            SELECT 
                SUM([VR_COTA]) as VALOR_INICIAL,
                COUNT(*) as QTD_PARCELAS,
                MAX([NOME_CONDOMINIO]) as NOME_CONDOMINIO
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB030_SISCALCULO_DADOS]
            WHERE [IMOVEL] = :contrato
            AND [DT_ATUALIZACAO] = (
                SELECT MAX([DT_ATUALIZACAO])
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB030_SISCALCULO_DADOS]
                WHERE [IMOVEL] = :contrato
            )
        """)
        result_valor = db.session.execute(sql_valor_inicial, {'contrato': contrato}).fetchone()

        valor_inicial = float(result_valor[0]) if result_valor and result_valor[0] else 0
        qtd_parcelas_inicial = result_valor[1] if result_valor and result_valor[1] else 0
        nome_condominio = result_valor[2] if result_valor and result_valor[2] else None

        # ===== BUSCA 4: Valor de Débito EXCLUÍDOS as Cotas Prescritas =====
        sql_valor_prescrito = text("""
            SELECT 
                SUM([VR_COTA]) as VALOR_PRESCRITO,
                COUNT(*) as QTD_PRESCRITO
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB032_SISCALCULO_PRESCRICOES]
            WHERE [IMOVEL] = :contrato
            AND [ID_INDICE_ECONOMICO] = (
                SELECT MIN([ID_INDICE_ECONOMICO])
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB032_SISCALCULO_PRESCRICOES]
                WHERE [IMOVEL] = :contrato
                AND [DT_ATUALIZACAO] = (
                    SELECT MAX([DT_ATUALIZACAO])
                    FROM [BDDASHBOARDBI].[BDG].[MOV_TB032_SISCALCULO_PRESCRICOES]
                    WHERE [IMOVEL] = :contrato
                )
            )
            AND [DT_ATUALIZACAO] = (
                SELECT MAX([DT_ATUALIZACAO])
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB032_SISCALCULO_PRESCRICOES]
                WHERE [IMOVEL] = :contrato
            )
        """)
        result_prescrito = db.session.execute(sql_valor_prescrito, {'contrato': contrato}).fetchone()

        valor_prescrito = float(result_prescrito[0]) if result_prescrito and result_prescrito[0] else 0
        qtd_prescrito = result_prescrito[1] if result_prescrito and result_prescrito[1] else 0

        # ===== BUSCA 5: Índices disponíveis no SISCalculo COM TOTAIS POR TIPO =====
        sql_indices = text("""
            SELECT 
                C.ID_INDICE_ECONOMICO,
                SUM(C.VR_TOTAL - (C.VR_TOTAL * C.PERC_HONORARIOS / 100.0)) as VR_DIVIDA,
                MAX(C.PERC_HONORARIOS) as PERC_HONORARIOS,
                SUM(C.VR_TOTAL * C.PERC_HONORARIOS / 100.0) as VR_HONORARIOS,
                SUM(C.VR_TOTAL) as VR_TOTAL_COM_HONORARIOS
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB031_SISCALCULO_CALCULOS] C
            WHERE C.[IMOVEL] = :contrato
            AND C.[DT_ATUALIZACAO] = (
                SELECT MAX([DT_ATUALIZACAO])
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB031_SISCALCULO_CALCULOS]
                WHERE [IMOVEL] = :contrato
            )
            GROUP BY C.ID_INDICE_ECONOMICO
            ORDER BY C.ID_INDICE_ECONOMICO
        """)
        result_indices = db.session.execute(sql_indices, {'contrato': contrato}).fetchall()

        indices_disponiveis = []
        if result_indices:
            from app.models.siscalculo import ParamIndicesEconomicos, SiscalculoCalculos, TipoParcela

            for idx, row in enumerate(result_indices):
                id_indice = row[0]
                vr_total_sem_honorarios = float(row[1]) if row[1] else 0
                perc_honorarios = float(row[2]) if row[2] else 0
                vr_honorarios = float(row[3]) if row[3] else 0
                vr_total_com_honorarios = float(row[4]) if row[4] else 0

                indice_obj = ParamIndicesEconomicos.query.get(id_indice)
                nome_indice = indice_obj.DSC_INDICE_ECONOMICO if indice_obj else f"Índice {id_indice}"

                # Buscar totais por tipo de parcela para este índice específico
                totais_por_tipo = db.session.query(
                    SiscalculoCalculos.ID_TIPO,
                    TipoParcela.DSC_TIPO,
                    db.func.count(SiscalculoCalculos.DT_VENCIMENTO).label('quantidade'),
                    db.func.sum(SiscalculoCalculos.VR_TOTAL).label('valor_total')
                ).join(
                    TipoParcela,
                    SiscalculoCalculos.ID_TIPO == TipoParcela.ID_TIPO
                ).filter(
                    SiscalculoCalculos.IMOVEL == contrato,
                    SiscalculoCalculos.ID_INDICE_ECONOMICO == id_indice,
                    SiscalculoCalculos.DT_ATUALIZACAO == db.session.query(
                        db.func.max(SiscalculoCalculos.DT_ATUALIZACAO)
                    ).filter(
                        SiscalculoCalculos.IMOVEL == contrato
                    ).scalar_subquery()
                ).group_by(
                    SiscalculoCalculos.ID_TIPO,
                    TipoParcela.DSC_TIPO
                ).order_by(
                    SiscalculoCalculos.ID_TIPO
                ).all()

                # Converter para lista de dicionários
                tipos_parcela_lista = [
                    {
                        'id_tipo': t.ID_TIPO,
                        'descricao': t.DSC_TIPO,
                        'quantidade': t.quantidade,
                        'valor_total': float(t.valor_total) if t.valor_total else 0
                    }
                    for t in totais_por_tipo
                ]

                indices_disponiveis.append({
                    'id_indice': id_indice,
                    'nome_indice': nome_indice,
                    'valor_divida': vr_total_sem_honorarios,
                    'valor_honorarios': vr_honorarios,
                    'perc_honorarios': perc_honorarios,
                    'valor_com_honorarios': vr_total_com_honorarios,
                    'totais_por_tipo': tipos_parcela_lista
                })

        # ===== BUSCA 6: Valor de Avaliação e Data do Laudo =====
        sql_avaliacao = text("""
            SELECT TOP 1
                [VR_LAUDO_AVALIACAO],
                [DT_LAUDO]
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB001_IMOVEIS_NAO_USO_STATUS]
            WHERE [NR_CONTRATO] = :contrato
            ORDER BY [DT_REFERENCIA] DESC
        """)
        result_avaliacao = db.session.execute(sql_avaliacao, {'contrato': contrato}).fetchone()

        vr_avaliacao = float(result_avaliacao[0]) if result_avaliacao and result_avaliacao[0] else None
        dt_laudo = result_avaliacao[1].strftime('%Y-%m-%d') if result_avaliacao and result_avaliacao[1] else None

        # ===== BUSCA 7: Status do Imóvel =====
        sql_status = text("""
            SELECT TOP 1
                [DSC_SITUACAO]
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB020_IMOVEIS_RM_TOTVS]
            WHERE [IMOVEL] = :contrato
        """)
        result_status = db.session.execute(sql_status, {'contrato': contrato}).fetchone()
        status_imovel = result_status[0] if result_status else None

        # ===== BUSCA 8: Dados de Venda + TIPO DE PAGAMENTO =====
        sql_venda = text("""
            SELECT TOP 1
                [VR_VENDA],
                [DT_VENDA],
                [NO_COMPRADOR],
                [VR_SD_DEVEDOR]
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB023_VENDA_IMOVEIS_RM_TOTVS]
            WHERE [NU_IMOVEL] = :contrato
            ORDER BY [DT_VENDA] DESC
        """)
        result_venda = db.session.execute(sql_venda, {'contrato': contrato}).fetchone()

        vr_venda = float(result_venda[0]) if result_venda and result_venda[0] else None
        dt_venda = result_venda[1].strftime('%Y-%m-%d') if result_venda and result_venda[1] else None
        nome_comprador = result_venda[2] if result_venda and result_venda[2] else None

        # Determinar tipo de pagamento
        tipo_pagamento_venda = None
        if result_venda and result_venda[3] is not None:
            vr_sd_devedor = float(result_venda[3])
            if vr_sd_devedor > 0:
                tipo_pagamento_venda = "PARCELADO"
            else:
                tipo_pagamento_venda = "A_VISTA"

        # ===== BUSCA 9: Processos Judiciais =====
        sql_processos = text("""
            SELECT [nrProcessoCnj]
            FROM [BDEMGEAODS].[SISJUD].[tblProcessoCredito]
            WHERE [nrContrato] = :contrato
        """)
        result_processos = db.session.execute(sql_processos, {'contrato': contrato}).fetchall()

        if not result_processos:
            try:
                sql_processos_alt = text("""
                    SELECT [nrProcessoCnj]
                    FROM [BDEMGEAODS].[SISJUD].[tblProcessoCredito]
                    WHERE CAST([fkContratoSISCTR] AS VARCHAR(50)) = :contrato
                """)
                result_processos = db.session.execute(sql_processos_alt, {'contrato': contrato}).fetchall()
            except:
                result_processos = []

        processos_judiciais = []
        if result_processos:
            for processo in result_processos:
                nr_processo = processo[0]

                sql_detalhes = text("""
                    SELECT 
                        [dsComarca],
                        [dsLocal],
                        [dsFaseProcessual]
                    FROM [BDEMGEAODS].[SISJUD].[tblProcessoEmgea]
                    WHERE [nrProcessoCnj] = :nr_processo
                """)
                result_detalhes = db.session.execute(sql_detalhes, {'nr_processo': nr_processo}).fetchone()

                if result_detalhes:
                    comarca = result_detalhes[0] if result_detalhes[0] else ''
                    local = result_detalhes[1] if result_detalhes[1] else ''
                    fase = result_detalhes[2] if result_detalhes[2] else ''

                    vara = f"{comarca}-{local}" if comarca and local else comarca or local or 'Não informado'

                    processos_judiciais.append({
                        'numero': nr_processo,
                        'vara': vara,
                        'fase': fase
                    })

        # ===== BUSCA 10: Débitos Pagos - SISDEX =====
        sql_sisdex = text("""
            SELECT SUM([VR_PAGO]) as TOTAL_SISDEX
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB017_DESPESAS_MANUTENCAO]
            WHERE [NU_CONTRATO] = :contrato
        """)
        result_sisdex = db.session.execute(sql_sisdex, {'contrato': contrato}).fetchone()
        vr_sisdex = float(result_sisdex[0]) if result_sisdex and result_sisdex[0] else 0

        # ===== BUSCA 11: Débitos Pagos - SISGEA =====
        sql_sisgea = text("""
            SELECT SUM([VR_DESPESA]) as TOTAL_SISGEA
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB004_DESPESAS_ANALITICO]
            WHERE [NR_CONTRATO] = :contrato
        """)
        result_sisgea = db.session.execute(sql_sisgea, {'contrato': contrato}).fetchone()
        vr_sisgea = float(result_sisgea[0]) if result_sisgea and result_sisgea[0] else 0

        # ===== BUSCA 12: Débitos Pagos - SISINC (NOVO) =====
        sql_sisinc = text("""
            SELECT SUM([VR_LANCAMENTO]) as TOTAL_SISINC
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB014_DESPESAS_EXECUCAO_SISINC]
            WHERE [NU_CONTRATO] = :contrato
        """)
        result_sisinc = db.session.execute(sql_sisinc, {'contrato': contrato}).fetchone()
        vr_sisinc = float(result_sisinc[0]) if result_sisinc and result_sisinc[0] else 0

        # TOTAL agora inclui SISINC
        vr_total_debitos = vr_sisdex + vr_sisgea + vr_sisinc

        # ===== RETORNAR TODOS OS DADOS =====
        return jsonify({
            'success': True,
            'dt_entrada_estoque': dt_entrada_estoque,
            'periodo_prescricao': periodo_prescricao,
            'dt_periodo_cobranca_inicio': dt_periodo_cobranca_inicio,
            'dt_periodo_cobranca_fim': dt_periodo_cobranca_fim,
            'periodo_cobranca_encontrado': periodo_cobranca_encontrado,
            'valor_inicial': valor_inicial,
            'qtd_parcelas_inicial': qtd_parcelas_inicial,
            'nome_condominio': nome_condominio,
            'valor_prescrito': valor_prescrito,
            'qtd_parcelas_prescrito': qtd_prescrito,
            'indices_disponiveis': indices_disponiveis,
            'vr_avaliacao': vr_avaliacao,
            'dt_laudo': dt_laudo,
            'status_imovel': status_imovel,
            'vr_venda': vr_venda,
            'dt_venda': dt_venda,
            'nome_comprador': nome_comprador,
            'tipo_pagamento_venda': tipo_pagamento_venda,
            'processos_judiciais': processos_judiciais,
            'vr_debitos_sisdex': vr_sisdex,
            'vr_debitos_sisgea': vr_sisgea,
            'vr_debitos_sisinc': vr_sisinc,  # NOVO
            'vr_debitos_total': vr_total_debitos
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao buscar dados: {str(e)}'
        }), 500


@sumov_bp.route('/deliberacao-pagamento/editar/<contrato>', methods=['GET', 'POST'])
@login_required
def deliberacao_pagamento_editar(contrato):
    """Editar uma Deliberação de Pagamento existente"""
    try:
        deliberacao = DeliberacaoPagamento.buscar_por_contrato(contrato)

        if not deliberacao:
            flash('Deliberação não encontrada.', 'warning')
            return redirect(url_for('sumov.deliberacao_pagamento'))

        # ===== POST - SALVAR EDIÇÕES =====
        if request.method == 'POST':
            try:
                # Capturar campos editáveis
                consideracoes_analista = request.form.get('consideracoes_analista', '').strip() or None
                consideracoes_gestor_geadi = request.form.get('consideracoes_gestor', '').strip() or None
                consideracoes_gestor_sumov = request.form.get('consideracoes_gestor_sumov', '').strip() or None
                gravame_matricula = request.form.get('gravame_matricula', '').strip() or None
                acoes_negociais_adm = request.form.get('acoes_negociais_administrativas', '').strip() or None
                nr_processos = request.form.get('nr_processos_judiciais', '').strip() or None
                vara_processo = request.form.get('vara_processo', '').strip() or None
                fase_processo = request.form.get('fase_processo', '').strip() or None
                relatorio_assessoria = request.form.get('relatorio_assessoria_juridica', '').strip() or None
                penalidade_ans = request.form.get('penalidade_ans_caixa', '').strip() or None
                prejuizo_financeiro = request.form.get('prejuizo_financeiro_caixa', '').strip() or None

                # ✅ NOVA LÓGICA: Verificar SE MUDOU ANTES de atualizar

                # Verificar se Gestor GEADI preencheu/mudou considerações
                if consideracoes_gestor_geadi and consideracoes_gestor_geadi != deliberacao.CONSIDERACOES_GESTOR_GEADI:
                    deliberacao.USUARIO_GESTOR_GEADI = current_user.nome
                    deliberacao.DT_DELIBERACAO_GEADI = datetime.utcnow()

                # Verificar se Gestor SUMOV preencheu/mudou considerações
                if consideracoes_gestor_sumov and consideracoes_gestor_sumov != deliberacao.CONSIDERACOES_GESTOR_SUMOV:
                    deliberacao.USUARIO_GESTOR_SUMOV = current_user.nome
                    deliberacao.DT_DELIBERACAO_SUMOV = datetime.utcnow()

                # AGORA SIM atualizar os campos editáveis (DEPOIS de verificar)
                deliberacao.CONSIDERACOES_ANALISTA_GEADI = consideracoes_analista
                deliberacao.CONSIDERACOES_GESTOR_GEADI = consideracoes_gestor_geadi
                deliberacao.CONSIDERACOES_GESTOR_SUMOV = consideracoes_gestor_sumov
                deliberacao.GRAVAME_MATRICULA = gravame_matricula
                deliberacao.ACOES_NEGOCIAIS_ADMINISTRATIVAS = acoes_negociais_adm
                deliberacao.NR_PROCESSOS_JUDICIAIS = nr_processos
                deliberacao.VARA_PROCESSO = vara_processo
                deliberacao.FASE_PROCESSO = fase_processo
                deliberacao.RELATORIO_ASSESSORIA_JURIDICA = relatorio_assessoria
                deliberacao.PENALIDADE_ANS_CAIXA = penalidade_ans
                deliberacao.PREJUIZO_FINANCEIRO_CAIXA = prejuizo_financeiro

                # LÓGICA: Se alguma consideração de gestor foi preenchida
                if consideracoes_gestor_geadi or consideracoes_gestor_sumov:
                    # Atualizar status para DELIBERADO
                    deliberacao.STATUS_DOCUMENTO = 'DELIBERADO'

                    # Registrar quem deliberou (apenas se ainda não tiver sido registrado)
                    if not deliberacao.USUARIO_DELIBEROU:
                        deliberacao.USUARIO_DELIBEROU = current_user.nome

                # Atualizar dados de auditoria
                deliberacao.USUARIO_ATUALIZACAO = current_user.nome
                deliberacao.UPDATED_AT = datetime.utcnow()

                if deliberacao.salvar():
                    registrar_log(
                        acao='editar',
                        entidade='deliberacao_pagamento',
                        entidade_id=contrato,
                        descricao=f'Deliberação de Pagamento editada: {contrato}'
                    )
                    flash('Deliberação atualizada com sucesso!', 'success')
                else:
                    flash('Erro ao atualizar deliberação.', 'danger')

                return redirect(url_for('sumov.deliberacao_pagamento'))

            except Exception as e:
                db.session.rollback()
                flash(f'Erro ao atualizar deliberação: {str(e)}', 'danger')
                import traceback
                traceback.print_exc()
                return redirect(url_for('sumov.deliberacao_pagamento'))

        # ===== GET - MOSTRAR FORMULÁRIO DE EDIÇÃO =====
        # Buscar totais por tipo de parcela se existir índice
        totais_por_tipo = []

        try:
            from app.models.siscalculo import SiscalculoCalculos, TipoParcela, ParamIndicesEconomicos

            indice_usado = deliberacao.INDICE_DEBITO_EMGEA

            if indice_usado:
                # Buscar pelo nome do índice
                indice_obj = ParamIndicesEconomicos.query.filter(
                    ParamIndicesEconomicos.DSC_INDICE_ECONOMICO.contains(indice_usado.split(' - ')[0])
                ).first()

                if indice_obj:
                    id_indice = indice_obj.ID_INDICE_ECONOMICO

                    # Query para buscar totais por tipo
                    totais_query = db.session.query(
                        SiscalculoCalculos.ID_TIPO,
                        TipoParcela.DSC_TIPO,
                        db.func.count(SiscalculoCalculos.DT_VENCIMENTO).label('quantidade'),
                        db.func.sum(SiscalculoCalculos.VR_TOTAL).label('valor_total')
                    ).join(
                        TipoParcela,
                        SiscalculoCalculos.ID_TIPO == TipoParcela.ID_TIPO
                    ).filter(
                        SiscalculoCalculos.IMOVEL == contrato,
                        SiscalculoCalculos.ID_INDICE_ECONOMICO == id_indice,
                        SiscalculoCalculos.DT_ATUALIZACAO == db.session.query(
                            db.func.max(SiscalculoCalculos.DT_ATUALIZACAO)
                        ).filter(
                            SiscalculoCalculos.IMOVEL == contrato
                        ).scalar_subquery()
                    ).group_by(
                        SiscalculoCalculos.ID_TIPO,
                        TipoParcela.DSC_TIPO
                    ).order_by(
                        SiscalculoCalculos.ID_TIPO
                    ).all()

                    # Converter para lista de dicionários
                    totais_por_tipo = [
                        {
                            'id_tipo': t.ID_TIPO,
                            'tipo': t.DSC_TIPO,
                            'descricao': t.DSC_TIPO,
                            'quantidade': t.quantidade,
                            'valor_total': float(t.valor_total) if t.valor_total else 0
                        }
                        for t in totais_query
                    ]

                    print(f"[DEBUG EDIÇÃO] Encontrados {len(totais_por_tipo)} tipos de parcela")
                    for t in totais_por_tipo:
                        print(f"  Tipo {t['id_tipo']}: {t['descricao']} - {t['quantidade']} parcelas")

        except Exception as e:
            # Se der erro ao buscar tipos de parcela, apenas não mostra (não quebra a página)
            print(f"Aviso: Não foi possível buscar tipos de parcela: {e}")
            import traceback
            traceback.print_exc()
            totais_por_tipo = []

        return render_template(
            'sumov/deliberacao_pagamento/editar.html',
            deliberacao=deliberacao,
            totais_por_tipo=totais_por_tipo
        )

    except Exception as e:
        flash(f'Erro ao carregar deliberação: {str(e)}', 'danger')
        return redirect(url_for('sumov.deliberacao_pagamento'))


@sumov_bp.route('/deliberacao-pagamento/excluir/<contrato>', methods=['POST'])
@login_required
def deliberacao_pagamento_excluir(contrato):
    """Excluir (soft delete) uma Deliberação de Pagamento"""
    try:
        deliberacao = DeliberacaoPagamento.buscar_por_contrato(contrato)

        if not deliberacao:
            flash('Deliberação não encontrada.', 'warning')
            return redirect(url_for('sumov.deliberacao_pagamento'))

        # Soft delete
        deliberacao.DELETED_AT = datetime.utcnow()
        deliberacao.USUARIO_ATUALIZACAO = current_user.nome
        deliberacao.UPDATED_AT = datetime.utcnow()

        if deliberacao.salvar():
            registrar_log(
                acao='excluir',
                entidade='deliberacao_pagamento',
                entidade_id=contrato,
                descricao=f'Deliberação de Pagamento excluída: {contrato}'
            )
            flash('Deliberação excluída com sucesso!', 'success')
        else:
            flash('Erro ao excluir deliberação.', 'danger')

        return redirect(url_for('sumov.deliberacao_pagamento'))

    except Exception as e:
        flash(f'Erro ao excluir deliberação: {str(e)}', 'danger')
        return redirect(url_for('sumov.deliberacao_pagamento'))


@sumov_bp.route('/deliberacao-pagamento/pdf/<contrato>')
@login_required
def deliberacao_pagamento_pdf(contrato):
    """Gerar PDF da Deliberação de Pagamento"""
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
        from reportlab.pdfgen import canvas
        from datetime import datetime
        from decimal import Decimal
        import html

        # Buscar deliberação
        deliberacao = DeliberacaoPagamento.buscar_por_contrato(contrato)

        if not deliberacao:
            flash('Deliberação não encontrada.', 'warning')
            return redirect(url_for('sumov.deliberacao_pagamento'))

        # Criar buffer
        buffer = BytesIO()

        # Criar documento
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2.5 * cm,
            bottomMargin=2 * cm
        )

        # Estilos
        styles = getSampleStyleSheet()

        style_titulo = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a5490'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        style_subtitulo = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#1a5490'),
            spaceAfter=12,
            spaceBefore=15,
            fontName='Helvetica-Bold'
        )

        style_normal = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_JUSTIFY,
            spaceAfter=8,
            fontName='Helvetica'
        )

        style_label = ParagraphStyle(
            'CustomLabel',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#666666'),
            spaceAfter=2,
            fontName='Helvetica-Bold'
        )

        style_valor = ParagraphStyle(
            'CustomValor',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=10,
            fontName='Helvetica'
        )

        # Helper functions
        def escape_text(text):
            if not text:
                return '-'
            return html.escape(str(text))

        def formatar_data(data):
            if not data:
                return '-'
            if isinstance(data, str):
                return data
            return data.strftime('%d/%m/%Y')

        def formatar_moeda(valor):
            if not valor:
                return 'R$ 0,00'
            return 'R$ {:,.2f}'.format(float(valor)).replace(',', 'X').replace('.', ',').replace('X', '.')

        def formatar_percentual(valor):
            if not valor:
                return '0,00%'
            return '{:,.2f}%'.format(float(valor)).replace('.', ',')

        # Elementos do PDF
        elements = []

        # ===== CABEÇALHO =====
        elements.append(Paragraph('DELIBERAÇÃO DE PAGAMENTO', style_titulo))
        elements.append(Paragraph(f'Contrato/Imóvel: {escape_text(deliberacao.NU_CONTRATO)}', style_normal))
        elements.append(Spacer(1, 0.5 * cm))

        # ===== SEÇÃO 1: IDENTIFICAÇÃO =====
        elements.append(Paragraph('1. IDENTIFICAÇÃO', style_subtitulo))

        dados_identificacao = [
            ['Colaborador que Analisou:', escape_text(deliberacao.COLABORADOR_ANALISOU)],
            ['Data da Análise:', formatar_data(deliberacao.DT_ANALISE)],
            ['Matrícula (Caixa/Emgea):', escape_text(deliberacao.MATRICULA_CAIXA_EMGEA)],
            ['Data da Arrematação/Aquisição:', formatar_data(deliberacao.DT_ARREMATACAO_AQUISICAO)],
            ['Data de Entrada no Estoque:', formatar_data(deliberacao.DT_ENTRADA_ESTOQUE)]
        ]

        tabela_identificacao = Table(dados_identificacao, colWidths=[7 * cm, 10 * cm])
        tabela_identificacao.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4F8')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1a5490')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(tabela_identificacao)
        elements.append(Spacer(1, 0.5 * cm))

        # ===== SEÇÃO 2: COBRANÇA E DÍVIDA =====
        elements.append(Paragraph('2. COBRANÇA E DÍVIDA', style_subtitulo))

        dados_cobranca = [
            ['Período Prescrito:', escape_text(deliberacao.PERIODO_PRESCRITO)],
        ]

        # NOVO: Adicionar período de cobrança
        if deliberacao.DT_PERIODO_COBRANCA_INICIO and deliberacao.DT_PERIODO_COBRANCA_FIM:
            periodo_cobranca_texto = f"{formatar_data(deliberacao.DT_PERIODO_COBRANCA_INICIO)} até {formatar_data(deliberacao.DT_PERIODO_COBRANCA_FIM)}"
            dados_cobranca.append(['Período da Cobrança dos Débitos:', periodo_cobranca_texto])

        dados_cobranca.extend([
            ['Valor Inicial da Dívida:', formatar_moeda(deliberacao.VR_DIVIDA_CONDOMINIO_1)],
        ])

        # NOVO: Adicionar valor manual se existir
        if deliberacao.VR_DIVIDA_CONDOMINIO_2:
            dados_cobranca.append(['Valor da Dívida (Manual):', formatar_moeda(deliberacao.VR_DIVIDA_CONDOMINIO_2)])

        dados_cobranca.extend([
            ['Valor Excluídos Cotas Prescritas:', formatar_moeda(deliberacao.VR_DEBITO_EXCLUIDO_PRESCRITAS)],
            ['Índice Econômico:', escape_text(deliberacao.INDICE_DEBITO_EMGEA)],
            ['Percentual de Honorários:', formatar_percentual(deliberacao.PERC_HONORARIOS_EMGEA)],
            ['Valor dos Honorários:', formatar_moeda(deliberacao.VR_HONORARIOS_EMGEA)],
            ['VALOR TOTAL (com Honorários):', formatar_moeda(deliberacao.VR_DEBITO_CALCULADO_EMGEA)]
        ])

        tabela_cobranca = Table(dados_cobranca, colWidths=[7 * cm, 10 * cm])
        tabela_cobranca.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4F8')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1a5490')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D4EDDA')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(tabela_cobranca)
        elements.append(Spacer(1, 0.5 * cm))

        # ✅ NOVO: Adicionar tabela de tipos de parcela (se houver dados)
        # Busca o índice diretamente do SISCalculo
        try:
            from app.models.siscalculo import SiscalculoCalculos, TipoParcela

            # Primeiro, buscar qual índice foi usado para este contrato
            indice_usado = db.session.query(
                SiscalculoCalculos.ID_INDICE_ECONOMICO
            ).filter(
                SiscalculoCalculos.IMOVEL == contrato,
                SiscalculoCalculos.DT_ATUALIZACAO == db.session.query(
                    db.func.max(SiscalculoCalculos.DT_ATUALIZACAO)
                ).filter(
                    SiscalculoCalculos.IMOVEL == contrato
                ).scalar_subquery()
            ).first()

            # Se encontrou o índice, buscar os totais por tipo
            if indice_usado:
                id_indice = indice_usado[0]

                totais_query = db.session.query(
                    SiscalculoCalculos.ID_TIPO,
                    TipoParcela.DSC_TIPO,
                    db.func.count(SiscalculoCalculos.DT_VENCIMENTO).label('quantidade'),
                    db.func.sum(SiscalculoCalculos.VR_TOTAL).label('valor_total')
                ).join(
                    TipoParcela,
                    SiscalculoCalculos.ID_TIPO == TipoParcela.ID_TIPO
                ).filter(
                    SiscalculoCalculos.IMOVEL == contrato,
                    SiscalculoCalculos.ID_INDICE_ECONOMICO == id_indice,
                    SiscalculoCalculos.DT_ATUALIZACAO == db.session.query(
                        db.func.max(SiscalculoCalculos.DT_ATUALIZACAO)
                    ).filter(
                        SiscalculoCalculos.IMOVEL == contrato
                    ).scalar_subquery()
                ).group_by(
                    SiscalculoCalculos.ID_TIPO,
                    TipoParcela.DSC_TIPO
                ).order_by(
                    SiscalculoCalculos.ID_TIPO
                ).all()

                if totais_query:
                    # Adicionar subtítulo
                    elements.append(Spacer(1, 0.3 * cm))
                    elements.append(Paragraph('2.1 RESUMO POR TIPO DE PARCELA', style_subtitulo))

                    # Criar dados da tabela
                    dados_tipos = [['ID', 'Descrição do Tipo', 'Quantidade', 'Valor Total']]

                    total_qtd = 0
                    total_valor = Decimal('0')

                    for t in totais_query:
                        valor = Decimal(str(t.valor_total)) if t.valor_total else Decimal('0')
                        total_qtd += t.quantidade
                        total_valor += valor

                        dados_tipos.append([
                            str(t.ID_TIPO),
                            escape_text(t.DSC_TIPO),
                            str(t.quantidade),
                            formatar_moeda(valor)
                        ])

                    # Adicionar linha de totais
                    dados_tipos.append([
                        '',
                        'TOTAL GERAL:',
                        str(total_qtd),
                        formatar_moeda(total_valor)
                    ])

                    # Criar tabela
                    tabela_tipos = Table(dados_tipos, colWidths=[2 * cm, 8 * cm, 3 * cm, 4 * cm])
                    tabela_tipos.setStyle(TableStyle([
                        # Cabeçalho
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),

                        # Corpo da tabela
                        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
                        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
                        ('ALIGN', (0, 1), (0, -2), 'CENTER'),  # ID centralizado
                        ('ALIGN', (1, 1), (1, -2), 'LEFT'),  # Descrição à esquerda
                        ('ALIGN', (2, 1), (2, -2), 'CENTER'),  # Quantidade centralizada
                        ('ALIGN', (3, 1), (3, -2), 'RIGHT'),  # Valor à direita
                        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -2), 9),

                        # Linha de totais (última linha)
                        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D4EDDA')),
                        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, -1), (-1, -1), 10),
                        ('ALIGN', (1, -1), (1, -1), 'RIGHT'),
                        ('ALIGN', (2, -1), (2, -1), 'CENTER'),
                        ('ALIGN', (3, -1), (3, -1), 'RIGHT'),

                        # Bordas e padding
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 6),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                        ('TOPPADDING', (0, 0), (-1, -1), 6),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ]))

                    elements.append(tabela_tipos)
                    elements.append(Spacer(1, 0.5 * cm))
        except Exception as e:
            # Se der erro ao buscar tipos de parcela, apenas não adiciona no PDF (não quebra)
            print(f"Aviso: Não foi possível adicionar tipos de parcela ao PDF: {e}")

        # ===== SEÇÃO 3: AVALIAÇÃO E VENDA =====
        elements.append(Paragraph('3. AVALIAÇÃO E VENDA', style_subtitulo))

        dados_avaliacao = [
            ['Valor de Avaliação:', formatar_moeda(deliberacao.VR_AVALIACAO)],
            ['Data do Laudo:', formatar_data(deliberacao.DT_LAUDO)],
            ['Status do Imóvel:', escape_text(deliberacao.STATUS_IMOVEL)],
        ]

        if deliberacao.STATUS_IMOVEL and 'VENDIDO' in deliberacao.STATUS_IMOVEL.upper():
            dados_avaliacao.extend([
                ['Valor de Venda:', formatar_moeda(deliberacao.VR_VENDA)],
                ['Data da Venda:', formatar_data(deliberacao.DT_VENDA)],
                ['Nome do Adquirente:', escape_text(deliberacao.NOME_COMPRADOR)],
            ])

            tipo_pagamento_texto = '-'
            if deliberacao.TIPO_PAGAMENTO_VENDA:
                if deliberacao.TIPO_PAGAMENTO_VENDA == 'A_VISTA':
                    tipo_pagamento_texto = 'À Vista'
                elif deliberacao.TIPO_PAGAMENTO_VENDA == 'PARCELADO':
                    tipo_pagamento_texto = 'Parcelado'

            dados_avaliacao.append(['Tipo de Pagamento:', tipo_pagamento_texto])

        dados_avaliacao.append(['Data do Registro:', formatar_data(deliberacao.DT_REGISTRO)])

        tabela_avaliacao = Table(dados_avaliacao, colWidths=[7 * cm, 10 * cm])
        tabela_avaliacao.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4F8')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1a5490')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(tabela_avaliacao)
        elements.append(Spacer(1, 0.5 * cm))

        # ===== SEÇÃO 4: AÇÕES E PROCESSOS JUDICIAIS =====
        elements.append(Paragraph('4. AÇÕES E PROCESSOS JUDICIAIS', style_subtitulo))

        if deliberacao.GRAVAME_MATRICULA:
            elements.append(Paragraph('<b>Gravame na Matrícula:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.GRAVAME_MATRICULA), style_valor))

        if deliberacao.ACOES_NEGOCIAIS_ADMINISTRATIVAS:
            elements.append(Paragraph('<b>Ações Negociais Administrativas:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.ACOES_NEGOCIAIS_ADMINISTRATIVAS), style_valor))

        if deliberacao.NR_PROCESSOS_JUDICIAIS:
            elements.append(Paragraph('<b>Nº da Ação Judicial:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.NR_PROCESSOS_JUDICIAIS), style_valor))

        if deliberacao.VARA_PROCESSO:
            elements.append(Paragraph('<b>Vara do Processo:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.VARA_PROCESSO), style_valor))

        if deliberacao.FASE_PROCESSO:
            elements.append(Paragraph('<b>Fase do Processo:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.FASE_PROCESSO), style_valor))

        if deliberacao.RELATORIO_ASSESSORIA_JURIDICA:
            elements.append(Paragraph('<b>Relatório da Assessoria Jurídica:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.RELATORIO_ASSESSORIA_JURIDICA), style_valor))

        elements.append(Spacer(1, 0.3 * cm))

        # ===== SEÇÃO 5: DÉBITOS E PENALIDADES =====
        elements.append(Paragraph('5. DÉBITOS E PENALIDADES', style_subtitulo))

        dados_debitos = [
            ['Débitos Sisdex:', formatar_moeda(deliberacao.VR_DEBITOS_SISDEX)],
            ['Débitos Sisgea:', formatar_moeda(deliberacao.VR_DEBITOS_SISGEA)],
            ['TOTAL de Débitos:', formatar_moeda(deliberacao.VR_DEBITOS_TOTAL)]
        ]

        tabela_debitos = Table(dados_debitos, colWidths=[7 * cm, 10 * cm])
        tabela_debitos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4F8')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1a5490')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF3CD')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(tabela_debitos)
        elements.append(Spacer(1, 0.3 * cm))

        if deliberacao.PENALIDADE_ANS_CAIXA:
            elements.append(Paragraph('<b>Penalidade de ANS - CAIXA:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.PENALIDADE_ANS_CAIXA), style_valor))

        if deliberacao.PREJUIZO_FINANCEIRO_CAIXA:
            elements.append(Paragraph('<b>Prejuízo Financeiro - CAIXA:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.PREJUIZO_FINANCEIRO_CAIXA), style_valor))

        elements.append(Spacer(1, 0.3 * cm))

        # ===== SEÇÃO 6: CONSIDERAÇÕES FINAIS =====
        elements.append(Paragraph('6. CONSIDERAÇÕES FINAIS', style_subtitulo))

        if deliberacao.CONSIDERACOES_ANALISTA_GEADI:
            elements.append(Paragraph('<b>Considerações da Analista GEADI:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.CONSIDERACOES_ANALISTA_GEADI), style_valor))
            elements.append(Spacer(1, 0.3 * cm))

        if deliberacao.CONSIDERACOES_GESTOR_GEADI:
            elements.append(Paragraph('<b>Considerações Finais do Gestor da GEADI:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.CONSIDERACOES_GESTOR_GEADI), style_valor))
            elements.append(Spacer(1, 0.3 * cm))

        if deliberacao.CONSIDERACOES_GESTOR_SUMOV:
            elements.append(Paragraph('<b>Considerações Finais do Gestor da SUMOV:</b>', style_label))
            elements.append(Paragraph(escape_text(deliberacao.CONSIDERACOES_GESTOR_SUMOV), style_valor))
            elements.append(Spacer(1, 0.3 * cm))

        # ===== RODAPÉ =====
        elements.append(Spacer(1, 1 * cm))
        elements.append(Paragraph(
            f'<i>Documento gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M")}</i>',
            ParagraphStyle('Rodape', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        ))

        # Construir PDF
        doc.build(elements)

        # Retornar PDF
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'Deliberacao_Pagamento_{contrato}.pdf'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Erro ao gerar PDF: {str(e)}', 'danger')
        return redirect(url_for('sumov.deliberacao_pagamento'))


# ==================== DIFERENÇA DESPESAS X SISCOR ====================

from app.models.diferenca_despesas_siscor import DiferencaDespesasSiscor


@sumov_bp.route('/despesas-pagamentos/diferenca-siscor')
@login_required
def diferenca_despesas_siscor():
    """Página de análise de diferenças entre Despesas e SISCOR"""

    # Buscar datas já formatadas em uma única operação
    datas_formatadas = DiferencaDespesasSiscor.listar_datas_disponiveis_formatadas()

    return render_template('sumov/despesas_pagamentos/diferenca_siscor.html',
                           datas_disponiveis=datas_formatadas)


@sumov_bp.route('/despesas-pagamentos/diferenca-siscor/buscar', methods=['POST'])
@login_required
def buscar_diferenca_siscor():
    """API para buscar diferenças com filtros"""
    try:
        data = request.get_json()

        # Obter filtros do request
        datas_selecionadas = data.get('datas', [])
        id_item = data.get('id_item', None)

        # Converter id_item para int se fornecido
        if id_item:
            try:
                id_item = int(id_item)
            except (ValueError, TypeError):
                id_item = None

        # Buscar registros com filtros
        registros = DiferencaDespesasSiscor.buscar_por_filtros(
            datas=datas_selecionadas if len(datas_selecionadas) > 0 else None,
            id_item=id_item
        )

        # Calcular totais
        totais = DiferencaDespesasSiscor.calcular_totais(registros)

        # Formatar registros para JSON
        registros_formatados = []
        for r in registros:
            registros_formatados.append({
                'dt_despesa': r.DT_DESPESA,
                'dt_despesa_formatada': DiferencaDespesasSiscor.formatar_data_mesano(r.DT_DESPESA),
                'id_item': r.ID_ITEM,
                'vr_siscor': float(r.VR_SISCOR) if r.VR_SISCOR else 0,
                'vr_despesa': float(r.VR_DESPESA) if r.VR_DESPESA else 0,
                'dif': float(r.DIF) if r.DIF else 0
            })

        # Retornar dados
        return jsonify({
            'success': True,
            'registros': registros_formatados,
            'totais': totais,
            'quantidade': len(registros_formatados)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'erro': str(e)
        }), 500


# =============================================================================
# ROTAS: PENALIDADES ANS
# =============================================================================

@sumov_bp.route('/penalidade-ans')
@login_required
def penalidade_ans_index():
    """
    Página inicial de Penalidades ANS
    Lista todas as deliberações salvas
    """
    try:
        from app.models.deliberacao_ans import DeliberacaoANS
        from app.models.penalidade_ans import PenalidadeANS

        # Buscar todas as deliberações salvas
        deliberacoes = DeliberacaoANS.query.order_by(
            DeliberacaoANS.CREATED_AT.desc()
        ).all()

        # Contar contratos ANS cadastrados (tabela de referência)
        total_contratos_ans = PenalidadeANS.query.count()

        return render_template('sumov/penalidade_ans/index.html',
                               deliberacoes=deliberacoes,
                               total_contratos_ans=total_contratos_ans)

    except Exception as e:
        flash(f'Erro ao carregar penalidades ANS: {str(e)}', 'danger')
        import traceback
        traceback.print_exc()
        return redirect(url_for('sumov.index'))


@sumov_bp.route('/penalidade-ans/nova', methods=['GET', 'POST'])
@login_required
def penalidade_ans_nova():
    """
    Página para cadastrar nova penalidade ANS
    """
    if request.method == 'POST':
        try:
            from app.models.penalidade_ans import PenalidadeANS
            from app.utils.audit import registrar_log
            from decimal import Decimal

            # Capturar dados do formulário
            nu_contrato = request.form.get('nu_contrato', '').strip()
            ini_vigencia_str = request.form.get('ini_vigencia', '').strip()
            fim_vigencia_str = request.form.get('fim_vigencia', '').strip()
            vr_tarifa_str = request.form.get('vr_tarifa', '').strip()
            prazo_dias_str = request.form.get('prazo_dias', '').strip()

            # Validações
            if not nu_contrato or not ini_vigencia_str or not fim_vigencia_str or not vr_tarifa_str:
                flash('Preencha todos os campos obrigatórios (Contrato, Datas de Vigência e Valor da Tarifa).',
                      'warning')
                return redirect(url_for('sumov.penalidade_ans_nova'))

            # Converter datas
            try:
                ini_vigencia = datetime.strptime(ini_vigencia_str, '%Y-%m-%d').date()
                fim_vigencia = datetime.strptime(fim_vigencia_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Datas de vigência inválidas.', 'danger')
                return redirect(url_for('sumov.penalidade_ans_nova'))

            # Validar se data fim é maior que data início
            if fim_vigencia <= ini_vigencia:
                flash('A data de fim da vigência deve ser posterior à data de início.', 'danger')
                return redirect(url_for('sumov.penalidade_ans_nova'))

            try:
                vr_tarifa = Decimal(vr_tarifa_str.replace('.', '').replace(',', '.'))
            except:
                flash('Valor da tarifa inválido.', 'danger')
                return redirect(url_for('sumov.penalidade_ans_nova'))

            prazo_dias = None
            if prazo_dias_str:
                try:
                    prazo_dias = int(prazo_dias_str)
                except ValueError:
                    flash('Prazo em dias inválido.', 'danger')
                    return redirect(url_for('sumov.penalidade_ans_nova'))

            # Verificar se já existe penalidade com mesmas chaves
            existe = PenalidadeANS.query.filter_by(
                NU_CONTRATO=nu_contrato,
                INI_VIGENCIA=ini_vigencia,
                FIM_VIGENCIA=fim_vigencia
            ).first()

            if existe:
                flash('Já existe uma penalidade cadastrada para este contrato com as mesmas datas de vigência.',
                      'warning')
                return redirect(url_for('sumov.penalidade_ans_nova'))

            # Criar nova penalidade (SEM campos de auditoria)
            penalidade = PenalidadeANS(
                NU_CONTRATO=nu_contrato,
                INI_VIGENCIA=ini_vigencia,
                FIM_VIGENCIA=fim_vigencia,
                VR_TARIFA=vr_tarifa,
                PRAZO_DIAS=prazo_dias
            )

            if penalidade.salvar():
                # Registrar log
                registrar_log(
                    acao='criar',
                    entidade='penalidade_ans',
                    entidade_id=f"{nu_contrato}_{ini_vigencia_str}_{fim_vigencia_str}",
                    descricao=f'Cadastro de Penalidade ANS - Contrato {nu_contrato}',
                    dados_novos={
                        'nu_contrato': nu_contrato,
                        'ini_vigencia': ini_vigencia_str,
                        'fim_vigencia': fim_vigencia_str,
                        'vr_tarifa': str(vr_tarifa),
                        'prazo_dias': prazo_dias
                    }
                )

                flash('Penalidade ANS cadastrada com sucesso!', 'success')
                return redirect(url_for('sumov.penalidade_ans_index'))
            else:
                flash('Erro ao salvar penalidade ANS no banco de dados.', 'danger')
                return redirect(url_for('sumov.penalidade_ans_nova'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao processar penalidade ANS: {str(e)}', 'danger')
            import traceback
            traceback.print_exc()
            return redirect(url_for('sumov.penalidade_ans_nova'))

    # GET - Exibir formulário
    return render_template('sumov/penalidade_ans/nova.html')


def verificar_tipo_despesa_contrato(nu_contrato):
    """
    Verifica se o contrato possui despesas de IPTU ou Condomínio
    PEGA O REGISTRO COM A MAIOR DT_LANCAMENTO_PAGAMENTO

    Retorna:
        'IPTU' - se tiver despesas de IPTU (prazo 120 dias)
        'Condomínio' - se tiver apenas despesas de Condomínio (prazo 150 dias)
        'Ambos' - se tiver os dois tipos
        None - se não encontrar despesas
    """
    try:
        print(f"[DEBUG] ===== VERIFICANDO TIPO DE DESPESA DO CONTRATO {nu_contrato} =====")

        # Buscar despesa mais recente (maior DT_LANCAMENTO_PAGAMENTO)
        sql_despesa_recente = text("""
            SELECT TOP 1
                D3.TIPO_DESPESA
            FROM [BDDASHBOARDBI].[BDG].[MOV_TB004_DESPESAS_ANALITICO] D4
            INNER JOIN [BDDASHBOARDBI].[BDG].[MOV_TB003_DESPESAS] D3
                ON D4.DSC_ITEM_SERVICO = D3.DSC_ITEM_SERVICO
            WHERE D4.NR_CONTRATO = :contrato
                AND D3.TIPO_DESPESA IN ('IPTU', 'Condomínio')
                AND D4.DT_LANCAMENTO_PAGAMENTO IS NOT NULL
            ORDER BY D4.DT_LANCAMENTO_PAGAMENTO DESC
        """)

        result = db.session.execute(sql_despesa_recente, {'contrato': nu_contrato}).fetchone()

        if not result:
            print(f"[DEBUG] ⚠️ NENHUMA DESPESA ENCONTRADA! Retornando None (usará 150 dias)")
            return None

        tipo_despesa = result[0]
        print(f"[DEBUG] ✓ Tipo de despesa mais recente: {tipo_despesa}")

        if tipo_despesa == 'IPTU':
            print(f"[DEBUG] ✓ É IPTU - Retornando 'IPTU'")
            return 'IPTU'
        elif tipo_despesa == 'Condomínio':
            print(f"[DEBUG] ✓ É Condomínio - Retornando 'Condomínio'")
            return 'Condomínio'
        else:
            print(f"[DEBUG] ⚠️ Tipo não identificado: {tipo_despesa} - Retornando None")
            return None

    except Exception as e:
        print(f"[ERRO] Erro ao verificar tipo de despesa: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def obter_prazo_contrato_03_2014(nu_contrato):
    """
    Retorna o prazo correto para o Contrato 03/2014 baseado no tipo de despesa

    Regras:
    - IPTU: 120 dias
    - Condomínio: 150 dias
    - Não encontrado: 150 dias (padrão)
    """
    print(f"\n[DEBUG] ===== OBTENDO PRAZO PARA CONTRATO 03/2014 =====")

    tipo_despesa = verificar_tipo_despesa_contrato(nu_contrato)

    if tipo_despesa == 'IPTU':
        print(f"[INFO] ✓ Contrato {nu_contrato} possui despesas de IPTU → Prazo: 120 dias\n")
        return 120
    else:
        # Condomínio ou não encontrado: usa 150 dias (padrão)
        print(f"[INFO] ⚠️ Contrato {nu_contrato} possui Condomínio ou não encontrado → Prazo: 150 dias (padrão)\n")
        return 150

@sumov_bp.route('/penalidade-ans/consultar', methods=['GET', 'POST'])
@login_required
def penalidade_ans_consultar():
    """
    Página para consultar e calcular penalidades ANS de um contrato
    Esta é apenas para VISUALIZAÇÃO dos cálculos
    """
    if request.method == 'POST':
        try:
            from app.models.penalidade_ans import PenalidadeANS
            from dateutil.relativedelta import relativedelta

            # Capturar número do contrato
            nu_contrato = request.form.get('nu_contrato', '').strip()

            if not nu_contrato:
                flash('Informe o número do contrato.', 'warning')
                return redirect(url_for('sumov.penalidade_ans_consultar'))

            # ===== BUSCAR DATA DE ENTRADA NO ESTOQUE =====
            sql_estoque = text("""
                SELECT TOP 1 
                    [DT_ENTRADA_ESTOQUE]
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB012_IMOVEIS_NAO_USO_ESTOQUE]
                WHERE [NR_CONTRATO] = :contrato
            """)
            result_estoque = db.session.execute(sql_estoque, {'contrato': nu_contrato}).fetchone()

            if not result_estoque or not result_estoque[0]:
                flash('Data de entrada no estoque não encontrada para este contrato.', 'warning')
                return redirect(url_for('sumov.penalidade_ans_consultar'))

            dt_entrada_estoque = result_estoque[0]

            # ===== BUSCAR DADOS DE VENDA =====
            sql_venda = text("""
                SELECT TOP 1
                    [VR_VENDA],
                    [DT_VENDA]
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB023_VENDA_IMOVEIS_RM_TOTVS]
                WHERE [NU_IMOVEL] = :contrato
                ORDER BY [DT_VENDA] DESC
            """)
            result_venda = db.session.execute(sql_venda, {'contrato': nu_contrato}).fetchone()

            vr_venda = float(result_venda[0]) if result_venda and result_venda[0] else None
            dt_venda = result_venda[1] if result_venda and result_venda[1] else None

            # ===== BUSCAR VALOR DE AVALIAÇÃO =====
            sql_avaliacao = text("""
                SELECT TOP 1
                    [VR_LAUDO_AVALIACAO],
                    [DT_LAUDO]
                FROM [BDDASHBOARDBI].[BDG].[MOV_TB001_IMOVEIS_NAO_USO_STATUS]
                WHERE [NR_CONTRATO] = :contrato
                ORDER BY [DT_REFERENCIA] DESC
            """)
            result_avaliacao = db.session.execute(sql_avaliacao, {'contrato': nu_contrato}).fetchone()

            vr_avaliacao = float(result_avaliacao[0]) if result_avaliacao and result_avaliacao[0] else None
            dt_laudo = result_avaliacao[1] if result_avaliacao and result_avaliacao[1] else None

            # ===== BUSCAR TODOS OS CONTRATOS ANS (TABELA DE REFERÊNCIA) =====
            contratos_ans = PenalidadeANS.query.order_by(
                PenalidadeANS.INI_VIGENCIA.asc()
            ).all()

            if not contratos_ans:
                flash('Nenhum contrato ANS cadastrado na tabela de referência.', 'warning')
                return redirect(url_for('sumov.penalidade_ans_consultar'))

            # ===== CALCULAR PENALIDADES PARA CADA CONTRATO ANS =====
            penalidades_calculadas = []
            total_penalidades = 0

            # ===== VERIFICAR TIPO DE DESPESA PARA CONTRATO 03/2014 =====
            prazo_contrato_03_2014 = obter_prazo_contrato_03_2014(nu_contrato)

            for contrato_ans in contratos_ans:
                # ===== DETERMINAR DATA LIMITE DE PAGAMENTO =====

                if contrato_ans.NU_CONTRATO == 'Contrato s/nº':
                    # Contrato s/nº: entrada estoque + 120 dias
                    data_limite_pagamento = dt_entrada_estoque + timedelta(days=120)

                elif contrato_ans.NU_CONTRATO == 'Contrato 03/2014':
                    # REGRA ESPECIAL: Se entrada < 13/03/2014, data limite é 31/12/2015
                    if dt_entrada_estoque < contrato_ans.INI_VIGENCIA:
                        # Imóvel entrou ANTES do contrato: prazo até 31/12/2015
                        data_limite_pagamento = datetime(2015, 12, 31).date()
                        print(
                            f"[INFO] Contrato 03/2014 - Imóvel em estoque antes do contrato: Data limite fixa 31/12/2015")
                    else:
                        # Imóvel entrou DEPOIS do contrato: usa prazo normal (120 IPTU ou 150 Condomínio)
                        data_limite_pagamento = dt_entrada_estoque + timedelta(days=prazo_contrato_03_2014)
                        print(
                            f"[INFO] Contrato 03/2014 - Imóvel em estoque depois do contrato: Prazo {prazo_contrato_03_2014} dias")

                elif contrato_ans.NU_CONTRATO == 'Contrato 13/2019':
                    # Contrato 13/2019: se entrada < início contrato, conta da assinatura
                    if dt_entrada_estoque < contrato_ans.INI_VIGENCIA:
                        data_limite_pagamento = contrato_ans.INI_VIGENCIA + timedelta(days=120)
                        print(
                            f"[INFO] Contrato 13/2019 - Imóvel em estoque antes do contrato: Prazo conta da assinatura")
                    else:
                        data_limite_pagamento = dt_entrada_estoque + timedelta(days=120)

                else:
                    # Outros contratos: usa prazo da tabela
                    data_limite_pagamento = dt_entrada_estoque + timedelta(days=contrato_ans.PRAZO_DIAS)

                # ===== CALCULAR DATA DE INÍCIO DA PENALIDADE =====

                if contrato_ans.NU_CONTRATO == 'Contrato s/nº':
                    # Contrato s/nº: só pode penalizar após 18 meses da assinatura
                    data_minima_penalizacao = contrato_ans.INI_VIGENCIA + relativedelta(months=18)
                    # Penalidade começa no dia seguinte ao maior entre data_limite e data_minima
                    data_inicio_penalidade = max(data_limite_pagamento, data_minima_penalizacao) + timedelta(days=1)
                    print(
                        f"[DEBUG] Contrato s/nº - Data mínima: {data_minima_penalizacao}, Data limite: {data_limite_pagamento}, Início penalidade: {data_inicio_penalidade}")
                else:
                    # Penalidade começa no dia seguinte ao vencimento do prazo
                    # E não pode começar antes do início da vigência do contrato
                    data_inicio_penalidade = max(data_limite_pagamento + timedelta(days=1), contrato_ans.INI_VIGENCIA)
                    print(
                        f"[DEBUG] {contrato_ans.NU_CONTRATO} - Data limite: {data_limite_pagamento}, Início penalidade: {data_inicio_penalidade}")

                # ===== DATA DE FIM DA PENALIDADE =====
                data_fim_penalidade = contrato_ans.FIM_VIGENCIA

                # ===== VERIFICAR SE HÁ PENALIDADE =====
                if data_inicio_penalidade >= data_fim_penalidade:
                    penalidades_calculadas.append({
                        'nome_contrato': contrato_ans.NU_CONTRATO,
                        'dt_inicio_contrato': contrato_ans.INI_VIGENCIA,
                        'dt_fim_contrato': contrato_ans.FIM_VIGENCIA,
                        'qtd_meses_atraso': 0,
                        'vr_unitario_tarifa': float(contrato_ans.VR_TARIFA),
                        'valor_penalidade': 0
                    })
                    continue

                # ===== CALCULAR QUANTIDADE DE MESES =====
                # USAR A MESMA LÓGICA DO ACCESS: Calcular dias e dividir por 30
                # Fórmula Access: IIf([Atr03]/30>=0.33 And [Atr03]/30<1,1,Int([Atr03]/30))

                # Calcular quantidade de DIAS entre as datas
                dias_atraso = (data_fim_penalidade - data_inicio_penalidade).days

                # Aplicar a fórmula do Access
                meses_calculado = dias_atraso / 30.0

                if meses_calculado >= 0.33 and meses_calculado < 1:
                    qtd_meses_atraso = 1
                else:
                    qtd_meses_atraso = int(dias_atraso / 30)

                print(
                    f"[DEBUG] {contrato_ans.NU_CONTRATO} - De {data_inicio_penalidade} até {data_fim_penalidade} = {dias_atraso} dias, {meses_calculado:.2f} meses calculado, {qtd_meses_atraso} meses final")

                # Calcular valor da penalidade
                valor_penalidade = float(contrato_ans.VR_TARIFA) * qtd_meses_atraso

                # Adicionar à lista
                penalidades_calculadas.append({
                    'nome_contrato': contrato_ans.NU_CONTRATO,
                    'dt_inicio_contrato': contrato_ans.INI_VIGENCIA,
                    'dt_fim_contrato': contrato_ans.FIM_VIGENCIA,
                    'qtd_meses_atraso': qtd_meses_atraso,
                    'vr_unitario_tarifa': float(contrato_ans.VR_TARIFA),
                    'valor_penalidade': valor_penalidade
                })

                total_penalidades += valor_penalidade

            # Renderizar página com resultados
            return render_template('sumov/penalidade_ans/consultar.html',
                                   nu_contrato=nu_contrato,
                                   dt_entrada_estoque=dt_entrada_estoque,
                                   vr_venda=vr_venda,
                                   dt_venda=dt_venda,
                                   vr_avaliacao=vr_avaliacao,
                                   dt_laudo=dt_laudo,
                                   penalidades=penalidades_calculadas,
                                   total_penalidades=total_penalidades,
                                   exibir_resultados=True)

        except Exception as e:
            flash(f'Erro ao calcular penalidades: {str(e)}', 'danger')
            import traceback
            traceback.print_exc()
            return redirect(url_for('sumov.penalidade_ans_consultar'))

    # GET - Exibir formulário de pesquisa
    return render_template('sumov/penalidade_ans/consultar.html',
                           exibir_resultados=False)


@sumov_bp.route('/penalidade-ans/visualizar/<nu_contrato>')
@login_required
def penalidade_ans_visualizar(nu_contrato):
    """
    Visualiza todas as penalidades de um contrato específico
    """
    try:
        from app.models.penalidade_ans import PenalidadeANS

        # Buscar penalidades do contrato
        penalidades = PenalidadeANS.query.filter_by(
            NU_CONTRATO=nu_contrato
        ).order_by(
            PenalidadeANS.INI_VIGENCIA.desc()
        ).all()

        if not penalidades:
            flash('Nenhuma penalidade encontrada para este contrato.', 'warning')
            return redirect(url_for('sumov.penalidade_ans_index'))

        return render_template('sumov/penalidade_ans/visualizar.html',
                               nu_contrato=nu_contrato,
                               penalidades=penalidades)

    except Exception as e:
        flash(f'Erro ao carregar penalidades: {str(e)}', 'danger')
        import traceback
        traceback.print_exc()
        return redirect(url_for('sumov.penalidade_ans_index'))


@sumov_bp.route('/penalidade-ans/excluir-deliberacao', methods=['POST'])
@login_required
def penalidade_ans_excluir_deliberacao():
    """
    Exclui uma deliberação ANS salva
    """
    try:
        from app.models.deliberacao_ans import DeliberacaoANS
        from app.utils.audit import registrar_log

        data = request.get_json()
        nu_contrato = data.get('nu_contrato')

        if not nu_contrato:
            return jsonify({'success': False, 'message': 'Contrato não informado'}), 400

        # Buscar deliberação
        deliberacao = DeliberacaoANS.buscar_por_contrato(nu_contrato)

        if not deliberacao:
            return jsonify({'success': False, 'message': 'Deliberação não encontrada'}), 404

        # Excluir
        try:
            db.session.delete(deliberacao)
            db.session.commit()

            # Registrar log
            registrar_log(
                acao='excluir',
                entidade='deliberacao_ans',
                entidade_id=nu_contrato,
                descricao=f'Exclusão de Deliberação ANS - Contrato {nu_contrato}',
                dados_antigos={
                    'nu_contrato': nu_contrato,
                    'vr_ans': str(deliberacao.VR_ANS)
                }
            )

            return jsonify({'success': True, 'message': 'Deliberação excluída com sucesso!'})

        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'Erro ao excluir: {str(e)}'}), 500

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500


@sumov_bp.route('/penalidade-ans/salvar-deliberacao', methods=['POST'])
@login_required
def penalidade_ans_salvar_deliberacao():
    """
    Salva a deliberação de penalidades ANS calculadas
    """
    try:
        from app.models.deliberacao_ans import DeliberacaoANS
        from app.utils.audit import registrar_log
        from decimal import Decimal

        data = request.get_json()

        nu_contrato = data.get('nu_contrato', '').strip()
        dt_estoque_str = data.get('dt_estoque', '').strip()
        vr_ans = data.get('vr_ans', 0)

        # Dados dos 3 contratos ANS
        penalidades = data.get('penalidades', [])

        if not nu_contrato or not dt_estoque_str:
            return jsonify({'success': False, 'message': 'Dados incompletos'}), 400

        # Converter data
        dt_estoque = datetime.strptime(dt_estoque_str, '%Y-%m-%d').date()

        # Organizar dados por contrato
        dados_contratos = {}
        for p in penalidades:
            nome = p['nome_contrato']
            dados_contratos[nome] = {
                'qtd_meses': p['qtd_meses_atraso'],
                'vr_penalidade': p['valor_penalidade']
            }

        # Criar deliberação
        deliberacao = DeliberacaoANS(
            NU_CONTRATO=nu_contrato,
            DT_ESTOQUE=dt_estoque,
            VR_ANS=Decimal(str(vr_ans)),

            # Contrato s/nº
            QTD_MESES_SN=dados_contratos.get('Contrato s/nº', {}).get('qtd_meses'),
            VR_PENALIDADE_SN=Decimal(str(dados_contratos.get('Contrato s/nº', {}).get('vr_penalidade', 0))),

            # Contrato 03/2014
            QTD_MESES_03_2014=dados_contratos.get('Contrato 03/2014', {}).get('qtd_meses'),
            VR_PENALIDADE_03_2014=Decimal(str(dados_contratos.get('Contrato 03/2014', {}).get('vr_penalidade', 0))),

            # Contrato 13/2019
            QTD_MESES_13_2019=dados_contratos.get('Contrato 13/2019', {}).get('qtd_meses'),
            VR_PENALIDADE_13_2019=Decimal(str(dados_contratos.get('Contrato 13/2019', {}).get('vr_penalidade', 0)))
        )

        if deliberacao.salvar():
            # Registrar log
            registrar_log(
                acao='criar',
                entidade='deliberacao_ans',
                entidade_id=nu_contrato,
                descricao=f'Deliberação de Penalidades ANS salva - Contrato {nu_contrato}',
                dados_novos={
                    'nu_contrato': nu_contrato,
                    'vr_ans': str(vr_ans)
                }
            )

            return jsonify({'success': True, 'message': 'Deliberação salva com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao salvar no banco de dados'}), 500

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500


# =====================================================
# ANÁLISE DE OCORRÊNCIAS DE FATURAMENTO
# =====================================================

@sumov_bp.route('/faturamento')
@login_required
def faturamento_index():
    """Dashboard principal de Faturamento"""
    return render_template('sumov/faturamento/index.html')


@sumov_bp.route('/faturamento/analise-ocorrencias')
@login_required
def analise_ocorrencias():
    """Lista de ocorrências para análise - agrupadas por STATUS"""
    from app.models.ocorrencias_faturamento import OcorrenciasFaturamento
    from sqlalchemy import text

    # Buscar ocorrências agrupadas por STATUS
    ocorrencias_por_status = OcorrenciasFaturamento.listar_por_status()

    # Buscar ocorrências analisadas (mantém como está)
    analisadas = OcorrenciasFaturamento.listar_analisadas()

    # Calcular total de ocorrências sem análise
    total_sem_analise = sum(len(ocorrencias) for ocorrencias in ocorrencias_por_status.values())

    # Buscar responsáveis distintos para o filtro
    sql_responsaveis = text("""
        SELECT DISTINCT RESPONSAVEL
        FROM BDG.MOV_TB039_SMART_OCORRENCIAS_ANALISAR
        WHERE RESPONSAVEL IS NOT NULL
        ORDER BY RESPONSAVEL
    """)
    resultado_responsaveis = db.session.execute(sql_responsaveis).fetchall()
    responsaveis = [row[0] for row in resultado_responsaveis]

    return render_template('sumov/faturamento/analise_ocorrencias.html',
                           ocorrencias_por_status=ocorrencias_por_status,
                           analisadas=analisadas,
                           total_sem_analise=total_sem_analise,
                           responsaveis=responsaveis)


@sumov_bp.route('/faturamento/analise-ocorrencias/exportar-excel', methods=['POST'])
@login_required
def exportar_excel_ocorrencias():
    """Exporta ocorrências selecionadas para Excel formatado"""
    from app.models.ocorrencias_faturamento import OcorrenciasFaturamento
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    from io import BytesIO
    from datetime import datetime
    from app.utils.audit import registrar_log

    try:
        # Capturar abas selecionadas
        abas_selecionadas = request.form.getlist('abas_selecionadas')

        if not abas_selecionadas:
            flash('Nenhuma aba foi selecionada para exportação.', 'warning')
            return redirect(url_for('sumov.analise_ocorrencias'))

        # Buscar dados
        ocorrencias_por_status = OcorrenciasFaturamento.listar_por_status()
        analisadas = OcorrenciasFaturamento.listar_analisadas()

        # Criar workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover sheet padrão

        # Estilos para formatação
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Função auxiliar para criar sheet formatada
        def criar_sheet(wb, nome_sheet, dados, incluir_houve_faturamento=False):
            ws = wb.create_sheet(title=nome_sheet[:31])  # Limite de 31 caracteres

            # Definir cabeçalhos
            if incluir_houve_faturamento:
                headers = ['Contrato', 'Ocorrência', 'Justificativa', 'Data Justificativa',
                           'Houve Faturamento?', 'Mês/Ano', 'Status', 'Item Serviço', 'OBS',
                           'Estado', 'Responsável']
            else:
                headers = ['Contrato', 'Ocorrência', 'Justificativa', 'Data Justificativa',
                           'Status', 'Item Serviço', 'OBS', 'Estado', 'Responsável']

            # Escrever cabeçalhos
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            # Escrever dados
            for row_num, ocorrencia in enumerate(dados, 2):
                ws.cell(row=row_num, column=1, value=int(ocorrencia.NR_CONTRATO))
                ws.cell(row=row_num, column=2, value=ocorrencia.nrOcorrencia)
                ws.cell(row=row_num, column=3, value=ocorrencia.dsJustificativa or '-')

                # Data Justificativa
                if ocorrencia.DT_JUSTIFICATIVA:
                    ws.cell(row=row_num, column=4, value=ocorrencia.DT_JUSTIFICATIVA.strftime('%d/%m/%Y'))
                else:
                    ws.cell(row=row_num, column=4, value='-')

                col_offset = 5

                # Se incluir houve faturamento (aba Analisadas)
                if incluir_houve_faturamento:
                    if hasattr(ocorrencia, 'ID_FATURAMENTO'):
                        if ocorrencia.ID_FATURAMENTO == 1:
                            ws.cell(row=row_num, column=5, value='Sim')
                        elif ocorrencia.ID_FATURAMENTO == 0:
                            ws.cell(row=row_num, column=5, value='Não')
                        else:
                            ws.cell(row=row_num, column=5, value='-')
                    else:
                        ws.cell(row=row_num, column=5, value='-')

                    ws.cell(row=row_num, column=6, value=ocorrencia.formatar_mes_ano())
                    col_offset = 7

                # Colunas comuns
                ws.cell(row=row_num, column=col_offset, value=ocorrencia.STATUS or 'Sem Status')
                ws.cell(row=row_num, column=col_offset + 1, value=ocorrencia.ITEM_SERVICO or '-')
                ws.cell(row=row_num, column=col_offset + 2, value=ocorrencia.OBS or '-')
                ws.cell(row=row_num, column=col_offset + 3, value=ocorrencia.DSC_ESTADO or '-')
                ws.cell(row=row_num, column=col_offset + 4, value=ocorrencia.RESPONSAVEL or '-')

                # Aplicar bordas e alinhamento
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = border
                    if col_num in [1, 2, 4, col_offset, col_offset + 3, col_offset + 4]:  # Colunas centralizadas
                        cell.alignment = center_alignment

            # Ajustar largura das colunas
            if incluir_houve_faturamento:
                column_widths = [15, 12, 50, 18, 18, 15, 15, 20, 20, 20, 25]
            else:
                column_widths = [15, 12, 50, 18, 15, 20, 20, 20, 25]

            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Congelar primeira linha
            ws.freeze_panes = 'A2'

        # Criar sheets para cada aba selecionada
        for aba_nome in abas_selecionadas:
            if aba_nome == 'Analisadas':
                criar_sheet(wb, 'Analisadas', analisadas, incluir_houve_faturamento=True)
            else:
                # Buscar dados do status específico
                if aba_nome in ocorrencias_por_status:
                    dados = ocorrencias_por_status[aba_nome]
                    criar_sheet(wb, aba_nome, dados, incluir_houve_faturamento=False)

        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Ocorrencias_Faturamento_{timestamp}.xlsx'

        # Registrar log
        registrar_log(
            acao='exportar',
            entidade='ocorrencias_faturamento_excel',
            entidade_id='batch',
            descricao=f'Exportação de ocorrências para Excel - Abas: {", ".join(abas_selecionadas)}'
        )

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Erro ao exportar para Excel: {str(e)}', 'danger')
        return redirect(url_for('sumov.analise_ocorrencias'))


@sumov_bp.route('/faturamento/analise-ocorrencias/analisar/<nr_contrato>/<int:nr_ocorrencia>/<identificador>',
                methods=['GET', 'POST'])
@login_required
def analisar_ocorrencia(nr_contrato, nr_ocorrencia, identificador):
    """Formulário para analisar uma ocorrência específica"""
    from app.models.ocorrencias_faturamento import OcorrenciasFaturamento
    from sqlalchemy import text
    from app.utils.audit import registrar_log

    # Buscar ocorrência específica pelo identificador
    ocorrencia = OcorrenciasFaturamento.buscar_por_identificador(nr_contrato, nr_ocorrencia, identificador)

    if not ocorrencia:
        flash('Ocorrência não encontrada.', 'danger')
        return redirect(url_for('sumov.analise_ocorrencias'))

    if request.method == 'POST':
        try:
            # Capturar dados do formulário
            houve_faturamento = request.form.get('houve_faturamento')
            mes_ano = request.form.get('mes_ano', '').strip()

            # Validações
            if not houve_faturamento:
                flash('Por favor, informe se houve faturamento.', 'warning')
                return redirect(request.url)

            # Determinar ID_FATURAMENTO e MES_ANO_FATURAMENTO
            if houve_faturamento == 'sim':
                # Se SIM: ID_FATURAMENTO = 1 e MES_ANO_FATURAMENTO preenchido
                if not mes_ano:
                    flash('Por favor, informe o mês/ano do faturamento quando houver faturamento.', 'warning')
                    return redirect(request.url)

                id_faturamento = 1

                # Converter mes_ano de MM/YYYY para YYYYMM (formato americano)
                # Exemplo: 05/2025 → 202505
                try:
                    mes, ano = mes_ano.split('/')
                    mes_ano_int = int(f"{ano}{mes}")  # ano primeiro, depois mês
                except:
                    flash('Formato de mês/ano inválido. Use MM/AAAA (Ex: 05/2025)', 'danger')
                    return redirect(request.url)

            else:  # houve_faturamento == 'nao'
                # Se NÃO: ID_FATURAMENTO = 0 e MES_ANO_FATURAMENTO = NULL
                id_faturamento = 0
                mes_ano_int = None

            # ===================================================================
            # CORREÇÃO: Usar SQL bruto puro - tabela não tem chave primária
            # Identificar registro por NR_CONTRATO, nrOcorrencia e dsJustificativa
            # ===================================================================

            # Pegar o valor da justificativa (pode ser NULL)
            justificativa_valor = ocorrencia.dsJustificativa

            # Construir UPDATE com tratamento adequado de NULL
            if justificativa_valor is None:
                # Se justificativa é NULL, usar IS NULL no WHERE
                sql_update = text("""
                    UPDATE BDG.MOV_TB039_SMART_OCORRENCIAS_ANALISAR
                    SET ID_FATURAMENTO = :id_faturamento,
                        MES_ANO_FATURAMENTO = :mes_ano_faturamento
                    WHERE NR_CONTRATO = :nr_contrato
                      AND nrOcorrencia = :nr_ocorrencia
                      AND dsJustificativa IS NULL
                """)

                db.session.execute(sql_update, {
                    'id_faturamento': id_faturamento,
                    'mes_ano_faturamento': mes_ano_int,
                    'nr_contrato': nr_contrato,
                    'nr_ocorrencia': nr_ocorrencia
                })
            else:
                # Se justificativa tem valor, usar comparação normal
                sql_update = text("""
                    UPDATE BDG.MOV_TB039_SMART_OCORRENCIAS_ANALISAR
                    SET ID_FATURAMENTO = :id_faturamento,
                        MES_ANO_FATURAMENTO = :mes_ano_faturamento
                    WHERE NR_CONTRATO = :nr_contrato
                      AND nrOcorrencia = :nr_ocorrencia
                      AND dsJustificativa = :ds_justificativa
                """)

                db.session.execute(sql_update, {
                    'id_faturamento': id_faturamento,
                    'mes_ano_faturamento': mes_ano_int,
                    'nr_contrato': nr_contrato,
                    'nr_ocorrencia': nr_ocorrencia,
                    'ds_justificativa': justificativa_valor
                })

            db.session.commit()

            # Registrar log
            if houve_faturamento == 'sim':
                descricao = f'Análise de ocorrência: Contrato {nr_contrato}, Ocorrência {nr_ocorrencia}, Faturamento: Sim, Mês/Ano: {mes_ano}'
            else:
                descricao = f'Análise de ocorrência: Contrato {nr_contrato}, Ocorrência {nr_ocorrencia}, Faturamento: Não'

            registrar_log(
                acao='editar',
                entidade='ocorrencias_faturamento',
                entidade_id=f"{nr_contrato}-{nr_ocorrencia}",
                descricao=descricao
            )

            flash('Ocorrência analisada com sucesso!', 'success')
            return redirect(url_for('sumov.analise_ocorrencias'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao analisar ocorrência: {str(e)}', 'danger')
            import traceback
            traceback.print_exc()
            return redirect(request.url)

    # GET - Exibir formulário
    return render_template('sumov/faturamento/analisar_form.html',
                           ocorrencia=ocorrencia)


@sumov_bp.route('/faturamento/analise-ocorrencias/editar/<nr_contrato>/<int:nr_ocorrencia>/<identificador>',
                methods=['GET', 'POST'])
@login_required
def editar_ocorrencia(nr_contrato, nr_ocorrencia, identificador):
    """Formulário para editar uma ocorrência já analisada"""
    from app.models.ocorrencias_faturamento import OcorrenciasFaturamento
    from sqlalchemy import text

    # Buscar ocorrência específica pelo identificador
    ocorrencia = OcorrenciasFaturamento.buscar_por_identificador(nr_contrato, nr_ocorrencia, identificador)

    if not ocorrencia:
        flash('Ocorrência não encontrada.', 'danger')
        return redirect(url_for('sumov.analise_ocorrencias'))

    if request.method == 'POST':
        try:
            # Capturar dados do formulário
            houve_faturamento = request.form.get('houve_faturamento')
            mes_ano = request.form.get('mes_ano', '').strip()

            # Validações
            if not houve_faturamento:
                flash('Por favor, informe se houve faturamento.', 'warning')
                return redirect(request.url)

            # Determinar ID_FATURAMENTO e MES_ANO_FATURAMENTO
            if houve_faturamento == 'sim':
                # Se SIM: ID_FATURAMENTO = 1 e MES_ANO_FATURAMENTO preenchido
                if not mes_ano:
                    flash('Por favor, informe o mês/ano do faturamento quando houver faturamento.', 'warning')
                    return redirect(request.url)

                id_faturamento = 1

                # Converter mes_ano de MM/YYYY para YYYYMM (formato americano)
                try:
                    mes, ano = mes_ano.split('/')
                    mes_ano_int = int(f"{ano}{mes}")
                except:
                    flash('Formato de mês/ano inválido. Use MM/AAAA (Ex: 05/2025)', 'danger')
                    return redirect(request.url)

            else:  # houve_faturamento == 'nao'
                # Se NÃO: ID_FATURAMENTO = 0 e MES_ANO_FATURAMENTO = NULL
                id_faturamento = 0
                mes_ano_int = None

            # Atualizar usando ORM diretamente
            ocorrencia.ID_FATURAMENTO = id_faturamento
            ocorrencia.MES_ANO_FATURAMENTO = mes_ano_int
            db.session.commit()

            # Registrar log
            if houve_faturamento == 'sim':
                descricao = f'Edição de ocorrência: Contrato {nr_contrato}, Ocorrência {nr_ocorrencia}, Faturamento: Sim, Mês/Ano: {mes_ano}'
            else:
                descricao = f'Edição de ocorrência: Contrato {nr_contrato}, Ocorrência {nr_ocorrencia}, Faturamento: Não'

            registrar_log(
                acao='editar',
                entidade='ocorrencias_faturamento',
                entidade_id=f"{nr_contrato}-{nr_ocorrencia}",
                descricao=descricao
            )

            flash('Ocorrência editada com sucesso!', 'success')
            return redirect(url_for('sumov.analise_ocorrencias'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao editar ocorrência: {str(e)}', 'danger')
            return redirect(request.url)

    # GET - Exibir formulário com dados preenchidos
    return render_template('sumov/faturamento/editar_form.html',
                           ocorrencia=ocorrencia)


@sumov_bp.route('/faturamento/analise-ocorrencias/atualizar', methods=['POST'])
@login_required
def atualizar_ocorrencias_faturamento():
    """Executa o UPDATE para sincronizar dados analisados com a tabela de faturamento"""
    from app.models.ocorrencias_faturamento import OcorrenciasFaturamento

    try:
        # Contar quantos registros serão sincronizados
        sql_count = text("""
            SELECT COUNT(*)
            FROM BDG.MOV_TB039_SMART_OCORRENCIAS_ANALISAR A
            WHERE A.ID_FATURAMENTO IS NOT NULL
              AND NOT EXISTS (
                  SELECT 1 
                  FROM BDDASHBOARDBI.[BDG].[MOV_TB034_SMART_FATURAMENTO] B
                  WHERE B.NR_CONTRATO = A.NR_CONTRATO
                    AND B.nrOcorrencia = A.nrOcorrencia
                    AND B.ID_FATURAMENTO = A.ID_FATURAMENTO
                    AND (
                        (B.MES_ANO_FATURAMENTO = A.MES_ANO_FATURAMENTO)
                        OR (B.MES_ANO_FATURAMENTO IS NULL AND A.MES_ANO_FATURAMENTO IS NULL)
                    )
              )
        """)

        total_para_sincronizar = db.session.execute(sql_count).scalar()

        if total_para_sincronizar == 0:
            flash('Não há novos registros para sincronizar. Todos os registros analisados já foram processados.',
                  'info')
            return redirect(url_for('sumov.analise_ocorrencias'))

        # Executar UPDATE em 2 etapas
        linhas_afetadas = OcorrenciasFaturamento.atualizar_faturamento_smart()

        # Registrar log
        registrar_log(
            acao='atualizar',
            entidade='faturamento_smart',
            entidade_id='batch',
            descricao=f'Atualização em lote de faturamento SMART - {linhas_afetadas} registros sincronizados (Etapa 1: ID_FATURAMENTO=1, Etapa 2: ID_FATURAMENTO=0)'
        )

        if linhas_afetadas > 0:
            flash(
                f'Sincronização concluída com sucesso! {linhas_afetadas} registro(s) atualizado(s) na tabela de faturamento. Os registros foram removidos da lista de Analisadas.',
                'success')
        else:
            flash('Nenhum registro foi atualizado. Pode haver um problema com a sincronização.', 'warning')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar ocorrências: {str(e)}', 'danger')

    return redirect(url_for('sumov.analise_ocorrencias'))


@sumov_bp.route('/faturamento/analise-ocorrencias/teste-duplicidade')
@login_required
def teste_duplicidade_faturamento():
    """Testa duplicidade de contratos na tabela de faturamento - apenas ID_FATURAMENTO = 1"""
    from sqlalchemy import text

    try:
        # Query para buscar contratos duplicados - APENAS ID_FATURAMENTO = 1
        sql = text("""
            SELECT 
                [ID], 
                [DT_REFERENCIA], 
                [fkContratoSISCTR], 
                [nrOcorrencia], 
                [NR_CONTRATO],
                [itemServico], 
                [NO_DESTINO], 
                [DT_ULTIMO_TRAMITE], 
                [NO_DEVEDOR],
                [DT_JUSTIF], 
                [JUST_APRESENT], 
                [ANO_MES_ABERTURA], 
                [ANO_MES_JUSTIF],
                [ID_FATURAMENTO], 
                [MES_ANO_FATURAMENTO], 
                [OBS]
            FROM BDDASHBOARDBI.[BDG].[MOV_TB034_SMART_FATURAMENTO]
            WHERE ID_FATURAMENTO = 1
              AND NR_CONTRATO IN (
                SELECT NR_CONTRATO 
                FROM BDDASHBOARDBI.[BDG].[MOV_TB034_SMART_FATURAMENTO] 
                WHERE ID_FATURAMENTO = 1
                  AND MES_ANO_FATURAMENTO IS NOT NULL 
                GROUP BY NR_CONTRATO 
                HAVING COUNT(*) > 1
            )
            ORDER BY NR_CONTRATO, nrOcorrencia
        """)

        resultado = db.session.execute(sql).fetchall()

        # Agrupar por contrato para facilitar visualização
        contratos_duplicados = {}
        for row in resultado:
            # NR_CONTRATO agora está na posição 4 (índice 4) - garantido pela query explícita
            nr_contrato = int(row[4]) if row[4] is not None else 0

            if nr_contrato not in contratos_duplicados:
                contratos_duplicados[nr_contrato] = []
            contratos_duplicados[nr_contrato].append(row)

        return render_template('sumov/faturamento/teste_duplicidade.html',
                               contratos_duplicados=contratos_duplicados,
                               total_duplicados=len(contratos_duplicados))

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Erro ao executar teste de duplicidade: {str(e)}', 'danger')
        return redirect(url_for('sumov.analise_ocorrencias'))


@sumov_bp.route('/faturamento/teste-duplicidade/editar/<nr_contrato>/<int:nr_ocorrencia>',
                methods=['GET', 'POST'])
@login_required
def editar_faturamento_duplicidade(nr_contrato, nr_ocorrencia):
    """Formulário para editar ID_FATURAMENTO e MES_ANO_FATURAMENTO diretamente na tabela SMART_FATURAMENTO"""
    from sqlalchemy import text
    from app.utils.audit import registrar_log

    try:
        # Buscar o registro específico na tabela MOV_TB034_SMART_FATURAMENTO
        sql_buscar = text("""
            SELECT TOP 1 
                [ID], [DT_REFERENCIA], [fkContratoSISCTR], [nrOcorrencia], [NR_CONTRATO],
                [itemServico], [NO_DESTINO], [DT_ULTIMO_TRAMITE], [NO_DEVEDOR],
                [DT_JUSTIF], [JUST_APRESENT], [ANO_MES_ABERTURA], [ANO_MES_JUSTIF],
                [ID_FATURAMENTO], [MES_ANO_FATURAMENTO], [OBS]
            FROM BDDASHBOARDBI.[BDG].[MOV_TB034_SMART_FATURAMENTO]
            WHERE NR_CONTRATO = :nr_contrato 
              AND nrOcorrencia = :nr_ocorrencia
            ORDER BY ID DESC
        """)

        resultado = db.session.execute(sql_buscar, {
            'nr_contrato': nr_contrato,
            'nr_ocorrencia': nr_ocorrencia
        }).fetchone()

        if not resultado:
            flash('Registro não encontrado na tabela de faturamento.', 'danger')
            return redirect(url_for('sumov.teste_duplicidade_faturamento'))

        # Converter resultado em dicionário
        registro = {
            'ID': resultado[0],
            'DT_REFERENCIA': resultado[1],
            'fkContratoSISCTR': resultado[2],
            'nrOcorrencia': resultado[3],
            'NR_CONTRATO': resultado[4],
            'itemServico': resultado[5],
            'NO_DESTINO': resultado[6],
            'DT_ULTIMO_TRAMITE': resultado[7],
            'NO_DEVEDOR': resultado[8],
            'DT_JUSTIF': resultado[9],
            'JUST_APRESENT': resultado[10],
            'ANO_MES_ABERTURA': resultado[11],
            'ANO_MES_JUSTIF': resultado[12],
            'ID_FATURAMENTO': resultado[13],
            'MES_ANO_FATURAMENTO': resultado[14],
            'OBS': resultado[15]
        }

        if request.method == 'POST':
            # Capturar dados do formulário
            houve_faturamento = request.form.get('houve_faturamento')
            mes_ano = request.form.get('mes_ano', '').strip()

            # Validações
            if not houve_faturamento:
                flash('Por favor, informe se houve faturamento.', 'warning')
                return redirect(request.url)

            # Determinar ID_FATURAMENTO e MES_ANO_FATURAMENTO
            if houve_faturamento == 'sim':
                # Se SIM: ID_FATURAMENTO = 1 e MES_ANO_FATURAMENTO preenchido
                if not mes_ano:
                    flash('Por favor, informe o mês/ano do faturamento quando houver faturamento.', 'warning')
                    return redirect(request.url)

                id_faturamento = 1

                # Converter mes_ano de MM/YYYY para YYYYMM (formato americano)
                try:
                    mes, ano = mes_ano.split('/')
                    mes_ano_int = int(f"{ano}{mes}")
                except:
                    flash('Formato de mês/ano inválido. Use MM/AAAA (Ex: 05/2025)', 'danger')
                    return redirect(request.url)

            else:  # houve_faturamento == 'nao'
                # Se NÃO: ID_FATURAMENTO = 0 e MES_ANO_FATURAMENTO = NULL
                id_faturamento = 0
                mes_ano_int = None

            # Atualizar o registro na tabela MOV_TB034_SMART_FATURAMENTO
            sql_atualizar = text("""
                UPDATE BDDASHBOARDBI.[BDG].[MOV_TB034_SMART_FATURAMENTO]
                SET [ID_FATURAMENTO] = :id_faturamento,
                    [MES_ANO_FATURAMENTO] = :mes_ano_faturamento,
                    [OBS] = 'EDITADO VIA DUPLICIDADE'
                WHERE NR_CONTRATO = :nr_contrato 
                  AND nrOcorrencia = :nr_ocorrencia
            """)

            db.session.execute(sql_atualizar, {
                'id_faturamento': id_faturamento,
                'mes_ano_faturamento': mes_ano_int,
                'nr_contrato': nr_contrato,
                'nr_ocorrencia': nr_ocorrencia
            })
            db.session.commit()

            # Registrar log
            if houve_faturamento == 'sim':
                descricao = f'Edição via duplicidade: Contrato {nr_contrato}, Ocorrência {nr_ocorrencia}, Faturamento: Sim, Mês/Ano: {mes_ano}'
            else:
                descricao = f'Edição via duplicidade: Contrato {nr_contrato}, Ocorrência {nr_ocorrencia}, Faturamento: Não'

            registrar_log(
                acao='editar',
                entidade='faturamento_duplicidade',
                entidade_id=f"{nr_contrato}-{nr_ocorrencia}",
                descricao=descricao
            )

            flash('Registro de faturamento editado com sucesso!', 'success')
            return redirect(url_for('sumov.teste_duplicidade_faturamento'))

        # GET - Exibir formulário com dados preenchidos
        return render_template('sumov/faturamento/editar_duplicidade_form.html',
                               registro=registro)

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao editar registro de faturamento: {str(e)}', 'danger')
        return redirect(url_for('sumov.teste_duplicidade_faturamento'))


@sumov_bp.route('/faturamento/analise-ocorrencias/inserir-tabela-final', methods=['POST'])
@login_required
def inserir_tabela_final_faturamento():
    """Insere registros analisados na tabela final com VR_TARIFA = 247.48"""
    from sqlalchemy import text

    try:
        # Query para inserir na tabela final - Gerando ID sequencial
        sql = text("""
            -- Pegar o maior ID da tabela final
            DECLARE @MaxID INT
            SELECT @MaxID = ISNULL(MAX(ID), 0) FROM BDDASHBOARDBI.[BDG].[MOV_TB035_SMART_FATURAMENTO_FINAL]

            -- Inserir com IDs sequenciais a partir do maior + 1
            INSERT INTO BDDASHBOARDBI.[BDG].[MOV_TB035_SMART_FATURAMENTO_FINAL]
            ([ID], [DT_REFERENCIA], [fkContratoSISCTR], [nrOcorrencia], [NR_CONTRATO], 
             [itemServico], [NO_DESTINO], [DT_ULTIMO_TRAMITE], [NO_DEVEDOR], 
             [DT_JUSTIF], [JUST_APRESENT], [ANO_MES_ABERTURA], [ANO_MES_JUSTIF], 
             [ID_FATURAMENTO], [MES_ANO_FATURAMENTO], [VR_TARIFA], [OBS])
            SELECT 
                @MaxID + ROW_NUMBER() OVER (ORDER BY [ID]) AS [ID],
                [DT_REFERENCIA], [fkContratoSISCTR], [nrOcorrencia], [NR_CONTRATO], 
                [itemServico], [NO_DESTINO], [DT_ULTIMO_TRAMITE], [NO_DEVEDOR], 
                [DT_JUSTIF], [JUST_APRESENT], [ANO_MES_ABERTURA], [ANO_MES_JUSTIF], 
                [ID_FATURAMENTO], [MES_ANO_FATURAMENTO], 247.48 AS VR_TARIFA, [OBS]
            FROM BDDASHBOARDBI.[BDG].[MOV_TB034_SMART_FATURAMENTO]
            WHERE MES_ANO_FATURAMENTO IS NOT NULL
        """)

        result = db.session.execute(sql)
        db.session.commit()

        linhas_inseridas = result.rowcount

        # Registrar log
        registrar_log(
            acao='inserir',
            entidade='faturamento_final',
            entidade_id='batch',
            descricao=f'Inserção em lote na tabela final de faturamento - {linhas_inseridas} registros inseridos com VR_TARIFA = 247.48'
        )

        flash(f'Inserção concluída com sucesso! {linhas_inseridas} registro(s) inserido(s) na tabela final.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao inserir na tabela final: {str(e)}', 'danger')

    return redirect(url_for('sumov.analise_ocorrencias'))